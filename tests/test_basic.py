import json
import os
from pathlib import Path
from datetime import date

import pytest

from modules.text_normalizer import normalize_text


def _kunming_niu_runtime_config() -> dict:
    """返回昆明牛项目的运行时配置，供单元测试使用。
    避免测试依赖当前 app_config 的运行状态。
    """
    return {
        "accounts": {
            "银康01": {
                "baidu_name": "银康01",
                "baidu_names": ["银康01"],
                "excel_name": "银康01",
                "kst_ids": ["72828178"],
                "kst_names": ["银康01"],
                "aliases": ["银康01"],
            },
            "银康银屑02": {
                "baidu_name": "银康银屑02",
                "baidu_names": ["银康银屑02"],
                "excel_name": "银康银屑02",
                "kst_ids": ["72828179"],
                "kst_names": ["银康银屑02"],
                "aliases": ["银康银屑02"],
            },
            "银康03": {
                "baidu_name": "baidu-银康03",
                "baidu_names": ["baidu-银康03", "银康03"],
                "excel_name": "银康03",
                "kst_ids": ["81509165"],
                "kst_names": ["银康03"],
                "aliases": ["银康03", "baidu-银康03"],
            },
        },
        "kst": {
            "promotion_id_accounts": {
                "72828178": "银康01",
                "72828179": "银康银屑02",
                "81509165": "银康03",
            },
        },
    }


def _replace_xlsx_text(path: Path, member: str, old: str, new: str) -> None:
    from zipfile import ZIP_DEFLATED, ZipFile

    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with ZipFile(path, "r") as source, ZipFile(tmp_path, "w", ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == member:
                data = data.decode("utf-8").replace(old, new).encode("utf-8")
            target.writestr(item, data)
    tmp_path.replace(path)


def test_restore_sheet_filter_protection_metadata_keeps_original_protection_attrs(tmp_path):
    from zipfile import ZipFile

    from openpyxl import Workbook, load_workbook

    class Logger:
        def info(self, *args, **kwargs):
            pass

        def warning(self, *args, **kwargs):
            pass

    excel_path = tmp_path / "current.xlsx"
    backup_path = tmp_path / "backup.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A3"] = "日期"
    ws["B3"] = "时段"
    ws.auto_filter.ref = "A3:B10"
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = True
    wb.save(backup_path)
    wb.save(excel_path)

    with ZipFile(backup_path) as zf:
        original_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    original_node_start = original_xml.index("<sheetProtection")
    original_node_end = original_xml.index("/>", original_node_start) + 2
    original_node = original_xml[original_node_start:original_node_end]
    rich_node = (
        '<sheetProtection selectLockedCells="0" selectUnlockedCells="0" '
        'algorithmName="SHA-512" sheet="1" objects="1" insertRows="1" '
        'autoFilter="0" scenarios="0" formatColumns="0" sort="1" />'
    )
    dropped_node = '<sheetProtection sheet="1" formatColumns="0" autoFilter="0" objects="1"/>'
    _replace_xlsx_text(backup_path, "xl/worksheets/sheet1.xml", original_node, rich_node)
    _replace_xlsx_text(excel_path, "xl/worksheets/sheet1.xml", original_node, dropped_node)

    restored = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        ["时段数据"],
        Logger(),
    )

    assert restored is True
    with ZipFile(excel_path) as zf:
        restored_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert rich_node in restored_xml
    assert dropped_node not in restored_xml
    assert load_workbook(excel_path)["时段数据"]["A3"].value == "日期"


def test_restore_sheet_filter_protection_metadata_restores_original_auto_filter(tmp_path):
    from openpyxl import Workbook, load_workbook

    class Logger:
        def info(self, *args, **kwargs):
            pass

        def warning(self, *args, **kwargs):
            pass

    excel_path = tmp_path / "current.xlsx"
    backup_path = tmp_path / "backup.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A3"] = "日期"
    ws["B3"] = "时段"
    ws.auto_filter.ref = "A3:XDP1464"
    wb.save(backup_path)

    ws.auto_filter.ref = None
    wb.save(excel_path)

    restored = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        ["时段数据"],
        Logger(),
    )

    assert restored is True
    assert load_workbook(excel_path)["时段数据"].auto_filter.ref == "A3:XDP1464"


def test_restore_sheet_filter_protection_metadata_restores_all_sheets_when_unspecified(tmp_path):
    from openpyxl import Workbook, load_workbook

    class Logger:
        def info(self, *args, **kwargs):
            pass

        def warning(self, *args, **kwargs):
            pass

    excel_path = tmp_path / "current.xlsx"
    backup_path = tmp_path / "backup.xlsx"

    wb = Workbook()
    sheet_filters = {
        "时段数据": "A3:XDP1464",
        "百度": "A2:A370",
        "大夜数据": "A2:A370",
    }
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.create_sheet("大夜数据")
    for sheet_name, auto_filter_ref in sheet_filters.items():
        ws = wb[sheet_name]
        ws["A1"] = sheet_name
        ws.auto_filter.ref = auto_filter_ref
    wb.save(backup_path)

    for ws in wb.worksheets:
        ws.auto_filter.ref = None
    wb.save(excel_path)
    wb.close()

    restored = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        None,
        Logger(),
    )

    assert restored is True
    restored_wb = load_workbook(excel_path)
    assert {
        sheet_name: restored_wb[sheet_name].auto_filter.ref
        for sheet_name in sheet_filters
    } == sheet_filters
    restored_wb.close()


def test_read_back_values_uses_read_only_workbook(tmp_path, monkeypatch):
    from openpyxl import Workbook

    import modules.excel_writer as writer

    excel_path = tmp_path / "verify.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["B8"] = 12
    ws["D8"] = "完成"
    wb.save(excel_path)

    real_load_workbook = writer.load_workbook
    calls = []

    def recording_load_workbook(*args, **kwargs):
        calls.append(dict(kwargs))
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(writer, "load_workbook", recording_load_workbook)

    values = writer._read_back_values(excel_path, "时段数据", ["B8", "D8"])

    assert values == {"B8": 12, "D8": "完成"}
    assert calls == [{"data_only": False, "read_only": True}]

from modules.kst_export_parser import find_latest_kst_export, parse_kst_export_file, write_empty_kst_export_result
from modules.kst_daily_parser import classify_daily_dialog_by_tags, parse_kst_daily_file, write_empty_kst_daily_result
from modules.kst_parser import aggregate_kst_export_rows, classify_dialog_by_tags, has_visitor_dialog
from modules.run_pipeline import run_daily_pipeline, run_half_auto_pipeline
from modules.excel_inspector import (
    _build_account_ranges,
    _build_merged_bounds_map,
    _build_merged_value_map,
    _find_account_titles,
    _get_merged_value,
    _scan_non_empty_cells,
)
from modules.excel_writer import (
    _find_target_row,
    _normalize_period_for_excel,
    _restore_sheet_filter_protection_metadata,
    _validate_write_target,
)
from modules.daily_excel_inspector import inspect_daily_worksheet
from modules.data_merger import build_merged_daily_data, build_merged_hourly_data
from modules.excel_writer import write_merged_daily_data, write_merged_hourly_data
from modules.baidu_parser import _build_account_map, _build_parse_debug, _map_account, _parse_number, extract_baidu_rows_from_visible_text, parse_baidu_table
from modules.baidu_browser import _extract_selected_date_from_text, _write_debug_artifacts
from modules.baidu_detector import classify_baidu_page
from modules.baidu_overview import is_search_promotion_overview, overview_text_has_account_table, should_open_cas_login, validate_overview_ready
from modules.baidu_validator import validate_baidu_account_data
from modules.baidu_auto import build_baidu_auto_report_from_visible_text
from modules.baidu_auto import fetch_baidu_auto
from modules.baidu_report_api import fetch_baidu_api_probe
from modules.baidu_api_simulation import simulate_baidu_api_hourly
from modules.baidu_oauth_bundle import import_baidu_oauth_bundle
from modules.baidu_daily import build_baidu_daily_report_from_visible_text, default_daily_date
from modules.credential_manager import build_login_failure_message, load_project_credentials
from modules.browser_manager import BrowserLaunchError, CONNECT_EXISTING_HELP, cleanup_extra_tabs, connect_existing_chrome, get_browser_settings
from modules.chrome_debug import ensure_chrome_debug_ready, find_chrome_executable, is_chrome_debug_port_alive, start_debug_chrome
from modules.excel_engine import format_openpyxl_save_error
from modules.project_config import (
    build_runtime_config_from_project,
    get_account_alias_maps,
    get_current_project,
    get_excel_path,
    get_project_accounts,
    list_projects,
    reload_current_project,
    set_current_project,
    validate_project_config,
)
from modules.doctor import run_doctor
from modules.validators import validate_baidu_report, validate_merged_daily_data, validate_merged_hourly_data
from menu import MENU_TEXT, build_confirmation_lines, build_runtime_config, dispatch_menu_task
from tools.build_release import build_release, should_include_file
import inspect
from datetime import date
from pathlib import Path


def test_merged_cell_lookup_uses_prebuilt_map():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "账户标题"
    ws.merge_cells("A1:C1")

    merged_values = _build_merged_value_map(ws)

    assert _get_merged_value(ws, 1, 1, merged_values) == "账户标题"
    assert _get_merged_value(ws, 1, 2, merged_values) == "账户标题"
    assert _get_merged_value(ws, 1, 3, merged_values) == "账户标题"


def test_project_config_loads_default_project_and_lists_projects(tmp_path):
    app_dir = tmp_path / "configs"
    projects_dir = app_dir / "projects"
    projects_dir.mkdir(parents=True)
    (app_dir / "app_config.json").write_text(
        """
{
  "default_project_id": "demo",
  "projects_dir": "configs/projects",
  "secrets_file": "secrets/secrets.json"
}
""",
        encoding="utf-8",
    )
    (projects_dir / "demo.json").write_text(
        """
{
  "project_id": "demo",
  "project_name": "演示项目",
  "excel_path": "samples/demo.xlsx",
  "sheets": {"hourly": "时段数据", "daily": "百度"},
  "kst": {"export_dir": "kst_exports"},
  "baidu": {
    "credential_profile": "demo_profile",
    "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]
  },
  "accounts": [
    {"standard_name": "银康01", "baidu_aliases": ["银康01"], "excel_name": "银康01", "kst_promotion_id": "72828178", "kst_aliases": ["银康01"]},
    {"standard_name": "银康银屑02", "baidu_aliases": ["银康银屑02"], "excel_name": "银康银屑02", "kst_promotion_id": "72828179", "kst_aliases": ["银康银屑02"]},
    {"standard_name": "银康03", "baidu_aliases": ["baidu-银康03", "银康03"], "excel_name": "银康03", "kst_promotion_id": "81509165", "kst_aliases": ["银康03"]}
  ],
  "hourly": {"periods": ["11点", "15点", "18点"]},
  "daily": {
    "write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"],
    "forbidden_fields": ["总对话", "预约", "到诊", "就诊"]
  }
}
""",
        encoding="utf-8",
    )

    current = get_current_project(tmp_path)
    projects = list_projects(tmp_path)

    assert current["project_id"] == "demo"
    assert current["project_name"] == "演示项目"
    assert current["_app_config"]["secrets_file"] == "secrets/secrets.json"
    assert projects == [{"project_id": "demo", "project_name": "演示项目", "path": str(projects_dir / "demo.json")}]
    assert validate_project_config(current) == []


def test_project_config_validation_reports_missing_required_fields():
    errors = validate_project_config({
        "project_id": "bad",
        "project_name": "坏配置",
        "accounts": [],
    })

    assert any("缺少字段：excel.path" in error for error in errors)
    assert any("项目至少需要 1 个账户" in error for error in errors)


def test_project_config_allows_two_account_single_source_project():
    project = {
        "project_id": "two_account_project",
        "project_name": "两账户项目",
        "excel": {"path": "samples/two.xlsx", "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
        "kst": {"export_dir": "kst_exports", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "two_profile", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "账户1", "baidu_names": ["账户1"], "excel_name": "账户1", "kst_ids": ["1001"], "kst_names": ["账户1"]},
            {"standard_name": "账户2", "baidu_names": ["账户2"], "excel_name": "账户2", "kst_ids": ["1002"], "kst_names": ["账户2"]},
        ],
        "hourly": {"periods": ["11点", "15点", "18点"]},
        "daily": {
            "write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"],
            "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"],
        },
    }

    assert validate_project_config(project) == []


def test_project_config_can_switch_current_project_and_normalize_new_schema(tmp_path):
    app_dir = tmp_path / "configs"
    projects_dir = app_dir / "projects"
    projects_dir.mkdir(parents=True)
    (app_dir / "app_config.json").write_text(
        json.dumps({
            "default_project_id": "demo_a",
            "projects_dir": "configs/projects",
            "secrets_file": "secrets/secrets.json",
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    base_project = {
        "project_name": "项目A",
        "excel": {"path": "samples/a.xlsx", "hourly_sheet": "时段数据A", "daily_sheet": "百度A", "engine": "openpyxl"},
        "kst": {"export_dir": "kst_a", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "profile_a", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "账户A1", "baidu_names": ["百度A1"], "excel_name": "账户A1", "kst_ids": ["1001"], "kst_names": ["商务A1"]},
            {"standard_name": "账户A2", "baidu_names": ["百度A2"], "excel_name": "账户A2", "kst_ids": ["1002"], "kst_names": ["商务A2"]},
            {"standard_name": "账户A3", "baidu_names": ["百度A3"], "excel_name": "账户A3", "kst_ids": ["1003"], "kst_names": ["商务A3"]},
        ],
        "hourly": {"periods": ["11点", "15点", "18点"]},
        "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
    }
    for project_id, name in [("demo_a", "项目A"), ("demo_b", "项目B")]:
        project = dict(base_project)
        project["project_id"] = project_id
        project["project_name"] = name
        project["excel"] = dict(base_project["excel"], path=f"samples/{project_id}.xlsx")
        (projects_dir / f"{project_id}.json").write_text(json.dumps(project, ensure_ascii=False), encoding="utf-8")

    assert get_current_project(tmp_path)["project_id"] == "demo_a"
    changed = set_current_project(tmp_path, "demo_b")
    reloaded = reload_current_project(tmp_path)

    assert changed["project_id"] == "demo_b"
    assert reloaded["project_name"] == "项目B"
    assert get_excel_path(reloaded, tmp_path).name == "demo_b.xlsx"
    assert get_project_accounts(reloaded) == ["账户A1", "账户A2", "账户A3"]
    alias_maps = get_account_alias_maps(reloaded)
    assert alias_maps["kst_id_to_account"]["1003"] == "账户A3"
    assert alias_maps["baidu_alias_to_account"]["百度A2"] == "账户A2"
    assert validate_project_config(reloaded) == []


def test_runtime_config_uses_current_project_accounts_not_kunming_defaults(tmp_path):
    project = {
        "project_id": "demo",
        "project_name": "演示",
        "excel": {"path": "target.xlsx", "hourly_sheet": "小时", "daily_sheet": "日报", "engine": "openpyxl"},
        "kst": {"export_dir": "exports", "auto_pick_latest": True, "max_file_age_hours": 4},
        "baidu": {"credential_profile": "profile_demo", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "项目账户1", "baidu_names": ["百度账户1"], "excel_name": "表格账户1", "kst_ids": ["9001"], "kst_names": ["商务账户1"]},
            {"standard_name": "项目账户2", "baidu_names": ["百度账户2"], "excel_name": "表格账户2", "kst_ids": ["9002"], "kst_names": ["商务账户2"]},
            {"standard_name": "项目账户3", "baidu_names": ["百度账户3"], "excel_name": "表格账户3", "kst_ids": ["9003"], "kst_names": ["商务账户3"]},
        ],
    }

    runtime = build_runtime_config_from_project(project, {"baidu": {}, "kst": {}})

    assert runtime["project_id"] == "demo"
    assert runtime["excel_path"] == "target.xlsx"
    assert runtime["sheet_name"] == "小时"
    assert runtime["daily_sheet_name"] == "日报"
    assert runtime["kst"]["export_dir"] == "exports"
    assert runtime["kst"]["max_file_age_minutes"] == 30
    assert runtime["kst"]["promotion_id_accounts"] == {"9001": "项目账户1", "9002": "项目账户2", "9003": "项目账户3"}
    assert set(runtime["accounts"]) == {"项目账户1", "项目账户2", "项目账户3"}
    assert "银康01" not in runtime["accounts"]


def test_global_data_source_preference_defaults_to_api_and_persists(tmp_path):
    from modules.project_config import get_data_source_preference, set_data_source_preference

    config_dir = tmp_path / "configs"
    config_dir.mkdir()
    (config_dir / "app_config.json").write_text(
        json.dumps({
            "default_project_id": "demo",
            "projects_dir": "configs/projects",
            "secrets_file": "secrets/secrets.json",
        }),
        encoding="utf-8",
    )

    assert get_data_source_preference(tmp_path) == "api"
    assert set_data_source_preference(tmp_path, "browser") == "browser"
    saved = json.loads((config_dir / "app_config.json").read_text("utf-8"))
    assert saved["baidu_data_source_preference"] == "browser"
    assert saved["default_project_id"] == "demo"


def test_runtime_config_uses_global_preference_without_destroying_project_mode():
    from modules.project_config import build_runtime_config_from_project

    project = {
        "project_id": "demo",
        "project_name": "demo",
        "excel": {"path": "target.xlsx", "hourly_sheet": "hourly", "daily_sheet": "daily", "engine": "openpyxl"},
        "kst": {"export_dir": "exports", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "demo", "data_source_mode": "api_shadow"},
        "accounts": [],
        "hourly": {"periods": []},
        "daily": {},
        "_app_config": {
            "secrets_file": "secrets/secrets.json",
            "baidu_data_source_preference": "browser",
        },
    }

    runtime = build_runtime_config_from_project(project, {})

    assert runtime["baidu"]["data_source_preference"] == "browser"
    assert runtime["baidu"]["configured_data_source_mode"] == "api_shadow"
    assert runtime["baidu"]["data_source_mode"] == "browser"


@pytest.mark.parametrize(
    ("configured_mode", "preference", "expected_mode"),
    [
        ("browser", "api", "api_preferred"),
        ("browser", "browser", "browser"),
        ("api_shadow", "api", "api_preferred"),
        ("api_shadow", "browser", "browser"),
        ("api_preferred", "api", "api_preferred"),
        ("api_preferred", "browser", "browser"),
    ],
)
def test_runtime_config_global_preference_overrides_project_mode(
    configured_mode, preference, expected_mode
):
    from modules.project_config import build_runtime_config_from_project

    project = {
        "project_id": "demo",
        "project_name": "demo",
        "excel": {"path": "target.xlsx", "hourly_sheet": "hourly", "daily_sheet": "daily", "engine": "openpyxl"},
        "kst": {"export_dir": "exports", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "demo", "data_source_mode": configured_mode},
        "accounts": [],
        "hourly": {"periods": []},
        "daily": {},
        "_app_config": {"baidu_data_source_preference": preference},
    }

    runtime = build_runtime_config_from_project(project, {})

    assert runtime["baidu"]["configured_data_source_mode"] == configured_mode
    assert runtime["baidu"]["data_source_preference"] == preference
    assert runtime["baidu"]["data_source_mode"] == expected_mode


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("api", "api"),
        ("browser", "browser"),
        (" API ", "api"),
        ("BROWSER", "browser"),
        (None, "api"),
        ("", "api"),
        ("   ", "api"),
        ("unsupported", "api"),
    ],
)
def test_normalize_data_source_preference_handles_valid_and_invalid_values(value, expected):
    from modules.project_config import normalize_data_source_preference

    assert normalize_data_source_preference(value) == expected


def test_load_app_config_warns_and_defaults_for_invalid_data_source_preference(tmp_path):
    from modules.project_config import load_app_config

    config_dir = tmp_path / "configs"
    config_dir.mkdir()
    (config_dir / "app_config.json").write_text(
        json.dumps({
            "default_project_id": "demo",
            "projects_dir": "configs/projects",
            "secrets_file": "secrets/secrets.json",
            "baidu_data_source_preference": "unsupported",
        }),
        encoding="utf-8",
    )

    with pytest.warns(RuntimeWarning, match="baidu_data_source_preference"):
        config = load_app_config(tmp_path)

    assert config["baidu_data_source_preference"] == "api"


def test_load_app_config_defaults_missing_data_source_preference_without_warning(tmp_path):
    import warnings
    from modules.project_config import load_app_config

    config_dir = tmp_path / "configs"
    config_dir.mkdir()
    (config_dir / "app_config.json").write_text(
        json.dumps({
            "default_project_id": "demo",
            "projects_dir": "configs/projects",
            "secrets_file": "secrets/secrets.json",
        }),
        encoding="utf-8",
    )

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        config = load_app_config(tmp_path)

    assert caught == []
    assert config["baidu_data_source_preference"] == "api"


def test_set_data_source_preference_preserves_original_config_when_atomic_replace_fails(tmp_path, monkeypatch):
    import os
    import modules.project_config as project_config

    config_dir = tmp_path / "configs"
    config_dir.mkdir()
    config_path = config_dir / "app_config.json"
    original = {
        "default_project_id": "demo",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
        "desktop_pet": "hidden",
    }
    config_path.write_text(json.dumps(original), encoding="utf-8")

    def fail_replace(*_args):
        raise OSError("replace failed")

    monkeypatch.setattr(os, "replace", fail_replace)

    with pytest.raises(OSError, match="replace failed"):
        project_config.set_data_source_preference(tmp_path, "browser")

    assert json.loads(config_path.read_text(encoding="utf-8")) == original
    assert not list(config_dir.glob(".app_config.json.*.tmp"))


def test_menu_runtime_config_converts_project_accounts():
    project = {
        "excel_path": "target.xlsx",
        "sheets": {"hourly": "时段数据", "daily": "百度"},
        "kst": {"export_dir": "kst_exports"},
        "baidu": {"credential_profile": "profile1"},
        "accounts": [
            {"standard_name": "银康01", "baidu_aliases": ["银康01"], "excel_name": "银康01", "kst_promotion_id": "72828178", "kst_aliases": ["银康01"]},
            {"standard_name": "银康银屑02", "baidu_aliases": ["银康银屑02"], "excel_name": "银康银屑02", "kst_promotion_id": "72828179", "kst_aliases": ["银康银屑02"]},
            {"standard_name": "银康03", "baidu_aliases": ["baidu-银康03", "银康03"], "excel_name": "银康03", "kst_promotion_id": "81509165", "kst_aliases": ["银康03"]},
        ],
    }

    runtime = build_runtime_config(project, {"browser": {"max_tabs": 3}, "baidu": {}, "kst": {}})

    assert runtime["excel_path"] == "target.xlsx"
    assert runtime["sheet_name"] == "时段数据"
    assert runtime["baidu"]["credential_project"] == "profile1"
    assert runtime["kst"]["promotion_id_accounts"]["81509165"] == "银康03"
    assert runtime["accounts"]["银康03"]["baidu_name"] == "baidu-银康03"
    assert "银康03" in runtime["accounts"]["银康03"]["aliases"]


def test_menu_confirmation_lines_include_project_task_and_latest_export(tmp_path):
    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    export = export_dir / "latest.xlsx"
    export.write_text("placeholder", encoding="utf-8")
    project = {
        "project_name": "昆明银康 NPX",
        "excel_path": "target.xlsx",
        "sheets": {"hourly": "时段数据", "daily": "百度"},
        "kst": {"export_dir": str(export_dir)},
    }

    lines = build_confirmation_lines(tmp_path, project, "运行15点小时报", period="15点")

    text = "\n".join(lines)
    assert "执行确认" in text
    assert "昆明银康 NPX" in text
    assert "运行15点小时报" in text
    assert "target.xlsx" in text
    assert "时段数据" in text
    assert "15点" in text
    assert f"{export}" in text


def test_menu_header_shows_project_config_and_excel_path(tmp_path):
    from menu import build_menu_header

    project = {
        "project_id": "demo",
        "project_name": "演示项目",
        "_config_path": str(tmp_path / "configs" / "projects" / "demo.json"),
        "excel": {"path": str(tmp_path / "target.xlsx")},
    }

    header = build_menu_header(tmp_path, project)

    assert "演示项目" in header
    assert f"{tmp_path / 'target.xlsx'}" in header


def test_menu_confirmation_accepts_kst_export_file_path(tmp_path):
    export = tmp_path / "export.xlsx"
    export.write_text("placeholder", encoding="utf-8")
    project = {
        "project_name": "昆明银康 NPX",
        "excel_path": "target.xlsx",
        "sheets": {"hourly": "时段数据", "daily": "百度"},
        "kst": {"export_dir": str(export)},
    }

    lines = build_confirmation_lines(tmp_path, project, "检查运行环境")

    text = "\n".join(lines)
    assert f"{export}" in text
    assert "执行确认" in text


def test_menu_dispatch_uses_existing_pipeline_functions(tmp_path):
    calls = []

    def fake_daily(**kwargs):
        calls.append(("daily", kwargs["target_date"]))
        return {"passed": True}

    dispatch_menu_task(
        "2",
        config={},
        root=tmp_path,
        logger=None,
        target_date="2026-05-07",
        kst_file=None,
        runners={"run_daily": fake_daily},
    )

    assert calls == [("daily", "2026-05-07")]


def test_menu_text_is_simplified_for_new_users():
    assert "1. 小时报" in MENU_TEXT
    assert "2. 日报" in MENU_TEXT
    assert "3. 切换项目" in MENU_TEXT
    assert "4. 检查条件项" in MENU_TEXT
    assert "5. 更多功能" in MENU_TEXT
    assert "0. 退出" in MENU_TEXT
    for advanced_text in ["预检与环境", "报告与日志", "配置与诊断", "HERMES / 夏思道帮助", "多百度来源摘要"]:
        assert advanced_text not in MENU_TEXT


def test_menu_hourly_dispatch_uses_internal_period_while_confirm_can_show_simple_period(tmp_path):
    calls = []

    def fake_hourly(**kwargs):
        calls.append(kwargs["period"])
        return {"passed": True}

    dispatch_menu_task(
        "hourly:15点",
        config={},
        root=tmp_path,
        logger=None,
        runners={"run_hourly": fake_hourly},
    )

    assert calls == ["15点"]


def test_menu_choice_1_must_not_be_valid_dispatch_choice(tmp_path):
    """菜单选项 1（小时报）已在 run_menu 中转为 hourly:XX，不应直接到达 dispatch_menu_task。"""
    raised = False
    try:
        dispatch_menu_task(
            "1",
            config={},
            root=tmp_path,
            logger=None,
        )
    except ValueError as exc:
        raised = True
        assert "不支持的菜单选项" in str(exc)
    assert raised, "dispatch_menu_task('1') 应该抛出 ValueError"


def test_menu_choice_2_binds_daily_pipeline(tmp_path):
    """菜单选项 2 必须绑定日报 pipeline。"""
    calls = []

    def fake_daily(**kwargs):
        calls.append(("daily", kwargs["target_date"]))
        return {"passed": True}

    dispatch_menu_task(
        "2",
        config={},
        root=tmp_path,
        logger=None,
        target_date="2026-05-07",
        runners={"run_daily": fake_daily},
    )

    assert calls == [("daily", "2026-05-07")]


def test_menu_task_meta_daily_mapped_to_choice_2(tmp_path):
    """_task_meta 中日报映射 key 为 '2'。"""
    from menu import _task_meta

    project = {
        "excel": {"hourly_sheet": "时段", "daily_sheet": "日报"},
        "sheets": {},
    }

    meta = _task_meta("2", project)

    assert meta["name"] == "运行日报"
    assert meta["sheet"] == "日报"
    assert meta["period"] is None
    assert meta["date"] is not None  # 有默认日期


def test_menu_task_meta_hourly_choices_still_work(tmp_path):
    """_task_meta 中 hourly:XX 映射仍然正常。"""
    from menu import _task_meta

    project = {
        "excel": {"hourly_sheet": "时", "daily_sheet": "日"},
        "sheets": {},
    }

    for period_key, period_name in [("hourly:11点", "11点"), ("hourly:15点", "15点"), ("hourly:18点", "18点")]:
        meta = _task_meta(period_key, project)
        assert meta["period"] == period_name
        assert meta["sheet"] == "时"
        assert meta["date"] is not None


def test_doctor_reports_project_excel_sheets_and_missing_secrets(tmp_path):
    from openpyxl import Workbook

    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "secrets").mkdir()
    (tmp_path / "kst_exports").mkdir()
    (tmp_path / "configs" / "app_config.json").write_text(
        '{"default_project_id":"demo","projects_dir":"configs/projects","secrets_file":"secrets/secrets.json"}',
        encoding="utf-8",
    )
    excel_path = tmp_path / "samples" / "demo.xlsx"
    excel_path.parent.mkdir()
    wb = Workbook()
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.save(excel_path)
    (tmp_path / "configs" / "projects" / "demo.json").write_text(
        f"""
{{
  "project_id": "demo",
  "project_name": "演示项目",
  "excel_path": "{str(excel_path).replace('\\', '/')}",
  "sheets": {{"hourly": "时段数据", "daily": "百度"}},
  "kst": {{"export_dir": "kst_exports"}},
  "baidu": {{"credential_profile": "demo", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]}},
  "accounts": [
    {{"standard_name": "银康01", "baidu_aliases": ["银康01"], "excel_name": "银康01", "kst_promotion_id": "72828178", "kst_aliases": ["银康01"]}},
    {{"standard_name": "银康银屑02", "baidu_aliases": ["银康银屑02"], "excel_name": "银康银屑02", "kst_promotion_id": "72828179", "kst_aliases": ["银康银屑02"]}},
    {{"standard_name": "银康03", "baidu_aliases": ["银康03"], "excel_name": "银康03", "kst_promotion_id": "81509165", "kst_aliases": ["银康03"]}}
  ],
  "hourly": {{"periods": ["11点", "15点", "18点"]}},
  "daily": {{"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "forbidden_fields": ["总对话", "预约", "到诊", "就诊"]}}
}}
""",
        encoding="utf-8",
    )

    report = run_doctor(tmp_path, {"browser": {"cdp_endpoint": "http://127.0.0.1:9", "auto_start_debug_chrome": False}, "kst": {}})

    assert report["summary"]["total"] >= 13
    assert report["checks"]["app_config"]["passed"] is True
    assert report["checks"]["project_config"]["passed"] is True
    assert report["checks"]["target_excel"]["passed"] is True
    assert report["checks"]["hourly_sheet"]["passed"] is True
    assert report["checks"]["daily_sheet"]["passed"] is True
    assert report["checks"]["secrets_json"]["passed"] is False
    assert "百度账号未配置" in report["checks"]["secrets_json"]["message"]


def test_doctor_openpyxl_engine_does_not_require_microsoft_excel(tmp_path):
    from openpyxl import Workbook

    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "secrets").mkdir()
    (tmp_path / "kst_exports").mkdir()
    (tmp_path / "configs" / "app_config.json").write_text(
        '{"default_project_id":"demo","projects_dir":"configs/projects","secrets_file":"secrets/secrets.json"}',
        encoding="utf-8",
    )
    excel_path = tmp_path / "samples" / "demo.xlsx"
    excel_path.parent.mkdir()
    wb = Workbook()
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.save(excel_path)
    (tmp_path / "configs" / "projects" / "demo.json").write_text(
        f"""
{{
  "project_id": "demo",
  "project_name": "演示项目",
  "excel_path": "{str(excel_path).replace('\\', '/')}",
  "excel": {{"path": "{str(excel_path).replace('\\', '/')}", "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"}},
  "sheets": {{"hourly": "时段数据", "daily": "百度"}},
  "kst": {{"export_dir": "kst_exports"}},
  "baidu": {{"credential_profile": "demo", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]}},
  "accounts": [
    {{"standard_name": "银康01", "baidu_aliases": ["银康01"], "excel_name": "银康01", "kst_promotion_id": "72828178", "kst_aliases": ["银康01"]}},
    {{"standard_name": "银康银屑02", "baidu_aliases": ["银康银屑02"], "excel_name": "银康银屑02", "kst_promotion_id": "72828179", "kst_aliases": ["银康银屑02"]}},
    {{"standard_name": "银康03", "baidu_aliases": ["银康03"], "excel_name": "银康03", "kst_promotion_id": "81509165", "kst_aliases": ["银康03"]}}
  ],
  "hourly": {{"periods": ["11点", "15点", "18点"]}},
  "daily": {{"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "forbidden_fields": ["总对话", "预约", "到诊", "就诊"]}}
}}
""",
        encoding="utf-8",
    )

    report = run_doctor(tmp_path, {"browser": {"cdp_endpoint": "http://127.0.0.1:9", "auto_start_debug_chrome": False}, "kst": {}})

    assert report["checks"]["excel_engine"]["passed"] is True
    assert "openpyxl" in report["checks"]["excel_engine"]["message"]
    assert "WPS" in report["checks"]["excel_engine"]["message"]
    assert report["checks"]["openpyxl"]["passed"] is True
    assert report["checks"]["openpyxl_save_test"]["passed"] is True
    assert "excel_app" not in report["checks"]


def test_openpyxl_save_error_mentions_closing_wps_file(tmp_path):
    message = format_openpyxl_save_error(tmp_path / "target.xlsx", PermissionError("locked"))

    assert "请关闭 WPS 中的目标文件后重试" in message
    assert "target.xlsx" in message


def test_doctor_openpyxl_mode_skips_excel_com_only_requirements(tmp_path, monkeypatch):
    from importlib.metadata import PackageNotFoundError
    from modules import doctor

    (tmp_path / "requirements.txt").write_text("openpyxl>=3.1.2\nxlwings>=0.30.0\npywin32>=306\n", encoding="utf-8")

    def fake_version(package_name):
        if package_name in {"xlwings", "pywin32"}:
            raise PackageNotFoundError(package_name)
        return "1.0"

    monkeypatch.setattr(doctor.importlib.metadata, "version", fake_version)

    openpyxl_report = doctor._check_requirements(tmp_path, excel_engine="openpyxl")
    excel_com_report = doctor._check_requirements(tmp_path, excel_engine="excel_com")

    assert openpyxl_report["passed"] is True
    assert "xlwings" in openpyxl_report["detail"]["skipped_optional"]
    assert excel_com_report["passed"] is False
    assert "xlwings" in excel_com_report["detail"]["missing"]


def test_doctor_excel_com_reads_optional_requirements_file(tmp_path, monkeypatch):
    import modules.doctor as doctor

    (tmp_path / "requirements.txt").write_text("openpyxl>=3.1.2\n", encoding="utf-8")
    (tmp_path / "requirements-excel-com.txt").write_text("xlwings>=0.30.0\npywin32>=306\n", encoding="utf-8")

    def fake_version(package_name):
        if package_name in {"xlwings", "pywin32"}:
            raise doctor.importlib.metadata.PackageNotFoundError
        return "1.0"

    monkeypatch.setattr(doctor.importlib.metadata, "version", fake_version)

    openpyxl_report = doctor._check_requirements(tmp_path, excel_engine="openpyxl")
    excel_com_report = doctor._check_requirements(tmp_path, excel_engine="excel_com")

    assert openpyxl_report["passed"] is True
    assert excel_com_report["passed"] is False
    assert set(excel_com_report["detail"]["missing"]) == {"xlwings", "pywin32"}


def test_default_install_excludes_optional_excel_com_dependencies():
    root = Path(__file__).resolve().parents[1]
    base = (root / "requirements.txt").read_text(encoding="utf-8")
    optional = (root / "requirements-excel-com.txt").read_text(encoding="utf-8")

    assert "xlwings" not in base
    assert "pywin32" not in base
    assert "xlwings" in optional
    assert "pywin32" in optional


def test_run_menu_installs_or_repairs_missing_dependencies_before_importing_menu():
    root = Path(__file__).resolve().parents[1]
    script = (root / "run_menu.bat").read_text(encoding="utf-8")

    assert 'if not exist ".venv\\Scripts\\python.exe"' in script
    assert 'set "NEED_INSTALL=1"' in script
    assert 'call "%~dp0install_env.bat"' in script
    assert 'import openpyxl, pandas, xlrd, dateutil, playwright, rich' in script


def test_install_env_prefers_runtime_lock_when_available():
    root = Path(__file__).resolve().parents[1]
    script = (root / "install_env.bat").read_text(encoding="utf-8")

    assert "requirements-runtime.lock.txt" in script
    assert 'set "RUNTIME_REQUIREMENTS=%~dp0requirements-runtime.txt"' in script
    assert 'if exist "%~dp0requirements-runtime.lock.txt" set "RUNTIME_REQUIREMENTS=%~dp0requirements-runtime.lock.txt"' in script
    assert '-r "%RUNTIME_REQUIREMENTS%"' in script


def test_release_builder_excludes_sensitive_and_runtime_files():
    assert should_include_file(Path("main.py")) is True
    assert should_include_file(Path("modules") / "doctor.py") is True
    assert should_include_file(Path("reports") / ".gitkeep") is True
    assert should_include_file(Path("logs") / ".gitkeep") is True
    assert should_include_file(Path("backups") / ".gitkeep") is True
    assert should_include_file(Path("reports") / "final_run_report.json") is False
    assert should_include_file(Path("logs") / "run.log") is False
    assert should_include_file(Path("backups") / "target.xlsx") is False
    assert should_include_file(Path("diagnostics") / "diagnostic_20260722_101257.zip") is False
    assert should_include_file(Path("kst_exports") / "export.xlsx") is False
    assert should_include_file(Path("secrets") / "secrets.json") is False
    assert should_include_file(Path("samples") / "真实业务.xlsx") is False
    assert should_include_file(Path(".venv") / "pyvenv.cfg") is False
    assert should_include_file(Path("tests") / "test_basic.py") is False
    assert should_include_file(Path("run_11.bat")) is False


def test_dependency_lock_file_contains_exact_runtime_versions(tmp_path):
    from modules.maintenance import build_runtime_dependency_lock

    (tmp_path / "requirements-runtime.txt").write_text(
        "openpyxl>=3.1.2\npandas>=2.0.0\n",
        encoding="utf-8",
    )

    lock_path = build_runtime_dependency_lock(
        tmp_path,
        installed_packages={
            "openpyxl": "3.1.5",
            "pandas": "3.0.2",
            "pytest": "9.0.3",
        },
    )

    content = lock_path.read_text(encoding="utf-8")
    assert lock_path.name == "requirements-runtime.lock.txt"
    assert "openpyxl==3.1.5" in content
    assert "pandas==3.0.2" in content
    assert "pytest" not in content


def test_diagnostic_bundle_redacts_sensitive_values_and_skips_secrets(tmp_path):
    import zipfile
    from modules.maintenance import create_diagnostic_bundle

    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "secrets").mkdir()
    (tmp_path / "logs").mkdir()
    (tmp_path / "reports").mkdir()
    (tmp_path / "configs" / "app_config.json").write_text(
        json.dumps({"current_project_id": "demo", "hmac_client_key": "real-key"}, ensure_ascii=False),
        encoding="utf-8",
    )
    (tmp_path / "configs" / "projects" / "demo.json").write_text(
        json.dumps({"project_name": "演示", "excel": {"path": "D:/data/demo.xlsx"}}, ensure_ascii=False),
        encoding="utf-8",
    )
    (tmp_path / "secrets" / "secrets.json").write_text(
        json.dumps({"password": "real-password", "accessToken": "real-token"}, ensure_ascii=False),
        encoding="utf-8",
    )
    (tmp_path / "logs" / "run.log").write_text("password=real-password accessToken=eyJabcdef.ghijklmnop.qrstuvwxyz\n", encoding="utf-8")
    (tmp_path / "reports" / "final_run_report.json").write_text(
        json.dumps({"authorization": "Bearer real-token", "passed": False}, ensure_ascii=False),
        encoding="utf-8",
    )

    bundle = create_diagnostic_bundle(tmp_path, now_label="20260722_101500")

    with zipfile.ZipFile(bundle) as archive:
        names = set(archive.namelist())
        payload = "\n".join(archive.read(name).decode("utf-8", errors="ignore") for name in names)

    assert "manifest.json" in names
    assert "configs/app_config.json" in names
    assert "secrets/secrets.json" not in names
    assert "real-password" not in payload
    assert "real-token" not in payload
    assert "eyJabcdef" not in payload
    assert "***" in payload


def test_diagnostic_bundle_skips_files_removed_during_collection(tmp_path, monkeypatch):
    import zipfile
    import modules.maintenance as maintenance

    logs = tmp_path / "logs"
    logs.mkdir()
    removed = logs / "removed.log"
    removed.write_text("old", encoding="utf-8")

    monkeypatch.setattr(maintenance, "_diagnostic_candidates", lambda root: [removed])
    removed.unlink()

    bundle = maintenance.create_diagnostic_bundle(tmp_path, now_label="20260722_102000")

    with zipfile.ZipFile(bundle) as archive:
        assert "manifest.json" in archive.namelist()
        assert "logs/removed.log" not in archive.namelist()


def test_archive_logs_moves_only_old_log_files(tmp_path):
    from datetime import datetime, timedelta
    import zipfile
    from modules.maintenance import archive_logs

    logs = tmp_path / "logs"
    logs.mkdir()
    old_log = logs / "old.log"
    current_log = logs / "run.log"
    old_log.write_text("old", encoding="utf-8")
    current_log.write_text("current", encoding="utf-8")
    old_ts = (datetime(2026, 7, 1) - datetime(1970, 1, 1)).total_seconds()
    os.utime(old_log, (old_ts, old_ts))

    result = archive_logs(
        tmp_path,
        older_than_days=7,
        now=datetime(2026, 7, 22, 10, 20, 0),
    )

    assert result["archived_count"] == 1
    assert not old_log.exists()
    assert current_log.exists()
    archive_path = Path(result["archive_path"])
    assert archive_path.exists()
    with zipfile.ZipFile(archive_path) as archive:
        assert "old.log" in archive.namelist()


def test_project_template_and_demo_project_are_complete():
    root = Path(__file__).resolve().parents[1]
    template = json.loads((root / "configs" / "projects" / "project_template.json").read_text(encoding="utf-8"))
    demo = json.loads((root / "configs" / "projects" / "demo_project.json").read_text(encoding="utf-8"))

    for project in [template, demo]:
        normalized_errors = validate_project_config(project)
        assert normalized_errors == []
        assert project["excel"]["engine"] == "openpyxl"
        assert "auto_pick_latest" in project["kst"]
        assert "max_file_age_minutes" in project["kst"]
        assert project["daily"]["do_not_write_fields"] == ["总对话", "预约", "到诊", "就诊"]
        for account in project["accounts"]:
            assert set(["standard_name", "baidu_names", "excel_name", "kst_ids", "kst_names"]) <= set(account)


def test_demo_project_is_template_and_not_in_project_list():
    """demo_project 是模板，不出现在项目列表中，但可以单独加载验证。"""
    root = Path(__file__).resolve().parents[1]

    projects = list_projects(root)
    ids = [p["project_id"] for p in projects]
    assert "demo_project" not in ids, "demo_project 是模板，不应出现在项目列表"

    # 但可以直接加载和校验
    from modules.project_config import load_project_config
    demo = load_project_config(root, "demo_project")
    assert demo["project_name"] == "演示项目"
    assert demo.get("is_template") is True
    assert validate_project_config(demo) == []


def test_release_builder_accepts_version_and_includes_project_template_docs(tmp_path):
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="2.0")

    assert release.name == "hourly_report_bot_release_v2.0.zip"
    import zipfile

    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())

    assert "configs/projects/project_template.json" in names
    assert "configs/projects/demo_project.json" in names
    assert "docs/如何新增一个项目.md" in names
    assert "secrets/secrets.json" not in names


def test_core_modules_do_not_hardcode_kunming_project_accounts():
    root = Path(__file__).resolve().parents[1]
    forbidden = [
        "银康01",
        "银康银屑02",
        "银康03",
        "baidu-银康03",
        "72828178",
        "72828179",
        "81509165",
    ]
    offenders = []
    for path in (root / "modules").glob("*.py"):
        text = path.read_text(encoding="utf-8")
        for token in forbidden:
            if token in text:
                offenders.append(f"{path.name}:{token}")

    assert offenders == []


def test_account_ranges_follow_horizontal_merged_title_blocks():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "每日时段统计数据"
    ws.merge_cells("A1:E1")
    ws["F1"] = "银康01-72828178"
    ws.merge_cells("F1:J1")
    ws["K1"] = "银康银屑02-72828179"
    ws.merge_cells("K1:O1")
    ws["P1"] = "baidu-银康03-81509165"
    ws.merge_cells("P1:T1")

    config = {
        "accounts": {
            "银康01": {"aliases": ["银康01"], "excel_name": "银康01", "baidu_name": "银康01"},
            "银康银屑02": {"aliases": ["银康银屑02"], "excel_name": "银康银屑02", "baidu_name": "银康银屑02"},
            "银康03": {"aliases": ["银康03", "baidu-银康03"], "excel_name": "银康03", "baidu_name": "baidu-银康03"},
        }
    }
    rows = _scan_non_empty_cells(ws, _build_merged_value_map(ws))
    titles = _find_account_titles(rows, config)
    ranges = _build_account_ranges(titles, ws, _build_merged_bounds_map(ws))

    assert ranges["银康01"]["range"]["min_col"] == 6
    assert ranges["银康01"]["range"]["max_col"] == 10
    assert ranges["银康银屑02"]["range"]["min_col"] == 11
    assert ranges["银康银屑02"]["range"]["max_col"] == 15
    assert ranges["银康03"]["range"]["min_col"] == 16
    assert ranges["银康03"]["range"]["max_col"] == 20


def test_scan_non_empty_cells_does_not_materialize_blank_rectangle():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "账户"
    ws["Z100"] = "末端标题"
    materialized_before = set(ws._cells)

    rows = _scan_non_empty_cells(ws, _build_merged_value_map(ws))

    assert set(ws._cells) == materialized_before
    assert [(item["address"], item["raw_text"]) for item in rows] == [
        ("A1", "账户"),
        ("Z100", "末端标题"),
    ]


def test_mock_write_helpers_find_row_and_reject_summary_columns():
    from datetime import date
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "每日时段统计数据"
    ws["P1"] = "银康01"
    ws["A3"] = "日期"
    ws["B3"] = "时段"
    ws["P3"] = "展现量"
    ws["A4"] = date(2026, 5, 6)
    ws.merge_cells("A4:A7")
    ws["B4"] = "昨天数据"
    ws["B5"] = "11点"
    ws["B6"] = "3点"
    ws["B7"] = "6点"

    merged_values = _build_merged_value_map(ws)

    assert _normalize_period_for_excel("15点") == "3点"
    assert _normalize_period_for_excel("15") == "3点"
    assert _find_target_row(ws, 1, 2, date(2026, 5, 6), "15点", merged_values) == 6
    assert _validate_write_target(
        "银康01",
        "展现",
        6,
        16,
        {"range": {"min_row": 1, "max_row": 20, "min_col": 16, "max_col": 20}},
        [{"range": {"min_row": 1, "max_row": 20, "min_col": 1, "max_col": 15}}],
    ) == []
    assert _validate_write_target(
        "银康01",
        "展现",
        6,
        3,
        {"range": {"min_row": 1, "max_row": 20, "min_col": 16, "max_col": 20}},
        [{"range": {"min_row": 1, "max_row": 20, "min_col": 1, "max_col": 15}}],
    )


def test_inspect_daily_worksheet_detects_baidu_sheet_blocks_and_write_rules():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "百度"
    ws["A1"] = "日期"
    ws["C1"] = "汇总数据"
    ws.merge_cells("C1:O1")
    ws["P1"] = "银康01-72828178"
    ws.merge_cells("P1:AB1")
    ws["AC1"] = "银康银屑02-72828179"
    ws.merge_cells("AC1:AO1")
    ws["AP1"] = "银康03-81509165"
    ws.merge_cells("AP1:BB1")
    headers = ["展现量", "点击", "消费", "acp", "总对话", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜", "预约", "就诊", "转潜成本"]
    for start in [16, 29, 42]:
        for offset, header in enumerate(headers):
            ws.cell(row=2, column=start + offset).value = header
    ws["A2"] = "日期"
    ws["A3"] = "2026-05-07"
    ws["A4"] = "2026-05-08"
    config = {
        "accounts": {
            "银康01": {"excel_name": "银康01", "baidu_name": "银康01", "aliases": ["银康01"]},
            "银康银屑02": {"excel_name": "银康银屑02", "baidu_name": "银康银屑02", "aliases": ["银康银屑02"]},
            "银康03": {"excel_name": "银康03", "baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    report = inspect_daily_worksheet(ws, config, "unit.xlsx")

    assert report["sheet_found"] is True
    assert report["date_column"]["found"] is True
    assert report["date_rows"]["count"] == 2
    assert report["accounts"]["银康01"]["range"]["min_col"] == 16
    assert report["accounts"]["银康银屑02"]["fields"]["有效对话"]["header_cell"] == "AH2"
    assert report["accounts"]["银康03"]["fields"]["消费"]["header_cell"] == "AR2"
    assert report["allowed_fields"] == ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"]
    assert report["forbidden_fields"] == ["总对话", "预约", "到诊", "就诊"]
    assert report["accounts"]["银康01"]["fields"]["预约"]["write_allowed"] is False
    assert report["errors"] == []


def test_normalize_text():
    assert normalize_text(" Baidu-银康03\n") == "baidu-银康03"
    assert normalize_text("银康 01") == "银康01"


def test_kst_tags():
    r = classify_dialog_by_tags("转潜-有效")
    assert r["总对话"] == 1
    assert r["有效对话"] == 1
    assert r["一般有效"] == 0
    assert r["有效转潜"] == 1
    assert r["总转潜"] == 1

    general = classify_dialog_by_tags("有效-一般")
    assert general["总对话"] == 1
    assert general["有效对话"] == 0
    assert general["一般有效"] == 1
    assert general["有效转潜"] == 0
    assert general["总转潜"] == 0

    three_messages = classify_dialog_by_tags("有效-三句话")
    assert three_messages["有效对话"] == 1
    assert three_messages["一般有效"] == 0
    assert three_messages["有效转潜"] == 0


def test_visitor_dialog_requires_message_count_at_least_one():
    assert has_visitor_dialog({"访客消息数": "1"}) is True
    assert has_visitor_dialog({"访客消息数": " 2 "}) is True
    assert has_visitor_dialog({"访客消息数": "0"}) is False
    assert has_visitor_dialog({"访客消息数": ""}) is False
    assert has_visitor_dialog({"访客消息数": "无"}) is False


def test_visitor_dialog_accepts_visitor_sent_count_alias():
    assert has_visitor_dialog({"访客发送数": "1"}) is True
    assert has_visitor_dialog({"访客发送数": "0"}) is False


def test_visitor_dialog_accepts_visitor_sent_message_count_alias():
    assert has_visitor_dialog({"访客发送消息数": "1"}) is True
    assert has_visitor_dialog({"访客发送消息数": "0"}) is False


def test_kst_daily_tags_do_not_count_invalid_as_valid():
    invalid = classify_daily_dialog_by_tags("转潜-无效, 无效")
    assert invalid["总对话"] == 1
    assert invalid["有效对话"] == 0
    assert invalid["无效对话"] == 1
    assert invalid["一般有效对话"] == 0
    assert invalid["有效转潜"] == 0
    assert invalid["总转潜"] == 1

    valid = classify_daily_dialog_by_tags("转潜-有效, 有效-一般")
    assert valid["总对话"] == 1
    assert valid["有效对话"] == 1
    assert valid["无效对话"] == 0
    assert valid["一般有效对话"] == 1
    assert valid["有效转潜"] == 1
    assert valid["总转潜"] == 1

    three_messages = classify_daily_dialog_by_tags("有效-三句话")
    assert three_messages["有效对话"] == 1
    assert three_messages["无效对话"] == 0
    assert three_messages["一般有效对话"] == 0

    general = classify_daily_dialog_by_tags("有效-一般")
    assert general["有效对话"] == 0
    assert general["无效对话"] == 0
    assert general["一般有效对话"] == 1


def test_daily_kst_validation_treats_general_as_independent_category():
    from modules.validators import validate_daily_kst_counts

    general_only = {
        "总对话": 1,
        "有效对话": 0,
        "无效对话": 0,
        "一般有效对话": 1,
        "有效转潜": 0,
        "总转潜": 0,
    }
    valid_and_general = {
        "总对话": 1,
        "有效对话": 1,
        "无效对话": 0,
        "一般有效对话": 1,
        "有效转潜": 1,
        "总转潜": 1,
    }

    assert validate_daily_kst_counts(general_only) == []
    assert validate_daily_kst_counts(valid_and_general) == []


def test_aggregate_kst_export_rows_maps_remark_promotion_ids_and_counts_tags():
    rows = [
        {"备注说明": "推广ID：72828178</br>关键词ID：x", "名片标签": "转潜-有效", "对话时间": "2026-05-07 10:00", "访客消息数": "1"},
        {"备注说明": "72828178 其他说明", "名片标签": "有效-一般", "对话时间": "2026-05-07 10:01", "访客消息数": "2"},
        {"备注说明": "72828179", "名片标签": "有效-三句", "对话时间": "2026-05-07 10:02", "访客消息数": "1"},
        {"备注说明": "81509165", "名片标签": "", "对话时间": "2026-05-07 10:03", "访客消息数": "1"},
        {"备注说明": "999999", "名片标签": "转潜-有效", "对话时间": "2026-05-07 10:04", "访客消息数": "1"},
    ]

    parsed = aggregate_kst_export_rows(rows, _kunming_niu_runtime_config())

    assert parsed["errors"] == []
    assert parsed["accounts"]["银康01"]["总对话"] == 2
    assert parsed["accounts"]["银康01"]["有效对话"] == 1
    assert parsed["accounts"]["银康01"]["一般有效"] == 1
    assert parsed["accounts"]["银康01"]["有效转潜"] == 1
    assert parsed["accounts"]["银康01"]["总转潜"] == 1
    assert parsed["accounts"]["银康银屑02"]["有效对话"] == 1
    assert parsed["accounts"]["银康银屑02"]["一般有效"] == 0
    assert parsed["accounts"]["银康03"]["总对话"] == 1
    assert parsed["accounts"]["银康03"]["有效对话"] == 0
    assert parsed["accounts"]["银康03"]["一般有效"] == 0
    assert parsed["summary"]["unmatched_rows"] == 1
    assert parsed["account_dialog_details"]["银康01"][0]["promotion_id"] == "72828178"


def test_aggregate_kst_export_rows_skips_rows_without_visitor_messages():
    rows = [
        {"备注说明": "72828178", "名片标签": "转潜-有效", "对话时间": "2026-05-07 10:00", "访客消息数": "0"},
        {"备注说明": "72828178", "名片标签": "有效-一般", "对话时间": "2026-05-07 10:01", "访客消息数": "1"},
    ]

    parsed = aggregate_kst_export_rows(rows, _kunming_niu_runtime_config())

    assert parsed["accounts"]["银康01"]["总对话"] == 1
    assert parsed["accounts"]["银康01"]["有效对话"] == 0
    assert parsed["accounts"]["银康01"]["一般有效"] == 1
    assert parsed["accounts"]["银康01"]["有效转潜"] == 0
    assert parsed["accounts"]["银康01"]["总转潜"] == 0
    assert parsed["summary"]["skipped_no_visitor_messages"] == 1
    assert parsed["account_dialog_details"]["银康01"][0]["counts"]["总对话"] == 0


def test_parse_kst_export_csv_outputs_reports(tmp_path):
    export = tmp_path / "kst.csv"
    today = date.today().isoformat()
    export.write_text(
        "对话时间,备注说明,名片标签,搜索词,访客消息数\n"
        f"{today} 10:00,72828178-abc,转潜-有效,银屑病,1\n"
        f"{today} 10:01,81509165,,皮肤病,1\n",
        encoding="utf-8-sig",
    )
    reports = tmp_path / "reports"
    config = _kunming_niu_runtime_config()
    config["kst"] = {"export_dir": str(tmp_path), "promotion_id_accounts": _kunming_niu_runtime_config()["kst"]["promotion_id_accounts"]}

    result = parse_kst_export_file(export, config, tmp_path, "15点")

    assert result["dialog_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["dialog_data"]["accounts"]["银康01"]["有效对话"] == 1
    assert result["dialog_data"]["accounts"]["银康01"]["一般有效"] == 0
    assert result["dialog_data"]["accounts"]["银康03"]["总对话"] == 1
    assert result["dialog_data"]["summary"]["raw_rows"] == 2
    assert (reports / "kst_dialog_data.json").exists()
    assert (reports / "kst_parse_report.json").exists()
    assert (reports / "kst_unmatched_rows.json").exists()


def test_parse_kst_export_accepts_visitor_sent_count_header(tmp_path):
    export = tmp_path / "kst_sent_count.csv"
    today = date.today().isoformat()
    export.write_text(
        "对话时间,备注说明,名片标签,访客发送数\n"
        f"{today} 10:00,72828178-abc,转潜-有效,1\n",
        encoding="utf-8-sig",
    )
    config = _kunming_niu_runtime_config()
    config["kst"] = {"export_dir": str(tmp_path), "promotion_id_accounts": _kunming_niu_runtime_config()["kst"]["promotion_id_accounts"]}

    result = parse_kst_export_file(export, config, tmp_path, "15点")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["field_info"]["has_visitor_messages"] is True
    assert result["dialog_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["dialog_data"]["accounts"]["银康01"]["有效对话"] == 1


def test_parse_kst_export_accepts_visitor_sent_message_count_header(tmp_path):
    export = tmp_path / "kst_sent_message_count.csv"
    today = date.today().isoformat()
    export.write_text(
        "对话时间,备注说明,名片标签,访客发送消息数\n"
        f"{today} 10:00,72828178-abc,转潜-有效,1\n",
        encoding="utf-8-sig",
    )
    config = _kunming_niu_runtime_config()
    config["kst"] = {"export_dir": str(tmp_path), "promotion_id_accounts": _kunming_niu_runtime_config()["kst"]["promotion_id_accounts"]}

    result = parse_kst_export_file(export, config, tmp_path, "15点")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["field_info"]["has_visitor_messages"] is True
    assert result["dialog_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["dialog_data"]["accounts"]["银康01"]["有效对话"] == 1


def test_parse_kst_export_filters_non_current_date_without_marking_unmatched(tmp_path):
    export = tmp_path / "kst.csv"
    export.write_text(
        "对话时间,备注说明,名片标签,访客消息数\n"
        "2026-05-06 10:00,72828178-abc,转潜-有效,1\n",
        encoding="utf-8-sig",
    )

    km_config = _kunming_niu_runtime_config()
    km_config["kst"] = {"export_dir": str(tmp_path), "promotion_id_accounts": _kunming_niu_runtime_config()["kst"]["promotion_id_accounts"]}
    result = parse_kst_export_file(export, km_config, tmp_path, "15")

    assert result["dialog_data"]["summary"]["raw_rows"] == 1
    assert result["dialog_data"]["summary"]["matched_rows"] == 0
    assert result["dialog_data"]["summary"]["unmatched_rows"] == 0
    assert result["dialog_data"]["summary"]["date_filtered_rows"] == 1
    assert result["unmatched_rows"] == []


def test_find_latest_kst_export_ignores_files_older_than_30_minutes(tmp_path):
    import os
    import time

    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    old_file = export_dir / "old.csv"
    old_file.write_text("old", encoding="utf-8")
    old_time = time.time() - 31 * 60
    os.utime(old_file, (old_time, old_time))

    assert find_latest_kst_export(tmp_path, {"kst": {"export_dir": "kst_exports"}}) is None

    fresh_file = export_dir / "fresh.csv"
    fresh_file.write_text("fresh", encoding="utf-8")

    assert find_latest_kst_export(tmp_path, {"kst": {"export_dir": "kst_exports"}}) == fresh_file


def test_empty_kst_export_result_writes_zero_dialog_report(tmp_path):
    config = _kunming_niu_runtime_config()

    result = write_empty_kst_export_result(config, tmp_path, "15点", "没有新导出")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["errors"] == []
    assert result["parse_report"]["warnings"] == ["没有新导出"]
    assert result["dialog_data"]["summary"]["no_export_file"] is True
    assert all(row["总对话"] == 0 for row in result["dialog_data"]["accounts"].values())
    assert (tmp_path / "reports" / "kst_dialog_data.json").exists()


def test_parse_kst_daily_file_filters_date_and_outputs_daily_counts(tmp_path):
    export = tmp_path / "kst_daily.csv"
    export.write_text(
        "对话时间,备注说明,名片标签,搜索词,访客消息数\n"
        "2026-05-07 10:00,72828178-abc,转潜-有效,银屑病,1\n"
        "2026-05-07 10:01,72828178-abc,有效-一般,皮肤病,2\n"
        "2026-05-07 10:02,72828179-abc,无效,银屑病,1\n"
        "2026-05-07 10:03,81509165-abc,转潜-无效,银屑病,1\n"
        "2026-05-06 10:03,81509165-abc,转潜-有效,旧日期,1\n",
        encoding="utf-8-sig",
    )

    result = parse_kst_daily_file(export, _kunming_niu_runtime_config(), tmp_path, "2026-05-07")

    assert result["daily_data"]["date"] == "2026-05-07"
    assert result["daily_data"]["source"] == "kst_daily_export"
    assert result["daily_data"]["accounts"]["银康01"]["总对话"] == 2
    assert result["daily_data"]["accounts"]["银康01"]["有效对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["无效对话"] == 0
    assert result["daily_data"]["accounts"]["银康01"]["一般有效对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["有效转潜"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["总转潜"] == 1
    assert result["daily_data"]["accounts"]["银康银屑02"]["总对话"] == 1
    assert result["daily_data"]["accounts"]["银康银屑02"]["有效对话"] == 0
    assert result["daily_data"]["accounts"]["银康银屑02"]["无效对话"] == 1
    assert result["daily_data"]["accounts"]["银康03"]["总转潜"] == 1
    assert result["daily_data"]["accounts"]["银康03"]["有效转潜"] == 0
    assert result["daily_data"]["summary"]["raw_rows"] == 5
    assert result["daily_data"]["summary"]["date_filtered_rows"] == 1
    assert result["parse_report"]["passed"] is True
    assert (tmp_path / "reports" / "kst_daily_data.json").exists()
    assert (tmp_path / "reports" / "kst_daily_parse_report.json").exists()


def test_parse_kst_daily_accepts_visitor_sent_count_header(tmp_path):
    export = tmp_path / "kst_daily_sent_count.csv"
    export.write_text(
        "对话时间,备注说明,名片标签,访客发送数\n"
        "2026-05-07 10:00,72828178-abc,转潜-有效,1\n",
        encoding="utf-8-sig",
    )

    result = parse_kst_daily_file(export, _kunming_niu_runtime_config(), tmp_path, "2026-05-07")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["field_info"]["has_visitor_messages"] is True
    assert result["daily_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["有效对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["无效对话"] == 0


def test_parse_kst_daily_accepts_visitor_sent_message_count_header(tmp_path):
    export = tmp_path / "kst_daily_sent_message_count.csv"
    export.write_text(
        "对话时间,备注说明,名片标签,访客发送消息数\n"
        "2026-05-07 10:00,72828178-abc,转潜-有效,1\n",
        encoding="utf-8-sig",
    )

    result = parse_kst_daily_file(export, _kunming_niu_runtime_config(), tmp_path, "2026-05-07")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["field_info"]["has_visitor_messages"] is True
    assert result["daily_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["有效对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["无效对话"] == 0


def test_parse_kst_daily_file_skips_rows_without_visitor_messages(tmp_path):
    export = tmp_path / "kst_daily.csv"
    export.write_text(
        "对话时间,备注说明,名片标签,访客消息数\n"
        "2026-05-07 10:00,72828178-abc,转潜-有效,0\n"
        "2026-05-07 10:01,72828178-abc,有效-一般,1\n",
        encoding="utf-8-sig",
    )

    result = parse_kst_daily_file(export, _kunming_niu_runtime_config(), tmp_path, "2026-05-07")

    assert result["daily_data"]["accounts"]["银康01"]["总对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["有效对话"] == 0
    assert result["daily_data"]["accounts"]["银康01"]["无效对话"] == 0
    assert result["daily_data"]["accounts"]["银康01"]["一般有效对话"] == 1
    assert result["daily_data"]["accounts"]["银康01"]["有效转潜"] == 0
    assert result["daily_data"]["summary"]["skipped_no_visitor_messages"] == 1


def test_empty_kst_daily_result_writes_zero_dialog_report(tmp_path):
    config = _kunming_niu_runtime_config()

    result = write_empty_kst_daily_result(config, tmp_path, "2026-05-07", "没有新日报导出")

    assert result["parse_report"]["passed"] is True
    assert result["parse_report"]["errors"] == []
    assert result["parse_report"]["warnings"] == ["没有新日报导出"]
    assert result["daily_data"]["summary"]["no_export_file"] is True
    assert all(row["总对话"] == 0 for row in result["daily_data"]["accounts"].values())
    assert (tmp_path / "reports" / "kst_daily_data.json").exists()


def test_parse_baidu_table_maps_accounts_and_numeric_fields():
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "1,001", "点击": "51", "消费": "301.17"},
        {"账户名称": "银康银屑02", "展现量": "2,002", "点击": "62", "消费": "402.28"},
        {"账户": "baidu-银康03", "展现": "3,003", "点击": "73", "花费": "503.39"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert parsed["errors"] == []
    assert parsed["accounts"]["银康01"]["展现"] == 1001
    assert parsed["accounts"]["银康银屑02"]["点击"] == 62
    assert parsed["accounts"]["银康03"]["消费"] == 503.39


def test_map_account_requires_exact_match_for_ningbo_suffix_accounts():
    config = {
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
            "宁波博润13": {"baidu_name": "宁波博润13", "aliases": ["宁波博润13"]},
        }
    }
    account_map = _build_account_map(config)

    assert _map_account("宁波博润12", account_map) == "宁波博润12"
    assert _map_account("宁波博润1", account_map) == "宁波博润1"
    assert _map_account(" 宁波博润1 ", account_map) == "宁波博润1"


def test_map_account_does_not_fallback_to_contains_match():
    config = {
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
        }
    }
    account_map = _build_account_map(config)

    assert _map_account("宁波博润12", account_map) is None


def test_parse_baidu_table_treats_unconfigured_suffix_account_as_unknown():
    config = {
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
        }
    }
    rows = [
        {"账户": "宁波博润12", "展现": "100", "点击": "10", "消费": "20.5"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert "宁波博润1" not in parsed["accounts"]
    assert parsed["unknown_accounts"][0]["account_name"] == "宁波博润12"


def test_parse_baidu_table_distinguishes_changsha_accounts_by_exact_name():
    config = {
        "accounts": {
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209", "aliases": ["竞网CS博润241209"]},
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304", "aliases": ["竞网CS博润240304"]},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218", "aliases": ["竞网CS博润251218"]},
        }
    }
    rows = [
        {"账户": "竞网CS博润241209", "展现": "100", "点击": "10", "消费": "20.5"},
        {"账户": "竞网CS博润240304", "展现": "200", "点击": "20", "消费": "30.5"},
        {"账户": "竞网CS博润251218", "展现": "300", "点击": "30", "消费": "40.5"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert parsed["errors"] == []
    assert parsed["accounts"]["竞网CS博润241209"]["展现"] == 100
    assert parsed["accounts"]["竞网CS博润240304"]["展现"] == 200
    assert parsed["accounts"]["竞网CS博润251218"]["展现"] == 300


def test_build_parse_debug_records_only_exact_account_matches():
    config = {
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        }
    }
    rows = [
        {"账户": " 宁波博润1 ", "展现": "100", "点击": "10", "消费": "20.5"},
        {"账户": "宁波博润12", "展现": "200", "点击": "20", "消费": "30.5"},
    ]

    debug = _build_parse_debug(rows, config, "dom")

    assert [item["matched_account"] for item in debug["account_match_details"]] == ["宁波博润1", "宁波博润12"]
    assert all(item["match_type"] == "exact" for item in debug["account_match_details"])


def test_browser_settings_default_to_chrome_without_edge_fallback():
    settings = get_browser_settings(
        {
            "browser_preference": "chrome",
            "browser_channel": "chrome",
            "chrome_executable_path": "C:/Program Files/Google/Chrome/Application/chrome.exe",
            "browser_profile_dir": "browser_profile/chrome",
            "browser_launch_mode": "cdp",
            "remote_debugging_port": 9222,
            "allow_edge_fallback": False,
        }
    )

    assert settings["browser_preference"] == "chrome"
    assert settings["browser_channel"] == "chrome"
    assert settings["chrome_executable_path"] == "C:/Program Files/Google/Chrome/Application/chrome.exe"
    assert settings["browser_profile_dir"] == "browser_profile/chrome"
    assert settings["browser_launch_mode"] == "connect_existing"
    assert settings["remote_debugging_port"] == 9222
    assert settings["allow_edge_fallback"] is False


def test_browser_settings_accept_nested_connect_existing_config():
    settings = get_browser_settings(
        {
            "browser": {
                "mode": "connect_existing",
                "cdp_endpoint": "http://127.0.0.1:9222",
                "prefer_existing_chrome": True,
                "allow_edge_fallback": False,
                "managed": {
                    "channel": "chrome",
                    "executable_path": "C:/Program Files/Google/Chrome/Application/chrome.exe",
                    "profile_dir": "browser_profile/chrome",
                    "headless": False,
                },
            }
        }
    )

    assert settings["mode"] == "connect_existing"
    assert settings["cdp_endpoint"] == "http://127.0.0.1:9222"
    assert settings["browser_channel"] == "chrome"
    assert settings["browser_profile_dir"] == "browser_profile/chrome"
    assert settings["allow_edge_fallback"] is False
    assert settings["silent_automation"] is True
    assert settings["window_state"] == "minimized"
    assert settings["show_on_manual_intervention"] is True
    assert settings["disable_password_manager"] is True


def test_browser_settings_accepts_silent_overrides():
    settings = get_browser_settings(
        {
            "browser": {
                "silent_automation": False,
                "window_state": "normal",
                "show_on_manual_intervention": False,
                "disable_password_manager": False,
            }
        }
    )

    assert settings["silent_automation"] is False
    assert settings["window_state"] == "normal"
    assert settings["show_on_manual_intervention"] is False
    assert settings["disable_password_manager"] is False


def test_prepare_automation_page_does_not_touch_window_in_silent_mode():
    from modules.browser_manager import prepare_automation_page

    class FakeContext:
        def new_cdp_session(self, page):
            raise AssertionError("silent automation must not activate the window through CDP")

    class FakePage:
        context = FakeContext()

        def bring_to_front(self):
            raise AssertionError("silent automation must not bring Chrome to front")

    prepare_automation_page(FakePage(), {"browser": {"silent_automation": True, "window_state": "minimized"}})


def test_prepare_automation_page_can_focus_when_silent_disabled():
    from modules.browser_manager import prepare_automation_page

    class FakePage:
        def __init__(self):
            self.front = False

        def bring_to_front(self):
            self.front = True

    page = FakePage()
    prepare_automation_page(page, {"browser": {"silent_automation": False}})

    assert page.front is True


def test_connect_existing_does_not_launch_managed_chrome_or_edge():
    class FakeChromium:
        def __init__(self):
            self.launch_persistent_context_called = False

        def connect_over_cdp(self, endpoint):
            assert endpoint == "http://127.0.0.1:9222"
            raise RuntimeError("connection refused")

        def launch_persistent_context(self, **kwargs):
            self.launch_persistent_context_called = True
            raise AssertionError("connect_existing must not launch a managed browser")

    class FakePlaywright:
        def __init__(self):
            self.chromium = FakeChromium()

    fake = FakePlaywright()
    config = {
        "browser": {
            "mode": "connect_existing",
            "cdp_endpoint": "http://127.0.0.1:9222",
            "allow_edge_fallback": False,
            "managed": {"channel": "chrome", "profile_dir": "browser_profile/chrome", "headless": False},
        }
    }

    try:
        connect_existing_chrome(fake, config)
    except BrowserLaunchError as exc:
        assert "remote-debugging-port=9222" in str(exc)
    else:
        raise AssertionError("connect_existing should fail clearly when CDP is unavailable")
    assert fake.chromium.launch_persistent_context_called is False


def test_connect_existing_help_mentions_running_chrome_blocks_debug_port():
    assert "Google Chrome" in CONNECT_EXISTING_HELP
    assert "chrome_debug" in CONNECT_EXISTING_HELP
    assert "不需要关闭日常 Chrome" in CONNECT_EXISTING_HELP
    assert "--remote-debugging-port=9222" in CONNECT_EXISTING_HELP
    assert "--start-minimized" in CONNECT_EXISTING_HELP
    assert "cas.baidu.com" in CONNECT_EXISTING_HELP
    assert "yingxiao.baidu.com" not in CONNECT_EXISTING_HELP


def test_default_chrome_startup_url_is_cas_login():
    import json

    from modules.browser_manager import get_browser_settings
    from modules.chrome_debug import DEFAULT_STARTUP_URL
    from modules.chrome_debug_launcher import START_URL

    settings = get_browser_settings({})
    example_config = json.loads((Path(__file__).resolve().parents[1] / "config.example.json").read_text(encoding="utf-8"))
    example_settings = get_browser_settings(example_config)

    assert "cas.baidu.com" in settings["startup_url"]
    assert "cas.baidu.com" in example_settings["startup_url"]
    assert "cas.baidu.com" in example_config["baidu"]["start_url"]
    assert "cas.baidu.com" in example_config["baidu"]["login_url"]
    assert "cas.baidu.com" in DEFAULT_STARTUP_URL
    assert "cas.baidu.com" in START_URL
    assert "yingxiao.baidu.com" not in settings["startup_url"]
    assert "yingxiao.baidu.com" not in example_settings["startup_url"]
    assert "yingxiao.baidu.com" not in example_config["browser"]["startup_url"]
    assert "qingge.baidu.com" not in example_config["baidu"]["login_url"]
    assert "yingxiao.baidu.com" not in DEFAULT_STARTUP_URL
    assert "yingxiao.baidu.com" not in START_URL


def test_select_context_repoints_legacy_yingxiao_page_to_cas():
    from modules.browser_manager import DEFAULT_BAIDU_START_URL, _select_context_and_page

    class FakePage:
        def __init__(self, url):
            self.url = url
            self.goto_calls = []
            self.front = False

        def goto(self, url, wait_until=None, timeout=None):
            self.goto_calls.append(url)
            self.url = url

        def bring_to_front(self):
            self.front = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

        def new_page(self):
            page = FakePage("about:blank")
            self.pages.append(page)
            return page

    class FakeBrowser:
        def __init__(self, contexts):
            self.contexts = contexts

    legacy_page = FakePage("https://yingxiao.baidu.com/home")
    context, page = _select_context_and_page(FakeBrowser([FakeContext([legacy_page])]), DEFAULT_BAIDU_START_URL)

    assert page is legacy_page
    assert context.pages == [legacy_page]
    assert legacy_page.goto_calls == [DEFAULT_BAIDU_START_URL]
    assert legacy_page.front is False


def test_select_context_keeps_existing_cc_report_page():
    from modules.browser_manager import DEFAULT_BAIDU_START_URL, _select_context_and_page

    class FakePage:
        def __init__(self, url):
            self.url = url
            self.goto_calls = []
            self.front = False

        def goto(self, url, wait_until=None, timeout=None):
            self.goto_calls.append(url)
            self.url = url

        def bring_to_front(self):
            self.front = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    class FakeBrowser:
        def __init__(self, contexts):
            self.contexts = contexts

    report_page = FakePage("https://cc.baidu.com/report")
    context, page = _select_context_and_page(FakeBrowser([FakeContext([report_page])]), DEFAULT_BAIDU_START_URL)

    assert page is report_page
    assert report_page.goto_calls == []
    assert report_page.front is False


def test_select_context_can_show_page_for_non_silent_mode():
    from modules.browser_manager import DEFAULT_BAIDU_START_URL, _select_context_and_page

    class FakePage:
        def __init__(self, url):
            self.url = url
            self.front = False

        def bring_to_front(self):
            self.front = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    class FakeBrowser:
        def __init__(self, contexts):
            self.contexts = contexts

    report_page = FakePage("https://cc.baidu.com/report")
    _select_context_and_page(FakeBrowser([FakeContext([report_page])]), DEFAULT_BAIDU_START_URL, silent=False)

    assert report_page.front is True


def test_cleanup_extra_tabs_keeps_baidu_page_and_limits_to_three():
    class FakePage:
        def __init__(self, url):
            self.url = url
            self.closed = False
            self.front = False

        def bring_to_front(self):
            self.front = True

        def close(self):
            self.closed = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    pages = [
        FakePage("https://old.example/1"),
        FakePage("https://old.example/2"),
        FakePage("https://old.example/3"),
        FakePage("https://cc.baidu.com/report"),
        FakePage("https://new.example/4"),
    ]
    context = FakeContext(pages)

    closed = cleanup_extra_tabs(context, pages[3], max_tabs=3)

    assert closed == ["https://old.example/1", "https://old.example/2"]
    assert pages[0].closed is True
    assert pages[1].closed is True
    assert pages[3].closed is False
    assert pages[3].front is False


def test_cleanup_extra_tabs_can_show_keep_page_for_non_silent_mode():
    class FakePage:
        def __init__(self, url):
            self.url = url
            self.closed = False
            self.front = False

        def bring_to_front(self):
            self.front = True

        def close(self):
            self.closed = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    pages = [FakePage("https://old.example/1"), FakePage("https://cc.baidu.com/report")]

    cleanup_extra_tabs(FakeContext(pages), pages[1], max_tabs=1, silent=False)

    assert pages[1].front is True


def test_cleanup_extra_tabs_does_nothing_when_page_count_is_three():
    class FakePage:
        def __init__(self, url):
            self.url = url
            self.closed = False

        def close(self):
            self.closed = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    pages = [FakePage("1"), FakePage("2"), FakePage("https://cc.baidu.com/report")]

    closed = cleanup_extra_tabs(FakeContext(pages), pages[2], max_tabs=3)

    assert closed == []
    assert all(page.closed is False for page in pages)


def test_cleanup_extra_tabs_limits_even_when_all_pages_are_baidu():
    class FakePage:
        def __init__(self, url):
            self.url = url
            self.closed = False

        def bring_to_front(self):
            pass

        def close(self):
            self.closed = True

    class FakeContext:
        def __init__(self, pages):
            self.pages = pages

    pages = [FakePage(f"https://cc.baidu.com/report?page={index}") for index in range(5)]

    closed = cleanup_extra_tabs(FakeContext(pages), pages[4], max_tabs=3)

    assert closed == ["https://cc.baidu.com/report?page=0", "https://cc.baidu.com/report?page=1"]
    assert [page.closed for page in pages] == [True, True, False, False, False]


def test_baidu_number_cleanup_and_report_validation():
    assert _parse_number(" ￥ 1,234.50 元 ") == 1234.5
    report = {
        "accounts": {
            "银康01": {"展现": 100, "点击": 10, "消费": 20.5},
            "银康银屑02": {"展现": 200, "点击": 20, "消费": 30},
            "银康03": {"展现": 300, "点击": 30, "消费": 40},
        }
    }

    km_accounts = list(_kunming_niu_runtime_config()["accounts"].keys())
    assert validate_baidu_report(report, required_accounts=km_accounts) == []

    bad_report = {"accounts": {"银康01": {"展现": "x", "点击": 1, "消费": 1}}}
    errors = validate_baidu_report(bad_report, required_accounts=km_accounts)
    assert any("账户数量不匹配" in error for error in errors)
    assert any("不是数字" in error for error in errors)


# ── baidu_parser 未知账户测试 ─────────────────────────────


def test_unknown_account_all_zero_goes_to_ignored():
    """未知账户 展现=0、点击=0、消费=0 → ignored_unknown_accounts。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知零账户", "展现": "0", "点击": "0", "消费": "0"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert "银康01" in parsed["accounts"]
    assert len(parsed["unknown_accounts"]) == 0
    assert len(parsed["ignored_unknown_accounts"]) == 1
    assert parsed["ignored_unknown_accounts"][0]["account_name"] == "未知零账户"
    assert parsed["ignored_unknown_accounts"][0]["展现"] == 0
    assert "已忽略" in parsed["ignored_unknown_accounts"][0]["reason"]


def test_unknown_account_with_impressions_goes_to_unknown():
    """未知账户 展现>0 → unknown_accounts。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知展现", "展现": "200", "点击": "0", "消费": "0"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知展现"
    assert parsed["unknown_accounts"][0]["展现"] == 200
    assert parsed["unknown_accounts"][0]["点击"] == 0
    assert len(parsed["ignored_unknown_accounts"]) == 0
    assert "未配置" in parsed["unknown_accounts"][0]["reason"]


def test_unknown_account_with_clicks_goes_to_unknown():
    """未知账户 点击>0 → unknown_accounts。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知点击", "展现": "0", "点击": "5", "消费": "0"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知点击"
    assert parsed["unknown_accounts"][0]["点击"] == 5


def test_unknown_account_with_cost_goes_to_unknown():
    """未知账户 消费>0 → unknown_accounts。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知消费", "展现": "0", "点击": "0", "消费": "10.5"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知消费"
    assert parsed["unknown_accounts"][0]["消费"] == 10.5


def test_known_accounts_still_in_accounts():
    """已配置账户仍然正常进入 accounts。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "银康银屑02", "展现": "200", "点击": "20", "消费": "30"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["accounts"]) == 2
    assert parsed["accounts"]["银康01"]["展现"] == 100
    assert parsed["accounts"]["银康银屑02"]["展现"] == 200
    assert len(parsed["unknown_accounts"]) == 0
    assert len(parsed["ignored_unknown_accounts"]) == 0


def test_unknown_accounts_do_not_affect_known_parsing():
    """unknown/ignored 不影响已配置账户解析。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知A", "展现": "999", "点击": "99", "消费": "99"},
        {"账户": "未知B", "展现": "0", "点击": "0", "消费": "0"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["accounts"]) == 1
    assert parsed["accounts"]["银康01"]["展现"] == 100
    assert len(parsed["unknown_accounts"]) == 1
    assert len(parsed["ignored_unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["展现"] == 999
    assert parsed["ignored_unknown_accounts"][0]["展现"] == 0


def test_unknown_accounts_not_in_errors():
    """未知账户不进入 errors。"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
        }
    }
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知有数据", "展现": "100", "点击": "10", "消费": "50"},
    ]

    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert not any("未知" in e for e in parsed["errors"]), "未知账户不应进入 errors"


def test_unknown_missing_impressions_goes_to_unknown():
    """未知账户展现字段缺失 → unknown_accounts，不进 ignored。"""
    config = {"accounts": {"银康01": {"baidu_name": "银康01"}}}
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知缺展现", "点击": "0", "消费": "0"},
    ]
    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知缺展现"
    assert len(parsed["ignored_unknown_accounts"]) == 0


def test_unknown_missing_clicks_goes_to_unknown():
    """未知账户点击字段缺失 → unknown_accounts。"""
    config = {"accounts": {"银康01": {"baidu_name": "银康01"}}}
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知缺点击", "展现": "0", "消费": "0"},
    ]
    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知缺点击"
    assert len(parsed["ignored_unknown_accounts"]) == 0


def test_unknown_missing_cost_goes_to_unknown():
    """未知账户消费字段缺失 → unknown_accounts。"""
    config = {"accounts": {"银康01": {"baidu_name": "银康01"}}}
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知缺消费", "展现": "0", "点击": "0"},
    ]
    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知缺消费"
    assert len(parsed["ignored_unknown_accounts"]) == 0


def test_unknown_non_numeric_field_goes_to_unknown():
    """未知账户字段为非数字 → unknown_accounts。"""
    config = {"accounts": {"银康01": {"baidu_name": "银康01"}}}
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "未知非数字", "展现": "N/A", "点击": "0", "消费": "0"},
    ]
    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 1
    assert parsed["unknown_accounts"][0]["account_name"] == "未知非数字"
    assert len(parsed["ignored_unknown_accounts"]) == 0


def test_unknown_all_zero_still_ignored():
    """原有未知账户 0/0/0 仍然进入 ignored_unknown_accounts。"""
    config = {"accounts": {"银康01": {"baidu_name": "银康01"}}}
    rows = [
        {"账户": "银康01", "展现": "100", "点击": "10", "消费": "50"},
        {"账户": "零账户", "展现": "0", "点击": "0", "消费": "0"},
    ]
    parsed = parse_baidu_table(rows, config)

    assert len(parsed["unknown_accounts"]) == 0
    assert len(parsed["ignored_unknown_accounts"]) == 1
    assert parsed["ignored_unknown_accounts"][0]["account_name"] == "零账户"


def test_extract_baidu_rows_from_visible_text_reports_missing_click_column():
    text = """
详细数据
自定义列
下载
重置列宽
账户
账户ID
展现
消费
点击率
平均点击价格
总计-3
-
6,791
2,794.83
5.26%
7.83
银康01
72828178
4,169
1,873.41
4.49%
10.02
银康银屑02
72828179
2,397
829.67
6.26%
5.53
baidu-银康03
81509165
225
91.75
8.89%
4.59
20条/页
"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03"},
        }
    }

    rows = extract_baidu_rows_from_visible_text(text)
    parsed = parse_baidu_table(rows, config)

    assert len(rows) == 4
    assert set(parsed["accounts"].keys()) == {"银康01", "银康银屑02", "银康03"}
    assert parsed["accounts"]["银康01"]["展现"] == 4169
    assert parsed["accounts"]["银康01"]["消费"] == 1873.41
    assert any("字段 点击" in error for error in parsed["errors"])


def test_extract_baidu_rows_from_visible_text_uses_data_width_when_group_header_has_no_cell():
    text = """
详细数据
账户
账户ID
展现
点击
消费
平均点击价格
落地页转化
一句话咨询量
留线索量
总计-3
-
7,990
1,113
2,989.19
2.69
14
1
竞网CS博润240304
53559272
720
58
186.85
3.22
0
0
竞网CS博润241209
61561000
1,568
174
1,779.96
10.23
5
0
竞网CS博润251218
77609531
5,702
881
1,022.38
1.16
9
1
20条/页
"""
    config = {
        "accounts": {
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304"},
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209"},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218"},
        }
    }

    rows = extract_baidu_rows_from_visible_text(text)
    parsed = parse_baidu_table(rows, config)

    assert len(rows) == 4
    assert parsed["errors"] == []
    assert parsed["accounts"]["竞网CS博润240304"]["展现"] == 720
    assert parsed["accounts"]["竞网CS博润240304"]["点击"] == 58
    assert parsed["accounts"]["竞网CS博润240304"]["消费"] == 186.85
    assert parsed["accounts"]["竞网CS博润241209"]["展现"] == 1568
    assert parsed["accounts"]["竞网CS博润241209"]["点击"] == 174
    assert parsed["accounts"]["竞网CS博润241209"]["消费"] == 1779.96
    assert parsed["accounts"]["竞网CS博润251218"]["展现"] == 5702
    assert parsed["accounts"]["竞网CS博润251218"]["点击"] == 881
    assert parsed["accounts"]["竞网CS博润251218"]["消费"] == 1022.38


def test_extract_baidu_rows_from_visible_text_rejects_unverifiable_data_width():
    text = """
账户
账户ID
展现
点击
消费
平均点击价格
落地页转化
一句话咨询量
留线索量
总计-3
-
7,990
1,113
2,989.19
2.69
14
1
竞网CS博润240304
53559272
720
58
186.85
3.22
0
0
竞网CS博润241209
61561000
1,568
174
1,779.96
10.23
5
20条/页
"""

    assert extract_baidu_rows_from_visible_text(text) == []


def test_extract_selected_date_from_baidu_visible_text():
    text = """
账户：
全部已绑定账户
2026/05/06
推广设备：全部
详细数据
"""

    assert _extract_selected_date_from_text(text) == "2026-05-06"


def test_classify_baidu_page_types_from_url_and_visible_text():
    login = classify_baidu_page("https://yingxiao.baidu.com/", "登录 百度营销 请输入账号 密码")
    assert login["login_status"] == "not_logged_in"
    assert login["page_type"] == "未登录页"

    public_home = classify_baidu_page("https://yingxiao.baidu.com/home", "首页 百度伴飞 登录 注册 立即推广")
    assert public_home["login_status"] == "not_logged_in"
    assert public_home["page_type"] == "未登录页"

    cas_login = classify_baidu_page("https://cas.baidu.com/?tpl=www2", "百度营销账号 百度账号 扫码登录 注册忘记密码 搜索推广")
    assert cas_login["login_status"] == "not_logged_in"
    assert cas_login["page_type"] == "未登录页"
    assert is_search_promotion_overview(cas_login) is False

    home = classify_baidu_page("https://yingxiao.baidu.com/home", "百度营销 客户中心 首页 进入")
    assert home["login_status"] == "logged_in"
    assert home["page_type"] == "百度营销首页"

    report = classify_baidu_page("https://yingxiao.baidu.com/report", "数据报告 数据概览 搜索推广 详细数据 展现 点击 消费")
    assert report["page_type"] == "搜索推广"
    assert report["signals"]["has_data_report"] is True
    assert report["signals"]["has_data_overview"] is True
    assert report["signals"]["has_search_promotion"] is True

    cc_report = classify_baidu_page("https://cc.baidu.com/report", "数据报告 数据概览 全部推广产品")
    assert cc_report["login_status"] == "logged_in"
    assert cc_report["page_type"] == "数据概览"
    assert cc_report["signals"]["has_data_report"] is True


def test_search_promotion_overview_detection_requires_overview_and_search_signals():
    search = classify_baidu_page("https://cc.baidu.com/report", "数据报告 数据概览 搜索推广 展现 点击 消费")
    cc_report_search = classify_baidu_page("https://cc.baidu.com/report", "数据报告 搜索推广 详细数据 账户 展现 点击 消费")
    overview_only = classify_baidu_page("https://cc.baidu.com/report", "数据报告 数据概览 信息流推广")
    old_yingxiao = classify_baidu_page("https://yingxiao.baidu.com/report", "数据报告 数据概览 搜索推广 展现 点击 消费")

    assert is_search_promotion_overview(search) is True
    assert is_search_promotion_overview(cc_report_search) is True
    assert is_search_promotion_overview(overview_only) is False
    assert is_search_promotion_overview(old_yingxiao) is False


def test_validate_overview_ready_checks_today_accounts_and_headers():
    text = """
数据报告
搜索推广
账户：
全部已绑定账户
2026/05/07
详细数据
账户
账户ID
展现
点击
消费
银康01
银康银屑02
baidu-银康03
"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01", "aliases": ["银康01"]},
            "银康银屑02": {"baidu_name": "银康银屑02", "aliases": ["银康银屑02"]},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    report = validate_overview_ready(text, "2026-05-07", config)

    assert report["passed"] is True
    assert report["checks"]["date_is_today"] is True
    assert all(report["accounts"].values())
    assert all(report["fields"].values())


def test_validate_overview_ready_rejects_wrong_date():
    report = validate_overview_ready("数据报告 搜索推广 2026/05/06 账户 展现 点击 消费 银康01 银康银屑02 baidu-银康03", "2026-05-07", {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03"]},
        }
    })

    assert report["passed"] is False
    assert report["checks"]["date_is_today"] is False
    assert any("日期不是今天" in error for error in report["errors"])


def test_overview_text_has_account_table_requires_header_and_account():
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    loading_text = "数据报告 搜索推广 2026/05/07 点击（次） - 展现（次） - 消费（元） - 详细数据"
    loaded_text = "数据报告 搜索推广 2026/05/07 详细数据 账户 账户ID 展现 点击 消费 银康01"

    assert overview_text_has_account_table(loading_text, config) is False
    assert overview_text_has_account_table(loaded_text, config) is True


def test_qingge_login_page_should_redirect_to_cas_form():
    assert should_open_cas_login("https://qingge.baidu.com/login") is True
    assert should_open_cas_login("https://yingxiao.baidu.com/") is True
    assert should_open_cas_login("https://cas.baidu.com/?tpl=www2") is False


def test_search_promotion_overview_requires_report_url_not_homepage():
    homepage = classify_baidu_page("https://cc.baidu.com/homepage", "首页 数据概览 搜索推广 账户 展现 点击 消费 银康01")
    report = classify_baidu_page("https://cc.baidu.com/report", "数据报告 搜索推广 详细数据 账户 展现 点击 消费 银康01")

    assert is_search_promotion_overview(homepage) is False
    assert is_search_promotion_overview(report) is True


def test_build_baidu_auto_report_from_overview_visible_text():
    today = date.today()
    visible_date = today.strftime("%Y/%m/%d")
    expected_date = today.isoformat()
    text = f"""
首页
数据报告
搜索推广
账户：
全部已绑定账户
{visible_date}
详细数据
账户
账户ID
展现
点击
消费
平均点击价格
点击率
搜索推广余额
总计-3
-
3,970
363
3,048.96
8.4
9.14%
-
银康01
72828178
1,958
206
2,339.88
11.36
10.52%
29,637.05
银康银屑02
72828179
1,849
141
693.96
4.92
7.63%
29,637.05
baidu-银康03
81509165
163
16
15.12
0.95
9.82%
29,637.05
20条/页
"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    report = build_baidu_auto_report_from_visible_text(text, config, "15点", visible_text_path="reports/baidu_visible_text.txt")

    assert report["date"] == expected_date
    assert report["period"] == "15点"
    assert report["parse_source"] == "visible_text"
    assert report["accounts"]["银康01"]["展现"] == 1958
    assert report["accounts"]["银康银屑02"]["点击"] == 141
    assert report["accounts"]["银康03"]["source_account"] == "baidu-银康03"
    assert report["accounts"]["银康03"]["消费"] == 15.12
    assert report["errors"] == []


def test_build_baidu_daily_report_accepts_specified_non_today_date():
    text = """
数据报告
搜索推广
账户：
全部已绑定账户
2026/05/07
详细数据
账户
账户ID
展现
点击
消费
平均点击价格
点击率
搜索推广余额
总计-3
-
309
30
127.7
4.26
9.71%
-
银康01
72828178
172
20
75.68
3.78
11.63%
29,507.53
银康银屑02
72828179
101
8
50.96
6.37
7.92%
29,507.53
baidu-银康03
81509165
36
2
1.06
0.53
5.56%
29,507.53
20条/页
"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    report = build_baidu_daily_report_from_visible_text(text, config, "2026-05-07", "dump.txt")

    assert report["date"] == "2026-05-07"
    assert report["source"] == "baidu_daily_report"
    assert report["accounts"]["银康01"]["点击"] == 20
    assert report["accounts"]["银康03"]["source_account"] == "baidu-银康03"
    assert report["errors"] == []
    assert report["self_check"]["selected_date_matches_target"] is True


def test_build_baidu_daily_report_rejects_page_date_mismatch():
    report = build_baidu_daily_report_from_visible_text("数据报告 搜索推广 2026/05/08", {"accounts": {}}, "2026-05-07")

    assert report["date"] == "2026-05-07"
    assert any("百度日报页面日期不匹配" in error for error in report["errors"])


def test_build_baidu_daily_report_rejects_non_search_promotion_data():
    text = """
数据报告
信息流推广
2026/05/07
账户
展现
点击
消费
银康01
100
10
50
银康银屑02
80
8
40
baidu-银康03
30
3
10
"""
    config = {
        "accounts": {
            "银康01": {"baidu_name": "银康01"},
            "银康银屑02": {"baidu_name": "银康银屑02"},
            "银康03": {"baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
        }
    }

    report = build_baidu_daily_report_from_visible_text(text, config, "2026-05-07")

    assert any("当前百度日报页面不是搜索推广数据" in error for error in report["errors"])


def test_default_daily_date_is_yesterday():
    today = date(2026, 5, 8)

    assert default_daily_date(today) == "2026-05-07"


def test_load_project_credentials_from_local_file(tmp_path):
    credentials = tmp_path / "credentials.local.json"
    credentials.write_text(
        """
{
  "baidu": {
    "yunnan_yinkang": {
      "username": "demo-user",
      "password": "demo-pass"
    }
  }
}
""",
        encoding="utf-8",
    )

    item = load_project_credentials(tmp_path, {"credentials_path": "credentials.local.json"}, "baidu", "yunnan_yinkang")

    assert item["username"] == "demo-user"
    assert item["password"] == "demo-pass"


def test_validate_baidu_account_data_accepts_three_standard_accounts(tmp_path):
    source = tmp_path / "baidu_account_data.json"
    output = tmp_path / "baidu_validate_report.json"
    source.write_text(
        """
{
  "date": "2026-05-07",
  "period": "15",
  "accounts": {
    "银康01": {"source_account": "银康01", "展现": 4169, "点击": 187, "消费": 1873.41},
    "银康银屑02": {"source_account": "银康银屑02", "展现": 2397, "点击": 150, "消费": 829.67},
    "银康03": {"source_account": "baidu-银康03", "展现": 225, "点击": 20, "消费": 91.75}
  },
  "errors": []
}
""",
        encoding="utf-8",
    )

    km_accounts = list(_kunming_niu_runtime_config()["accounts"].keys())
    report = validate_baidu_account_data(source, output, expected_accounts=km_accounts)

    assert report["passed"] is True
    assert report["checks"]["exactly_three_standard_accounts"] is True
    assert report["checks"]["all_source_accounts_present"] is True
    assert report["checks"]["impressions_and_clicks_are_integers"] is True
    assert output.exists()


def test_validate_baidu_account_data_rejects_visible_dump_date_mismatch(tmp_path):
    dump = tmp_path / "baidu_page_text_dump.txt"
    dump.write_text("账户：\n全部已绑定账户\n2026/05/06\n推广设备：全部\n", encoding="utf-8")
    source = tmp_path / "baidu_account_data.json"
    output = tmp_path / "baidu_validate_report.json"
    source.write_text(
        f"""
{{
  "date": "2026-05-07",
  "period": "15",
  "accounts": {{
    "银康01": {{"source_account": "银康01", "展现": 4169, "点击": 187, "消费": 1873.41}},
    "银康银屑02": {{"source_account": "银康银屑02", "展现": 2397, "点击": 150, "消费": 829.67}},
    "银康03": {{"source_account": "baidu-银康03", "展现": 225, "点击": 20, "消费": 91.75}}
  }},
  "exceptions": [
    {{"type": "visible_text_dump", "path": "{str(dump).replace('\\', '\\\\')}"}}
  ],
  "errors": []
}}
""",
        encoding="utf-8",
    )

    report = validate_baidu_account_data(source, output)

    assert report["passed"] is False
    assert report["checks"]["source_date_matches_visible_dump"] is False
    assert any("页面实际日期" in error for error in report["errors"])


def test_baidu_debug_artifacts_do_not_write_screenshot_by_default(tmp_path):
    class FakePage:
        def __init__(self):
            self.screenshot_called = False

        def content(self):
            return "<html>debug</html>"

        def screenshot(self, **kwargs):
            self.screenshot_called = True

    page = FakePage()
    report = {"exceptions": []}

    _write_debug_artifacts(tmp_path, page, report)

    assert (tmp_path / "reports" / "baidu_debug.html").exists()
    assert not (tmp_path / "reports" / "baidu_debug.png").exists()
    assert page.screenshot_called is False
    assert [item["type"] for item in report["exceptions"]] == ["debug_html"]


def test_merge_hourly_data_requires_three_accounts_and_strict_field_types():
    baidu = {
        "date": "2026-05-07",
        "period": "15",
        "accounts": {
            "银康01": {"展现": 4169, "点击": 187, "消费": 1873.41},
            "银康银屑02": {"展现": 2397, "点击": 150, "消费": 829.67},
            "银康03": {"展现": 225, "点击": 20, "消费": 91.75},
        },
    }
    kst = {
        "date": "2026-05-07",
        "period": "15点",
        "accounts": {
            "银康01": {"总对话": 8, "有效对话": 4, "一般有效": 2, "有效转潜": 1, "总转潜": 2},
            "银康银屑02": {"总对话": 9, "有效对话": 2, "一般有效": 1, "有效转潜": 1, "总转潜": 1},
            "银康03": {"总对话": 3, "有效对话": 1, "一般有效": 0, "有效转潜": 1, "总转潜": 2},
        },
    }

    km_accounts = list(_kunming_niu_runtime_config()["accounts"].keys())
    merged = build_merged_hourly_data(baidu, kst, "15点", required_accounts=km_accounts)

    assert merged["date"] == "2026-05-07"
    assert merged["period"] == "15点"
    assert list(merged["accounts"].keys()) == ["银康01", "银康银屑02", "银康03"]
    assert merged["accounts"]["银康03"]["消费"] == 91.75
    assert merged["accounts"]["银康银屑02"]["总对话"] == 9
    assert validate_merged_hourly_data(merged, baidu, kst, required_accounts=km_accounts) == []

    merged["accounts"]["银康01"]["点击"] = 187.5
    errors = validate_merged_hourly_data(merged, baidu, kst, required_accounts=km_accounts)
    assert any("点击 必须是整数" in error for error in errors)


def test_merge_daily_data_combines_baidu_and_kst_daily_fields():
    baidu = {
        "date": "2026-05-07",
        "accounts": {
            "银康01": {"展现": 1958, "点击": 206, "消费": 2339.88},
            "银康银屑02": {"展现": 1849, "点击": 141, "消费": 693.96},
            "银康03": {"展现": 163, "点击": 16, "消费": 15.12},
        },
    }
    kst = {
        "date": "2026-05-07",
        "accounts": {
            "银康01": {"总对话": 11, "有效对话": 1, "无效对话": 10, "一般有效对话": 0, "有效转潜": 0, "总转潜": 1},
            "银康银屑02": {"总对话": 4, "有效对话": 0, "无效对话": 4, "一般有效对话": 0, "有效转潜": 0, "总转潜": 0},
            "银康03": {"总对话": 0, "有效对话": 0, "无效对话": 0, "一般有效对话": 0, "有效转潜": 0, "总转潜": 0},
        },
    }

    km_accounts = list(_kunming_niu_runtime_config()["accounts"].keys())
    merged = build_merged_daily_data(baidu, kst, required_accounts=km_accounts)

    assert merged["date"] == "2026-05-07"
    assert merged["source"]["baidu"] == "reports/baidu_daily_data.json"
    assert merged["source"]["kst"] == "reports/kst_daily_data.json"
    assert merged["accounts"]["银康01"]["展现"] == 1958
    assert merged["accounts"]["银康01"]["有效对话"] == 1
    assert merged["accounts"]["银康01"]["无效对话"] == 10
    assert merged["accounts"]["银康03"]["总对话"] == 0
    assert validate_merged_daily_data(merged, baidu, kst, required_accounts=km_accounts) == []

    merged["accounts"]["银康01"]["无效对话"] = 9
    errors = validate_merged_daily_data(merged, baidu, kst, required_accounts=km_accounts)
    assert any("有效、一般与无效对话未覆盖总对话" in error for error in errors)


def test_write_merged_hourly_data_backs_up_writes_and_verifies(tmp_path):
    import logging
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A1"] = "每日时段统计数据"
    ws.merge_cells("A1:E1")
    ws["F1"] = "银康01"
    ws.merge_cells("F1:M1")
    ws["N1"] = "银康银屑02"
    ws.merge_cells("N1:U1")
    ws["V1"] = "银康03"
    ws.merge_cells("V1:AC1")
    ws["A2"] = "日期"
    ws["B2"] = "时段"
    headers = ["展现", "点击", "消费", "总对话", "有效对话", "一般有效", "有效转潜", "总转潜"]
    for offset, header in enumerate(headers):
        ws.cell(row=2, column=6 + offset).value = header
        ws.cell(row=2, column=14 + offset).value = header
        ws.cell(row=2, column=22 + offset).value = header
    ws["A3"] = "2026-05-07"
    ws.merge_cells("A3:A6")
    ws["B3"] = "昨日数据"
    ws["B4"] = "11点"
    ws["B5"] = "3点"
    ws["B6"] = "6点"
    ws.auto_filter.ref = "A2:AC6"
    notes_ws = wb.create_sheet("说明")
    notes_ws["A1"] = "无关工作表"
    excel_path = tmp_path / "target.xlsx"
    wb.save(excel_path)

    reports = tmp_path / "reports"
    reports.mkdir()
    (reports / "merged_hourly_data.json").write_text(
        """
{
  "date": "2026-05-07",
  "period": "15点",
  "source": {"baidu": "reports/baidu_account_data.json", "kst": "reports/kst_dialog_data.json"},
  "accounts": {
    "银康01": {"展现": 4169, "点击": 187, "消费": 1873.41, "总对话": 8, "有效对话": 4, "一般有效": 2, "有效转潜": 1, "总转潜": 2},
    "银康银屑02": {"展现": 2397, "点击": 150, "消费": 829.67, "总对话": 9, "有效对话": 2, "一般有效": 1, "有效转潜": 1, "总转潜": 1},
    "银康03": {"展现": 225, "点击": 20, "消费": 91.75, "总对话": 3, "有效对话": 1, "一般有效": 0, "有效转潜": 1, "总转潜": 2}
  }
}
""",
        encoding="utf-8",
    )
    config = {
        "excel_path": str(excel_path),
        "sheet_name": "时段数据",
        "accounts": {
            "银康01": {"aliases": ["银康01"], "excel_name": "银康01", "baidu_name": "银康01"},
            "银康银屑02": {"aliases": ["银康银屑02"], "excel_name": "银康银屑02", "baidu_name": "银康银屑02"},
            "银康03": {"aliases": ["银康03", "baidu-银康03"], "excel_name": "银康03", "baidu_name": "baidu-银康03"},
        },
        "field_aliases": {
            "日期": ["日期"],
            "时段": ["时段"],
            "展现": ["展现"],
            "点击": ["点击"],
            "消费": ["消费"],
            "总对话": ["总对话"],
            "有效对话": ["有效对话"],
            "一般有效": ["一般有效"],
            "有效转潜": ["有效转潜"],
            "总转潜": ["总转潜"],
        },
    }

    report = write_merged_hourly_data(config, tmp_path, logging.getLogger("test"), "15点")

    assert report["errors"] == []
    assert report["self_check"]["backup_created"] is True
    assert report["self_check"]["verification_passed"] is True
    assert report["overwrite_summary"]["overwrite_count"] == 0
    assert len(report["writes"]) == 24
    assert (tmp_path / "reports" / "write_report.json").exists()
    assert (tmp_path / "backups").exists()

    verify_wb = load_workbook(excel_path, data_only=False, read_only=False)
    verify_ws = verify_wb["时段数据"]
    assert verify_ws["F5"].value == 4169
    assert verify_ws["H5"].value == 1873.41
    assert verify_ws["V5"].value == 225
    assert verify_ws["AC5"].value == 2
    assert verify_ws.auto_filter.ref == "A2:AC6"
    assert verify_wb["说明"].auto_filter.ref is None


def test_write_merged_daily_data_backs_up_writes_allowed_fields_only(tmp_path):
    import logging
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "百度"
    ws["A1"] = "日期"
    ws["P1"] = "银康01"
    ws.merge_cells("P1:AB1")
    ws["AC1"] = "银康银屑02"
    ws.merge_cells("AC1:AO1")
    ws["AP1"] = "银康03"
    ws.merge_cells("AP1:BB1")
    headers = ["展现", "点击", "消费", "acp", "总对话", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜", "预约", "就诊", "转潜成本"]
    for offset, header in enumerate(headers):
        ws.cell(row=2, column=16 + offset).value = header
        ws.cell(row=2, column=29 + offset).value = header
        ws.cell(row=2, column=42 + offset).value = header
    ws["A3"] = "2026-05-07"
    ws["T3"] = 999
    ws["Z3"] = 3
    ws["AA3"] = 2
    ws.auto_filter.ref = "A2:BB3"
    excel_path = tmp_path / "daily.xlsx"
    wb.save(excel_path)

    reports = tmp_path / "reports"
    reports.mkdir()
    (reports / "merged_daily_data.json").write_text(
        """
{
  "date": "2026-05-07",
  "source": {"baidu": "reports/baidu_daily_data.json", "kst": "reports/kst_daily_data.json"},
  "accounts": {
    "银康01": {"展现": 1958, "点击": 206, "消费": 2339.88, "总对话": 11, "有效对话": 1, "无效对话": 10, "一般有效对话": 0, "有效转潜": 0, "总转潜": 1},
    "银康银屑02": {"展现": 1849, "点击": 141, "消费": 693.96, "总对话": 4, "有效对话": 0, "无效对话": 4, "一般有效对话": 0, "有效转潜": 0, "总转潜": 0},
    "银康03": {"展现": 163, "点击": 16, "消费": 15.12, "总对话": 0, "有效对话": 0, "无效对话": 0, "一般有效对话": 0, "有效转潜": 0, "总转潜": 0}
  }
}
""",
        encoding="utf-8",
    )

    report = write_merged_daily_data(
        config={
            "excel_path": str(excel_path),
            "accounts": {
                "银康01": {"excel_name": "银康01", "baidu_name": "银康01", "aliases": ["银康01"]},
                "银康银屑02": {"excel_name": "银康银屑02", "baidu_name": "银康银屑02", "aliases": ["银康银屑02"]},
                "银康03": {"excel_name": "银康03", "baidu_name": "baidu-银康03", "aliases": ["银康03", "baidu-银康03"]},
            },
        },
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date="2026-05-07",
    )

    assert report["errors"] == []
    assert report["self_check"]["verification_passed"] is True
    assert len(report["writes"]) == 24
    assert all(item["field"] not in {"总对话", "预约", "就诊"} for item in report["writes"])
    assert (tmp_path / "reports" / "daily_write_report.json").exists()
    assert report["backup_path"]
    wb2 = load_workbook(excel_path, data_only=False)
    ws2 = wb2["百度"]
    assert ws2["P3"].value == 1958
    assert ws2["U3"].value == 1
    assert ws2["V3"].value == 10
    assert ws2["T3"].value == 999
    assert ws2["Z3"].value == 3
    assert ws2["AA3"].value == 2
    assert ws2.auto_filter.ref == "A2:BB3"


def test_write_merged_hourly_data_reports_overwrites(tmp_path):
    import logging
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A1"] = "每日时段统计数据"
    ws.merge_cells("A1:E1")
    ws["F1"] = "银康01"
    ws.merge_cells("F1:M1")
    ws["N1"] = "银康银屑02"
    ws.merge_cells("N1:U1")
    ws["V1"] = "银康03"
    ws.merge_cells("V1:AC1")
    ws["A2"] = "日期"
    ws["B2"] = "时段"
    headers = ["展现", "点击", "消费", "总对话", "有效对话", "一般有效", "有效转潜", "总转潜"]
    for offset, header in enumerate(headers):
        ws.cell(row=2, column=6 + offset).value = header
        ws.cell(row=2, column=14 + offset).value = header
        ws.cell(row=2, column=22 + offset).value = header
    ws["A3"] = "2026-05-07"
    ws.merge_cells("A3:A6")
    ws["B5"] = "3点"
    ws["F5"] = 999
    excel_path = tmp_path / "target.xlsx"
    wb.save(excel_path)

    (tmp_path / "reports").mkdir()
    (tmp_path / "reports" / "merged_hourly_data.json").write_text(
        """
{
  "date": "2026-05-07",
  "period": "15点",
  "accounts": {
    "银康01": {"展现": 4169, "点击": 187, "消费": 1873.41, "总对话": 8, "有效对话": 4, "一般有效": 2, "有效转潜": 1, "总转潜": 2},
    "银康银屑02": {"展现": 2397, "点击": 150, "消费": 829.67, "总对话": 9, "有效对话": 2, "一般有效": 1, "有效转潜": 1, "总转潜": 1},
    "银康03": {"展现": 225, "点击": 20, "消费": 91.75, "总对话": 3, "有效对话": 1, "一般有效": 0, "有效转潜": 1, "总转潜": 2}
  }
}
""",
        encoding="utf-8",
    )
    config = {
        "excel_path": str(excel_path),
        "sheet_name": "时段数据",
        "accounts": {
            "银康01": {"aliases": ["银康01"], "excel_name": "银康01", "baidu_name": "银康01"},
            "银康银屑02": {"aliases": ["银康银屑02"], "excel_name": "银康银屑02", "baidu_name": "银康银屑02"},
            "银康03": {"aliases": ["银康03", "baidu-银康03"], "excel_name": "银康03", "baidu_name": "baidu-银康03"},
        },
        "field_aliases": {
            "日期": ["日期"],
            "时段": ["时段"],
            "展现": ["展现"],
            "点击": ["点击"],
            "消费": ["消费"],
            "总对话": ["总对话"],
            "有效对话": ["有效对话"],
            "一般有效": ["一般有效"],
            "有效转潜": ["有效转潜"],
            "总转潜": ["总转潜"],
        },
    }

    report = write_merged_hourly_data(config, tmp_path, logging.getLogger("test"), "15点")

    assert report["errors"] == []
    assert report["overwrite_summary"]["overwrite_count"] == 1
    assert report["overwrite_summary"]["items"][0]["cell"] == "F5"
    assert report["overwrite_summary"]["items"][0]["old_value"] == 999


def test_run_pipeline_stops_on_failed_baidu_step(tmp_path):
    import logging

    def failed_baidu(**kwargs):
        return {"errors": ["未读取到三个百度账户"], "date": "2026-05-07", "period": "15点"}

    report = run_half_auto_pipeline(
        config={"excel_path": "target.xlsx"},
        root=tmp_path,
        logger=logging.getLogger("test"),
        period="15点",
        kst_file=tmp_path / "kst.xlsx",
        fetch_baidu_func=failed_baidu,
        parse_kst_func=lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("不应继续解析快商通")),
    )

    assert report["passed"] is False
    assert report["failed_step"] == "fetch-baidu-auto"
    assert len(report["steps"]) == 1
    assert report["steps"][0]["passed"] is False
    assert (tmp_path / "reports" / "final_run_report.json").exists()


def test_pipeline_defaults_to_resilient_baidu_fetchers():
    from modules.baidu_data_source import fetch_baidu_resilient_daily, fetch_baidu_resilient_hourly

    hourly_signature = inspect.signature(run_half_auto_pipeline)
    daily_signature = inspect.signature(run_daily_pipeline)

    assert hourly_signature.parameters["fetch_baidu_func"].default is fetch_baidu_resilient_hourly
    assert daily_signature.parameters["fetch_baidu_func"].default is fetch_baidu_resilient_daily


def test_task_stop_gate_keeps_the_first_atomic_decision(tmp_path):
    from modules.task_stop_gate import (
        claim_excel_write,
        pipeline_exit_code,
        read_task_stop_decision,
        request_task_stop,
    )

    stop_first = tmp_path / "stop-first.gate"
    assert request_task_stop(stop_first) is True
    assert claim_excel_write(stop_first) is False
    assert read_task_stop_decision(stop_first) == "cancel"

    excel_first = tmp_path / "excel-first.gate"
    assert claim_excel_write(excel_first) is True
    assert request_task_stop(excel_first) is False
    assert read_task_stop_decision(excel_first) == "excel"
    assert pipeline_exit_code({"passed": False, "cancelled": True}) == 130
    assert pipeline_exit_code({"passed": False, "cancelled": False}) == 1
    assert pipeline_exit_code({"passed": True, "cancelled": False}) == 0


@pytest.mark.parametrize("task", ["hourly", "daily"])
def test_run_pipeline_honors_stop_request_before_excel(tmp_path, monkeypatch, task):
    import logging

    from modules.task_stop_gate import STOP_GATE_ENV, request_task_stop

    gate_path = tmp_path / "reports" / ".test-stop.gate"
    monkeypatch.setenv(STOP_GATE_ENV, str(gate_path))
    export_file = tmp_path / "kst.xlsx"
    export_file.write_text("placeholder", encoding="utf-8")
    write_calls = []

    def ok_baidu(**_kwargs):
        return {"date": "2026-07-19", "period": "15点", "accounts": {}, "errors": []}

    def ok_parse(*_args, **_kwargs):
        return {"parse_report": {"passed": True, "errors": []}, "outputs": {}}

    def request_stop_during_merge(**_kwargs):
        assert request_task_stop(gate_path) is True
        return {
            "merged": {"date": "2026-07-19", "period": "15点"},
            "validate_report": {"passed": True, "errors": []},
            "outputs": {},
        }

    def forbidden_write(**_kwargs):
        write_calls.append(True)
        raise AssertionError("停止请求生效后不应进入 Excel 写入")

    common = {
        "config": {"project_id": "demo", "project_name": "演示项目", "excel_path": "target.xlsx"},
        "root": tmp_path,
        "logger": logging.getLogger(f"safe-stop-{task}"),
        "kst_file": export_file,
        "fetch_baidu_func": ok_baidu,
        "parse_kst_func": ok_parse,
        "merge_func": request_stop_during_merge,
        "write_func": forbidden_write,
    }
    if task == "hourly":
        report = run_half_auto_pipeline(period="15点", assume_yes=True, **common)
    else:
        report = run_daily_pipeline(target_date="2026-07-19", **common)

    assert report["passed"] is False
    assert report["cancelled"] is True
    assert report["failed_step"] == "cancelled-before-excel"
    assert write_calls == []


def test_run_pipeline_reports_success_summary(tmp_path):
    class Logger:
        def __init__(self):
            self.messages = []

        def info(self, message, *args):
            self.messages.append(message % args if args else message)

        def error(self, message, *args):
            self.messages.append(message % args if args else message)

    logger = Logger()

    kst_file = tmp_path / "kst.xlsx"
    kst_file.write_text("placeholder", encoding="utf-8")
    excel_file = tmp_path / "target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")

    def ok_baidu(**kwargs):
        return {
            "date": "2026-05-07",
            "period": "15",
            "accounts": {
                "银康01": {"展现": 1, "点击": 1, "消费": 1.0},
                "银康银屑02": {"展现": 2, "点击": 2, "消费": 2.0},
                "银康03": {"展现": 3, "点击": 3, "消费": 3.0},
            },
            "data_source": "browser_fallback",
            "api_attempts": 2,
            "fallback_reason": "network_error",
            "self_heal_actions": ["network_retry"],
            "errors": [],
        }

    def ok_kst(export_file, config, root, period):
        return {
            "parse_report": {"passed": True, "errors": []},
            "dialog_data": {"export_file": str(export_file)},
            "outputs": {
                "dialog_data": str(root / "reports" / "kst_dialog_data.json"),
                "parse_report": str(root / "reports" / "kst_parse_report.json"),
            },
        }

    def ok_merge(**kwargs):
        return {
            "merged": {"date": "2026-05-07", "period": "15点"},
            "validate_report": {"passed": True, "errors": []},
            "outputs": {
                "merged": str(tmp_path / "reports" / "merged_hourly_data.json"),
                "validate_report": str(tmp_path / "reports" / "merge_validate_report.json"),
            },
        }

    def ok_write(**kwargs):
        return {
            "date": "2026-05-07",
            "period": "15点",
            "excel_path": str(excel_file),
            "backup_path": str(tmp_path / "backups" / "target_backup.xlsx"),
            "writes": [{"account": "银康01", "field": "展现", "cell": "P5", "verified": True}],
            "self_check": {"verification_passed": True},
            "errors": [],
        }

    report = run_half_auto_pipeline(
        config={"project_id": "demo", "project_name": "演示项目", "excel_path": str(excel_file)},
        root=tmp_path,
        logger=logger,
        period="15点",
        kst_file=kst_file,
        assume_yes=True,
        fetch_baidu_func=ok_baidu,
        parse_kst_func=ok_kst,
        merge_func=ok_merge,
        write_func=ok_write,
    )

    assert report["passed"] is True
    assert report["project_id"] == "demo"
    assert report["project_name"] == "演示项目"
    assert report["date"] == "2026-05-07"
    assert report["period"] == "15点"
    assert report["excel_path"] == str(excel_file)
    assert report["kst_export_file"] == str(kst_file)
    assert report["baidu_source_ok"] is True
    assert report["data_source"] == "browser_fallback"
    assert report["api_attempts"] == 2
    assert report["fallback_reason"] == "network_error"
    assert report["self_heal_actions"] == ["network_retry"]
    assert "[实际来源] 百度数据：浏览器降级" in logger.messages
    assert "成功" in report["summary_text"]
    assert report["target_sheet"] == "时段数据"
    assert report["kst_export"]["file_name"] == "kst.xlsx"
    assert report["write_summary"]["write_count"] == 1
    assert [step["name"] for step in report["steps"]] == ["fetch-baidu-auto", "parse-kst-export", "merge-data", "write-excel"]


def test_run_pipeline_uses_latest_kst_export_when_file_is_omitted(tmp_path):
    import logging
    import os
    import time

    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    older = export_dir / "older.csv"
    latest = export_dir / "latest.xlsx"
    older.write_text("old", encoding="utf-8")
    latest.write_text("new", encoding="utf-8")
    now = time.time()
    os.utime(older, (now - 20 * 60, now - 20 * 60))
    os.utime(latest, (now - 10 * 60, now - 10 * 60))
    excel_file = tmp_path / "target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")
    seen = {}

    def ok_baidu(**kwargs):
        return {"date": "2026-05-07", "period": "15", "accounts": {}, "errors": []}

    def ok_kst(export_file, config, root, period):
        seen["export_file"] = export_file
        return {"parse_report": {"passed": True, "errors": []}, "outputs": {}}

    report = run_half_auto_pipeline(
        config={"excel_path": str(excel_file), "kst": {"export_dir": "kst_exports"}},
        root=tmp_path,
        logger=logging.getLogger("test"),
        period="15点",
        kst_file=None,
        assume_yes=True,
        fetch_baidu_func=ok_baidu,
        parse_kst_func=ok_kst,
        merge_func=lambda **kwargs: {"merged": {"date": "2026-05-07", "period": "15点"}, "validate_report": {"passed": True, "errors": []}, "outputs": {}},
        write_func=lambda **kwargs: {"date": "2026-05-07", "period": "15点", "excel_path": str(excel_file), "writes": [], "self_check": {"verification_passed": True}, "errors": []},
    )

    assert report["passed"] is True
    assert seen["export_file"] == latest
    assert report["kst_export_file"] == str(latest)
    assert report["kst_export"]["file_name"] == "latest.xlsx"
    assert report["kst_export"]["full_path"] == str(latest)
    assert report["kst_export"]["last_modified"]


def test_run_daily_pipeline_defaults_to_yesterday_and_reports_write_summary(tmp_path):
    import logging

    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    latest = export_dir / "daily.xlsx"
    latest.write_text("placeholder", encoding="utf-8")
    excel_file = tmp_path / "daily-target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")

    def ok_baidu(**kwargs):
        assert kwargs["target_date"] == "2026-05-07"
        return {"date": "2026-05-07", "accounts": {"银康01": {}, "银康银屑02": {}, "银康03": {}}, "errors": []}

    def ok_kst(export_file, config, root, target_date):
        assert export_file == latest
        assert target_date == "2026-05-07"
        return {
            "parse_report": {"passed": True, "errors": []},
            "daily_data": {"date": "2026-05-07"},
            "outputs": {"daily_data": str(root / "reports" / "kst_daily_data.json")},
        }

    def ok_merge(**kwargs):
        assert kwargs["target_date"] == "2026-05-07"
        return {
            "merged": {"date": "2026-05-07"},
            "validate_report": {"passed": True, "errors": []},
            "outputs": {"merged": str(kwargs["root"] / "reports" / "merged_daily_data.json")},
        }

    def ok_write(**kwargs):
        assert kwargs["target_date"] == "2026-05-07"
        return {
            "date": "2026-05-07",
            "excel_path": str(excel_file),
            "backup_path": str(tmp_path / "backups" / "daily_backup.xlsx"),
            "writes": [{"cell": "P132"}, {"cell": "Q132"}],
            "overwrite_summary": {"overwrite_count": 1},
            "self_check": {"verification_passed": True},
            "errors": [],
        }

    report = run_daily_pipeline(
        config={"project_id": "demo", "project_name": "演示项目", "excel_path": str(excel_file), "kst": {"export_dir": "kst_exports"}},
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date=None,
        kst_file=None,
        today=date(2026, 5, 8),
        fetch_baidu_func=ok_baidu,
        parse_kst_func=ok_kst,
        merge_func=ok_merge,
        write_func=ok_write,
    )

    assert report["passed"] is True
    assert report["project_id"] == "demo"
    assert report["project_name"] == "演示项目"
    assert report["date"] == "2026-05-07"
    assert report["excel_path"] == str(excel_file)
    assert report["backup_path"].endswith("daily_backup.xlsx")
    assert report["kst_export_file"] == str(latest)
    assert report["write_summary"]["write_count"] == 2
    assert report["write_summary"]["overwrite_count"] == 1
    assert report["write_summary"]["verification_passed"] is True
    assert [step["name"] for step in report["steps"]] == ["fetch-baidu-daily", "parse-kst-daily", "merge-daily", "write-daily"]
    assert (tmp_path / "reports" / "daily_final_run_report.json").exists()


def test_run_daily_pipeline_treats_stale_auto_discovered_kst_file_as_zero_dialogs(tmp_path):
    import logging
    import os
    import time

    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    old_export = export_dir / "daily-old.xlsx"
    old_export.write_text("old", encoding="utf-8")
    old_time = time.time() - 31 * 60
    os.utime(old_export, (old_time, old_time))
    excel_file = tmp_path / "daily-target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")
    config = _kunming_niu_runtime_config()
    config.update({"excel_path": str(excel_file), "kst": {"export_dir": "kst_exports", "promotion_id_accounts": config["kst"]["promotion_id_accounts"]}})

    def ok_merge(**kwargs):
        data = json.loads((tmp_path / "reports" / "kst_daily_data.json").read_text(encoding="utf-8"))
        assert all(row["总对话"] == 0 for row in data["accounts"].values())
        return {"merged": {"date": "2026-05-07"}, "validate_report": {"passed": True, "errors": []}, "outputs": {}}

    report = run_daily_pipeline(
        config=config,
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date="2026-05-07",
        kst_file=None,
        fetch_baidu_func=lambda **kwargs: {"date": "2026-05-07", "accounts": {}, "errors": []},
        parse_kst_func=lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("过期自动发现文件不应进入日报解析")),
        merge_func=ok_merge,
        write_func=lambda **kwargs: {"date": "2026-05-07", "excel_path": str(excel_file), "writes": [], "overwrite_summary": {}, "self_check": {"verification_passed": True}, "errors": []},
    )

    assert report["passed"] is True
    assert report["kst_export_file"] == ""
    assert report["steps"][1]["passed"] is True


def test_run_daily_pipeline_stops_when_kst_parse_fails(tmp_path):
    import logging

    export = tmp_path / "daily.xlsx"
    export.write_text("placeholder", encoding="utf-8")

    def ok_baidu(**kwargs):
        return {"date": "2026-05-07", "errors": []}

    def bad_kst(export_file, config, root, target_date):
        return {"parse_report": {"passed": False, "errors": ["商务通日报解析失败"]}, "outputs": {}}

    def should_not_merge(**kwargs):
        raise AssertionError("解析失败后不应继续合并")

    report = run_daily_pipeline(
        config={"excel_path": "target.xlsx"},
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date="2026-05-07",
        kst_file=export,
        fetch_baidu_func=ok_baidu,
        parse_kst_func=bad_kst,
        merge_func=should_not_merge,
    )

    assert report["passed"] is False
    assert report["failed_step"] == "parse-kst-daily"
    assert any("商务通日报解析失败" in error for error in report["errors"])


def test_run_pipeline_confirmation_can_quit_before_steps(tmp_path):
    import logging

    export = tmp_path / "kst.xlsx"
    export.write_text("placeholder", encoding="utf-8")

    report = run_half_auto_pipeline(
        config={"excel_path": "target.xlsx"},
        root=tmp_path,
        logger=logging.getLogger("test"),
        period="15点",
        kst_file=export,
        confirm_before_run=True,
        input_func=lambda prompt: "0",
        fetch_baidu_func=lambda **kwargs: (_ for _ in ()).throw(AssertionError("返回后不应抓百度")),
    )

    assert report["passed"] is False
    assert report["failed_step"] == "preflight-confirm"
    assert report["steps"] == []
    assert any("用户返回主菜单" in error for error in report["errors"])


def test_run_pipeline_treats_stale_auto_discovered_kst_file_as_zero_dialogs(tmp_path):
    import logging
    import os
    import time

    export_dir = tmp_path / "kst_exports"
    export_dir.mkdir()
    old_export = export_dir / "old.xlsx"
    old_export.write_text("old", encoding="utf-8")
    old_time = time.time() - 31 * 60
    os.utime(old_export, (old_time, old_time))
    excel_file = tmp_path / "target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")

    def ok_merge(**kwargs):
        data = json.loads((tmp_path / "reports" / "kst_dialog_data.json").read_text(encoding="utf-8"))
        assert all(row["总对话"] == 0 for row in data["accounts"].values())
        return {"merged": {"date": "2026-05-07", "period": "15点"}, "validate_report": {"passed": True, "errors": []}, "outputs": {}}

    report = run_half_auto_pipeline(
        config={**_kunming_niu_runtime_config(), "excel_path": str(excel_file), "sheet_name": "时段数据", "kst": {"export_dir": "kst_exports", "promotion_id_accounts": _kunming_niu_runtime_config()["kst"]["promotion_id_accounts"]}},
        root=tmp_path,
        logger=logging.getLogger("test"),
        period="15点",
        kst_file=None,
        assume_yes=True,
        fetch_baidu_func=lambda **kwargs: {"date": "2026-05-07", "period": "15点", "accounts": {}, "errors": []},
        parse_kst_func=lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("过期自动发现文件不应进入解析")),
        merge_func=ok_merge,
        write_func=lambda **kwargs: {"date": "2026-05-07", "period": "15点", "excel_path": str(excel_file), "writes": [], "self_check": {"verification_passed": True}, "errors": []},
    )

    assert report["passed"] is True
    assert report["kst_export_file"] == ""
    assert report["steps"][1]["passed"] is True


# ── 凭据文件路径修复 ──


def test_runtime_config_credentials_path_from_app_config_secrets():
    project = {
        "project_id": "demo",
        "project_name": "演示",
        "excel": {"path": "target.xlsx", "hourly_sheet": "时段", "daily_sheet": "日", "engine": "openpyxl"},
        "kst": {"export_dir": "exports"},
        "baidu": {"credential_profile": "demo_p", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "A1", "baidu_names": ["B1"], "excel_name": "A1", "kst_ids": ["1"], "kst_names": ["K1"]},
            {"standard_name": "A2", "baidu_names": ["B2"], "excel_name": "A2", "kst_ids": ["2"], "kst_names": ["K2"]},
            {"standard_name": "A3", "baidu_names": ["B3"], "excel_name": "A3", "kst_ids": ["3"], "kst_names": ["K3"]},
        ],
        "_app_config": {"secrets_file": "secrets/secrets.json"},
    }

    runtime = build_runtime_config_from_project(project, {"baidu": {}, "kst": {}})

    assert runtime["credentials_path"] == "secrets/secrets.json"
    assert runtime["baidu"]["credential_project"] == "demo_p"


def test_runtime_config_credentials_path_not_overridden_when_no_app_config():
    project = {
        "project_id": "x",
        "project_name": "y",
        "excel": {"path": "t.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
        "kst": {"export_dir": "e"},
        "baidu": {"credential_profile": "p", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "a1", "baidu_names": ["b1"], "excel_name": "a1", "kst_ids": ["1"], "kst_names": ["k1"]},
            {"standard_name": "a2", "baidu_names": ["b2"], "excel_name": "a2", "kst_ids": ["2"], "kst_names": ["k2"]},
            {"standard_name": "a3", "baidu_names": ["b3"], "excel_name": "a3", "kst_ids": ["3"], "kst_names": ["k3"]},
        ],
    }

    runtime = build_runtime_config_from_project(project, {"credentials_path": "old.json", "baidu": {}, "kst": {}})

    assert runtime["credentials_path"] == "old.json"


def test_load_credentials_from_secrets_json_structure(tmp_path):
    secrets_dir = tmp_path / "secrets"
    secrets_dir.mkdir()
    secrets_file = secrets_dir / "secrets.json"
    secrets_file.write_text(
        json.dumps({
            "baidu": {
                "kunming_niu_baidu": {
                    "username": "test_user",
                    "password": "test_pass",
                },
            },
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    credentials = load_project_credentials(
        tmp_path,
        {"credentials_path": "secrets/secrets.json", "baidu": {"credential_project": "kunming_niu_baidu"}},
        "baidu",
        "kunming_niu_baidu",
    )

    assert credentials is not None
    assert credentials["username"] == "test_user"
    assert credentials["password"] == "test_pass"


def test_load_credentials_falls_back_to_local_json(tmp_path):
    old_file = tmp_path / "credentials.local.json"
    old_file.write_text(
        json.dumps({
            "baidu": {
                "old_profile": {
                    "username": "fallback_user",
                    "password": "fallback_pass",
                },
            },
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    credentials = load_project_credentials(
        tmp_path,
        {"credentials_path": "secrets/secrets.json", "baidu": {"credential_project": "old_profile"}},
        "baidu",
        "old_profile",
    )

    assert credentials is not None
    assert credentials["username"] == "fallback_user"


def test_load_credentials_returns_none_when_profile_missing(tmp_path):
    secrets_dir = tmp_path / "secrets"
    secrets_dir.mkdir()
    (secrets_dir / "secrets.json").write_text(
        json.dumps({"baidu": {"other": {"username": "u", "password": "p"}}}),
        encoding="utf-8",
    )

    credentials = load_project_credentials(
        tmp_path,
        {"credentials_path": "secrets/secrets.json", "baidu": {"credential_project": "missing_profile"}},
        "baidu",
        "missing_profile",
    )

    assert credentials is None


def test_build_login_failure_message_shows_actual_path_and_profile():
    msg = build_login_failure_message({
        "credentials_path": "secrets/secrets.json",
        "baidu": {"credential_project": "kunming_niu_baidu"},
    })

    assert "credentials.local.json" not in msg
    assert "secrets/secrets.json" in msg
    assert "kunming_niu_baidu" in msg
    assert "验证码" in msg


def test_build_login_failure_message_defaults_when_config_empty():
    msg = build_login_failure_message({})

    assert "credentials.local.json" in msg


def _baidu_credential_test_config(*profiles: str) -> dict:
    config = {
        "project_id": "demo",
        "project_name": "演示项目",
        "credentials_path": "secrets/secrets.json",
        "baidu": {"credential_profile": profiles[0] if profiles else ""},
    }
    if len(profiles) > 1:
        config["baidu_sources"] = [
            {"source_id": f"source_{index}", "source_name": f"来源{index}", "credential_profile": profile, "accounts": []}
            for index, profile in enumerate(profiles, start=1)
        ]
    return config


def test_preflight_credentials_fails_for_invalid_secrets_json_without_values(tmp_path):
    from modules.preflight import check_baidu_credentials

    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text('{"baidu": invalid', encoding="utf-8")

    report = check_baidu_credentials(tmp_path, _baidu_credential_test_config("demo_baidu"))
    output = json.dumps(report, ensure_ascii=False)

    assert report["passed"] is False
    assert "不是合法 JSON" in report["errors"][0]
    assert "line" in report["errors"][0]
    assert "invalid-secret-value" not in output


def test_preflight_credentials_fails_for_missing_profile(tmp_path):
    from modules.preflight import check_baidu_credentials

    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text('{"baidu": {}}', encoding="utf-8")

    report = check_baidu_credentials(tmp_path, _baidu_credential_test_config("missing_baidu"))

    assert report["passed"] is False
    assert "缺少 credential_profile：missing_baidu" in report["errors"]


def test_preflight_credentials_fails_for_empty_username_or_password(tmp_path):
    from modules.preflight import check_baidu_credentials

    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text(
        json.dumps({"baidu": {"empty_user": {"username": "", "password": "value"}, "empty_password": {"username": "value", "password": ""}}}),
        encoding="utf-8",
    )

    report = check_baidu_credentials(tmp_path, _baidu_credential_test_config("empty_user", "empty_password"))

    assert report["passed"] is False
    assert "profile empty_user 的 username 为空" in report["errors"]
    assert "profile empty_password 的 password 为空" in report["errors"]


def test_preflight_credentials_checks_every_multi_source_profile_without_leaking_values(tmp_path):
    from modules.preflight import check_baidu_credentials, print_credential_report

    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text(
        json.dumps({"baidu": {"profile_a": {"username": "secret-user-a", "password": "secret-pass-a"}, "profile_b": {"username": "secret-user-b", "password": "secret-pass-b"}}}),
        encoding="utf-8",
    )

    report = check_baidu_credentials(tmp_path, _baidu_credential_test_config("profile_a", "profile_b"))
    lines = []
    print_credential_report(report, output_func=lines.append)
    output = "\n".join(lines) + json.dumps(report, ensure_ascii=False)

    assert report["passed"] is True
    assert [item["credential_profile"] for item in report["profiles"]] == ["profile_a", "profile_b"]
    assert all(item["username_nonempty"] and item["password_nonempty"] for item in report["profiles"])
    assert "secret-user-a" not in output
    assert "secret-pass-b" not in output


def _prepare_daily_preflight_files(tmp_path, credentials: dict) -> dict:
    (tmp_path / "main.py").write_text("", encoding="utf-8")
    (tmp_path / "kst_exports").mkdir()
    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text(
        json.dumps({"baidu": credentials}, ensure_ascii=False),
        encoding="utf-8",
    )
    excel_path = tmp_path / "daily.xlsx"
    excel_path.write_text("", encoding="utf-8")
    config = _baidu_credential_test_config(*credentials.keys())
    config["excel_path"] = str(excel_path)
    config["kst"] = {"export_dir": "kst_exports"}
    return config


def test_preflight_api_preference_skips_chrome_start(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "test-user", "password": "test-password"}},
    )
    config["baidu"]["data_source_preference"] = "api"
    called = []
    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])

    def forbidden(*args, **kwargs):
        called.append(True)
        raise AssertionError("A 模式不应提前启动 Chrome")

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        quick=True,
        chrome_ready_func=forbidden,
    )

    assert called == []
    assert report["passed"] is True
    assert any(item.get("skipped") and "API" in item["message"] for item in report["checks"])
    assert any(item["passed"] and "浏览器降级" in item["message"] for item in report["checks"])
    assert report["api_profiles"]["passed"] is False


def test_preflight_browser_preference_checks_chrome(tmp_path):
    from modules.preflight import run_preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "test-user", "password": "test-password"}},
    )
    config["baidu"]["data_source_preference"] = "browser"
    called = []

    def ready(*args, **kwargs):
        called.append(True)
        return {"ready": True, "started_new_chrome": False}

    run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        quick=True,
        chrome_ready_func=ready,
    )

    assert called == [True]


def _write_api_profile_test_secrets(tmp_path, api_profiles: dict, browser_profiles: dict | None = None) -> None:
    secrets_dir = tmp_path / "secrets"
    secrets_dir.mkdir()
    (secrets_dir / "secrets.json").write_text(
        json.dumps({"baidu": browser_profiles or {}, "baidu_api": api_profiles}),
        encoding="utf-8",
    )


def test_check_baidu_api_profiles_supports_single_source(tmp_path):
    from modules.preflight import check_baidu_api_profiles

    _write_api_profile_test_secrets(
        tmp_path,
        {"source_a": {"access_token": "test-token-a", "refresh_token": "test-refresh-a"}},
    )
    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu": {"api_profile": "source_a"},
    }

    report = check_baidu_api_profiles(tmp_path, config)

    assert report["passed"] is True
    assert report["required_profiles"] == ["source_a"]
    assert report["missing_source_mappings"] == []
    assert "test-token-a" not in json.dumps(report)
    assert "test-refresh-a" not in json.dumps(report)


def test_check_baidu_api_profiles_supports_multi_source(tmp_path):
    from modules.preflight import check_baidu_api_profiles

    _write_api_profile_test_secrets(
        tmp_path,
        {
            "source_a": {"access_token": "test-token-a", "refresh_token": "test-refresh-a"},
            "source_b": {"access_token": "test-token-b", "refresh_token": "test-refresh-b"},
        },
    )
    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu_sources": [
            {"source_id": "a", "api_profile": "source_a"},
            {"source_id": "b", "api_profile": "source_b"},
        ],
    }

    report = check_baidu_api_profiles(tmp_path, config)

    assert report["passed"] is True
    assert report["required_profiles"] == ["source_a", "source_b"]
    assert report["missing_source_mappings"] == []
    assert "test-token-a" not in json.dumps(report)
    assert "test-refresh-b" not in json.dumps(report)


def test_check_baidu_api_profiles_fails_for_multi_source_missing_mapping(tmp_path):
    from modules.preflight import check_baidu_api_profiles

    _write_api_profile_test_secrets(
        tmp_path,
        {"source_b": {"access_token": "test-token-b", "refresh_token": "test-refresh-b"}},
    )
    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu_sources": [
            {"source_id": "a", "source_name": "来源 A"},
            {"source_id": "b", "source_name": "来源 B", "api_profile": "source_b"},
        ],
    }

    report = check_baidu_api_profiles(tmp_path, config)

    assert report["passed"] is False
    assert report["missing_source_mappings"] == [{"source_id": "a", "source_name": "来源 A"}]
    assert report["required_profiles"] == ["source_b"]
    assert "test-token-b" not in json.dumps(report)


def test_check_baidu_api_profiles_fails_for_missing_or_empty_tokens(tmp_path):
    from modules.preflight import check_baidu_api_profiles

    _write_api_profile_test_secrets(
        tmp_path,
        {
            "empty_access": {"access_token": "", "refresh_token": "test-refresh"},
            "empty_refresh": {"access_token": "test-token", "refresh_token": ""},
        },
    )
    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu_sources": [
            {"source_id": "missing", "api_profile": "missing_profile"},
            {"source_id": "access", "api_profile": "empty_access"},
            {"source_id": "refresh", "api_profile": "empty_refresh"},
        ],
    }

    report = check_baidu_api_profiles(tmp_path, config)

    assert report["passed"] is False
    assert [item["exists"] for item in report["profiles"]] == [False, True, True]
    assert report["profiles"][1]["access_token_nonempty"] is False
    assert report["profiles"][2]["refresh_token_nonempty"] is False
    assert "test-token" not in json.dumps(report)
    assert "test-refresh" not in json.dumps(report)


@pytest.mark.parametrize(
    "browser_profiles",
    [{}, {"profile_a": {"username": "", "password": "test-password"}}],
)
def test_preflight_api_preference_requires_browser_credentials_for_fallback(tmp_path, monkeypatch, browser_profiles):
    import modules.preflight as preflight

    _write_api_profile_test_secrets(
        tmp_path,
        {"source_a": {"access_token": "test-token", "refresh_token": "test-refresh"}},
        browser_profiles,
    )
    (tmp_path / "main.py").write_text("", encoding="utf-8")
    (tmp_path / "kst_exports").mkdir()
    excel_path = tmp_path / "daily.xlsx"
    excel_path.write_text("", encoding="utf-8")
    config = _baidu_credential_test_config("profile_a")
    config["excel_path"] = str(excel_path)
    config["kst"] = {"export_dir": "kst_exports"}
    config["baidu"]["data_source_preference"] = "api"
    config["baidu"]["api_profile"] = "source_a"
    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        quick=True,
        chrome_ready_func=lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("不应检查 Chrome")),
    )

    assert report["api_profiles"]["passed"] is True
    assert report["credentials"]["passed"] is False
    assert report["passed"] is False


def test_daily_preflight_selects_daily_sheet_check_and_hides_credentials(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "daily-secret-user", "password": "daily-secret-password"}},
    )
    selected = []

    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])
    monkeypatch.setattr(preflight, "inspect_excel_structure", lambda **kwargs: (_ for _ in ()).throw(AssertionError("日报不得检查小时报 sheet")))
    monkeypatch.setattr(
        preflight,
        "inspect_daily_excel_structure",
        lambda **kwargs: selected.append("daily") or {"errors": []},
        raising=False,
    )

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        task="daily",
        chrome_check_func=lambda **kwargs: True,
    )
    output = json.dumps(report, ensure_ascii=False)

    assert report["passed"] is True
    assert report["task"] == "daily"
    assert selected == ["daily"]
    assert "daily-secret-user" not in output
    assert "daily-secret-password" not in output


def test_daily_preflight_checks_all_multi_source_profiles(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {
            "profile_a": {"username": "u-a", "password": "p-a"},
            "profile_b": {"username": "u-b", "password": "p-b"},
        },
    )

    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])
    monkeypatch.setattr(
        preflight,
        "inspect_daily_excel_structure",
        lambda **kwargs: {"errors": []},
        raising=False,
    )

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        task="daily",
        chrome_check_func=lambda **kwargs: True,
    )

    assert report["passed"] is True
    assert [item["credential_profile"] for item in report["credentials"]["profiles"]] == ["profile_a", "profile_b"]


def test_preflight_defaults_to_hourly_sheet_check(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "hourly-user", "password": "hourly-password"}},
    )
    selected = []

    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])
    monkeypatch.setattr(
        preflight,
        "inspect_excel_structure",
        lambda **kwargs: selected.append("hourly") or {"errors": []},
    )
    monkeypatch.setattr(
        preflight,
        "inspect_daily_excel_structure",
        lambda **kwargs: (_ for _ in ()).throw(AssertionError("默认预检不得检查日报 sheet")),
    )

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        chrome_check_func=lambda **kwargs: True,
    )

    assert report["passed"] is True
    assert report["task"] == "hourly"
    assert selected == ["hourly"]


def test_quick_preflight_skips_excel_structure_scan(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "quick-user", "password": "quick-password"}},
    )

    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])
    monkeypatch.setattr(
        preflight,
        "inspect_excel_structure",
        lambda **kwargs: (_ for _ in ()).throw(AssertionError("快速预检不得扫描小时报 sheet")),
    )
    monkeypatch.setattr(
        preflight,
        "inspect_daily_excel_structure",
        lambda **kwargs: (_ for _ in ()).throw(AssertionError("快速预检不得扫描日报 sheet")),
    )

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "演示项目"},
        config,
        quick=True,
        chrome_check_func=lambda **kwargs: True,
    )

    assert report["passed"] is True
    assert report["quick"] is True
    assert any(item.get("skipped") for item in report["checks"])


def test_preflight_auto_starts_chrome_debug_when_default_check_is_used(tmp_path, monkeypatch):
    import modules.preflight as preflight

    config = _prepare_daily_preflight_files(
        tmp_path,
        {"profile_a": {"username": "quick-user", "password": "quick-password"}},
    )
    config["browser"] = {
        "auto_start_debug_chrome": True,
        "debug_startup_wait_seconds": 2,
    }
    calls = []

    monkeypatch.setattr(preflight, "validate_project_config", lambda project: [])
    monkeypatch.setattr(
        preflight,
        "ensure_chrome_debug_ready",
        lambda root, cfg, **kwargs: calls.append((root, cfg, kwargs)) or {
            "ready": True,
            "started_new_chrome": True,
            "debug_endpoint": "http://127.0.0.1:9222",
        },
    )

    report = preflight.run_preflight(
        tmp_path,
        {"project_id": "demo", "project_name": "demo"},
        config,
        quick=True,
    )

    assert report["passed"] is True
    assert calls
    assert calls[0][2]["auto_start"] is True
    assert calls[0][2]["wait_seconds"] == 2
    assert any("已自动启动并连接" in item["message"] for item in report["checks"])


def test_cli_preflight_accepts_daily_task_and_passes_it_to_runner(tmp_path, monkeypatch):
    import main as cli_main

    calls = []
    monkeypatch.setattr(cli_main, "ROOT", tmp_path)
    monkeypatch.setattr(cli_main, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(cli_main, "get_current_project", lambda root: {"project_id": "demo", "project_name": "演示"})
    monkeypatch.setattr(cli_main, "build_runtime_config_from_project", lambda current, base: {})
    monkeypatch.setattr(
        cli_main,
        "run_preflight",
        lambda root, project, config, task="hourly", quick=False: calls.append((task, quick)) or {"passed": True, "checks": [], "credentials": {}},
    )
    monkeypatch.setattr("sys.argv", ["main.py", "--mode", "preflight", "--task", "daily", "--quick"])

    result = cli_main.main()

    assert result == 0
    assert calls == [("daily", True)]


def test_cli_run_fails_before_baidu_pipeline_when_credential_precheck_fails(tmp_path, monkeypatch):
    import main as cli_main

    project = {"project_id": "demo", "project_name": "演示"}
    config = _baidu_credential_test_config("missing_baidu")
    called = []

    monkeypatch.setattr(cli_main, "ROOT", tmp_path)
    monkeypatch.setattr(cli_main, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(cli_main, "get_current_project", lambda root: project)
    monkeypatch.setattr(cli_main, "build_runtime_config_from_project", lambda current, base: config)
    monkeypatch.setattr(cli_main, "run_half_auto_pipeline", lambda **kwargs: called.append(kwargs))
    monkeypatch.setattr("sys.argv", ["main.py", "--mode", "run", "--period", "15点", "--yes"])

    result = cli_main.main()

    assert result == 1
    assert called == []


def test_cli_run_daily_fails_before_baidu_pipeline_when_credential_precheck_fails(tmp_path, monkeypatch):
    import main as cli_main

    project = {"project_id": "demo", "project_name": "演示"}
    config = _baidu_credential_test_config("missing_baidu")
    called = []

    monkeypatch.setattr(cli_main, "ROOT", tmp_path)
    monkeypatch.setattr(cli_main, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(cli_main, "get_current_project", lambda root: project)
    monkeypatch.setattr(cli_main, "build_runtime_config_from_project", lambda current, base: config)
    monkeypatch.setattr(cli_main, "run_daily_pipeline", lambda **kwargs: called.append(kwargs))
    monkeypatch.setattr("sys.argv", ["main.py", "--mode", "run-daily", "--yes"])

    result = cli_main.main()

    assert result == 1
    assert called == []


@pytest.mark.parametrize("mode", ["run", "run-daily"])
def test_cli_report_run_returns_dedicated_cancelled_exit_code(tmp_path, monkeypatch, mode):
    import main as cli_main

    monkeypatch.setattr(cli_main, "ROOT", tmp_path)
    monkeypatch.setattr(cli_main, "setup_logger", lambda *_args, **_kwargs: object())
    monkeypatch.setattr(cli_main, "load_config", lambda *_args, **_kwargs: {})
    monkeypatch.setattr(cli_main, "get_current_project", lambda _root: {"project_id": "demo"})
    monkeypatch.setattr(cli_main, "build_runtime_config_from_project", lambda *_args: {})
    monkeypatch.setattr(cli_main, "check_baidu_credentials", lambda *_args: {"passed": True})
    cancelled_report = {"passed": False, "cancelled": True, "failed_step": "cancelled-before-excel"}
    monkeypatch.setattr(cli_main, "run_half_auto_pipeline", lambda **_kwargs: cancelled_report)
    monkeypatch.setattr(cli_main, "run_daily_pipeline", lambda **_kwargs: cancelled_report)
    argv = ["main.py", "--mode", mode, "--yes"]
    if mode == "run":
        argv.extend(["--period", "15点"])
    monkeypatch.setattr("sys.argv", argv)

    assert cli_main.main() == 130


# ── Chrome 调试端口自动启动 ──


def test_is_chrome_debug_port_alive_returns_false_when_port_closed():
    assert is_chrome_debug_port_alive(host="127.0.0.1", port=1, timeout=1.0) is False


def test_find_chrome_executable_returns_none_when_not_found(monkeypatch):
    import modules.chrome_debug as cd

    monkeypatch.setattr(cd, "CHROME_CANDIDATES", [Path("X:/nonexistent/chrome.exe")])
    result = find_chrome_executable(None)
    assert result is None


def test_ensure_chrome_debug_ready_reports_port_ready_when_already_open(monkeypatch):
    def fake_is_alive(host="127.0.0.1", port=9222, timeout=3.0):
        return True

    monkeypatch.setattr("modules.chrome_debug.is_chrome_debug_port_alive", fake_is_alive)

    result = ensure_chrome_debug_ready(Path("."), None)
    assert result["ready"] is True
    assert result["port_already_open"] is True
    assert result["started_new_chrome"] is False


def test_ensure_chrome_debug_ready_returns_error_when_auto_start_disabled(monkeypatch):
    def fake_is_alive(host="127.0.0.1", port=9222, timeout=3.0):
        return False

    monkeypatch.setattr("modules.chrome_debug.is_chrome_debug_port_alive", fake_is_alive)

    result = ensure_chrome_debug_ready(Path("."), None, auto_start=False)
    assert result["ready"] is False
    assert "未就绪" in (result.get("error") or "")


def test_ensure_chrome_debug_ready_returns_error_when_chrome_not_found(monkeypatch, tmp_path):
    def fake_is_alive(host="127.0.0.1", port=9222, timeout=3.0):
        return False

    def fake_chrome_exists():
        return False

    def fake_find(config=None):
        return None

    monkeypatch.setattr("modules.chrome_debug.is_chrome_debug_port_alive", fake_is_alive)
    monkeypatch.setattr("modules.chrome_debug._chrome_process_exists", fake_chrome_exists)
    monkeypatch.setattr("modules.chrome_debug.find_chrome_executable", fake_find)

    result = ensure_chrome_debug_ready(tmp_path, None, auto_start=True)
    assert result["ready"] is False
    assert result.get("error")
    assert "未找到 Google Chrome" in result["error"]


def test_ensure_chrome_debug_ready_can_start_parallel_debug_profile(monkeypatch, tmp_path):
    import modules.chrome_debug as cd

    calls = {"alive": 0}

    def fake_is_alive(host="127.0.0.1", port=9222, timeout=3.0):
        calls["alive"] += 1
        return calls["alive"] > 1

    chrome = tmp_path / "chrome.exe"
    chrome.write_text("", encoding="utf-8")
    monkeypatch.setattr(cd, "is_chrome_debug_port_alive", fake_is_alive)
    monkeypatch.setattr(cd, "_chrome_process_exists", lambda: (_ for _ in ()).throw(AssertionError("daily Chrome should not block debug profile")))
    monkeypatch.setattr(cd, "find_chrome_executable", lambda config=None: chrome)
    monkeypatch.setattr(cd.subprocess, "Popen", lambda *args, **kwargs: object())

    result = ensure_chrome_debug_ready(tmp_path, {"browser": {"auto_start_debug_chrome": True}}, wait_seconds=2)

    assert result["ready"] is True
    assert result["started_new_chrome"] is True
    assert result["profile_dir"].endswith("browser_profile\\chrome_debug") or result["profile_dir"].endswith("browser_profile/chrome_debug")


def test_start_debug_chrome_uses_minimized_debug_profile_and_disables_password_manager(monkeypatch, tmp_path):
    import json
    import modules.chrome_debug as cd

    captured = {}

    class FakeProcess:
        pass

    def fake_popen(args, stdout=None, stderr=None, startupinfo=None):
        captured["args"] = args
        captured["startupinfo"] = startupinfo
        return FakeProcess()

    chrome = tmp_path / "chrome.exe"
    chrome.write_text("", encoding="utf-8")
    monkeypatch.setattr(cd, "find_chrome_executable", lambda config=None: chrome)
    monkeypatch.setattr(cd.subprocess, "Popen", fake_popen)

    result = start_debug_chrome(tmp_path, {"browser": {"debug_profile_dir": "browser_profile/chrome_debug"}})

    assert result["started"] is True
    args = captured["args"]
    assert "--start-minimized" in args
    assert "--disable-save-password-bubble" in args
    assert "--disable-features=PasswordManagerOnboarding,PasswordLeakDetection" in args
    assert captured["startupinfo"] is not None
    assert captured["startupinfo"].wShowWindow == 7
    assert not any("cas.baidu.com" in item or "cc.baidu.com" in item for item in args)
    assert any(str(tmp_path / "browser_profile" / "chrome_debug") in item for item in args)
    prefs = json.loads((tmp_path / "browser_profile" / "chrome_debug" / "Default" / "Preferences").read_text(encoding="utf-8"))
    assert prefs["credentials_enable_service"] is False
    assert prefs["profile"]["password_manager_enabled"] is False


def test_chrome_debug_launcher_reuses_existing_debug_port_without_killing_chrome(monkeypatch, tmp_path):
    import modules.chrome_debug_launcher as launcher

    chrome = tmp_path / "chrome.exe"
    chrome.write_text("", encoding="utf-8")
    monkeypatch.setattr(launcher, "CHROME_EXE", chrome)
    monkeypatch.setattr(launcher, "_is_port_open", lambda port: True)
    monkeypatch.setattr(launcher, "_run_connect_test", lambda: 0)
    monkeypatch.setattr(
        launcher.subprocess,
        "run",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("launcher must not inspect or close existing Chrome")),
    )

    assert launcher.main() == 0


def test_menu_chrome_check_prints_status(tmp_path, capsys):
    from menu import _check_chrome_debug

    def fake_is_alive(host="127.0.0.1", port=9222, timeout=3.0):
        return True

    import modules.chrome_debug as cd
    original = cd.is_chrome_debug_port_alive
    cd.is_chrome_debug_port_alive = fake_is_alive
    try:
        ok = _check_chrome_debug(tmp_path, {"browser": {"auto_start_debug_chrome": True}}, print)
        assert ok is True
    finally:
        cd.is_chrome_debug_port_alive = original


def test_browser_settings_includes_new_auto_start_fields():
    settings = get_browser_settings(
        {
            "browser": {
                "mode": "connect_existing",
                "cdp_endpoint": "http://127.0.0.1:9222",
                "auto_start_debug_chrome": True,
                "remote_debugging_host": "127.0.0.1",
                "remote_debugging_port": 9222,
                "profile_dir": "browser_profile/chrome",
                "startup_url": "https://yingxiao.baidu.com/",
                "allow_kill_existing_chrome": False,
            }
        }
    )

    assert settings["auto_start_debug_chrome"] is True
    assert settings["remote_debugging_host"] == "127.0.0.1"
    assert settings["startup_url"] == "https://yingxiao.baidu.com/"
    assert settings["allow_kill_existing_chrome"] is False
    assert settings["silent_automation"] is True
    assert settings["disable_password_manager"] is True


# ── 终端输出分层（v0.4.8）──


def test_console_ui_set_verbose_and_is_verbose():
    from modules.console_ui import is_verbose, set_verbose

    set_verbose(False)
    assert is_verbose() is False
    set_verbose(True)
    assert is_verbose() is True
    set_verbose(False)
    assert is_verbose() is False


def test_console_ui_print_check_result_does_not_raise():
    from io import StringIO

    from modules.console_ui import print_check_result, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_check_result("测试项", "pass", "正常")
        print_check_result("测试项", "fail", "有问题")
        print_check_result("测试项", "skip", "已跳过")
        print_check_result("测试项", "warn", "需关注")
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "[通过]" in output
    assert "[失败]" in output
    assert "[跳过]" in output
    assert "[注意]" in output


def test_console_ui_verbose_print_only_when_verbose():
    from io import StringIO

    from modules.console_ui import set_output_func, set_verbose, verbose_print

    buf = StringIO()
    set_output_func(buf.write)
    try:
        set_verbose(False)
        verbose_print("这条不应出现")
        assert "这条不应出现" not in buf.getvalue()

        set_verbose(True)
        verbose_print("这条应该出现")
        assert "这条应该出现" in buf.getvalue()
    finally:
        set_output_func(print)
        set_verbose(False)


def test_console_ui_step_output_format():
    from io import StringIO

    from modules.console_ui import print_final_failure, print_final_success, print_step, print_step_failure, print_step_success, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_step(2, 4, "解析数据")
        print_step_success("数据已解析")
        print_step_failure("解析失败", suggestion="文件格式不正确", report_path="reports/err.json")
        print_final_success("写入 8 个单元格")
        print_final_failure("失败步骤：fetch-baidu")
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "[2/4]" in output
    assert "[通过]" in output
    assert "[失败]" in output
    assert "文件格式不正确" in output
    assert "reports/err.json" in output
    assert "写入 8 个单元格" in output


def test_main_accepts_verbose_flag():
    import subprocess
    import sys

    import main

    result = main.ROOT  # verify main is importable
    assert result is not None


def test_doctor_check_labels_cover_all_checks():
    from modules.doctor import _DOCTOR_CHECK_LABELS

    assert "chrome" in _DOCTOR_CHECK_LABELS
    assert "chrome_debug_port" in _DOCTOR_CHECK_LABELS
    assert "target_excel" in _DOCTOR_CHECK_LABELS
    assert "secrets_json" in _DOCTOR_CHECK_LABELS
    assert "latest_kst_export" in _DOCTOR_CHECK_LABELS
    assert all(isinstance(v, str) for v in _DOCTOR_CHECK_LABELS.values())


def test_run_bat_files_support_dragged_kst_file_argument():
    from pathlib import Path

    root = Path(__file__).resolve().parents[1]
    for name, period in [("run_11.bat", "11"), ("run_15.bat", "15"), ("run_18.bat", "18")]:
        text = (root / name).read_text(encoding="utf-8")
        assert f"--period {period}" in text
        assert 'set "KST_FILE=%~1"' in text
        assert '--file "%KST_FILE%"' in text


def test_hermes_hourly_bat_fixes_utf8_and_runs_preflight_before_hourly_pipeline():
    root = Path(__file__).resolve().parents[1]
    script = (root / "run_hermes_hourly.bat").read_text(encoding="utf-8")

    assert 'cd /d "%~dp0"' in script
    assert "chcp 65001" in script
    assert "PYTHONUTF8=1" in script
    assert "PYTHONIOENCODING=utf-8" in script
    assert "20260710" in script
    assert "install_env.bat" in script
    assert ".venv\\Scripts\\python.exe main.py --mode preflight --quick" in script
    assert "main.py --mode run --period" in script


def test_hermes_hourly_sop_documents_preflight_credentials_and_password_rule():
    root = Path(__file__).resolve().parents[1]
    content = (root / "docs" / "hermes_hourly_sop.md").read_text(encoding="utf-8")

    assert "HERMES-20260710" in content
    assert "preflight --quick" in content
    assert "不索要或输出密码" in content
    for period in ["11点", "15点", "18点"]:
        assert f"run_hermes_hourly.bat {period}" in content
    assert "UTF-8" in content


def test_hermes_daily_bat_runs_daily_preflight_before_daily_pipeline():
    root = Path(__file__).resolve().parents[1]
    script = (root / "run_hermes_daily.bat").read_text(encoding="utf-8")

    assert 'cd /d "%~dp0"' in script
    assert "chcp 65001" in script
    assert "PYTHONUTF8=1" in script
    assert "PYTHONIOENCODING=utf-8" in script
    assert "main.py --mode preflight --task daily --quick" in script
    assert "main.py --mode run-daily --yes" in script
    assert 'main.py --mode run-daily --date "%~1" --yes' in script


def test_hermes_daily_sop_documents_preflight_password_and_write_boundaries():
    root = Path(__file__).resolve().parents[1]
    content = (root / "docs" / "hermes_daily_sop.md").read_text(encoding="utf-8")

    assert "HERMES 日报自动化执行手册" in content
    assert "run_hermes_daily.bat" in content
    assert "preflight --task daily --quick" in content
    assert "不索要或输出密码" in content
    assert "不改无关 sheet、公式区、汇总区或截图区" in content
    assert "networkidle" in content


def test_documentation_api_mode_uses_global_preference_and_current_product_name():
    root = Path(__file__).resolve().parents[1]
    documents = {
        "AGENTS.md": (root / "AGENTS.md").read_text(encoding="utf-8"),
        "README.md": (root / "README.md").read_text(encoding="utf-8"),
        "xia_sidao使用说明.md": (root / "xia_sidao使用说明.md").read_text(encoding="utf-8"),
        "hermes_hourly_sop.md": (root / "docs" / "hermes_hourly_sop.md").read_text(encoding="utf-8"),
        "hermes_daily_sop.md": (root / "docs" / "hermes_daily_sop.md").read_text(encoding="utf-8"),
    }

    for name, content in documents.items():
        assert "蚁之力 · 竞价数据自动化" in content, f"{name} 缺少当前产品名"
        assert "baidu_data_source_preference" in content, f"{name} 缺少应用级偏好键"
        assert "A" in content and "API 优先" in content, f"{name} 缺少 A 模式说明"
        assert "B" in content and "强制浏览器" in content, f"{name} 缺少 B 模式说明"


def test_hermes_documentation_shares_global_api_mode_and_delays_chrome_fallback():
    root = Path(__file__).resolve().parents[1]
    hourly = (root / "docs" / "hermes_hourly_sop.md").read_text(encoding="utf-8")
    daily = (root / "docs" / "hermes_daily_sop.md").read_text(encoding="utf-8")
    xia = (root / "xia_sidao使用说明.md").read_text(encoding="utf-8")

    assert "run_hermes_hourly.bat 11点" in hourly
    assert "HERMES Hourly - fixed entry - 20260710" in hourly
    assert "run_hermes_daily.bat" in daily
    assert "HERMES Daily - fixed entry - 20260710" in daily
    for content in (hourly, daily, xia):
        assert "共享同一应用级偏好" in content
        assert "API 模式的 preflight 不提前启动 Chrome" in content
        assert "实际降级" in content and "延迟启动" in content
        assert "普通 GUI" in content and "不得自动调用" in content


def test_documentation_api_mode_requires_readiness_and_atomic_dual_source_fallback():
    root = Path(__file__).resolve().parents[1]
    paths = [
        root / "AGENTS.md",
        root / "README.md",
        root / "xia_sidao使用说明.md",
        root / "docs" / "hermes_hourly_sop.md",
        root / "docs" / "hermes_daily_sop.md",
    ]
    documents = [path.read_text(encoding="utf-8") for path in paths]
    combined = "\n".join(documents)

    for content in documents:
        assert "九个项目、十一个授权" in content
        assert "test-baidu-api-readiness" in content
        assert "不读写 Excel" in content
        assert "Token" in content and "原子更新" in content
        assert "备份" in content and "敏感文件" in content
    assert "两路 API 全部成功后才合并" in combined
    assert "任一路失败" in combined and "整项目降级" in combined
    assert "禁止混合 API 与浏览器的部分数据" in combined
    assert "多项目模式只允许 API 并行准备" in combined
    assert "Excel 写入必须按选择顺序串行" in combined

    stale_phrases = [
        "其余十个超管仍需逐个授权",
        "昆明牛当前为 `api_shadow`",
        "当前九个项目均保持 `browser`",
        "双来源 API 合并及多项目并行尚未投入生产",
        "双来源 API 尚未投入生产",
    ]
    for phrase in stale_phrases:
        assert phrase not in combined


def test_hermes_documentation_never_skips_dual_source_browser_fallback():
    root = Path(__file__).resolve().parents[1]
    documents = [
        (root / "xia_sidao使用说明.md").read_text(encoding="utf-8"),
        (root / "docs" / "hermes_hourly_sop.md").read_text(encoding="utf-8"),
        (root / "docs" / "hermes_daily_sop.md").read_text(encoding="utf-8"),
    ]

    for content in documents:
        assert "API 任一路失败先丢弃临时结果并整项目降级浏览器" in content
        assert "仅 API 与浏览器均失败时停止，禁止写 Excel" in content
        assert "任一来源失败时停止" not in content
        assert "任一来源或稳定性校验失败时，不继续写 Excel" not in content


# ── console_ui 新增测试 ──────────────────────────────────


def test_print_check_table_shows_pass_fail_warn_status():
    from io import StringIO

    from modules.console_ui import print_check_table, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_check_table("测试检查", [
            {"name": "项目A", "status": "pass", "message": "正常"},
            {"name": "项目B", "status": "fail", "message": "有问题"},
            {"name": "项目C", "status": "warn", "message": "需关注"},
        ])
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "[通过]" in output
    assert "[失败]" in output
    assert "[注意]" in output
    assert "项目A" in output
    assert "正常" in output
    assert "有问题" in output
    assert "需关注" in output
    assert "2/3" in output or "1/3" in output  # 通过/总数


def test_console_ui_no_colorama_fallback_does_not_raise():
    import sys

    from modules.console_ui import set_output_func

    # 模拟 colorama 不可用
    original_modules = sys.modules.copy()
    try:
        # 先清除可能的缓存
        for key in list(sys.modules.keys()):
            if "colorama" in key.lower():
                del sys.modules[key]

        sys.modules["colorama"] = type(sys)("fake_colorama")
        sys.modules["colorama"].__spec__ = None

        # 强制重新加载 console_ui（带 fallback）
        import importlib
        import modules.console_ui as cui

        importlib.reload(cui)

        assert cui._HAS_COLOR is False
        # 确保核心函数仍然可用
        buf = __import__("io").StringIO()
        cui.set_output_func(buf.write)
        try:
            cui.print_success("测试")
            cui.print_error("测试")
            cui.print_warning("测试")
        finally:
            cui.set_output_func(print)

        output = buf.getvalue()
        assert "[通过]" in output
        assert "[失败]" in output
        assert "[注意]" in output
    finally:
        # 恢复
        for key in list(sys.modules.keys()):
            if "colorama" in key.lower():
                del sys.modules[key]
        for key in list(original_modules.keys()):
            if "colorama" in key.lower() and key not in sys.modules:
                pass  # 不恢复 colorama mock


def test_print_banner_shows_project_and_version():
    from io import StringIO
    from pathlib import Path

    from modules.console_ui import print_banner, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_banner({
            "project_name": "测试项目",
            "excel": {"path": "/path/to/test.xlsx"},
        }, version="1.0.0")
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "百度竞价自动化控制台" in output
    assert "1.0.0" in output
    assert "测试项目" in output
    assert "test.xlsx" in output
    assert "logs/run.log" in output


def test_print_project_info_shows_all_fields():
    from io import StringIO

    from modules.console_ui import print_project_info, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_project_info({
            "project_name": "演示项目",
            "project_id": "demo",
            "_config_path": "/path/to/config.json",
            "excel": {
                "path": "/path/to/excel.xlsx",
                "hourly_sheet": "时段数据",
                "daily_sheet": "百度",
            },
            "kst": {"export_dir": "/path/to/exports"},
            "baidu": {"credential_profile": "my_profile"},
        })
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "项目信息" in output
    assert "演示项目" in output
    assert "excel.xlsx" in output
    assert "时段数据" in output
    assert "百度" in output
    assert "my_profile" in output


def test_default_output_mode_does_not_print_json_blobs(capsys):
    """默认模式（非 verbose）不应输出大段 JSON。"""
    from modules.console_ui import print_quiet_line, set_verbose, verbose_print

    set_verbose(False)
    verbose_print('{"this": "should not appear"}')
    print_quiet_line("正常信息")

    captured = capsys.readouterr()
    assert "should not appear" not in captured.out
    assert "正常信息" in captured.out


def test_verbose_mode_still_outputs_debug_info(capsys):
    from modules.console_ui import set_verbose, verbose_print

    set_verbose(True)
    verbose_print("调试信息：详细数据")

    captured = capsys.readouterr()
    assert "调试信息" in captured.out

    set_verbose(False)


def test_doctor_report_no_longer_prints_internal_field_names(tmp_path, capsys):
    """doctor 输出不再包含 _DOCTOR_CHECK_LABELS key 名等内部字段。"""
    from modules.doctor import print_doctor_report

    report = {
        "project_name": "测试",
        "checks": {
            "python": {"passed": True, "message": "Python 3.14.4"},
            "chrome": {"passed": False, "level": "warning", "message": "未找到 Chrome"},
        },
    }
    print_doctor_report(report)

    captured = capsys.readouterr()
    # 应该显示中文标签，而不是内部 key 名
    assert "Python 版本" in captured.out
    assert "未找到 Chrome" in captured.out
    # 不应输出 JSON 结构
    assert '"passed"' not in captured.out
    assert '"checks"' not in captured.out


def test_menu_new_text_entries_exist():
    """菜单显示新文案。"""
    assert "1. 小时报" in MENU_TEXT
    assert "2. 日报" in MENU_TEXT
    assert "3. 切换项目" in MENU_TEXT
    assert "4. 检查条件项" in MENU_TEXT
    assert "5. 更多功能" in MENU_TEXT
    assert "0. 退出" in MENU_TEXT


def test_more_features_menu_contains_console_sections():
    from menu import MORE_FEATURES_MENU_TEXT

    for text in ["报告与日志", "配置诊断", "HERMES / 夏思道帮助", "多百度来源摘要", "项目信息详情", "高级分步调试"]:
        assert text in MORE_FEATURES_MENU_TEXT


def test_condition_menu_uses_colleague_facing_labels():
    from menu import CONDITION_MENU_TEXT

    for text in ["一键检查当前项目能否运行", "检查百度账号凭据", "检查小时报条件", "检查日报条件"]:
        assert text in CONDITION_MENU_TEXT
    assert "doctor" not in CONDITION_MENU_TEXT
    assert "preflight" not in CONDITION_MENU_TEXT


def test_diagnostic_menu_exposes_sheet_text_dump_without_write_action():
    from menu import DIAGNOSTIC_MENU_TEXT

    assert "导出 sheet 文本诊断" in DIAGNOSTIC_MENU_TEXT
    assert "写入" not in DIAGNOSTIC_MENU_TEXT


def test_hermes_menu_help_includes_bat_commands_and_password_rule():
    from menu import build_hermes_help_lines

    text = "\n".join(build_hermes_help_lines())

    assert "run_hermes_hourly.bat 11点" in text
    assert "run_hermes_daily.bat" in text
    assert "2026-07-10" in text
    assert "不得询问或输出真实百度密码" in text
    assert "预约、到诊、就诊等禁止字段不由本工具填写" in text


def test_multi_source_menu_summary_only_displays_safe_configuration_metadata():
    from menu import build_baidu_source_summary_lines

    project = {
        "project_name": "多来源项目",
        "excel_accounts": [{"standard_name": "写入A"}],
        "baidu_sources": [
            {
                "source_id": "a",
                "source_name": "来源A",
                "credential_profile": "profile_a",
                "accounts": [{"standard_name": "写入A"}],
                "username": "secret-user",
                "password": "secret-password",
            },
            {
                "source_id": "b",
                "source_name": "来源B",
                "credential_profile": "profile_b",
                "accounts": [{"standard_name": "候选B"}],
            },
        ],
    }

    text = "\n".join(build_baidu_source_summary_lines(project))

    assert "百度来源数量：2" in text
    assert "来源A" in text
    assert "profile_a" in text
    assert "写入A" in text
    assert "候选B" in text
    assert "ignored_inactive_accounts" in text
    assert "skipped_unmapped_accounts" in text
    assert "secret-user" not in text
    assert "secret-password" not in text


def test_single_source_menu_summary_is_brief():
    from menu import build_baidu_source_summary_lines

    assert build_baidu_source_summary_lines({"project_name": "单来源"}) == [
        "当前项目为单百度来源项目。"
    ]


def test_console_home_context_shows_project_id_source_type_account_count_condition_and_short_excel_name(tmp_path):
    from io import StringIO

    from modules.console_ui import print_console_context, set_output_func

    (tmp_path / "reports").mkdir()
    (tmp_path / "reports" / "doctor_report.json").write_text(
        json.dumps({
            "project_id": "shenyang_niu",
            "summary": {"all_passed": True},
            "checks": {},
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_console_context({
            "project_name": "沈阳牛",
            "project_id": "shenyang_niu",
            "excel": {"path": r"D:\very\long\path\沈阳YXB2026竞价数据.xlsx"},
            "baidu_sources": [{}, {}],
            "excel_accounts": [{"standard_name": "A"}, {"standard_name": "B"}, {"standard_name": "C"}],
        }, root=tmp_path)
    finally:
        set_output_func(print)

    text = buf.getvalue()
    assert "沈阳牛" in text
    assert "shenyang_niu" in text
    assert "多百度来源 x2" in text
    assert "写入账户：3 个" in text
    assert "百度来源：2 个" not in text
    assert "条件项：通过" in text
    assert "沈阳YXB2026竞价数据.xlsx" in text
    assert r"D:\very\long\path" not in text


def test_console_home_condition_status_defaults_to_unchecked_and_maps_cached_results(tmp_path):
    from modules.console_ui import get_condition_status

    assert get_condition_status(tmp_path, "demo") == "未检查"
    reports = tmp_path / "reports"
    reports.mkdir()
    report_path = reports / "doctor_report.json"
    report_path.write_text(json.dumps({
        "project_id": "demo",
        "summary": {"all_passed": False},
        "checks": {"chrome": {"passed": False, "level": "warning"}},
    }), encoding="utf-8")
    assert get_condition_status(tmp_path, "demo") == "注意"
    report_path.write_text(json.dumps({
        "project_id": "demo",
        "summary": {"all_passed": False},
        "checks": {"excel": {"passed": False}},
    }), encoding="utf-8")
    assert get_condition_status(tmp_path, "demo") == "未通过"
    report_path.write_text(json.dumps({
        "project_id": "other-project",
        "summary": {"all_passed": True},
        "checks": {},
    }), encoding="utf-8")
    assert get_condition_status(tmp_path, "demo") == "未检查"


def test_console_home_has_plain_text_fallback_when_rich_is_disabled(tmp_path, monkeypatch):
    from io import StringIO

    import modules.console_ui as console_ui

    monkeypatch.setattr(console_ui, "_HAS_RICH", False)
    buf = StringIO()
    console_ui.set_output_func(buf.write)
    try:
        console_ui.print_console_context({"project_name": "演示", "project_id": "demo"}, root=tmp_path)
        console_ui.print_task_status_header({"project_id": "demo"}, root=tmp_path)
        console_ui.print_main_menu({"project_id": "demo"}, root=tmp_path)
    finally:
        console_ui.set_output_func(print)

    output = buf.getvalue()
    assert "百度竞价自动化控制台" in output
    assert "当前项目：演示" in output
    assert "条件项：未检查" in output
    assert "今日任务" in output
    assert "日报" in output
    assert "11点" in output
    assert "15点" in output
    assert "18点" in output
    assert "=" not in output


def test_rich_home_renders_one_project_panel_and_one_task_table_when_available(tmp_path, monkeypatch):
    import modules.console_ui as console_ui

    if not console_ui._HAS_RICH:
        return
    renderables = []
    monkeypatch.setattr(console_ui, "_emit_rich", lambda renderable: renderables.append(renderable) or True)
    try:
        console_ui.print_console_context({"project_name": "演示", "project_id": "demo"}, root=tmp_path)
        console_ui.print_task_status_header({"project_id": "demo"}, root=tmp_path)
    finally:
        console_ui.set_output_func(print)

    assert [type(item).__name__ for item in renderables] == ["Panel", "Table"]
    assert "百度竞价自动化控制台" in str(renderables[0].title)
    assert "今日任务" in str(renderables[1].title)


def test_menu_enhancement_does_not_introduce_textual_dependency():
    root = Path(__file__).resolve().parents[1]
    content = (
        (root / "menu.py").read_text(encoding="utf-8")
        + (root / "modules" / "console_ui.py").read_text(encoding="utf-8")
        + (root / "requirements.txt").read_text(encoding="utf-8")
    )

    assert "import textual" not in content.lower()
    assert "from textual" not in content.lower()
    assert "textual" not in (root / "requirements.txt").read_text(encoding="utf-8").lower()


def test_home_output_is_colleague_facing_and_keeps_advanced_sections_hidden(tmp_path, monkeypatch):
    import menu

    output = []
    project = {
        "project_id": "demo_bai",
        "project_name": "演示白",
        "excel": {"path": r"D:\data\【演示】2026竞价数据.xlsx"},
        "baidu_sources": [{}, {}],
        "excel_accounts": [{"standard_name": "A"}, {"standard_name": "B"}],
    }
    monkeypatch.setattr(menu, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(menu, "get_current_project", lambda root: project)
    monkeypatch.setattr(menu, "setup_logger", lambda path: object())

    menu.run_menu(root=tmp_path, input_func=lambda prompt: "0", output_func=output.append)

    home = "\n".join(output)
    assert home.count("百度竞价自动化控制台") == 1
    for text in ["百度竞价自动化控制台", "当前项目", "演示白", "demo_bai", "条件项", "日报", "11点", "15点", "18点"]:
        assert text in home
    for text in ["HERMES / 夏思道帮助", "配置诊断", "报告与日志", "多百度来源摘要", "高级分步调试"]:
        assert text not in home
    for text in ["doctor", "preflight", "credential_profile", "baidu_sources", "debug"]:
        assert text not in home.lower()


def test_submenu_labels_use_consistent_return_choice():
    from menu import (
        ADVANCED_DEBUG_MENU_TEXT,
        CONDITION_MENU_TEXT,
        DIAGNOSTIC_MENU_TEXT,
        HERMES_MENU_TEXT,
        MORE_FEATURES_MENU_TEXT,
        REPORT_MENU_TEXT,
    )

    for content in [
        ADVANCED_DEBUG_MENU_TEXT,
        CONDITION_MENU_TEXT,
        DIAGNOSTIC_MENU_TEXT,
        HERMES_MENU_TEXT,
        MORE_FEATURES_MENU_TEXT,
        REPORT_MENU_TEXT,
    ]:
        assert "0. 返回" in content


def test_menu_preflight_execution_writes_report_and_logs_result(tmp_path, monkeypatch):
    import menu

    entries = []

    class Logger:
        def info(self, *args):
            entries.append(args)

    monkeypatch.setattr(
        menu,
        "run_preflight",
        lambda root, project, config, task, quick: {
            "passed": True,
            "task": task,
            "quick": quick,
            "checks": [],
            "credentials": {},
        },
    )

    report = menu._execute_preflight(tmp_path, {}, {}, "daily", Logger())

    assert report["passed"] is True
    assert report["quick"] is True
    assert json.loads((tmp_path / "reports" / "preflight_report.json").read_text(encoding="utf-8"))["task"] == "daily"
    assert entries


def test_menu_refresh_choice_does_not_dispatch_task(tmp_path, monkeypatch):
    import menu

    project = {
        "project_id": "demo",
        "project_name": "演示",
        "excel": {"path": "demo.xlsx"},
    }
    answers = iter(["r", "0"])
    monkeypatch.setattr(menu, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(menu, "get_current_project", lambda root: project)
    monkeypatch.setattr(menu, "build_runtime_config", lambda current, base: {})
    monkeypatch.setattr(menu, "setup_logger", lambda path: object())
    monkeypatch.setattr(menu, "_check_chrome_debug", lambda *args, **kwargs: True)
    monkeypatch.setattr(menu, "dispatch_menu_task", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("刷新不得执行任务")))

    menu.run_menu(root=tmp_path, input_func=lambda prompt: next(answers), output_func=lambda text: None)


def test_menu_no_longer_routes_hidden_developer_shortcuts(tmp_path, monkeypatch):
    import menu

    project = {
        "project_id": "demo",
        "project_name": "演示",
        "excel": {"path": "demo.xlsx"},
    }
    calls = []
    answers = iter(["p", "l", "o", "0"])
    monkeypatch.setattr(menu, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(menu, "get_current_project", lambda root: project)
    monkeypatch.setattr(menu, "build_runtime_config", lambda current, base: {})
    monkeypatch.setattr(menu, "setup_logger", lambda path: object())
    monkeypatch.setattr(menu, "_check_chrome_debug", lambda *args, **kwargs: True)
    monkeypatch.setattr(menu, "_run_condition_menu", lambda *args: calls.append("condition"))
    monkeypatch.setattr(menu, "_run_report_menu", lambda *args: calls.append("reports"))
    monkeypatch.setattr(menu, "_run_hermes_menu", lambda *args: calls.append("hermes"))
    monkeypatch.setattr(menu, "dispatch_menu_task", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("快捷键不得执行任务")))

    menu.run_menu(root=tmp_path, input_func=lambda prompt: next(answers), output_func=lambda text: None)

    assert calls == []


def test_menu_routes_switch_project_and_condition_check_from_home(tmp_path, monkeypatch):
    import menu

    project = {
        "project_id": "demo",
        "project_name": "演示",
        "excel": {"path": "demo.xlsx"},
    }
    calls = []
    answers = iter(["3", "", "4", "0"])
    monkeypatch.setattr(menu, "load_config", lambda *args, **kwargs: {})
    monkeypatch.setattr(menu, "get_current_project", lambda root: project)
    monkeypatch.setattr(menu, "build_runtime_config", lambda current, base: {})
    monkeypatch.setattr(menu, "setup_logger", lambda path: object())
    monkeypatch.setattr(menu, "_check_chrome_debug", lambda *args, **kwargs: True)
    monkeypatch.setattr(menu, "_select_project_from_list", lambda *args: calls.append("switch") or None)
    monkeypatch.setattr(menu, "_run_condition_menu", lambda *args: calls.append("condition"))

    menu.run_menu(root=tmp_path, input_func=lambda prompt: next(answers), output_func=lambda text: None)

    assert calls == ["switch", "condition"]


# ── 返回逻辑测试 ──────────────────────────────────────────


def test_confirm_panel_shows_return_hint():
    """确认面板显示"回车执行 / 0 返回"。"""
    from io import StringIO
    from modules.console_ui import print_confirm_panel, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_confirm_panel({"task_name": "测试任务", "project_name": "测试项目"})
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "回车执行" in output
    assert "0 返回" in output
    assert "执行确认" in output


def test_hourly_submenu_shows_return():
    """小时报子菜单包含 0. 返回。"""
    from io import StringIO
    from modules.console_ui import print_sub_menu_hourly, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_sub_menu_hourly()
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "0. 返回" in output
    assert "11点" in output
    assert "15点" in output
    assert "18点" in output


def test_project_info_shows_return_hint():
    """项目信息页显示 0. 返回。"""
    from io import StringIO
    from modules.console_ui import print_project_info, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_project_info({"project_name": "测试", "project_id": "test"})
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "0. 返回" in output
    assert "项目信息" in output


# ── 自动打开 Excel 测试 ───────────────────────────────────


def test_try_open_excel_returns_false_for_nonexistent_file(tmp_path):
    """目标文件不存在时 try_open_excel 返回 False。"""
    from modules.console_ui import try_open_excel

    result = try_open_excel(tmp_path / "不存在.xlsx")
    assert result is False


def test_try_open_excel_returns_false_for_empty_path():
    """空路径时 try_open_excel 返回 False。"""
    from modules.console_ui import try_open_excel

    assert try_open_excel("") is False


def test_print_auto_open_result_shows_status():
    """自动打开结果输出正确状态。"""
    from io import StringIO
    from modules.console_ui import print_auto_open_result, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_auto_open_result(True)
        print_auto_open_result(False)
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "[通过]" in output
    assert "[注意]" in output
    assert "Excel 文件自动打开失败" in output


def test_auto_open_failure_does_not_raise():
    """自动打开失败不抛出异常。"""
    from modules.console_ui import try_open_excel

    result = try_open_excel("Z:/不存在的路径/文件.xlsx")
    assert result is False  # 不抛异常，只返回 False


# ── 默认输出降噪测试 ──────────────────────────────────────


def test_default_output_hides_report_path_in_pipeline_summary():
    """默认模式不输出 report/json 路径到终端（verbose_print 默认关闭）。"""
    from io import StringIO
    from modules.console_ui import set_output_func, set_verbose, verbose_print

    buf = StringIO()
    set_output_func(buf.write)
    set_verbose(False)
    try:
        verbose_print("报告：reports/final_run_report.json")
        verbose_print("reports/write_report.json")
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert output == ""  # 默认模式不应输出这些路径


def test_verbose_mode_shows_report_path():
    from io import StringIO
    from modules.console_ui import set_output_func, set_verbose, verbose_print

    buf = StringIO()
    set_output_func(buf.write)
    set_verbose(True)
    try:
        verbose_print("报告：reports/final_run_report.json")
    finally:
        set_output_func(print)
        set_verbose(False)

    output = buf.getvalue()
    assert "reports/final_run_report.json" in output


# ── 项目列表排除模板 ──────────────────────────────────────


def test_list_projects_excludes_project_template_json():
    """list_projects 不包含 project_template.json。"""
    root = Path(__file__).resolve().parents[1]
    projects = list_projects(root)

    ids = [p["project_id"] for p in projects]
    assert "your_project_id" not in ids, "列表不应包含模板 project_id"
    for p in projects:
        assert "project_template.json" not in p["path"], f"不应包含模板文件：{p['path']}"


def test_list_projects_excludes_is_template_true():
    """list_projects 排除 is_template=true 的配置。"""
    tmp = Path(__file__).resolve().parents[1] / "tests" / "__tmp_test_projects"
    tmp.mkdir(parents=True, exist_ok=True)
    try:
        (tmp / "configs").mkdir(exist_ok=True)
        (tmp / "configs" / "app_config.json").write_text(
            '{"default_project_id":"real_proj","projects_dir":"projects","secrets_file":"s.json"}',
            encoding="utf-8",
        )
        (tmp / "projects").mkdir(exist_ok=True)
        (tmp / "projects" / "real_proj.json").write_text(json.dumps({
            "project_id": "real_proj",
            "project_name": "真项目",
            "excel": {"path": "a.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
            "kst": {"export_dir": "e", "auto_pick_latest": True, "max_file_age_hours": 2},
            "baidu": {"credential_profile": "p", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
            "accounts": [
                {"standard_name": "A1", "baidu_names": ["B1"], "excel_name": "A1", "kst_ids": ["1"], "kst_names": ["K1"]},
                {"standard_name": "A2", "baidu_names": ["B2"], "excel_name": "A2", "kst_ids": ["2"], "kst_names": ["K2"]},
                {"standard_name": "A3", "baidu_names": ["B3"], "excel_name": "A3", "kst_ids": ["3"], "kst_names": ["K3"]},
            ],
            "hourly": {"periods": ["11点", "15点", "18点"]},
            "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
        }, ensure_ascii=False), encoding="utf-8")
        (tmp / "projects" / "template_proj.json").write_text(json.dumps({
            "is_template": True,
            "project_id": "template_proj",
            "project_name": "模板",
            "excel": {"path": "b.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
            "kst": {"export_dir": "e2", "auto_pick_latest": True, "max_file_age_hours": 2},
            "baidu": {"credential_profile": "p2", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
            "accounts": [
                {"standard_name": "T1", "baidu_names": ["B1"], "excel_name": "T1", "kst_ids": ["1"], "kst_names": ["K1"]},
                {"standard_name": "T2", "baidu_names": ["B2"], "excel_name": "T2", "kst_ids": ["2"], "kst_names": ["K2"]},
                {"standard_name": "T3", "baidu_names": ["B3"], "excel_name": "T3", "kst_ids": ["3"], "kst_names": ["K3"]},
            ],
            "hourly": {"periods": ["11点", "15点", "18点"]},
            "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
        }, ensure_ascii=False), encoding="utf-8")

        projects = list_projects(tmp)
        ids = [p["project_id"] for p in projects]
        assert "real_proj" in ids
        assert "template_proj" not in ids, "is_template=true 的项目不应出现在列表中"
    finally:
        import shutil
        shutil.rmtree(tmp)
        # clean up test dir


def test_list_projects_excludes_your_project_id():
    """list_projects 排除 project_id 为 your_project_id 的配置。"""
    tmp = Path(__file__).resolve().parents[1] / "tests" / "__tmp_test_projects_2"
    tmp.mkdir(parents=True, exist_ok=True)
    try:
        (tmp / "configs").mkdir(exist_ok=True)
        (tmp / "configs" / "app_config.json").write_text(
            '{"default_project_id":"good","projects_dir":"projects","secrets_file":"s.json"}',
            encoding="utf-8",
        )
        (tmp / "projects").mkdir(exist_ok=True)
        (tmp / "projects" / "good.json").write_text(json.dumps({
            "project_id": "good",
            "project_name": "好项目",
            "excel": {"path": "a.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
            "kst": {"export_dir": "e", "auto_pick_latest": True, "max_file_age_hours": 2},
            "baidu": {"credential_profile": "p", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
            "accounts": [
                {"standard_name": "A1", "baidu_names": ["B1"], "excel_name": "A1", "kst_ids": ["1"], "kst_names": ["K1"]},
                {"standard_name": "A2", "baidu_names": ["B2"], "excel_name": "A2", "kst_ids": ["2"], "kst_names": ["K2"]},
                {"standard_name": "A3", "baidu_names": ["B3"], "excel_name": "A3", "kst_ids": ["3"], "kst_names": ["K3"]},
            ],
            "hourly": {"periods": ["11点", "15点", "18点"]},
            "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
        }, ensure_ascii=False), encoding="utf-8")
        # 第二个文件：文件名不同，project_id 是 your_project_id
        (tmp / "projects" / "not_a_template.json").write_text(json.dumps({
            "project_id": "your_project_id",
            "project_name": "伪模板",
            "excel": {"path": "b.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
            "kst": {"export_dir": "e2", "auto_pick_latest": True, "max_file_age_hours": 2},
            "baidu": {"credential_profile": "p2", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
            "accounts": [
                {"standard_name": "T1", "baidu_names": ["B1"], "excel_name": "T1", "kst_ids": ["1"], "kst_names": ["K1"]},
                {"standard_name": "T2", "baidu_names": ["B2"], "excel_name": "T2", "kst_ids": ["2"], "kst_names": ["K2"]},
                {"standard_name": "T3", "baidu_names": ["B3"], "excel_name": "T3", "kst_ids": ["3"], "kst_names": ["K3"]},
            ],
            "hourly": {"periods": ["11点", "15点", "18点"]},
            "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
        }, ensure_ascii=False), encoding="utf-8")

        projects = list_projects(tmp)
        ids = [p["project_id"] for p in projects]
        assert "good" in ids
        assert "your_project_id" not in ids, "project_id=your_project_id 不应出现在列表中"
    finally:
        import shutil
        shutil.rmtree(tmp)
        # clean up test dir


def test_select_project_from_list_shows_real_projects_only(tmp_path):
    """_select_project_from_list 只展示真实项目，不含模板。"""
    from menu import _select_project_from_list

    projects_dir = tmp_path / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text(
        json.dumps({
            "default_project_id": "good",
            "projects_dir": "configs/projects",
            "secrets_file": "secrets/secrets.json",
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    def make_project(pid, name, excel_path="a.xlsx"):
        return json.dumps({
            "project_id": pid,
            "project_name": name,
            "excel": {"path": excel_path, "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
            "kst": {"export_dir": "e", "auto_pick_latest": True, "max_file_age_hours": 2},
            "baidu": {"credential_profile": "p", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
            "accounts": [
                {"standard_name": "A1", "baidu_names": ["B1"], "excel_name": "A1", "kst_ids": ["1"], "kst_names": ["K1"]},
                {"standard_name": "A2", "baidu_names": ["B2"], "excel_name": "A2", "kst_ids": ["2"], "kst_names": ["K2"]},
                {"standard_name": "A3", "baidu_names": ["B3"], "excel_name": "A3", "kst_ids": ["3"], "kst_names": ["K3"]},
            ],
            "hourly": {"periods": ["11点", "15点", "18点"]},
            "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
        }, ensure_ascii=False)

    (projects_dir / "good.json").write_text(make_project("good", "好项目"), encoding="utf-8")
    # 模板不应出现在列表中
    (projects_dir / "project_template.json").write_text(json.dumps({
        "is_template": True, "project_id": "your_project_id", "project_name": "模板",
        "excel": {"path": "t.xlsx", "hourly_sheet": "H", "daily_sheet": "D", "engine": "openpyxl"},
        "kst": {"export_dir": "e2", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "p2", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "accounts": [
            {"standard_name": "T1", "baidu_names": ["B1"], "excel_name": "T1", "kst_ids": ["1"], "kst_names": ["K1"]},
            {"standard_name": "T2", "baidu_names": ["B2"], "excel_name": "T2", "kst_ids": ["2"], "kst_names": ["K2"]},
            {"standard_name": "T3", "baidu_names": ["B3"], "excel_name": "T3", "kst_ids": ["3"], "kst_names": ["K3"]},
        ],
        "hourly": {"periods": ["11点", "15点", "18点"]},
        "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
    }, ensure_ascii=False), encoding="utf-8")

    # 选择项目 1（好项目）
    result = _select_project_from_list(tmp_path, lambda prompt: "1", lambda s: None)
    assert result is not None
    assert result["project_id"] == "good"


def test_select_project_list_returns_none_on_zero_input():
    """输入 0 返回主菜单。"""
    from menu import _select_project_from_list

    root = Path(__file__).resolve().parents[1]
    # 用真实 repo 测试；输入 0 应返回 None
    result = _select_project_from_list(root, lambda prompt: "0", lambda s: None)
    assert result is None


def test_select_project_list_handles_invalid_number():
    """非法编号不会 traceback，返回 None。"""
    from menu import _select_project_from_list

    root = Path(__file__).resolve().parents[1]
    result = _select_project_from_list(root, lambda prompt: "999", lambda s: None)
    assert result is None


def test_list_projects_excludes_template_in_real_repo():
    """真实仓库 list_projects 不包含 project_template。"""
    root = Path(__file__).resolve().parents[1]
    projects = list_projects(root)
    for p in projects:
        assert p["project_id"] != "your_project_id", "真实项目列表不应包含 your_project_id"
        assert "project_template.json" not in p["path"]


# ── 任务完成状态测试 ──────────────────────────────────────


def test_task_status_file_not_exists_auto_init(tmp_path):
    """状态文件不存在时自动初始化。"""
    from modules.task_status import load_task_status

    (tmp_path / "reports").mkdir(exist_ok=True)
    data = load_task_status(tmp_path)
    assert data["date"] == date.today().isoformat()
    assert data["projects"] == {}


def test_task_status_new_day_resets(tmp_path):
    """跨天自动重置状态。"""
    from modules.task_status import load_task_status, save_task_status

    (tmp_path / "reports").mkdir(exist_ok=True)
    old_data = {"date": "2026-01-01", "projects": {"demo": {"daily": {"done": True}}}}
    save_task_status(tmp_path, old_data)

    data = load_task_status(tmp_path)
    assert data["date"] == date.today().isoformat()
    assert data["projects"] == {}


def test_mark_daily_done_and_check():
    """mark_daily_done 后 is_daily_done 返回 True。"""
    import tempfile
    from modules.task_status import is_daily_done, mark_daily_done

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "reports").mkdir(exist_ok=True)
        assert is_daily_done(root, "test_proj") is False
        mark_daily_done(root, "test_proj")
        assert is_daily_done(root, "test_proj") is True


def test_mark_hourly_done_and_check():
    """mark_hourly_done 后 is_hourly_done 返回 True。"""
    import tempfile
    from modules.task_status import is_hourly_done, mark_hourly_done

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "reports").mkdir(exist_ok=True)
        assert is_hourly_done(root, "test_proj", "11点") is False
        mark_hourly_done(root, "test_proj", "11点")
        assert is_hourly_done(root, "test_proj", "11点") is True
        # 其他时段不变
        assert is_hourly_done(root, "test_proj", "15点") is False
        assert is_hourly_done(root, "test_proj", "18点") is False


def test_completed_status_does_not_block_repeat():
    """已完成状态不阻止重复执行 — 状态只是提示。"""
    import tempfile
    from modules.task_status import is_hourly_done, mark_hourly_done

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "reports").mkdir(exist_ok=True)
        mark_hourly_done(root, "proj", "11点")
        assert is_hourly_done(root, "proj", "11点") is True
        # 再次标记同一时段 — 不报错
        mark_hourly_done(root, "proj", "11点")
        assert is_hourly_done(root, "proj", "11点") is True


def test_mark_daily_does_not_affect_hourly():
    """日报标记不影响小时报状态。"""
    import tempfile
    from modules.task_status import is_daily_done, is_hourly_done, mark_daily_done

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "reports").mkdir(exist_ok=True)
        mark_daily_done(root, "proj")
        assert is_daily_done(root, "proj") is True
        assert is_hourly_done(root, "proj", "11点") is False


def test_doctor_does_not_mark_completion():
    """doctor 不标记完成 — 用 mock 模拟 doctor 执行。"""
    from modules.task_status import is_daily_done, mark_daily_done
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        root = Path(tmp)
        (root / "reports").mkdir(exist_ok=True)

        # 直接验证 task_status 逻辑，而非通过 dispatch_menu_task(5) 触发真实 doctor
        # doctor 的 dispatch 路径(choice='5')不会调用 mark 函数 — 这是菜单 run_menu 的职责
        assert is_daily_done(root, "proj") is False
        mark_daily_done(root, "proj")
        assert is_daily_done(root, "proj") is True

        # 验证：之前的 mark 成功，状态文件存在
        from modules.task_status import load_task_status
        data = load_task_status(root)
        assert data["projects"]["proj"]["daily"]["done"] is True


# ── 完成时间显示测试 ──────────────────────────────────────


def test_format_done_time_standard_format():
    """format_done_time 把 '2026-05-10 10:02:33' 转成 '10:02'。"""
    from modules.console_ui import format_done_time

    assert format_done_time("2026-05-10 10:02:33") == "10:02"
    assert format_done_time("2026-05-10 09:28:00") == "09:28"
    assert format_done_time("2026-05-10 14:15:59") == "14:15"


def test_format_done_time_iso_format():
    """format_done_time 把 '2026-05-10T10:02:33' 转成 '10:02'。"""
    from modules.console_ui import format_done_time

    assert format_done_time("2026-05-10T10:02:33") == "10:02"
    assert format_done_time("2026-05-10T09:28:00") == "09:28"


def test_format_done_time_already_short():
    """format_done_time 遇到已经是 '10:02' 格式的保持不变。"""
    from modules.console_ui import format_done_time

    assert format_done_time("10:02") == "10:02"


def test_format_done_time_empty_and_none():
    """format_done_time 遇到空值返回空字符串。"""
    from modules.console_ui import format_done_time

    assert format_done_time("") == ""
    assert format_done_time(None) == ""
    assert format_done_time("  ") == ""


def test_format_done_time_bad_format():
    """format_done_time 遇到异常格式不报错，返回空字符串。"""
    from modules.console_ui import format_done_time

    assert format_done_time("not-a-time") == ""
    assert format_done_time("abcdefg") == ""
    assert format_done_time("2026/05/10") == ""


def test_main_menu_keeps_daily_entry_free_of_status_duplication(tmp_path):
    """首页菜单只承担导航，完成状态只在今日任务表中展示。"""
    from io import StringIO
    from modules.console_ui import print_main_menu, set_output_func
    from modules.task_status import mark_daily_done

    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_daily_done(tmp_path, "test_proj")

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_main_menu({"project_id": "test_proj"}, root=tmp_path)
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "2. 日报" in output
    assert "[已完成]" not in output
    assert "完成于：" not in output


def test_task_status_header_shows_daily_completion_time(tmp_path):
    """今日任务表展示日报完成状态和时间。"""
    from io import StringIO
    from modules.console_ui import print_task_status_header, set_output_func
    from modules.task_status import mark_daily_done

    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_daily_done(tmp_path, "test_proj")

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_task_status_header({"project_id": "test_proj"}, root=tmp_path)
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "今日任务" in output
    assert "日报" in output
    assert "已完成" in output
    assert ":" in output


def test_hourly_submenu_shows_completion_time(tmp_path):
    """小时报子菜单已完成时段显示完成时间。"""
    from io import StringIO
    from modules.console_ui import print_sub_menu_hourly, set_output_func
    from modules.task_status import mark_hourly_done

    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_hourly_done(tmp_path, "test_proj", "11点")

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_sub_menu_hourly(root=tmp_path, project_id="test_proj")
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "11点" in output
    assert "[已完成]" in output
    assert "完成于：" in output
    # 15点未完成
    assert "15点" in output
    assert "[未完成]" in output


def test_main_menu_daily_entry_has_no_undone_status_duplication(tmp_path):
    """未完成状态也不在导航菜单重复展示。"""
    from io import StringIO
    from modules.console_ui import print_main_menu, set_output_func

    (tmp_path / "reports").mkdir(exist_ok=True)

    buf = StringIO()
    set_output_func(buf.write)
    try:
        print_main_menu({"project_id": "nonexistent"}, root=tmp_path)
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "2. 日报" in output
    assert "[未完成]" not in output
    assert "完成于：" not in output


def test_missing_time_does_not_error(tmp_path):
    """last_success_time 缺失时不报错。"""
    from modules.task_status import load_task_status, save_task_status

    (tmp_path / "reports").mkdir(exist_ok=True)
    # 手动构造没有 last_success_time 的状态
    data = {
        "date": __import__("datetime").date.today().isoformat(),
        "projects": {
            "test_proj": {
                "daily": {"done": True},  # 没有 last_success_time
                "hourly": {
                    "11点": {"done": False, "last_success_time": None},
                    "15点": {"done": False, "last_success_time": None},
                    "18点": {"done": False, "last_success_time": None},
                },
            }
        },
    }
    save_task_status(tmp_path, data)

    from io import StringIO
    from modules.console_ui import print_task_status_header, set_output_func

    buf = StringIO()
    set_output_func(buf.write)
    try:
        # 不应报错
        print_task_status_header({"project_id": "test_proj"}, root=tmp_path)
    finally:
        set_output_func(print)

    output = buf.getvalue()
    assert "已完成" in output


# ── 四项目预置测试 ────────────────────────────────────────


def test_list_projects_includes_four_niu_projects():
    """list_projects 能识别四个真实项目。"""
    root = Path(__file__).resolve().parents[1]
    projects = list_projects(root)
    ids = {p["project_id"] for p in projects}
    assert "kunming_niu" in ids
    assert "nanjing_niu" in ids
    assert "ningbo_niu" in ids
    assert "changsha_niu" in ids


def test_list_projects_excludes_old_kunming():
    """list_projects 不再识别 kunming_npx。"""
    root = Path(__file__).resolve().parents[1]
    projects = list_projects(root)
    ids = {p["project_id"] for p in projects}
    assert "kunming_npx" not in ids


def test_current_project_is_listed_and_valid():
    """当前选中项目可以随运行切换，但必须存在且配置有效。"""
    root = Path(__file__).resolve().parents[1]
    current = get_current_project(root)
    project_ids = {project["project_id"] for project in list_projects(root)}
    assert current["project_id"] in project_ids
    assert current["project_name"]
    assert validate_project_config(current) == []


def test_kunming_niu_accounts():
    """昆明牛账户映射为银康01、银康银屑02、银康03。"""
    root = Path(__file__).resolve().parents[1]
    from modules.project_config import build_runtime_config_from_project, load_project_config
    proj = load_project_config(root, "kunming_niu")
    runtime = build_runtime_config_from_project(proj, {})
    accounts = proj["accounts"]
    assert accounts[0]["standard_name"] == "银康01"
    assert accounts[0]["kst_ids"] == ["72828178"]
    assert accounts[1]["standard_name"] == "银康银屑02"
    assert accounts[1]["kst_ids"] == ["72828179"]
    assert accounts[2]["standard_name"] == "银康03"
    assert accounts[2]["kst_ids"] == ["81509165"]
    assert "baidu-银康03" in accounts[2]["baidu_names"]
    assert runtime["baidu"]["api_profile"] == "kunming_niu_baidu"


def test_data_source_mode_defaults_to_browser_and_validates_allowed_values():
    from copy import deepcopy
    from modules.project_config import build_runtime_config_from_project, load_project_config, validate_project_config

    root = Path(__file__).resolve().parents[1]
    project = load_project_config(root, "kunming_niu")
    without_mode = deepcopy(project)
    without_mode["baidu"].pop("data_source_mode", None)
    runtime = build_runtime_config_from_project(without_mode, {})
    assert runtime["baidu"]["configured_data_source_mode"] == "browser"
    assert runtime["baidu"]["data_source_preference"] == "api"
    assert runtime["baidu"]["data_source_mode"] == "api_preferred"

    shadow = deepcopy(project)
    shadow["baidu"]["data_source_mode"] = "api_shadow"
    assert validate_project_config(shadow) == []
    shadow_runtime = build_runtime_config_from_project(shadow, {})["baidu"]
    assert shadow_runtime["configured_data_source_mode"] == "api_shadow"
    assert shadow_runtime["data_source_mode"] == "api_preferred"

    invalid = deepcopy(project)
    invalid["baidu"]["data_source_mode"] = "fastest"
    assert any("data_source_mode" in error for error in validate_project_config(invalid))


def test_multi_source_project_allows_api_preferred_mode():
    from copy import deepcopy
    from modules.project_config import load_project_config, validate_project_config

    root = Path(__file__).resolve().parents[1]
    project = load_project_config(root, "shenyang_niu")
    api_project = deepcopy(project)
    api_project["baidu"]["data_source_mode"] = "api_preferred"

    errors = validate_project_config(api_project)

    assert not any("双来源项目" in error and "browser" in error for error in errors)


def test_multi_source_project_requires_api_profile_per_source():
    from copy import deepcopy
    from modules.project_config import load_project_config, validate_project_config

    root = Path(__file__).resolve().parents[1]
    project = deepcopy(load_project_config(root, "shenyang_niu"))
    project["baidu_sources"][1].pop("api_profile")

    errors = validate_project_config(project)

    assert any("baidu_sources[2].api_profile" in error for error in errors)


def test_project_template_remains_browser_and_only_kunming_uses_api_shadow():
    root = Path(__file__).resolve().parents[1]
    template = json.loads((root / "configs" / "projects" / "project_template.json").read_text(encoding="utf-8"))
    kunming = json.loads((root / "configs" / "projects" / "kunming_niu.json").read_text(encoding="utf-8"))

    assert template["baidu"]["data_source_mode"] == "browser"
    assert kunming["baidu"]["data_source_mode"] == "api_shadow"
    for path in (root / "configs" / "projects").glob("*.json"):
        if path.name in {"project_template.json", "kunming_niu.json"}:
            continue
        project = json.loads(path.read_text(encoding="utf-8"))
        assert project["baidu"].get("data_source_mode", "browser") == "browser"


def test_secrets_example_has_six_profiles():
    """secrets.example.json 包含含沈阳双来源在内的六个 profile，密码为空。"""
    root = Path(__file__).resolve().parents[1]
    data = json.loads((root / "secrets" / "secrets.example.json").read_text(encoding="utf-8"))
    baidu = data["baidu"]
    for profile in [
        "kunming_niu_baidu",
        "nanjing_niu_baidu",
        "ningbo_niu_baidu",
        "changsha_niu_baidu",
        "shenyang_niu_zhongya_baidu",
        "shenyang_niu_yinkang_baidu",
    ]:
        assert profile in baidu
        assert baidu[profile]["username"] == ""
        assert baidu[profile]["password"] == ""
    assert data["baidu_api"]["daily_automation"]["access_token"] == ""


def test_shenzhen_bai_has_real_account_mapping_and_example_credentials_profile():
    root = Path(__file__).resolve().parents[1]
    project = json.loads(
        (root / "configs" / "projects" / "shenzhen_bai.json").read_text(encoding="utf-8")
    )
    accounts = project["accounts"]

    assert project["excel"]["path"] == (
        "D:\\Seafile\\【竞价】\\【❤深圳组】\\【2026】【深圳】竞价数据\\【深圳】2026竞价数据.xlsx"
    )
    assert [account["standard_name"] for account in accounts] == ["益尚8", "益尚66", "益尚888"]
    assert [account["excel_name"] for account in accounts] == ["益尚8", "益尚66", "益尚888"]
    assert [account["kst_ids"] for account in accounts] == [["5289605"], ["5401012"], ["28492104"]]
    assert all("TODO_" not in json.dumps(account, ensure_ascii=False) for account in accounts)

    example = json.loads((root / "secrets" / "secrets.example.json").read_text(encoding="utf-8"))
    assert example["baidu"]["shenzhen_bai_baidu"] == {"username": "", "password": ""}


def test_nanjing_bai_has_real_account_mapping_and_example_credentials_profile():
    root = Path(__file__).resolve().parents[1]
    project = json.loads(
        (root / "configs" / "projects" / "nanjing_bai.json").read_text(encoding="utf-8")
    )
    accounts = project["accounts"]

    assert project["excel"]["path"] == (
        "D:\\Seafile\\【竞价】\\【❤南京白】\\【2026年】【南京白】竞价数据\\【南京华厦bdf】2026竞价数据.xlsx"
    )
    expected_names = ["华厦04", "华厦05"]
    assert [account["standard_name"] for account in accounts] == expected_names
    assert [account["excel_name"] for account in accounts] == expected_names
    assert [account["kst_ids"] for account in accounts] == [
        ["65700427"],
        ["65742504"],
    ]
    assert [account["baidu_names"] for account in accounts] == [["华厦B04"], ["华厦B05"]]
    assert all("TODO_" not in json.dumps(account, ensure_ascii=False) for account in accounts)

    example = json.loads((root / "secrets" / "secrets.example.json").read_text(encoding="utf-8"))
    assert example["baidu"]["nanjing_bai_baidu"] == {"username": "", "password": ""}


def test_baidu_api_identity_prefers_oauth_master_name(tmp_path):
    from modules.baidu_report_api import _load_api_identity

    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu": {
            "credential_profile": "nanjing_niu_baidu",
            "api_profile": "nanjing_niu_baidu",
        },
    }
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu": {"nanjing_niu_baidu": {"username": "browser-login"}},
        "baidu_api": {"nanjing_niu_baidu": {"master_name": "oauth-master"}},
    }), encoding="utf-8")

    username, api_profile, _secrets = _load_api_identity(tmp_path, config)

    assert username == "oauth-master"
    assert api_profile == "nanjing_niu_baidu"


def test_baidu_api_identity_falls_back_to_browser_username_for_legacy_oauth_record(tmp_path):
    from modules.baidu_report_api import _load_api_identity

    config = {
        "credentials_path": "secrets/secrets.json",
        "baidu": {
            "credential_profile": "legacy_baidu",
            "api_profile": "legacy_baidu",
        },
    }
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu": {"legacy_baidu": {"username": "legacy-manager"}},
        "baidu_api": {"legacy_baidu": {}},
    }), encoding="utf-8")

    username, api_profile, _secrets = _load_api_identity(tmp_path, config)

    assert username == "legacy-manager"
    assert api_profile == "legacy_baidu"


def test_baidu_api_probe_maps_ids_and_validates_summary(tmp_path):
    config = _kunming_niu_runtime_config()
    config.update({
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "credentials_path": "secrets/secrets.json",
        "baidu": {
            "credential_profile": "kunming_niu_baidu",
            "api_profile": "daily_automation",
        },
    })
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu": {"kunming_niu_baidu": {"username": "manager", "password": "unused"}},
        "baidu_api": {"daily_automation": {"access_token": "header.payload.signature"}},
    }, ensure_ascii=False), encoding="utf-8")

    def fake_transport(url, payload, timeout):
        assert url.endswith("OpenApiReportService/getReportData")
        assert payload["header"]["userName"] == "manager"
        assert payload["header"]["accessToken"] == "header.payload.signature"
        assert payload["body"]["userIds"] == [72828178, 72828179, 81509165]
        assert timeout == 30
        return {
            "header": {"status": 0, "desc": "success", "failures": []},
            "body": {"data": [{
                "rowCount": 3,
                "totalRowCount": 3,
                "rows": [
                    {"userId": 81509165, "userName": "baidu-银康03", "impression": 12, "click": 2, "cost": 3.5},
                    {"userId": 72828179, "userName": "银康银屑02", "impression": 20, "click": 4, "cost": 8.25},
                    {"userId": 72828178, "userName": "银康01", "impression": 100, "click": 10, "cost": 50},
                ],
                "summary": {"impression": 132, "click": 16, "cost": 61.75},
            }]},
        }

    import logging
    report = fetch_baidu_api_probe(
        config,
        tmp_path,
        logging.getLogger("test"),
        target_date="2026-07-14",
        period="15点",
        transport=fake_transport,
    )

    assert report["errors"] == []
    assert report["accounts"]["银康01"]["消费"] == 50.0
    assert report["accounts"]["银康03"]["source_user_id"] == 81509165
    assert report["diagnostics"]["account_totals"] == {"impression": 132, "click": 16, "cost": 61.75}
    assert report["self_check"]["wrote_excel"] is False
    serialized = (tmp_path / "reports" / "baidu_api_probe_report.json").read_text(encoding="utf-8")
    assert "header.payload.signature" not in serialized


def test_baidu_api_probe_rejects_summary_mismatch(tmp_path):
    config = _kunming_niu_runtime_config()
    config.update({
        "credentials_path": "secrets/secrets.json",
        "baidu": {"credential_profile": "kunming_niu_baidu", "api_profile": "daily_automation"},
    })
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu": {"kunming_niu_baidu": {"username": "manager"}},
        "baidu_api": {"daily_automation": {"access_token": "header.payload.signature"}},
    }), encoding="utf-8")

    def fake_transport(_url, _payload, _timeout):
        rows = [
            {"userId": 72828178, "userName": "银康01", "impression": 1, "click": 1, "cost": 1},
            {"userId": 72828179, "userName": "银康银屑02", "impression": 1, "click": 1, "cost": 1},
            {"userId": 81509165, "userName": "银康03", "impression": 1, "click": 1, "cost": 1},
        ]
        return {
            "header": {"status": 0, "desc": "success", "failures": []},
            "body": {"data": [{"rows": rows, "summary": {"impression": 99, "click": 3, "cost": 3}}]},
        }

    import logging
    report = fetch_baidu_api_probe(config, tmp_path, logging.getLogger("test"), transport=fake_transport)

    assert any("汇总校验失败" in error for error in report["errors"])
    assert report["self_check"]["passed"] is False


def _api_production_config(tmp_path):
    config = _kunming_niu_runtime_config()
    config.update({
        "project_id": "kunming_niu",
        "project_name": "昆明牛",
        "credentials_path": "secrets/secrets.json",
        "baidu": {
            "credential_profile": "kunming_niu_baidu",
            "api_profile": "kunming_niu_baidu",
            "api_timeout_seconds": 30,
        },
    })
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu": {"kunming_niu_baidu": {"username": "manager", "password": "unused"}},
        "baidu_api": {"kunming_niu_baidu": {"access_token": "unused.by.provider"}},
    }, ensure_ascii=False), encoding="utf-8")
    return config


def _api_success_response(target_date, *, rows=None, summary=None):
    rows = rows if rows is not None else [
        {"date": target_date, "userId": 72828178, "userName": "银康01", "impression": 100, "click": 10, "cost": 50},
        {"date": target_date, "userId": 72828179, "userName": "银康银屑02", "impression": 20, "click": 4, "cost": 8.25},
        {"date": target_date, "userId": 81509165, "userName": "baidu-银康03", "impression": 12, "click": 2, "cost": 3.5},
    ]
    summary = summary if summary is not None else {"impression": 132, "click": 16, "cost": 61.75}
    return {
        "header": {"status": 0, "desc": "success", "failures": []},
        "body": {"data": [{"rowCount": len(rows), "totalRowCount": len(rows), "rows": rows, "summary": summary}]},
    }


@pytest.mark.parametrize(
    ("status_code", "expected_category"),
    [(400, "api_error"), (401, "authorization_error"), (403, "authorization_error"), (500, "network_error"), (503, "network_error")],
)
def test_baidu_report_http_status_retry_categories(monkeypatch, status_code, expected_category):
    import io
    import urllib.error
    import modules.baidu_report_api as report_api

    def raise_http_error(_request, timeout):
        assert timeout == 5
        raise urllib.error.HTTPError(
            report_api.REPORT_API_URL,
            status_code,
            "failure",
            {},
            io.BytesIO(b'{"message":"header.payload.signature"}'),
        )

    monkeypatch.setattr(report_api.urllib.request, "urlopen", raise_http_error)
    with pytest.raises(report_api.BaiduReportApiError) as exc_info:
        report_api._post_json(report_api.REPORT_API_URL, {}, 5)

    assert exc_info.value.category == expected_category
    assert "header.payload.signature" not in str(exc_info.value)


def test_baidu_api_hourly_and_daily_write_compatible_standard_reports(tmp_path):
    import logging
    from datetime import date
    from modules.baidu_report_api import fetch_baidu_api_daily, fetch_baidu_api_hourly

    config = _api_production_config(tmp_path)
    today = date.today().isoformat()
    token_calls = []

    def token_provider(_config, _root, profile, **kwargs):
        token_calls.append((profile, kwargs.get("force_refresh", False)))
        return "header.payload.signature", {
            "api_profile": profile,
            "token_refresh": "not_needed",
            "expires_time": "2026-07-18 09:00:00",
        }

    def transport(_url, payload, _timeout):
        target_date = payload["body"]["startDate"]
        assert payload["body"]["timeUnit"] == "DAY"
        return _api_success_response(target_date)

    hourly = fetch_baidu_api_hourly(
        config, tmp_path, logging.getLogger("api-hourly"), "15点",
        token_provider=token_provider, transport=transport,
    )
    daily = fetch_baidu_api_daily(
        config, tmp_path, logging.getLogger("api-daily"), "2026-07-16",
        token_provider=token_provider, transport=transport,
    )

    assert hourly["source"] == "baidu_open_api"
    assert hourly["date"] == today
    assert hourly["period"] == "15点"
    assert hourly["accounts"]["银康01"]["消费"] == 50.0
    assert daily["source"] == "baidu_open_api"
    assert daily["date"] == "2026-07-16"
    assert daily["target_date"] == "2026-07-16"
    assert json.loads((tmp_path / "reports" / "baidu_account_data.json").read_text("utf-8"))["accounts"] == hourly["accounts"]
    assert json.loads((tmp_path / "reports" / "baidu_daily_data.json").read_text("utf-8"))["accounts"] == daily["accounts"]
    assert token_calls == [("kunming_niu_baidu", False), ("kunming_niu_baidu", False)]


def test_baidu_api_expired_access_token_forces_one_refresh_and_retries(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    token_calls = []
    responses = [
        {"header": {"status": 1, "desc": "failure", "failures": [{"code": 894061}]}},
        _api_success_response("2026-07-16"),
    ]

    def token_provider(_config, _root, _profile, **kwargs):
        forced = kwargs.get("force_refresh", False)
        token_calls.append(forced)
        return ("new.token.value" if forced else "old.token.value"), {
            "api_profile": "kunming_niu_baidu", "token_refresh": "refreshed" if forced else "not_needed", "expires_time": "x"
        }

    report = fetch_baidu_api_daily(
        config, tmp_path, logging.getLogger("api-refresh"), "2026-07-16",
        token_provider=token_provider, transport=lambda *_args: responses.pop(0),
    )

    assert report["errors"] == []
    assert token_calls == [False, True]


@pytest.mark.parametrize("failure_code", [894062, 894063, 894064])
def test_baidu_api_marks_revoked_authorization_for_reauthorization(tmp_path, failure_code):
    import logging
    from modules.baidu_report_api import BaiduReportApiError, fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    with pytest.raises(BaiduReportApiError) as exc_info:
        fetch_baidu_api_daily(
            config, tmp_path, logging.getLogger("api-reauth"), "2026-07-16",
            token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
            transport=lambda *_args: {"header": {"status": 1, "desc": "failure", "failures": [{"code": failure_code}]}},
        )
    assert exc_info.value.category == "reauthorization_required"
    assert exc_info.value.reauthorization_required is True


@pytest.mark.parametrize("failure_code", [89405, 89406, 89407])
def test_baidu_api_classifies_other_authorization_errors_without_token_leak(tmp_path, failure_code):
    import logging
    from modules.baidu_report_api import BaiduReportApiError, fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    with pytest.raises(BaiduReportApiError) as exc_info:
        fetch_baidu_api_daily(
            config, tmp_path, logging.getLogger("api-auth-error"), "2026-07-16",
            token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
            transport=lambda *_args: {
                "header": {
                    "status": 1,
                    "desc": "header.payload.signature",
                    "failures": [{"code": failure_code}],
                }
            },
        )
    assert exc_info.value.category == "authorization_error"
    assert exc_info.value.reauthorization_required is False
    assert "header.payload.signature" not in str(exc_info.value)


@pytest.mark.parametrize("case", ["date", "missing_date", "unknown", "duplicate", "negative", "nonfinite", "summary"])
def test_baidu_api_integrity_failure_preserves_previous_standard_report(tmp_path, case):
    import logging
    from modules.baidu_report_api import BaiduReportApiError, fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    standard_path = tmp_path / "reports" / "baidu_daily_data.json"
    standard_path.parent.mkdir(parents=True)
    standard_path.write_bytes(b'{"previous":true}')
    response = _api_success_response("2026-07-16")
    rows = response["body"]["data"][0]["rows"]
    if case == "date":
        rows[0]["date"] = "2026-07-15"
    elif case == "missing_date":
        rows[0].pop("date")
    elif case == "unknown":
        rows[0]["userId"] = 99999999
    elif case == "duplicate":
        rows[1]["userId"] = rows[0]["userId"]
    elif case == "negative":
        rows[0]["cost"] = -50
        response["body"]["data"][0]["summary"]["cost"] = -38.25
    elif case == "nonfinite":
        rows[0]["cost"] = float("nan")
    elif case == "summary":
        response["body"]["data"][0]["summary"]["cost"] = 99

    with pytest.raises(BaiduReportApiError) as exc_info:
        fetch_baidu_api_daily(
            config, tmp_path, logging.getLogger("api-integrity"), "2026-07-16",
            token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
            transport=lambda *_args: response,
        )

    assert exc_info.value.category == "integrity_error"
    assert standard_path.read_bytes() == b'{"previous":true}'
    attempt = json.loads((tmp_path / "reports" / "baidu_api_attempt_report.json").read_text("utf-8"))
    assert attempt["passed"] is False
    assert "header.payload.signature" not in json.dumps(attempt)


def test_baidu_api_accepts_complete_all_zero_accounts(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    rows = [
        {"date": "2026-07-16", "userId": user_id, "userName": name, "impression": 0, "click": 0, "cost": 0}
        for user_id, name in ((72828178, "银康01"), (72828179, "银康银屑02"), (81509165, "银康03"))
    ]
    report = fetch_baidu_api_daily(
        config, tmp_path, logging.getLogger("api-zero"), "2026-07-16",
        token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
        transport=lambda *_args: _api_success_response(
            "2026-07-16", rows=rows, summary={"impression": 0, "click": 0, "cost": 0}
        ),
    )
    assert report["errors"] == []
    assert all(account["消费"] == 0 for account in report["accounts"].values())


def test_baidu_api_zero_fills_omitted_requested_account_after_summary_reconciliation(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    rows = [
        {"date": "2026-07-16", "userId": 72828178, "userName": "银康01", "impression": 100, "click": 10, "cost": 50},
        {"date": "2026-07-16", "userId": 72828179, "userName": "银康银屑02", "impression": 20, "click": 4, "cost": 8.25},
    ]
    report = fetch_baidu_api_daily(
        config,
        tmp_path,
        logging.getLogger("api-zero-fill-partial"),
        "2026-07-16",
        token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
        transport=lambda *_args: _api_success_response(
            "2026-07-16",
            rows=rows,
            summary={"impression": 120, "click": 14, "cost": 58.25},
        ),
        commit_standard_report=False,
    )

    assert set(report["accounts"]) == {"银康01", "银康银屑02", "银康03"}
    assert report["accounts"]["银康03"] == {
        "source_account": "银康03",
        "source_user_id": 81509165,
        "展现": 0,
        "点击": 0,
        "消费": 0.0,
        "synthetic_zero": True,
    }
    assert report["diagnostics"]["zero_filled_accounts"] == ["银康03"]
    assert report["diagnostics"]["zero_filled_count"] == 1
    assert report["diagnostics"]["account_totals"] == {
        "impression": 120,
        "click": 14,
        "cost": 58.25,
    }


def test_baidu_api_zero_fills_all_requested_accounts_when_rows_empty_and_summary_is_zero(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    report = fetch_baidu_api_daily(
        config,
        tmp_path,
        logging.getLogger("api-zero-fill-empty"),
        "2026-07-16",
        token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
        transport=lambda *_args: _api_success_response(
            "2026-07-16",
            rows=[],
            summary={"impression": 0, "click": 0, "cost": 0},
        ),
        commit_standard_report=False,
    )

    assert list(report["accounts"]) == ["银康01", "银康银屑02", "银康03"]
    assert all(row["synthetic_zero"] is True for row in report["accounts"].values())
    assert all(row["source_user_id"] in {72828178, 72828179, 81509165} for row in report["accounts"].values())
    assert report["diagnostics"]["zero_filled_accounts"] == ["银康01", "银康银屑02", "银康03"]
    assert report["diagnostics"]["zero_filled_count"] == 3


@pytest.mark.parametrize(
    ("summary", "expected_error"),
    [
        ({}, "未返回完整汇总指标"),
        ({"impression": 121, "click": 14, "cost": 58.25}, "汇总校验失败"),
    ],
)
def test_baidu_api_does_not_zero_fill_when_summary_is_missing_or_mismatched(tmp_path, summary, expected_error):
    from modules.baidu_report_api import _account_user_ids, _parse_api_response

    config = _api_production_config(tmp_path)
    _user_ids, account_by_id = _account_user_ids(config)
    rows = [
        {"date": "2026-07-16", "userId": 72828178, "userName": "银康01", "impression": 100, "click": 10, "cost": 50},
        {"date": "2026-07-16", "userId": 72828179, "userName": "银康银屑02", "impression": 20, "click": 4, "cost": 8.25},
    ]

    accounts, diagnostics, errors = _parse_api_response(
        _api_success_response("2026-07-16", rows=rows, summary=summary),
        config=config,
        account_by_id=account_by_id,
        expected_date="2026-07-16",
    )

    assert set(accounts) == {"银康01", "银康银屑02"}
    assert not any(row.get("synthetic_zero") for row in accounts.values())
    assert diagnostics["zero_filled_accounts"] == []
    assert diagnostics["zero_filled_count"] == 0
    assert any(expected_error in error for error in errors)


@pytest.mark.parametrize("invalid_case", ["unknown", "duplicate", "number", "date"])
def test_baidu_api_invalid_rows_never_trigger_zero_fill(tmp_path, invalid_case):
    from modules.baidu_report_api import _account_user_ids, _parse_api_response

    config = _api_production_config(tmp_path)
    _user_ids, account_by_id = _account_user_ids(config)
    rows = [
        {"date": "2026-07-16", "userId": 72828178, "userName": "银康01", "impression": 100, "click": 10, "cost": 50},
        {"date": "2026-07-16", "userId": 72828179, "userName": "银康银屑02", "impression": 20, "click": 4, "cost": 8.25},
    ]
    if invalid_case == "unknown":
        rows[0]["userId"] = 99999999
    elif invalid_case == "duplicate":
        rows[1]["userId"] = rows[0]["userId"]
    elif invalid_case == "number":
        rows[0]["cost"] = "not-a-number"
    else:
        rows[0]["date"] = "2026-07-15"

    accounts, diagnostics, errors = _parse_api_response(
        _api_success_response(
            "2026-07-16",
            rows=rows,
            summary={"impression": 120, "click": 14, "cost": 58.25},
        ),
        config=config,
        account_by_id=account_by_id,
        expected_date="2026-07-16",
    )

    assert errors
    assert diagnostics["zero_filled_accounts"] == []
    assert diagnostics["zero_filled_count"] == 0
    assert not any(row.get("synthetic_zero") for row in accounts.values())


def test_baidu_api_shadow_read_does_not_commit_standard_report(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    report = fetch_baidu_api_daily(
        config, tmp_path, logging.getLogger("api-shadow"), "2026-07-16",
        token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
        transport=lambda *_args: _api_success_response("2026-07-16"),
        commit_standard_report=False,
    )

    assert report["errors"] == []
    assert report["self_check"]["production_output_replaced"] is False
    assert not (tmp_path / "reports" / "baidu_daily_data.json").exists()


def _route_report(source="baidu_open_api", cost=10.0, errors=None):
    return {
        "date": "2026-07-16",
        "period": "15点",
        "source": source,
        "accounts": {
            "银康01": {"source_user_id": 72828178, "展现": 100, "点击": 10, "消费": cost},
        },
        "errors": list(errors or []),
    }


def _two_source_route_config():
    def account(name, promotion_id):
        return {
            "standard_name": name,
            "baidu_names": [name],
            "excel_name": name,
            "kst_ids": [promotion_id],
            "kst_names": [name],
        }

    return {
        "project_id": "two_source_demo",
        "project_name": "双来源演示",
        "baidu": {"data_source_mode": "api_preferred"},
        "accounts": {"账户A": {}, "账户B": {}},
        "baidu_sources": [
            {
                "source_id": "a",
                "source_name": "来源A",
                "credential_profile": "browser_a",
                "api_profile": "api_a",
                "accounts": [account("账户A", "1")],
            },
            {
                "source_id": "b",
                "source_name": "来源B",
                "credential_profile": "browser_b",
                "api_profile": "api_b",
                "accounts": [account("账户B", "2")],
            },
        ],
    }


def _two_source_browser_report(errors=None):
    return {
        "date": "2026-07-17",
        "period": "18点",
        "source": "baidu_multi_source",
        "accounts": {
            "账户A": {"展现": 2, "点击": 2, "消费": 2.0},
            "账户B": {"展现": 3, "点击": 3, "消费": 3.0},
        },
        "errors": list(errors or []),
    }


def test_source_runtime_config_copies_api_profile():
    from modules.baidu_multi_source import build_source_runtime_config

    runtime = build_source_runtime_config(
        {"baidu": {}, "accounts": {}},
        {
            "source_id": "a",
            "credential_profile": "browser_a",
            "api_profile": "api_a",
            "accounts": [{
                "standard_name": "账户A",
                "baidu_names": ["账户A"],
                "excel_name": "账户A",
                "kst_ids": ["1"],
                "kst_names": ["账户A"],
            }],
        },
    )

    assert runtime["baidu"]["credential_profile"] == "browser_a"
    assert runtime["baidu"]["api_profile"] == "api_a"


def test_api_preferred_multi_source_commits_only_after_all_sources_pass(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    calls = []

    def api_fetcher(*, config, commit_standard_report, deadline, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        calls.append((source_id, config["baidu"]["api_profile"], commit_standard_report, deadline))
        account = next(iter(config["accounts"]))
        return {
            "date": "2026-07-17",
            "period": "18点",
            "accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}},
            "errors": [],
        }

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: pytest.fail("API 成功不应调用浏览器"),
        clock=lambda: 10.0,
        sleep=lambda _seconds: None,
    )

    assert calls == [
        ("a", "api_a", False, 30.0),
        ("b", "api_b", False, 30.0),
    ]
    assert result["data_source"] == "api"
    standard = json.loads((tmp_path / "reports" / "baidu_account_data.json").read_text("utf-8"))
    assert standard["data_source"] == "api"
    assert standard["accounts"] == result["accounts"]


def test_multi_source_api_uses_independent_refresh_contexts(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    contexts = {"a": [], "b": []}
    refresh_counts = {"a": 0, "b": 0}

    def api_fetcher(*, config, task_context, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        contexts[source_id].append(task_context)
        if not task_context["refresh_attempted"]:
            task_context["refresh_attempted"] = True
            task_context["self_heal_actions"].append("token_refresh")
            refresh_counts[source_id] += 1
            raise BaiduReportApiError("retry after refresh", category="network_error")
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: pytest.fail("不应调用浏览器"),
        clock=lambda: 0.0,
        sleep=lambda _seconds: None,
    )

    assert refresh_counts == {"a": 1, "b": 1}
    assert len(contexts["a"]) == len(contexts["b"]) == 2
    assert contexts["a"][0] is contexts["a"][1]
    assert contexts["b"][0] is contexts["b"][1]
    assert contexts["a"][0] is not contexts["b"][0]
    assert result["api_attempts"] == 4
    assert result["self_heal_actions"] == ["token_refresh", "network_retry"]


def test_multi_source_api_shares_one_project_deadline(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    now = [0.0]
    api_calls = []
    browser_calls = []

    def api_fetcher(*, config, deadline, **_kwargs):
        api_calls.append((config["baidu_source"]["source_id"], deadline))
        now[0] = 20.0
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(),
        clock=lambda: now[0],
        sleep=lambda _seconds: None,
    )

    assert api_calls == [("a", 20.0)]
    assert browser_calls == [True]
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "api_budget_exhausted"


def test_multi_source_api_discards_last_source_success_after_deadline(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    now = [0.0]
    api_calls = []
    browser_calls = []

    def api_fetcher(*, config, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        api_calls.append(source_id)
        if source_id == "b":
            now[0] = 20.1
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(),
        clock=lambda: now[0],
        sleep=lambda _seconds: None,
    )

    assert api_calls == ["a", "b"]
    assert browser_calls == [True]
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "api_budget_exhausted"


def test_multi_source_api_rechecks_deadline_before_atomic_commit(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    clock_values = iter([0.0, 0.0, 0.0, 19.0, 21.0])
    browser_calls = []

    def api_fetcher(*, config, **_kwargs):
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(),
        clock=lambda: next(clock_values),
        sleep=lambda _seconds: None,
    )

    assert browser_calls == [True]
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "api_budget_exhausted"


def test_api_preferred_multi_source_failure_discards_partial_and_falls_back(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    standard = tmp_path / "reports" / "baidu_account_data.json"
    standard.parent.mkdir()
    standard.write_text('{"sentinel": true}', encoding="utf-8")
    api_calls = []
    browser_calls = []

    def api_fetcher(*, config, commit_standard_report, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        api_calls.append((source_id, commit_standard_report))
        if source_id == "b":
            raise RuntimeError("source b failed")
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    def browser_fetcher(*, config, **_kwargs):
        browser_calls.append([source["source_id"] for source in config["baidu_sources"]])
        return _two_source_browser_report()

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=browser_fetcher,
        sleep=lambda _seconds: None,
    )

    assert api_calls == [("a", False), ("b", False)]
    assert browser_calls == [["a", "b"]]
    assert result["data_source"] == "browser_fallback"
    saved = json.loads(standard.read_text("utf-8"))
    assert "sentinel" not in saved
    assert saved["data_source"] == "browser_fallback"


def test_multi_source_api_account_conflict_falls_back_without_api_commit(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser_calls = []

    def api_fetcher(*, config, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        account = "账户A" if source_id == "b" else next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(),
        sleep=lambda _seconds: None,
    )

    assert browser_calls == [True]
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "integrity_error"
    assert json.loads((tmp_path / "reports" / "baidu_account_data.json").read_text("utf-8"))["data_source"] == "browser_fallback"


def test_multi_source_api_cost_total_anomaly_falls_back(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser_calls = []

    def api_fetcher(*, config, **_kwargs):
        source_id = config["baidu_source"]["source_id"]
        account = next(iter(config["accounts"]))
        accounts = {account: {"展现": 1, "点击": 1, "消费": 1.0}}
        if source_id == "b":
            accounts["未映射消费"] = {"展现": 1, "点击": 1, "消费": 5.0}
        return {"accounts": accounts, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(),
        sleep=lambda _seconds: None,
    )

    assert browser_calls == [True]
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "integrity_error"


def test_multi_source_api_and_browser_failure_returns_failed_without_replacing_standard(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    standard = tmp_path / "reports" / "baidu_account_data.json"
    standard.parent.mkdir()
    standard.write_text('{"sentinel": true}', encoding="utf-8")
    browser_calls = []

    def api_fetcher(*, config, **_kwargs):
        if config["baidu_source"]["source_id"] == "b":
            raise RuntimeError("source b failed")
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: browser_calls.append(True) or _two_source_browser_report(["browser failed"]),
        sleep=lambda _seconds: None,
    )

    assert browser_calls == [True]
    assert result["data_source"] == "failed"
    assert result["errors"]
    assert json.loads(standard.read_text("utf-8")) == {"sentinel": True}


def test_api_preferred_multi_source_browser_side_effect_failure_preserves_hourly_canonical(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    canonical = tmp_path / "reports" / "baidu_account_data.json"
    canonical.parent.mkdir()
    canonical.write_text('{"sentinel": "hourly"}', encoding="utf-8")
    staged_paths = []

    def api_fetcher(*, config, **_kwargs):
        if config["baidu_source"]["source_id"] == "b":
            raise RuntimeError("source b failed")
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    def browser_fetcher(*, config, **_kwargs):
        staged = Path(config["baidu"]["output_path"])
        staged_paths.append(staged)
        staged.parent.mkdir(parents=True, exist_ok=True)
        staged.write_text('{"errors": ["browser failed"]}', encoding="utf-8")
        return _two_source_browser_report(["browser failed"])

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        None,
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=browser_fetcher,
        sleep=lambda _seconds: None,
    )

    assert result["data_source"] == "failed"
    assert json.loads(canonical.read_text("utf-8")) == {"sentinel": "hourly"}
    assert len(staged_paths) == 1
    assert staged_paths[0] != canonical
    assert staged_paths[0].parent == canonical.parent
    assert not staged_paths[0].exists()


def test_api_preferred_multi_source_daily_success_commits_only_daily_canonical(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_daily

    def api_fetcher(*, config, **_kwargs):
        account = next(iter(config["accounts"]))
        return {
            "date": "2026-07-16",
            "target_date": "2026-07-16",
            "accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}},
            "errors": [],
        }

    result = fetch_baidu_resilient_daily(
        _two_source_route_config(),
        tmp_path,
        None,
        "2026-07-16",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: pytest.fail("API 成功不应调用浏览器"),
        clock=lambda: 0.0,
        sleep=lambda _seconds: None,
    )

    daily_path = tmp_path / "reports" / "baidu_daily_data.json"
    assert result["data_source"] == "api"
    assert json.loads(daily_path.read_text("utf-8"))["data_source"] == "api"
    assert not (tmp_path / "reports" / "baidu_account_data.json").exists()


def test_api_preferred_multi_source_daily_double_failure_preserves_daily_canonical(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_daily

    canonical = tmp_path / "reports" / "baidu_daily_data.json"
    canonical.parent.mkdir()
    canonical.write_text('{"sentinel": "daily"}', encoding="utf-8")
    staged_paths = []

    def api_fetcher(*, config, **_kwargs):
        if config["baidu_source"]["source_id"] == "b":
            raise RuntimeError("source b failed")
        account = next(iter(config["accounts"]))
        return {"accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}}, "errors": []}

    def browser_fetcher(*, config, **_kwargs):
        staged = Path(config["baidu"]["daily_output_path"])
        staged_paths.append(staged)
        staged.parent.mkdir(parents=True, exist_ok=True)
        staged.write_text('{"errors": ["browser failed"]}', encoding="utf-8")
        return _two_source_browser_report(["browser failed"])

    result = fetch_baidu_resilient_daily(
        _two_source_route_config(),
        tmp_path,
        None,
        "2026-07-16",
        api_fetcher=api_fetcher,
        browser_fetcher=browser_fetcher,
        sleep=lambda _seconds: None,
    )

    assert result["data_source"] == "failed"
    assert json.loads(canonical.read_text("utf-8")) == {"sentinel": "daily"}
    assert len(staged_paths) == 1
    assert staged_paths[0].parent == canonical.parent
    assert not staged_paths[0].exists()
    assert not (tmp_path / "reports" / "baidu_account_data.json").exists()


def test_baidu_resilient_browser_mode_stages_success_and_cleans_temp(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    configured_output = tmp_path / "custom_reports" / "hourly_canonical.json"
    config = {
        "baidu": {
            "data_source_mode": "browser",
            "output_path": str(configured_output),
        }
    }
    staged_paths = []

    def browser_fetcher(*, config, **_kwargs):
        staged = Path(config["baidu"]["output_path"])
        staged_paths.append(staged)
        staged.parent.mkdir(parents=True, exist_ok=True)
        staged.write_text('{"source": "browser-side-effect"}', encoding="utf-8")
        return _route_report("baidu_auto_overview")

    result = fetch_baidu_resilient_hourly(
        config,
        tmp_path,
        None,
        "15点",
        api_fetcher=lambda **_kwargs: pytest.fail("B 模式不应调用 API"),
        browser_fetcher=browser_fetcher,
    )

    assert result["data_source"] == "browser"
    assert json.loads(configured_output.read_text("utf-8"))["data_source"] == "browser"
    assert len(staged_paths) == 1
    assert staged_paths[0] != configured_output
    assert staged_paths[0].parent == configured_output.parent
    assert not staged_paths[0].exists()
    assert config["baidu"]["output_path"] == str(configured_output)
    assert not (tmp_path / "reports" / "baidu_account_data.json").exists()


def test_baidu_resilient_browser_mode_failure_preserves_canonical_and_cleans_temp(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    canonical = tmp_path / "reports" / "baidu_account_data.json"
    canonical.parent.mkdir()
    canonical.write_text('{"sentinel": "browser"}', encoding="utf-8")
    staged_paths = []

    def browser_fetcher(*, config, **_kwargs):
        staged = Path(config["baidu"]["output_path"])
        staged_paths.append(staged)
        staged.write_text('{"errors": ["browser failed"]}', encoding="utf-8")
        return _route_report("baidu_auto_overview", errors=["browser failed"])

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "browser"}},
        tmp_path,
        None,
        "15点",
        browser_fetcher=browser_fetcher,
    )

    assert result["data_source"] == "failed"
    assert json.loads(canonical.read_text("utf-8")) == {"sentinel": "browser"}
    assert len(staged_paths) == 1
    assert not staged_paths[0].exists()


def test_browser_mode_never_calls_api(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "browser"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: (_ for _ in ()).throw(AssertionError("API must not run")),
        browser_fetcher=lambda **_kwargs: _route_report("baidu_auto_overview"),
        sleep=lambda _seconds: None,
    )
    assert result["data_source"] == "browser"
    assert result["api_attempts"] == 0


def test_api_preferred_success_never_calls_browser(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: _route_report(),
        browser_fetcher=lambda **_kwargs: (_ for _ in ()).throw(AssertionError("browser must not run")),
        sleep=lambda _seconds: None,
    )
    assert result["data_source"] == "api"
    assert result["api_attempts"] == 1


def test_api_network_error_retries_twice_then_falls_back(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    calls = {"api": 0, "browser": 0}

    def api_fetcher(**_kwargs):
        calls["api"] += 1
        raise BaiduReportApiError("network unavailable", category="network_error")

    def browser_fetcher(**_kwargs):
        calls["browser"] += 1
        return _route_report("baidu_auto_overview")

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=api_fetcher, browser_fetcher=browser_fetcher, sleep=lambda _seconds: None,
    )
    assert calls == {"api": 3, "browser": 1}
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "network_error"
    assert result["self_heal_actions"] == ["network_retry", "network_retry"]


def test_api_integrity_error_retries_once_then_falls_back(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    calls = []

    def api_fetcher(**_kwargs):
        calls.append("api")
        raise BaiduReportApiError("unstable data", category="integrity_error")

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: calls.append("browser") or _route_report("baidu_auto_overview"),
        sleep=lambda _seconds: None,
    )
    assert calls == ["api", "api", "browser"]
    assert result["self_heal_actions"] == ["integrity_retry"]


def test_refresh_required_falls_back_without_network_retry(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    calls = []

    def api_fetcher(**_kwargs):
        calls.append("api")
        raise BaiduReportApiError(
            "reauthorize", category="reauthorization_required", reauthorization_required=True
        )

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: calls.append("browser") or _route_report("baidu_auto_overview"),
        sleep=lambda _seconds: None,
    )
    assert calls == ["api", "browser"]
    assert result["fallback_reason"] == "reauthorization_required"


def test_api_and_browser_failure_returns_errors_and_no_success_report(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: (_ for _ in ()).throw(
            BaiduReportApiError("API failed", category="reauthorization_required", reauthorization_required=True)
        ),
        browser_fetcher=lambda **_kwargs: _route_report("baidu_auto_overview", errors=["browser failed"]),
        sleep=lambda _seconds: None,
    )
    assert result["errors"]
    assert result["data_source"] == "failed"
    assert result["fallback_reason"] == "reauthorization_required"
    assert not (tmp_path / "reports" / "baidu_account_data.json").exists()


def test_shadow_mode_uses_browser_output_and_writes_comparison(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser = _route_report("baidu_auto_overview", cost=11.0)
    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_shadow"}}, tmp_path, None, "15点",
        api_fetcher=lambda **kwargs: (
            _route_report(cost=10.0)
            if kwargs.get("commit_standard_report") is False
            else (_ for _ in ()).throw(AssertionError("shadow API must not commit"))
        ),
        browser_fetcher=lambda **_kwargs: browser,
        sleep=lambda _seconds: None,
    )
    comparison = json.loads((tmp_path / "reports" / "baidu_api_shadow_comparison.json").read_text("utf-8"))
    assert result["accounts"] == browser["accounts"]
    assert result["data_source"] == "browser_shadow"
    assert comparison["passed"] is False
    assert comparison["differences"]


def test_shadow_comparison_accepts_browser_report_without_source_user_id(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser = _route_report("baidu_auto_overview")
    browser["accounts"]["银康01"].pop("source_user_id")
    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_shadow"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: _route_report(),
        browser_fetcher=lambda **_kwargs: browser,
        sleep=lambda _seconds: None,
    )

    comparison = json.loads((tmp_path / "reports" / "baidu_api_shadow_comparison.json").read_text("utf-8"))
    assert result["data_source"] == "browser_shadow"
    assert comparison["passed"] is True
    assert comparison["differences"] == []


def test_shadow_comparison_rejects_source_user_id_when_both_reports_provide_it(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser = _route_report("baidu_auto_overview")
    browser["accounts"]["银康01"]["source_user_id"] = 999
    fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_shadow"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: _route_report(),
        browser_fetcher=lambda **_kwargs: browser,
        sleep=lambda _seconds: None,
    )

    comparison = json.loads((tmp_path / "reports" / "baidu_api_shadow_comparison.json").read_text("utf-8"))
    assert comparison["passed"] is False
    assert comparison["differences"] == [
        {"account": "银康01", "field": "source_user_id", "api": 72828178, "browser": 999}
    ]


def test_shadow_comparison_normalizes_equivalent_source_user_id_types(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    browser = _route_report("baidu_auto_overview")
    browser["accounts"]["银康01"]["source_user_id"] = "72828178"
    fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_shadow"}}, tmp_path, None, "15点",
        api_fetcher=lambda **_kwargs: _route_report(),
        browser_fetcher=lambda **_kwargs: browser,
        sleep=lambda _seconds: None,
    )

    comparison = json.loads((tmp_path / "reports" / "baidu_api_shadow_comparison.json").read_text("utf-8"))
    assert comparison["passed"] is True
    assert comparison["differences"] == []


def test_twenty_second_budget_stops_api_retries(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    calls = []
    clock_values = iter([0.0, 21.0, 21.0])

    def api_fetcher(**_kwargs):
        calls.append("api")
        raise BaiduReportApiError("network unavailable", category="network_error")

    result = fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}}, tmp_path, None, "15点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: calls.append("browser") or _route_report("baidu_auto_overview"),
        clock=lambda: next(clock_values), sleep=lambda _seconds: None,
    )
    assert calls == ["api", "browser"]
    assert result["fallback_reason"] == "network_error"


def test_api_budget_caps_each_network_call_and_counts_internal_requests(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_daily

    config = _api_production_config(tmp_path)
    now = [0.0]
    token_calls = []
    report_timeouts = []
    responses = [
        {"header": {"status": 1, "desc": "failure", "failures": [{"code": 894061}]}},
        _api_success_response("2026-07-16"),
    ]

    def token_provider(_config, _root, _profile, **kwargs):
        token_calls.append((kwargs.get("force_refresh"), kwargs.get("timeout_seconds")))
        if kwargs.get("force_refresh"):
            now[0] = 19.5
        return "header.payload.signature", {"token_refresh": "refreshed" if kwargs.get("force_refresh") else "not_needed"}

    def transport(_url, _payload, timeout):
        report_timeouts.append(timeout)
        if len(report_timeouts) == 1:
            now[0] = 19.0
        return responses.pop(0)

    report = fetch_baidu_api_daily(
        config,
        tmp_path,
        logging.getLogger("api-budget"),
        "2026-07-16",
        token_provider=token_provider,
        transport=transport,
        deadline=20.0,
        clock=lambda: now[0],
    )

    assert token_calls == [(False, 20.0), (True, 1.0)]
    assert report_timeouts == [20.0, 0.5]
    assert report["diagnostics"]["api_request_count"] == 2
    assert report["diagnostics"]["self_heal_actions"] == ["token_refresh"]


def test_router_allows_only_one_token_refresh_across_api_retries(tmp_path):
    import logging
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError, fetch_baidu_api_hourly

    config = _api_production_config(tmp_path)
    config["baidu"]["data_source_mode"] = "api_preferred"
    token_calls = []
    report_calls = ["expired", "network", "expired"]

    def token_provider(_config, _root, _profile, **kwargs):
        token_calls.append(bool(kwargs.get("force_refresh")))
        return "header.payload.signature", {"token_refresh": "refreshed" if kwargs.get("force_refresh") else "not_needed"}

    def transport(_url, _payload, _timeout):
        action = report_calls.pop(0)
        if action == "expired":
            return {"header": {"status": 1, "desc": "failure", "failures": [{"code": 894061}]}}
        raise BaiduReportApiError("network", category="network_error")

    def api_fetcher(**kwargs):
        return fetch_baidu_api_hourly(
            **kwargs,
            token_provider=token_provider,
            transport=transport,
        )

    result = fetch_baidu_resilient_hourly(
        config,
        tmp_path,
        logging.getLogger("one-refresh"),
        "18点",
        api_fetcher=api_fetcher,
        browser_fetcher=lambda **_kwargs: _route_report("baidu_auto_overview"),
        sleep=lambda _seconds: None,
    )

    assert token_calls.count(True) == 1
    assert result["data_source"] == "browser_fallback"
    assert result["fallback_reason"] == "authorization_error"
    assert result["api_attempts"] == 3
    assert result["self_heal_actions"] == ["token_refresh", "network_retry"]


def test_baidu_api_hourly_simulation_keeps_outputs_isolated(tmp_path, monkeypatch):
    import logging
    from datetime import date

    import modules.baidu_api_simulation as simulation

    config = _kunming_niu_runtime_config()
    config.update({"project_id": "kunming_niu", "project_name": "昆明牛"})
    selected_date = date.today().isoformat()
    api_report = {
        "date": selected_date,
        "period": "18点",
        "source": "baidu_open_api_probe",
        "accounts": {
            "银康01": {"展现": 10, "点击": 1, "消费": 2.0},
            "银康银屑02": {"展现": 20, "点击": 2, "消费": 3.0},
            "银康03": {"展现": 30, "点击": 3, "消费": 4.0},
        },
        "diagnostics": {},
        "errors": [],
    }
    monkeypatch.setattr(simulation, "fetch_baidu_api_probe", lambda **_kwargs: api_report)
    monkeypatch.setattr(simulation, "find_latest_kst_export", lambda *_args: None)
    monkeypatch.setattr(
        simulation,
        "_preview_excel_targets",
        lambda *_args: ([{"account": "银康01", "field": "消费", "cell": "A1", "old_value": None, "new_value": 2.0}], [], {"excel_path": "target.xlsx", "sheet_name": "时段数据"}),
    )

    report = simulate_baidu_api_hourly(config, tmp_path, logging.getLogger("test"), "18点", selected_date)

    assert report["errors"] == []
    assert report["self_check"]["passed"] is True
    assert report["self_check"]["wrote_excel"] is False
    assert report["self_check"]["created_backup"] is False
    assert report["kst"]["no_export_file"] is True
    assert report["merged"]["accounts"]["银康01"]["总对话"] == 0
    assert (tmp_path / "reports" / "baidu_api_hourly_simulated_merged.json").exists()
    assert not (tmp_path / "reports" / "merged_hourly_data.json").exists()
    assert not (tmp_path / "backups").exists()


def test_baidu_oauth_bundle_import_is_atomic_and_does_not_expose_tokens(tmp_path):
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu_api": {"daily_automation": {"app_id": "app-1", "access_token": "old.token.value"}},
    }), encoding="utf-8")
    bundle_path = tmp_path / "download.baidu-auth"
    bundle_path.write_text(json.dumps({
        "format": "baidu-oauth-export-v1",
        "app_id": "app-1",
        "authorization": {
            "access_token": "new.access.token",
            "refresh_token": "new.refresh.token",
            "open_id": "open-1",
            "user_id": 123,
            "master_uid": 123,
            "master_name": "manager",
            "user_account_type": 2,
            "sub_accounts": [{"user_id": 456, "user_name": "child"}],
        },
    }), encoding="utf-8")

    report = import_baidu_oauth_bundle(tmp_path, bundle_path, "changsha_niu_baidu")

    saved = json.loads(secrets_path.read_text(encoding="utf-8"))
    assert report["passed"] is True
    assert report["sub_account_count"] == 1
    assert "access_token" not in report
    assert saved["baidu_api"]["changsha_niu_baidu"]["access_token"] == "new.access.token"
    assert saved["baidu_api"]["changsha_niu_baidu"]["refresh_token"] == "new.refresh.token"
    assert len(list((tmp_path / "backups").glob("secrets_before_oauth_*.json"))) == 1


def test_baidu_oauth_bundle_can_sync_profile_to_cloud_token_store(tmp_path):
    import hashlib
    import hmac

    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu_api_gateway": {
            "app_id": "app-1",
            "client_key": "client-key",
            "token_url": "https://example.invalid/baidu/oauth/token",
        },
        "baidu_api": {"existing": {"app_id": "app-1"}},
    }), encoding="utf-8")
    bundle_path = tmp_path / "download.baidu-auth"
    bundle_path.write_text(json.dumps({
        "format": "baidu-oauth-export-v1",
        "app_id": "app-1",
        "authorization": {
            "access_token": "new.access.token",
            "refresh_token": "new.refresh.token",
            "open_id": "open-1",
            "user_id": 123,
            "master_uid": 123,
            "master_name": "manager",
            "user_account_type": 2,
            "sub_accounts": [{"user_id": 456, "user_name": "child"}],
        },
    }), encoding="utf-8")
    calls = []

    def transport(url, payload, headers, timeout):
        calls.append((url, payload, headers, timeout))
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        expected = hmac.new(
            b"client-key",
            (headers["X-Baidu-Refresh-Timestamp"] + "\n" + canonical).encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        assert headers["X-Baidu-Refresh-Signature"] == expected
        return {"status": "ok", "profile": {"api_profile": "changsha_niu_baidu"}}

    report = import_baidu_oauth_bundle(
        tmp_path,
        bundle_path,
        "changsha_niu_baidu",
        sync_cloud_token_store=True,
        transport=transport,
    )

    assert calls[0][0] == "https://example.invalid/baidu/oauth/store-profile"
    assert calls[0][1]["apiProfile"] == "changsha_niu_baidu"
    assert calls[0][1]["authorization"]["refresh_token"] == "new.refresh.token"
    assert report["cloud_sync"]["passed"] is True
    assert "new.access.token" not in json.dumps(report, ensure_ascii=False)
    assert "new.refresh.token" not in json.dumps(report, ensure_ascii=False)


def test_baidu_oauth_cloud_sync_all_profiles_skips_incomplete_records(tmp_path):
    import hashlib
    import hmac
    from modules.baidu_oauth_bundle import sync_baidu_oauth_profiles_to_cloud

    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu_api_gateway": {
            "app_id": "app-1",
            "client_key": "client-key",
            "store_profile_url": "https://example.invalid/baidu/oauth/store-profile",
        },
        "baidu_api": {
            "complete_baidu": {
                "app_id": "app-1",
                "access_token": "secret.access.token",
                "refresh_token": "secret.refresh.token",
                "open_id": "open",
                "user_id": 123,
                "sub_accounts": [{"user_id": 456}],
            },
            "legacy_placeholder": {
                "app_id": "app-1",
                "access_token": "placeholder",
            },
        },
    }), encoding="utf-8")
    calls = []

    def transport(url, payload, headers, timeout):
        calls.append((url, payload, headers, timeout))
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        expected = hmac.new(
            b"client-key",
            (headers["X-Baidu-Refresh-Timestamp"] + "\n" + canonical).encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        assert headers["X-Baidu-Refresh-Signature"] == expected
        return {"status": "ok", "profile": {"api_profile": payload["apiProfile"]}}

    report = sync_baidu_oauth_profiles_to_cloud(tmp_path, transport=transport)

    assert report["passed"] is True
    assert report["synced_count"] == 1
    assert report["skipped_count"] == 1
    assert calls[0][0] == "https://example.invalid/baidu/oauth/store-profile"
    assert calls[0][1]["apiProfile"] == "complete_baidu"
    serialized_report = json.dumps(report, ensure_ascii=False)
    assert "secret.access.token" not in serialized_report
    assert "secret.refresh.token" not in serialized_report


def test_oauth_import_rereads_secrets_inside_shared_lock(tmp_path, monkeypatch):
    from contextlib import contextmanager
    import modules.baidu_oauth_bundle as oauth_bundle

    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu_api_gateway": {"app_id": "app-1"},
        "baidu_api": {"existing": {"app_id": "app-1", "access_token": "old"}},
    }), encoding="utf-8")
    bundle_path = tmp_path / "download.baidu-auth"
    bundle_path.write_text(json.dumps(_oauth_match_bundle([456])), encoding="utf-8")

    @contextmanager
    def fake_lock(path):
        assert path == secrets_path
        latest = json.loads(secrets_path.read_text(encoding="utf-8"))
        latest["baidu_api"]["concurrent_refresh"] = {"app_id": "app-1", "access_token": "preserve-me"}
        secrets_path.write_text(json.dumps(latest), encoding="utf-8")
        yield

    monkeypatch.setattr(oauth_bundle, "secrets_file_lock", fake_lock)
    oauth_bundle.import_baidu_oauth_bundle(tmp_path, bundle_path, "changsha_niu_baidu")

    saved = json.loads(secrets_path.read_text(encoding="utf-8"))
    assert saved["baidu_api"]["concurrent_refresh"]["access_token"] == "preserve-me"
    assert saved["baidu_api"]["changsha_niu_baidu"]["access_token"] == "secret.access.token"


def _oauth_match_bundle(user_ids, app_id="app-1"):
    return {
        "format": "baidu-oauth-export-v1",
        "app_id": app_id,
        "authorization": {
            "access_token": "secret.access.token",
            "refresh_token": "secret.refresh.token",
            "open_id": "open-1",
            "user_id": 123,
            "sub_accounts": [{"user_id": user_id, "user_name": "account"} for user_id in user_ids],
        },
    }


def _write_oauth_match_root(root, projects):
    projects_dir = root / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    for project in projects:
        (projects_dir / f"{project['project_id']}.json").write_text(
            json.dumps(project, ensure_ascii=False), encoding="utf-8"
        )
    secrets_path = root / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text(json.dumps({
        "baidu_api_gateway": {"app_id": "app-1", "refresh_url": "https://example.invalid", "client_key": "key"},
        "baidu_api": {},
    }), encoding="utf-8")
    return secrets_path


def test_oauth_profile_match_finds_unique_single_source_project(tmp_path):
    from modules.baidu_oauth_bundle import match_baidu_oauth_profile

    _write_oauth_match_root(tmp_path, [{
        "project_id": "changsha_niu",
        "project_name": "长沙牛",
        "baidu": {"api_profile": "changsha_niu_baidu"},
        "accounts": [
            {"standard_name": "A", "baidu_user_ids": [111]},
            {"standard_name": "B", "kst_ids": ["222"]},
            {"standard_name": "C", "kst_ids": ["333"]},
        ],
    }])

    match = match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([333, 111, 222]))

    assert match == {
        "api_profile": "changsha_niu_baidu",
        "project_id": "changsha_niu",
        "source_id": None,
        "promotion_ids": [111, 222, 333],
    }
    assert "secret.access.token" not in json.dumps(match)


def test_oauth_profile_match_allows_unique_project_with_extra_inactive_authorized_account(tmp_path):
    from modules.baidu_oauth_bundle import match_baidu_oauth_profile

    _write_oauth_match_root(tmp_path, [{
        "project_id": "nanjing_niu",
        "project_name": "南京牛",
        "baidu": {"api_profile": "nanjing_niu_baidu"},
        "accounts": [
            {"standard_name": "npx1", "kst_ids": ["111"]},
            {"standard_name": "npx6", "kst_ids": ["666"]},
        ],
    }])

    match = match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([111, 222, 666]))

    assert match["api_profile"] == "nanjing_niu_baidu"
    assert match["promotion_ids"] == [111, 666]
    assert match["ignored_authorized_promotion_ids"] == [222]


def test_oauth_profile_match_rejects_authorized_superset_covering_multiple_projects(tmp_path):
    from modules.baidu_oauth_bundle import BaiduOAuthImportError, match_baidu_oauth_profile

    _write_oauth_match_root(tmp_path, [
        {
            "project_id": "project_a", "project_name": "A",
            "baidu": {"api_profile": "project_a_baidu"},
            "accounts": [{"standard_name": "A", "kst_ids": ["111"]}],
        },
        {
            "project_id": "project_b", "project_name": "B",
            "baidu": {"api_profile": "project_b_baidu"},
            "accounts": [{"standard_name": "B", "kst_ids": ["222"]}],
        },
    ])

    with pytest.raises(BaiduOAuthImportError, match="同时匹配多个项目来源"):
        match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([111, 222, 999]))


def test_oauth_profile_match_finds_exact_dual_source_profile(tmp_path):
    from modules.baidu_oauth_bundle import match_baidu_oauth_profile

    _write_oauth_match_root(tmp_path, [{
        "project_id": "shenyang_bai",
        "project_name": "沈阳白",
        "baidu": {},
        "accounts": [],
        "baidu_sources": [
            {
                "source_id": "source_a", "api_profile": "shenyang_bai_source_a_baidu",
                "accounts": [{"standard_name": "A", "kst_ids": ["101", "ignored-second-id"]}],
            },
            {
                "source_id": "source_b", "api_profile": "shenyang_bai_source_b_baidu",
                "accounts": [{"standard_name": "B", "baidu_user_ids": [202]}],
            },
        ],
    }])

    match = match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([202]))

    assert match["api_profile"] == "shenyang_bai_source_b_baidu"
    assert match["project_id"] == "shenyang_bai"
    assert match["source_id"] == "source_b"
    assert match["promotion_ids"] == [202]


def test_oauth_profile_match_rejects_zero_multiple_and_wrong_app_without_token_leak(tmp_path):
    from modules.baidu_oauth_bundle import BaiduOAuthImportError, match_baidu_oauth_profile

    project = {
        "project_id": "project_a", "project_name": "A",
        "baidu": {"api_profile": "project_a_baidu"},
        "accounts": [{"standard_name": "A", "kst_ids": ["111"]}],
    }
    _write_oauth_match_root(tmp_path, [project])
    sensitive = ("secret.access.token", "secret.refresh.token")

    with pytest.raises(BaiduOAuthImportError) as no_match:
        match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([999]))
    with pytest.raises(BaiduOAuthImportError) as wrong_app:
        match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([111], app_id="app-2"))

    duplicate = dict(project)
    duplicate["project_id"] = "project_b"
    duplicate["baidu"] = {"api_profile": "project_b_baidu"}
    (tmp_path / "configs" / "projects" / "project_b.json").write_text(
        json.dumps(duplicate, ensure_ascii=False), encoding="utf-8"
    )
    with pytest.raises(BaiduOAuthImportError) as multiple:
        match_baidu_oauth_profile(tmp_path, _oauth_match_bundle([111]))

    serialized = " ".join(str(item.value) for item in (no_match, wrong_app, multiple))
    assert all(token not in serialized for token in sensitive)


def test_import_baidu_oauth_auto_profile_uses_unique_match(tmp_path):
    project = {
        "project_id": "changsha_niu", "project_name": "长沙牛",
        "baidu": {"api_profile": "changsha_niu_baidu"},
        "accounts": [{"standard_name": "A", "kst_ids": ["111"]}],
    }
    secrets_path = _write_oauth_match_root(tmp_path, [project])
    bundle_path = tmp_path / "download.baidu-auth"
    bundle_path.write_text(json.dumps(_oauth_match_bundle([111])), encoding="utf-8")

    report = import_baidu_oauth_bundle(tmp_path, bundle_path, "auto")

    saved = json.loads(secrets_path.read_text(encoding="utf-8"))
    assert report["api_profile"] == "changsha_niu_baidu"
    assert report["matched_project_id"] == "changsha_niu"
    assert saved["baidu_api"]["changsha_niu_baidu"]["access_token"] == "secret.access.token"
    assert "secret.access.token" not in json.dumps(report)


def test_import_baidu_oauth_auto_reports_ignored_inactive_accounts(tmp_path):
    project = {
        "project_id": "nanjing_niu", "project_name": "南京牛",
        "baidu": {"api_profile": "nanjing_niu_baidu"},
        "accounts": [{"standard_name": "A", "kst_ids": ["111"]}],
    }
    _write_oauth_match_root(tmp_path, [project])
    bundle_path = tmp_path / "download.baidu-auth"
    bundle_path.write_text(json.dumps(_oauth_match_bundle([111, 222])), encoding="utf-8")

    report = import_baidu_oauth_bundle(tmp_path, bundle_path, "auto")

    assert report["matched_project_id"] == "nanjing_niu"
    assert report["ignored_authorized_promotion_ids"] == [222]


def test_secrets_package_round_trip_fully_replaces_target_and_backs_up(tmp_path):
    from modules.secrets_package import export_secrets_package, import_secrets_package

    source = tmp_path / "source.json"
    package = tmp_path / "team.baidu-secrets"
    target = tmp_path / "receiver" / "secrets" / "secrets.json"
    source_payload = {
        "baidu": {"demo": {"username": "fake-user", "password": "fake-password"}},
        "baidu_api": {"demo": {"access_token": "fake.access.token"}},
        "shared_setting": {"enabled": True},
    }
    old_payload = {
        "baidu": {"old": {"username": "old-user", "password": "old-password"}},
        "local_only": True,
    }
    source.write_text(json.dumps(source_payload, ensure_ascii=False), encoding="utf-8")
    target.parent.mkdir(parents=True)
    target.write_text(json.dumps(old_payload, ensure_ascii=False), encoding="utf-8")
    old_bytes = target.read_bytes()

    export_report = export_secrets_package(source, package)
    import_report = import_secrets_package(package, target, tmp_path / "backups")

    assert export_report["package_path"] == str(package)
    assert json.loads(target.read_text(encoding="utf-8")) == source_payload
    backup_path = Path(import_report["backup_path"])
    assert backup_path.name.startswith("secrets_before_package_import_")
    assert backup_path.read_bytes() == old_bytes
    assert json.loads(backup_path.read_text(encoding="utf-8")) == old_payload
    assert import_report["baidu_profile_count"] == 1
    assert import_report["api_profile_count"] == 1
    assert "fake-password" not in json.dumps(import_report, ensure_ascii=False)
    assert "fake.access.token" not in json.dumps(import_report, ensure_ascii=False)


def test_secrets_package_rejects_checksum_mismatch_without_changing_target(tmp_path):
    import pytest

    from modules.secrets_package import SecretsPackageError, import_secrets_package

    target = tmp_path / "secrets.json"
    target.write_text('{"baidu":{"keep":{}}}', encoding="utf-8")
    before = target.read_bytes()
    package = tmp_path / "bad.baidu-secrets"
    package.write_text(json.dumps({
        "format": "baidu-secrets-package-v1",
        "exported_at": "2026-07-15T15:30:00",
        "payload_sha256": "0" * 64,
        "secrets": {"baidu": {}},
    }), encoding="utf-8")

    with pytest.raises(SecretsPackageError, match="校验"):
        import_secrets_package(package, target, tmp_path / "backups")

    assert target.read_bytes() == before
    assert not (tmp_path / "backups").exists()


def test_secrets_package_import_creates_missing_target_and_rejects_bad_structure(tmp_path):
    import pytest

    from modules.secrets_package import SecretsPackageError, export_secrets_package, import_secrets_package

    invalid_source = tmp_path / "invalid.json"
    invalid_source.write_text('{"baidu": []}', encoding="utf-8")
    with pytest.raises(SecretsPackageError, match="baidu"):
        export_secrets_package(invalid_source, tmp_path / "invalid.baidu-secrets")

    source = tmp_path / "source.json"
    package = tmp_path / "valid.baidu-secrets"
    target = tmp_path / "new-root" / "secrets" / "secrets.json"
    source.write_text('{"baidu": {}, "baidu_api": {}}', encoding="utf-8")
    export_secrets_package(source, package)

    report = import_secrets_package(package, target, tmp_path / "backups")

    assert json.loads(target.read_text(encoding="utf-8")) == {"baidu": {}, "baidu_api": {}}
    assert report["backup_path"] is None


def test_secrets_package_rejects_malformed_or_wrong_format_package(tmp_path):
    import pytest

    from modules.secrets_package import SecretsPackageError, import_secrets_package

    target = tmp_path / "secrets.json"
    target.write_text('{"baidu":{"keep":{}}}', encoding="utf-8")
    before = target.read_bytes()
    malformed = tmp_path / "malformed.baidu-secrets"
    malformed.write_text('{"format":', encoding="utf-8")
    wrong_format = tmp_path / "wrong.baidu-secrets"
    wrong_format.write_text(json.dumps({
        "format": "unknown-v1",
        "payload_sha256": "0" * 64,
        "secrets": {"baidu": {}},
    }), encoding="utf-8")

    with pytest.raises(SecretsPackageError, match="合法 JSON"):
        import_secrets_package(malformed, target, tmp_path / "backups")
    with pytest.raises(SecretsPackageError, match="格式或版本"):
        import_secrets_package(wrong_format, target, tmp_path / "backups")

    assert target.read_bytes() == before


def test_secrets_package_atomic_replace_failure_keeps_original_target(tmp_path, monkeypatch):
    import pytest
    import modules.secrets_package as secrets_package

    source = tmp_path / "source.json"
    package = tmp_path / "team.baidu-secrets"
    target = tmp_path / "secrets.json"
    source.write_text('{"baidu":{"new":{}}}', encoding="utf-8")
    target.write_text('{"baidu":{"keep":{}}}', encoding="utf-8")
    secrets_package.export_secrets_package(source, package)
    before = target.read_bytes()
    monkeypatch.setattr(secrets_package.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("locked")))

    with pytest.raises(secrets_package.SecretsPackageError, match="无法写入"):
        secrets_package.import_secrets_package(package, target, tmp_path / "backups")

    assert target.read_bytes() == before
    assert not list(target.parent.glob("secrets.json.*.tmp"))


def test_secrets_package_backup_directory_error_is_actionable_and_keeps_target(tmp_path):
    import pytest

    from modules.secrets_package import SecretsPackageError, export_secrets_package, import_secrets_package

    source = tmp_path / "source.json"
    package = tmp_path / "team.baidu-secrets"
    target = tmp_path / "secrets.json"
    blocked_backup_dir = tmp_path / "backups"
    source.write_text('{"baidu":{"new":{}}}', encoding="utf-8")
    target.write_text('{"baidu":{"keep":{}}}', encoding="utf-8")
    blocked_backup_dir.write_text("not a directory", encoding="utf-8")
    export_secrets_package(source, package)
    before = target.read_bytes()

    with pytest.raises(SecretsPackageError, match="备份"):
        import_secrets_package(package, target, blocked_backup_dir)

    assert target.read_bytes() == before


def test_secrets_package_cleanup_error_does_not_hide_atomic_write_failure(tmp_path, monkeypatch):
    import pytest
    import modules.secrets_package as secrets_package

    source = tmp_path / "source.json"
    package = tmp_path / "team.baidu-secrets"
    target = tmp_path / "secrets.json"
    source.write_text('{"baidu":{"new":{}}}', encoding="utf-8")
    target.write_text('{"baidu":{"keep":{}}}', encoding="utf-8")
    secrets_package.export_secrets_package(source, package)
    monkeypatch.setattr(secrets_package.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("replace locked")))
    monkeypatch.setattr(secrets_package.Path, "unlink", lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("unlink locked")))

    with pytest.raises(secrets_package.SecretsPackageError, match="无法写入"):
        secrets_package.import_secrets_package(package, target, tmp_path / "backups")


def test_baidu_oauth_callback_exchanges_code_and_exports_subaccounts(tmp_path, monkeypatch):
    import importlib.util

    callback_path = Path(__file__).resolve().parents[1] / "cloud" / "baidu_oauth_callback" / "index.py"
    spec = importlib.util.spec_from_file_location("baidu_oauth_callback_test", callback_path)
    callback = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(callback)
    monkeypatch.setattr(callback, "_verify_signature", lambda query, secret: True)
    now_ms = str(int(__import__("time").time() * 1000))
    query = {
        "appId": "app-1",
        "authCode": "temporary-code",
        "state": "expected-state",
        "userId": "123",
        "timestamp": now_ms,
        "signature": "signed",
    }
    calls = []

    def fake_transport(url, payload, timeout):
        calls.append((url, payload, timeout))
        if url.endswith("/accessToken"):
            return {"code": "0", "data": {
                "accessToken": "new.access.token",
                "refreshToken": "new.refresh.token",
                "openId": "open-1",
                "userId": 123,
                "expiresIn": 86400,
                "refreshExpiresIn": 2592000,
            }}
        return {"code": "0", "data": {
            "masterUid": 123,
            "masterName": "manager",
            "userAcctType": 2,
            "hasNext": False,
            "subUserList": [{"ucId": 456, "ucName": "child"}],
        }}

    bundle = callback.process_oauth_callback(
        query,
        {"app_id": "app-1", "secret_key": "1234567890abcdef-extra", "allowed_states": {"expected-state"}, "max_timestamp_skew_seconds": 600},
        fake_transport,
    )

    assert bundle["format"] == "baidu-oauth-export-v1"
    assert bundle["authorization"]["sub_accounts"] == [{"user_id": 456, "user_name": "child"}]
    assert calls[0][1]["grantType"] == "auth_code"
    assert calls[1][1]["needSubList"] is True
    assert calls[1][1]["lastPageMaxUcId"] == 1


def _load_baidu_oauth_callback_module(module_name="baidu_oauth_refresh_test"):
    import importlib.util

    callback_path = Path(__file__).resolve().parents[1] / "cloud" / "baidu_oauth_callback" / "index.py"
    spec = importlib.util.spec_from_file_location(module_name, callback_path)
    callback = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(callback)
    return callback


def _signed_refresh_headers(callback, payload, client_key, timestamp):
    timestamp = str(timestamp)
    return {
        "X-Baidu-Refresh-Timestamp": timestamp,
        "X-Baidu-Refresh-Signature": callback._refresh_signature(timestamp, payload, client_key),
    }


def test_cloud_token_store_loads_missing_store_as_empty_and_upserts_profile():
    callback = _load_baidu_oauth_callback_module("baidu_oauth_cloud_store_test")
    objects = {}
    config = {
        "app_id": "app-1",
        "token_store_bucket": "hourlyreport-1300869225",
        "token_store_region": "ap-nanjing",
        "token_store_key": "baidu-oauth/token-store/baidu_oauth_tokens.json",
    }

    def fake_cos(method, bucket, region, key, body=None):
        assert bucket == "hourlyreport-1300869225"
        assert region == "ap-nanjing"
        if method == "GET":
            if key not in objects:
                raise callback.OAuthCallbackError("token_store_not_found", "missing", 404)
            return json.loads(objects[key])
        if method == "PUT":
            objects[key] = body
            return {"ok": True}
        raise AssertionError(method)

    store = callback.load_token_store(config, fake_cos)
    assert store["format"] == "baidu-token-store-v1"
    callback.upsert_token_profile(
        store,
        "ningbo_niu_baidu",
        {
            "access_token": "a.b.c",
            "refresh_token": "d.e.f",
            "open_id": "open",
            "user_id": 45187067,
            "expires_time": "2026-07-23 11:17:33",
            "refresh_expires_time": "2026-08-21 11:17:33",
        },
        "app-1",
    )
    callback.save_token_store(config, store, fake_cos)
    saved = json.loads(objects["baidu-oauth/token-store/baidu_oauth_tokens.json"])
    assert saved["profiles"]["ningbo_niu_baidu"]["user_id"] == 45187067


def test_cloud_token_endpoint_returns_cached_access_token_without_refresh():
    callback = _load_baidu_oauth_callback_module("baidu_oauth_cloud_token_cached_test")
    now = 1784700000
    payload = {"apiProfile": "ningbo_niu_baidu", "forceRefresh": False}
    headers = _signed_refresh_headers(callback, payload, "client-key", now)
    store = {
        "format": "baidu-token-store-v1",
        "profiles": {
            "ningbo_niu_baidu": {
                "app_id": "app-1",
                "access_token": "cached.access.token",
                "refresh_token": "cached.refresh.token",
                "open_id": "open",
                "user_id": 45187067,
                "expires_time": "2026-07-23 11:17:33",
                "refresh_expires_time": "2026-08-21 11:17:33",
            }
        },
    }

    def fake_cos(method, bucket, region, key, body=None):
        assert method == "GET"
        return store

    def fail_oauth(*_args):
        raise AssertionError("should not refresh")

    result = callback.process_cloud_token_request(
        payload,
        headers,
        {
            "app_id": "app-1",
            "secret_key": "server-secret-key",
            "refresh_client_key": "client-key",
            "refresh_max_timestamp_skew_seconds": 300,
            "token_store_bucket": "hourlyreport-1300869225",
            "token_store_region": "ap-nanjing",
            "token_store_key": "baidu-oauth/token-store/baidu_oauth_tokens.json",
        },
        fake_cos,
        fail_oauth,
        now,
    )
    assert result["access_token"] == "cached.access.token"
    assert result["token_refresh"] == "not_needed"
    assert result["api_profile"] == "ningbo_niu_baidu"
    assert "refresh_token" not in json.dumps(result)


def test_cloud_token_endpoint_refreshes_and_persists_rotated_token():
    callback = _load_baidu_oauth_callback_module("baidu_oauth_cloud_token_refresh_test")
    now = 1784700000
    payload = {"apiProfile": "ningbo_niu_baidu", "forceRefresh": True}
    headers = _signed_refresh_headers(callback, payload, "client-key", now)
    objects = {
        "baidu-oauth/token-store/baidu_oauth_tokens.json": json.dumps({
            "format": "baidu-token-store-v1",
            "profiles": {
                "ningbo_niu_baidu": {
                    "app_id": "app-1",
                    "access_token": "old.access.token",
                    "refresh_token": "old.refresh.token",
                    "open_id": "open",
                    "user_id": 45187067,
                    "expires_time": "2026-07-22 11:18:00",
                    "refresh_expires_time": "2026-08-21 11:17:33",
                }
            },
        })
    }
    oauth_calls = []

    def fake_cos(method, bucket, region, key, body=None):
        if method == "GET":
            return json.loads(objects[key])
        if method == "PUT":
            objects[key] = body
            return {"ok": True}
        raise AssertionError(method)

    def fake_oauth(url, body, timeout):
        oauth_calls.append((url, body, timeout))
        return {
            "code": "0",
            "data": {
                "accessToken": "new.access.token",
                "refreshToken": "new.refresh.token",
                "openId": "open-2",
                "expiresTime": "2026-07-23 11:17:33",
                "refreshExpiresTime": "2026-08-21 11:17:33",
                "expiresIn": 86400,
                "refreshExpiresIn": 2592000,
            },
        }

    result = callback.process_cloud_token_request(
        payload,
        headers,
        {
            "app_id": "app-1",
            "secret_key": "server-secret-key",
            "refresh_client_key": "client-key",
            "refresh_max_timestamp_skew_seconds": 300,
            "token_store_bucket": "hourlyreport-1300869225",
            "token_store_region": "ap-nanjing",
            "token_store_key": "baidu-oauth/token-store/baidu_oauth_tokens.json",
        },
        fake_cos,
        fake_oauth,
        now,
    )

    assert oauth_calls[0][1]["refreshToken"] == "old.refresh.token"
    saved = json.loads(objects["baidu-oauth/token-store/baidu_oauth_tokens.json"])
    assert saved["profiles"]["ningbo_niu_baidu"]["refresh_token"] == "new.refresh.token"
    assert result["access_token"] == "new.access.token"
    assert result["token_refresh"] == "refreshed"
    assert "refresh_token" not in json.dumps(result)


def test_store_profile_endpoint_upserts_authorization_without_leaking_tokens():
    callback = _load_baidu_oauth_callback_module("baidu_oauth_store_profile_test")
    now = 1784700000
    payload = {
        "apiProfile": "ningbo_niu_baidu",
        "authorization": {
            "access_token": "new.access.token",
            "refresh_token": "new.refresh.token",
            "open_id": "open",
            "user_id": 45187067,
            "expires_time": "2026-07-23 11:17:33",
            "refresh_expires_time": "2026-08-21 11:17:33",
            "master_name": "BDCC-test",
            "sub_accounts": [{"user_id": 45144300, "user_name": "ningbo-1"}],
        },
    }
    headers = _signed_refresh_headers(callback, payload, "client-key", now)
    objects = {}

    def fake_cos(method, bucket, region, key, body=None):
        if method == "GET":
            if key not in objects:
                raise callback.OAuthCallbackError("token_store_not_found", "missing", 404)
            return json.loads(objects[key])
        if method == "PUT":
            objects[key] = body
            return {"ok": True}
        raise AssertionError(method)

    result = callback.process_store_profile_request(
        payload,
        headers,
        {
            "app_id": "app-1",
            "refresh_client_key": "client-key",
            "refresh_max_timestamp_skew_seconds": 300,
            "token_store_bucket": "hourlyreport-1300869225",
            "token_store_region": "ap-nanjing",
            "token_store_key": "baidu-oauth/token-store/baidu_oauth_tokens.json",
        },
        fake_cos,
        now,
    )

    saved = json.loads(objects["baidu-oauth/token-store/baidu_oauth_tokens.json"])
    assert saved["profiles"]["ningbo_niu_baidu"]["refresh_token"] == "new.refresh.token"
    assert result["api_profile"] == "ningbo_niu_baidu"
    assert result["user_id"] == 45187067
    assert result["sub_account_count"] == 1
    assert "new.access.token" not in json.dumps(result)
    assert "new.refresh.token" not in json.dumps(result)


def test_baidu_oauth_refresh_accepts_signed_request_and_hides_server_secrets():
    callback = _load_baidu_oauth_callback_module()
    timestamp = "1800000000"
    payload = {"appId": "app-1", "userId": 123, "refreshToken": "old.refresh.token"}
    headers = {
        "X-Baidu-Refresh-Timestamp": timestamp,
        "X-Baidu-Refresh-Signature": callback._refresh_signature(timestamp, payload, "client-key"),
    }
    calls = []

    def fake_transport(url, body, timeout):
        calls.append((url, body, timeout))
        return {
            "code": 0,
            "message": "success",
            "data": {
                "accessToken": "new.access.token",
                "refreshToken": "new.refresh.token",
                "openId": "open-1",
                "expiresTime": "2026-07-18 09:00:00",
                "refreshExpiresTime": "2026-08-16 09:00:00",
                "expiresIn": 86400,
                "refreshExpiresIn": 2592000,
            },
        }

    result = callback.process_refresh_request(
        payload,
        headers,
        {
            "app_id": "app-1",
            "secret_key": "server-secret-key-123456",
            "refresh_client_key": "client-key",
            "refresh_max_timestamp_skew_seconds": 300,
        },
        fake_transport,
        now_timestamp=1800000000,
    )

    assert calls == [(
        "https://u.baidu.com/oauth/refreshToken",
        {
            "appId": "app-1",
            "refreshToken": "old.refresh.token",
            "secretKey": "server-secret-key-123456",
            "userId": 123,
        },
        20,
    )]
    assert result == {
        "access_token": "new.access.token",
        "refresh_token": "new.refresh.token",
        "open_id": "open-1",
        "expires_time": "2026-07-18 09:00:00",
        "refresh_expires_time": "2026-08-16 09:00:00",
        "expires_in": 86400,
        "refresh_expires_in": 2592000,
    }
    serialized = json.dumps(result)
    assert "server-secret-key" not in serialized
    assert "client-key" not in serialized
    assert "old.refresh.token" not in serialized


def test_baidu_oauth_refresh_rejects_bad_signature_expired_timestamp_and_wrong_app():
    import pytest

    callback = _load_baidu_oauth_callback_module("baidu_oauth_refresh_rejection_test")
    payload = {"appId": "app-1", "userId": 123, "refreshToken": "old.refresh.token"}
    config = {
        "app_id": "app-1",
        "secret_key": "server-secret-key-123456",
        "refresh_client_key": "client-key",
        "refresh_max_timestamp_skew_seconds": 300,
    }

    with pytest.raises(callback.OAuthCallbackError, match="签名"):
        callback.process_refresh_request(
            payload,
            {"X-Baidu-Refresh-Timestamp": "1800000000", "X-Baidu-Refresh-Signature": "wrong"},
            config,
            now_timestamp=1800000000,
        )

    valid_signature = callback._refresh_signature("1799999000", payload, "client-key")
    with pytest.raises(callback.OAuthCallbackError, match="过期"):
        callback.process_refresh_request(
            payload,
            {"X-Baidu-Refresh-Timestamp": "1799999000", "X-Baidu-Refresh-Signature": valid_signature},
            config,
            now_timestamp=1800000000,
        )

    wrong_app = {**payload, "appId": "app-2"}
    wrong_app_signature = callback._refresh_signature("1800000000", wrong_app, "client-key")
    with pytest.raises(callback.OAuthCallbackError, match="应用 ID"):
        callback.process_refresh_request(
            wrong_app,
            {"X-Baidu-Refresh-Timestamp": "1800000000", "X-Baidu-Refresh-Signature": wrong_app_signature},
            config,
            now_timestamp=1800000000,
        )


def test_baidu_oauth_refresh_rejects_upstream_failure_without_leaking_tokens():
    import pytest

    callback = _load_baidu_oauth_callback_module("baidu_oauth_refresh_upstream_test")
    timestamp = "1800000000"
    payload = {"appId": "app-1", "userId": 123, "refreshToken": "old.refresh.token"}
    headers = {
        "X-Baidu-Refresh-Timestamp": timestamp,
        "X-Baidu-Refresh-Signature": callback._refresh_signature(timestamp, payload, "client-key"),
    }

    with pytest.raises(callback.OAuthCallbackError, match="更新授权令牌失败") as exc_info:
        callback.process_refresh_request(
            payload,
            headers,
            {
                "app_id": "app-1",
                "secret_key": "server-secret-key-123456",
                "refresh_client_key": "client-key",
                "refresh_max_timestamp_skew_seconds": 300,
            },
            lambda *_args: {"code": 600001, "message": "bad old.refresh.token"},
            now_timestamp=1800000000,
        )

    assert "old.refresh.token" not in str(exc_info.value)
    assert exc_info.value.code == "reauthorization_required"
    assert exc_info.value.status_code == 401


def test_baidu_oauth_wsgi_routes_refresh_post_and_rejects_refresh_get(monkeypatch):
    import importlib.util
    import io
    import sys

    callback = _load_baidu_oauth_callback_module("baidu_oauth_refresh_wsgi_index")
    app_path = Path(__file__).resolve().parents[1] / "cloud" / "baidu_oauth_callback" / "app.py"
    spec = importlib.util.spec_from_file_location("baidu_oauth_refresh_wsgi_app", app_path)
    app = importlib.util.module_from_spec(spec)
    previous_index = sys.modules.get("index")
    sys.modules["index"] = callback
    try:
        assert spec.loader is not None
        spec.loader.exec_module(app)
    finally:
        if previous_index is None:
            sys.modules.pop("index", None)
        else:
            sys.modules["index"] = previous_index

    captured = {}

    def fake_refresh_handler(event, _context):
        captured.update(event)
        return callback._response(200, {"status": "ok"})

    monkeypatch.setattr(app, "refresh_handler", fake_refresh_handler)
    body = json.dumps({"appId": "app-1", "userId": 123, "refreshToken": "old.refresh.token"}).encode("utf-8")
    environ = {
        "REQUEST_METHOD": "POST",
        "PATH_INFO": "/baidu/oauth/refresh",
        "CONTENT_LENGTH": str(len(body)),
        "CONTENT_TYPE": "application/json",
        "HTTP_X_BAIDU_REFRESH_TIMESTAMP": "1800000000",
        "HTTP_X_BAIDU_REFRESH_SIGNATURE": "signed",
        "wsgi.input": io.BytesIO(body),
    }
    statuses = []
    result = app.application(environ, lambda status, _headers: statuses.append(status))

    assert statuses == ["200 OK"]
    assert json.loads(b"".join(result).decode("utf-8")) == {"status": "ok"}
    assert captured["httpMethod"] == "POST"
    assert captured["path"] == "/baidu/oauth/refresh"
    assert captured["headers"] == {
        "X-Baidu-Refresh-Timestamp": "1800000000",
        "X-Baidu-Refresh-Signature": "signed",
    }
    assert json.loads(captured["body"])["userId"] == 123

    statuses.clear()
    result = app.application(
        {"REQUEST_METHOD": "GET", "PATH_INFO": "/baidu/oauth/refresh", "wsgi.input": io.BytesIO(b"")},
        lambda status, _headers: statuses.append(status),
    )
    assert statuses == ["405 Method Not Allowed"]
    assert json.loads(b"".join(result).decode("utf-8"))["code"] == "method_not_allowed"


def _write_token_manager_secrets(
    root,
    *,
    expires_time="2026-07-17 09:20:00",
    other_profile=None,
    token_url=None,
):
    secrets_path = root / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    payload = {
        "baidu_api_gateway": {
            "refresh_url": "https://example.invalid/baidu/oauth/refresh",
            "client_key": "fake-client-key",
            "app_id": "app-1",
        },
        "baidu_api": {
            "kunming_niu_baidu": {
                "app_id": "app-1",
                "user_id": 123,
                "access_token": "old.access.token",
                "refresh_token": "old.refresh.token",
                "expires_time": expires_time,
                "refresh_expires_time": "2026-08-16 09:00:00",
            },
        },
    }
    if token_url is not None:
        payload["baidu_api_gateway"]["token_url"] = token_url
    if other_profile is not None:
        payload["baidu_api"]["other_baidu"] = other_profile
    secrets_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return secrets_path


def _token_refresh_response():
    return {
        "status": "ok",
        "authorization": {
            "access_token": "new.access.token",
            "refresh_token": "new.refresh.token",
            "open_id": "open-1",
            "expires_time": "2026-07-18 09:00:00",
            "refresh_expires_time": "2026-08-16 09:00:00",
            "expires_in": 86400,
            "refresh_expires_in": 2592000,
        },
    }


def test_token_manager_keeps_valid_token_without_refresh(tmp_path):
    from datetime import datetime
    from modules.baidu_token_manager import ensure_valid_access_token

    _write_token_manager_secrets(tmp_path)
    token, metadata = ensure_valid_access_token(
        {},
        tmp_path,
        "kunming_niu_baidu",
        now=datetime(2026, 7, 17, 9, 0, 0),
        transport=lambda *_args: (_ for _ in ()).throw(AssertionError("transport must not run")),
    )

    assert token == "old.access.token"
    assert metadata == {
        "api_profile": "kunming_niu_baidu",
        "token_refresh": "not_needed",
        "expires_time": "2026-07-17 09:20:00",
    }


def test_token_manager_refreshes_within_ten_minute_window(tmp_path):
    import hashlib
    import hmac
    from datetime import datetime
    from modules.baidu_token_manager import ensure_valid_access_token

    _write_token_manager_secrets(tmp_path, expires_time="2026-07-17 09:05:00")
    calls = []

    def transport(url, payload, headers, timeout):
        calls.append((url, payload, headers, timeout))
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        expected = hmac.new(
            b"fake-client-key",
            (headers["X-Baidu-Refresh-Timestamp"] + "\n" + canonical).encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        assert headers["X-Baidu-Refresh-Signature"] == expected
        return _token_refresh_response()

    token, metadata = ensure_valid_access_token(
        {}, tmp_path, "kunming_niu_baidu", now=datetime(2026, 7, 17, 9, 0, 0), transport=transport
    )

    assert token == "new.access.token"
    assert metadata["token_refresh"] == "refreshed"
    assert calls[0][0] == "https://example.invalid/baidu/oauth/refresh"
    assert calls[0][1] == {"appId": "app-1", "userId": 123, "refreshToken": "old.refresh.token"}
    assert calls[0][3] == 20


def test_token_manager_cloud_first_uses_token_endpoint_without_local_refresh_token(tmp_path):
    import hashlib
    import hmac
    from datetime import datetime
    from modules.baidu_token_manager import ensure_valid_access_token_cloud_first

    _write_token_manager_secrets(
        tmp_path,
        token_url="https://example.invalid/baidu/oauth/token",
    )
    calls = []

    def transport(url, payload, headers, timeout):
        calls.append((url, payload, headers, timeout))
        canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        expected = hmac.new(
            b"fake-client-key",
            (headers["X-Baidu-Refresh-Timestamp"] + "\n" + canonical).encode("utf-8"),
            hashlib.sha256,
        ).hexdigest()
        assert headers["X-Baidu-Refresh-Signature"] == expected
        return {
            "status": "ok",
            "authorization": {
                "access_token": "cloud.access.token",
                "api_profile": "kunming_niu_baidu",
                "expires_time": "2026-07-18 09:00:00",
                "token_refresh": "not_needed",
            },
        }

    token, metadata = ensure_valid_access_token_cloud_first(
        {},
        tmp_path,
        "kunming_niu_baidu",
        now=datetime(2026, 7, 17, 9, 0, 0),
        transport=transport,
    )

    assert token == "cloud.access.token"
    assert metadata["token_source"] == "cloud"
    assert calls[0][0] == "https://example.invalid/baidu/oauth/token"
    assert calls[0][1] == {"apiProfile": "kunming_niu_baidu", "forceRefresh": False}
    assert "old.refresh.token" not in json.dumps(calls[0], ensure_ascii=False)


def test_token_manager_deducts_lock_wait_from_refresh_timeout(tmp_path, monkeypatch):
    from contextlib import contextmanager
    from datetime import datetime
    import modules.baidu_token_manager as token_manager

    _write_token_manager_secrets(tmp_path, expires_time="2026-07-17 09:05:00")
    now = [0.0]

    @contextmanager
    def delayed_lock(_path, **kwargs):
        assert kwargs["timeout_seconds"] == 7
        now[0] = 5.0
        yield

    def transport(_url, _payload, _headers, timeout):
        assert timeout == 2.0
        return _token_refresh_response()

    monkeypatch.setattr(token_manager, "secrets_file_lock", delayed_lock)
    token_manager.ensure_valid_access_token(
        {},
        tmp_path,
        "kunming_niu_baidu",
        now=datetime(2026, 7, 17, 9, 0, 0),
        transport=transport,
        timeout_seconds=7,
        clock=lambda: now[0],
    )


def test_token_manager_rotates_both_tokens_atomically(tmp_path):
    from datetime import datetime
    from modules.baidu_token_manager import ensure_valid_access_token

    secrets_path = _write_token_manager_secrets(tmp_path, expires_time="2026-07-17 09:05:00")
    ensure_valid_access_token(
        {},
        tmp_path,
        "kunming_niu_baidu",
        now=datetime(2026, 7, 17, 9, 0, 0),
        transport=lambda *_args: _token_refresh_response(),
    )

    profile = json.loads(secrets_path.read_text(encoding="utf-8"))["baidu_api"]["kunming_niu_baidu"]
    assert profile["access_token"] == "new.access.token"
    assert profile["refresh_token"] == "new.refresh.token"
    assert len(list((tmp_path / "backups").glob("secrets_before_token_refresh_kunming_niu_baidu_*.json"))) == 1
    assert not list((tmp_path / "secrets").glob("*.tmp"))


def test_token_manager_refresh_failure_preserves_original_secrets(tmp_path):
    from datetime import datetime
    from modules.baidu_token_manager import BaiduTokenError, ensure_valid_access_token

    secrets_path = _write_token_manager_secrets(tmp_path, expires_time="2026-07-17 09:05:00")
    original = secrets_path.read_bytes()
    with pytest.raises(BaiduTokenError) as exc_info:
        ensure_valid_access_token(
            {},
            tmp_path,
            "kunming_niu_baidu",
            now=datetime(2026, 7, 17, 9, 0, 0),
            transport=lambda *_args: (_ for _ in ()).throw(OSError("old.refresh.token network failure")),
        )

    assert exc_info.value.category == "token_refresh_error"
    assert "old.refresh.token" not in str(exc_info.value)
    assert secrets_path.read_bytes() == original
    assert not list((tmp_path / "secrets").glob("*.tmp"))


def test_token_manager_maps_refresh_http_errors_without_exposing_response(monkeypatch):
    import io
    import urllib.error
    import modules.baidu_token_manager as token_manager

    def raise_http_error(_request, timeout):
        assert timeout == 7
        raise urllib.error.HTTPError(
            "https://example.invalid/refresh",
            502,
            "bad",
            {},
            io.BytesIO(b'{"status":"error","code":"refresh_failed","message":"secret token"}'),
        )

    monkeypatch.setattr(token_manager.urllib.request, "urlopen", raise_http_error)
    with pytest.raises(token_manager.BaiduTokenError) as exc_info:
        token_manager._post_refresh(
            "https://example.invalid/refresh", {}, {}, 7
        )

    assert exc_info.value.category == "reauthorization_required"
    assert exc_info.value.reauthorization_required is True
    assert "secret token" not in str(exc_info.value)


@pytest.mark.parametrize(
    "response_code",
    ["app_id_mismatch", "refresh_signature_mismatch", "expired_refresh_request", "invalid_refresh_request"],
)
def test_token_manager_does_not_misclassify_refresh_configuration_errors(monkeypatch, response_code):
    import io
    import urllib.error
    import modules.baidu_token_manager as token_manager

    body = json.dumps({"status": "error", "code": response_code, "message": "not a token problem"}).encode()

    def raise_http_error(_request, timeout):
        raise urllib.error.HTTPError(
            "https://example.invalid/refresh", 400, "bad", {}, io.BytesIO(body)
        )

    monkeypatch.setattr(token_manager.urllib.request, "urlopen", raise_http_error)
    with pytest.raises(token_manager.BaiduTokenError) as exc_info:
        token_manager._post_refresh("https://example.invalid/refresh", {}, {}, 7)

    assert exc_info.value.category == "configuration_error"
    assert exc_info.value.reauthorization_required is False


def test_token_manager_merge_update_preserves_other_profiles(tmp_path):
    from datetime import datetime
    from modules.baidu_token_manager import ensure_valid_access_token

    other = {"access_token": "other.new.value", "refresh_token": "other.refresh.value"}
    secrets_path = _write_token_manager_secrets(
        tmp_path, expires_time="2026-07-17 09:05:00", other_profile=other
    )
    ensure_valid_access_token(
        {},
        tmp_path,
        "kunming_niu_baidu",
        now=datetime(2026, 7, 17, 9, 0, 0),
        transport=lambda *_args: _token_refresh_response(),
    )

    saved = json.loads(secrets_path.read_text(encoding="utf-8"))
    assert saved["baidu_api"]["other_baidu"] == other
    assert saved["baidu_api"]["kunming_niu_baidu"]["access_token"] == "new.access.token"


def test_token_manager_report_never_contains_credentials(tmp_path):
    from datetime import datetime
    from modules.baidu_token_manager import BaiduTokenError, ensure_valid_access_token

    _write_token_manager_secrets(tmp_path)
    _token, metadata = ensure_valid_access_token(
        {}, tmp_path, "kunming_niu_baidu", now=datetime(2026, 7, 17, 9, 0, 0)
    )
    serialized = json.dumps(metadata, ensure_ascii=False) + repr(metadata)
    for sensitive in ("old.access.token", "old.refresh.token", "fake-client-key", "fake-password"):
        assert sensitive not in serialized

    with pytest.raises(BaiduTokenError) as exc_info:
        ensure_valid_access_token(
            {}, tmp_path, "kunming_niu_baidu", now=datetime(2026, 7, 17, 9, 0, 0), force_refresh=True,
            transport=lambda *_args: {"status": "error", "message": "old.refresh.token fake-client-key"},
        )
    assert "old.refresh.token" not in str(exc_info.value)
    assert "fake-client-key" not in str(exc_info.value)


def test_regular_build_excludes_secrets_json():
    """普通 build_release 不包含 secrets/secrets.json。"""
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="0.4.15")
    assert "hourly_report_bot_release_v0.4.15" in release.name

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "secrets/secrets.json" not in names
    assert "secrets/secrets.example.json" in names
    for pid in ["kunming_niu", "nanjing_niu", "ningbo_niu", "changsha_niu", "shenyang_niu"]:
        assert f"configs/projects/{pid}.json" in names
    assert "configs/projects/kunming_npx.json" not in names
    assert "reports/menu_task_status.json" not in names
    assert "configs/projects/project_template.json" in names


def test_online_update_build_contains_program_but_excludes_user_data(tmp_path):
    from tools.build_release import release_name

    root = Path(__file__).resolve().parents[1]
    release = build_release(
        root,
        version="2026.7.22.109",
        online_update=True,
        output_dir=tmp_path,
    )

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())

    assert release.name == "Hourlyreport_automation_v2026.7.22.109.zip"
    assert release.parent == tmp_path
    assert release_name("2026.7.22.109", online_update=True) == release.name
    assert "hourlyreport_automation.exe" in names
    assert "main.py" in names
    assert "gui/version.py" in names
    assert not any(name.startswith("configs/") for name in names)
    assert not any(name.startswith("secrets/") for name in names)
    assert not any(name.startswith("logs/") for name in names)
    assert not any(name.startswith("reports/") for name in names)
    assert not any(name.startswith("backups/") for name in names)
    assert not any(name.startswith("browser_profile/") for name in names)
    assert not any(name.endswith(".lock") for name in names)


def test_first_install_build_is_standalone_but_excludes_real_secrets(tmp_path):
    import zipfile

    from tools.build_desktop_exe import write_build_manifest
    from tools.build_release import build_release, release_name

    (tmp_path / "dist").mkdir()
    (tmp_path / "dist" / "hourlyreport_automation.exe").write_bytes(b"exe")
    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text(json.dumps({
        "default_project_id": "demo",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
    }), encoding="utf-8")
    (tmp_path / "configs" / "projects" / "demo.json").write_text('{"project_id":"demo"}', encoding="utf-8")
    (tmp_path / "configs" / "current_project.json").write_text(
        '{"current_project_id":"private-local-state"}', encoding="utf-8"
    )
    (tmp_path / "configs" / "multi_project_selection.json").write_text(
        '{"project_ids":["private-local-state"]}', encoding="utf-8"
    )
    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.example.json").write_text('{"baidu":{}}', encoding="utf-8")
    (tmp_path / "secrets" / "secrets.json").write_text('{"secret":"must-not-ship"}', encoding="utf-8")
    (tmp_path / "main.py").write_text("print('ok')", encoding="utf-8")
    (tmp_path / "gui").mkdir()
    (tmp_path / "gui" / "version.py").write_text(
        'CURRENT_VERSION = "2026.7.17.103"\n', encoding="utf-8"
    )
    (tmp_path / "install_env.bat").write_text("@echo off\r\n", encoding="ascii")
    (tmp_path / "requirements-runtime.txt").write_text("openpyxl\n", encoding="utf-8")
    write_build_manifest(tmp_path, tmp_path / "dist" / "hourlyreport_automation.exe", "2026.7.17.103")

    release = build_release(tmp_path, version="2026.7.17.103", first_install=True)
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())

    assert release.name == "Hourlyreport_automation_first_install_v2026.7.17.103.zip"
    assert release_name("2026.7.17.103", first_install=True) == release.name
    assert "hourlyreport_automation.exe" in names
    assert "configs/app_config.json" in names
    assert "configs/projects/demo.json" in names
    assert "configs/current_project.json" not in names
    assert "configs/multi_project_selection.json" not in names
    assert "secrets/secrets.example.json" in names
    assert "secrets/secrets.json" not in names
    assert "install_env.bat" in names
    assert "requirements-runtime.txt" in names


def test_windows_installer_definition_allows_path_selection_and_preserves_user_data():
    from tools.build_windows_installer import installer_name

    root = Path(__file__).resolve().parents[1]
    source = (root / "tools" / "hourlyreport_automation_installer.iss").read_text(encoding="utf-8")

    assert installer_name("2026.7.19.106") == "Hourlyreport_automation_setup_v2026.7.19.106.exe"
    assert "DisableDirPage=no" in source
    assert "PrivilegesRequired=lowest" in source
    assert "{localappdata}\\Programs\\Hourlyreport Automation" in source
    assert "{autodesktop}" in source
    assert "{autoprograms}" in source
    assert "onlyifdoesntexist" in source
    assert "uninsneveruninstall" in source
    assert 'Excludes: "configs\\*,secrets\\*,logs\\*,reports\\*,backups\\*,kst_exports\\*"' in source


def test_first_install_build_refuses_incomplete_source_tree(tmp_path):
    import pytest

    from tools.build_release import build_release

    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text('{}', encoding="utf-8")
    (tmp_path / "configs" / "projects" / "demo.json").write_text('{}', encoding="utf-8")
    (tmp_path / "main.py").write_text("print('ok')", encoding="utf-8")

    with pytest.raises(ValueError, match="首次安装包源文件不完整.*exe"):
        build_release(tmp_path, version="2026.7.17.103", first_install=True)

    assert not (tmp_path / "dist" / "Hourlyreport_automation_first_install_v2026.7.17.103.zip").exists()


def test_online_release_version_counter_never_resets_with_date():
    from datetime import date

    from tools.build_release import next_online_version, validate_online_version

    assert validate_online_version("2026.7.15.101") == "2026.7.15.101"
    assert next_online_version("2026.7.15.101", date(2026, 7, 16)) == "2026.7.16.102"
    assert next_online_version("2026.7.16.102", date(2026, 7, 16)) == "2026.7.16.103"


def test_online_release_version_rejects_invalid_date_or_counter():
    import pytest

    from tools.build_release import validate_online_version

    with pytest.raises(ValueError, match="日期"):
        validate_online_version("2026.2.30.102")
    with pytest.raises(ValueError, match="100"):
        validate_online_version("2026.7.16.99")


def test_online_update_file_filter_never_includes_user_configuration():
    assert should_include_file(Path("main.py"), online_update=True) is True
    assert should_include_file(Path("gui") / "main_window.py", online_update=True) is True
    assert should_include_file(Path("dist") / "hourlyreport_automation.exe", online_update=True) is True
    assert should_include_file(Path("configs") / "app_config.json", online_update=True) is False
    assert should_include_file(Path("configs") / "projects" / "kunming_niu.json", online_update=True) is False
    assert should_include_file(Path("secrets") / "secrets.json", online_update=True) is False


def test_release_filter_excludes_exported_authorization_packages():
    assert should_include_file(Path("百度授权配置.baidu-secrets"), internal=True) is False
    assert should_include_file(Path("exports") / "team.baidu-secrets", online_update=True) is False
    assert should_include_file(Path("nested") / "team.baidu-secrets") is False
    assert should_include_file(Path("Downloads") / "baidu_oauth_123.baidu-auth", first_install=True) is False
    assert should_include_file(Path("baidu_oauth_456.baidu-auth"), online_update=True) is False


def test_gitignore_excludes_exported_authorization_packages():
    root = Path(__file__).resolve().parents[1]
    ignored_patterns = {
        line.strip()
        for line in (root / ".gitignore").read_text(encoding="utf-8").splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }
    assert "*.baidu-secrets" in ignored_patterns
    assert "secrets/*.lock" in ignored_patterns


def test_internal_build_excludes_secrets_json():
    """内部 build_release 也不再包含 secrets/secrets.json。"""
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="0.4.15", internal=True)
    assert "hourly_report_bot_internal_v0.4.15" in release.name

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "secrets/secrets.json" not in names
    assert not any(name.endswith(".lock") for name in names)
    assert "secrets/secrets.example.json" in names


def test_built_archive_excludes_authorization_package_and_real_secrets(tmp_path):
    import zipfile

    (tmp_path / "main.py").write_text("print('demo')\n", encoding="utf-8")
    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.json").write_text('{"baidu": {}}', encoding="utf-8")
    (tmp_path / "secrets" / "secrets.example.json").write_text('{"baidu": {}}', encoding="utf-8")
    (tmp_path / "team.baidu-secrets").write_text("sensitive-placeholder", encoding="utf-8")

    release = build_release(tmp_path, version="config-safe", internal=True)

    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "main.py" in names
    assert "secrets/secrets.example.json" in names
    assert "secrets/secrets.json" not in names
    assert "team.baidu-secrets" not in names


def test_internal_build_includes_desktop_exe_when_available():
    root = Path(__file__).resolve().parents[1]
    exe = root / "dist" / "hourlyreport_automation.exe"
    exe.parent.mkdir(parents=True, exist_ok=True)
    if not exe.exists():
        exe.write_bytes(b"placeholder exe")

    release = build_release(root, version="2.0", internal=True)

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "hourlyreport_automation.exe" in names
    assert not any(name.startswith("dist/") for name in names)
    assert not any("_internal" in name for name in names)
    assert release.name == "hourly_report_bot_internal_v2.0.zip"


def test_release_resets_machine_specific_desktop_pet_position():
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="pet-portable", internal=True)

    import zipfile
    with zipfile.ZipFile(release) as archive:
        config = json.loads(archive.read("configs/app_config.json").decode("utf-8"))

    assert config["desktop_pet"] == "clawd"
    assert config["desktop_pet_scale"] == 1.0
    assert "desktop_pet_position" not in config


def test_default_internal_release_name_uses_hermes_date_marker():
    from tools.build_release import release_name

    assert release_name(internal=True) == "hourly_report_bot_internal_hermes_20260710.zip"


# ── 夏思道说明文件测试 ────────────────────────────────────


def test_xia_sidao_readme_exists():
    """xia_sidao使用说明.md 存在。"""
    root = Path(__file__).resolve().parents[1]
    path = root / "xia_sidao使用说明.md"
    assert path.exists(), "xia_sidao使用说明.md 不存在"
    content = path.read_text(encoding="utf-8")
    assert "唯一自动入口" in content
    assert "HERMES-20260710" in content
    assert "run_hermes_hourly.bat" in content
    assert "run_hermes_daily.bat" in content
    for name in ["昆明牛", "南京牛", "宁波牛", "长沙牛", "沈阳牛", "青岛白", "深圳白", "南京白", "沈阳白"]:
        assert name in content, f"缺少项目：{name}"
    assert "双百度来源" in content


def test_xia_sidao_readme_tracks_hermes_only_fixed_entries():
    """夏思道说明只保留 HERMES 固定入口。"""
    root = Path(__file__).resolve().parents[1]
    content = (root / "xia_sidao使用说明.md").read_text(encoding="utf-8")

    assert "HERMES-20260710" in content
    assert "run_hermes_hourly.bat 11点" in content
    assert "run_hermes_daily.bat" in content
    retired_names = ("".join(("Open", "Claw")), "".join(("run_", "open", "claw")))
    assert all(name not in content for name in retired_names)


def test_xia_sidao_readme_tracks_current_scope_without_retired_workflows():
    """夏思道说明只保留当前能力与安全规则。"""
    root = Path(__file__).resolve().parents[1]
    content = (root / "xia_sidao使用说明.md").read_text(encoding="utf-8")

    assert "正式项目" in content
    for name in ["昆明牛", "南京牛", "宁波牛", "长沙牛", "沈阳牛", "青岛白", "深圳白", "南京白", "沈阳白"]:
        assert name in content, f"缺少项目：{name}"
    for text in ["Excel 写入前必须备份原文件", "不重建工作簿", "失败后禁止手工补数字"]:
        assert text in content
    for text in ["Chrome 调试端口 `9222`", "不自动改用 Edge", "cookie/storage"]:
        assert text in content
    for retired in ["腾讯文档", "fill_daily_visit.py", "cron", "v0.4.19", "v0.4.21", "v1.0 内部发布版"]:
        assert retired not in content


def test_regular_build_includes_xia_sidao_readme():
    """普通包包含 xia_sidao使用说明.md。"""
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="0.4.16")

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "xia_sidao使用说明.md" in names


def test_internal_build_includes_xia_sidao_readme():
    """内部包包含 xia_sidao使用说明.md。"""
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="0.4.16", internal=True)

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "xia_sidao使用说明.md" in names


def test_desktop_gui_project_store_excludes_templates(tmp_path):
    projects_dir = tmp_path / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text(json.dumps({
        "default_project_id": "alpha",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
    }, ensure_ascii=False), encoding="utf-8")
    (projects_dir / "alpha.json").write_text(json.dumps({
        "project_id": "alpha",
        "project_name": "项目A",
    }, ensure_ascii=False), encoding="utf-8")
    (projects_dir / "project_template.json").write_text(json.dumps({
        "project_id": "your_project_id",
        "project_name": "模板",
        "is_template": True,
    }, ensure_ascii=False), encoding="utf-8")

    from gui.project_store import load_project_summaries

    projects = load_project_summaries(tmp_path)

    assert [(item.project_id, item.project_name) for item in projects] == [("alpha", "项目A")]


def test_desktop_gui_command_builder_uses_existing_main_entry():
    from gui.command_builder import build_daily_command, build_hourly_command, build_preflight_command

    root = Path("D:/app")

    assert build_hourly_command(root, "15", project_id="qingdao_bai") == [
        str(root / ".venv" / "Scripts" / "pythonw.exe"),
        "-u",
        str(root / "main.py"),
        "--mode",
        "run",
        "--project",
        "qingdao_bai",
        "--period",
        "15点",
        "--yes",
    ]
    assert build_daily_command(root, "2026-06-06", project_id="shenyang_bai") == [
        str(root / ".venv" / "Scripts" / "pythonw.exe"),
        "-u",
        str(root / "main.py"),
        "--mode",
        "run-daily",
        "--project",
        "shenyang_bai",
        "--date",
        "2026-06-06",
        "--yes",
    ]
    assert build_preflight_command(root, "daily", project_id="shenyang_bai") == [
        str(root / ".venv" / "Scripts" / "pythonw.exe"),
        "-u",
        str(root / "main.py"),
        "--mode",
        "preflight",
        "--project",
        "shenyang_bai",
        "--task",
        "daily",
        "--quick",
    ]


def test_desktop_gui_environment_subprocess_is_hidden():
    from gui.environment_check import hidden_subprocess_kwargs

    kwargs = hidden_subprocess_kwargs()

    if os.name == "nt":
        assert kwargs["creationflags"] != 0
        assert kwargs["startupinfo"].dwFlags != 0
    else:
        assert kwargs == {}


def test_desktop_gui_task_runner_forces_utf8_environment():
    from gui.task_runner import build_process_environment

    env = build_process_environment()

    assert env.value("PYTHONUTF8") == "1"
    assert env.value("PYTHONIOENCODING") == "utf-8"
    assert env.value("PYTHONUNBUFFERED") == "1"


def test_gui_commands_use_unbuffered_python(tmp_path):
    from gui.command_builder import build_hourly_command

    command = build_hourly_command(tmp_path, "18点", project_id="kunming_niu")

    assert command[1] == "-u"


def test_stream_output_buffers_partial_utf8_lines():
    from gui.task_runner import split_stream_output

    lines, pending = split_stream_output("", "[API] 正在读")
    assert lines == []
    assert pending == "[API] 正在读"

    lines, pending = split_stream_output(pending, "取数据\n[通过] 完成\n")
    assert lines == ["[API] 正在读取数据", "[通过] 完成"]
    assert pending == ""

    lines, pending = split_stream_output("末行", "", final=True)
    assert lines == ["末行"]
    assert pending == ""


def test_stream_output_keeps_crlf_split_across_chunks_as_one_newline():
    from gui.task_runner import split_stream_output

    lines, pending = split_stream_output("", "第一行\r")
    assert lines == []
    assert pending == "第一行\r"

    lines, pending = split_stream_output(pending, "\n第二行\r")
    assert lines == ["第一行"]
    assert pending == "第二行\r"

    lines, pending = split_stream_output(pending, "", final=True)
    assert lines == ["第二行"]
    assert pending == ""


def test_task_runner_flushes_pending_output_and_clears_it_for_new_task(tmp_path, monkeypatch):
    from types import SimpleNamespace

    import gui.task_runner as task_runner

    class SignalRecorder:
        def __init__(self):
            self.values = []

        def emit(self, value):
            self.values.append(value)

    class Process:
        def __init__(self):
            self.chunks = ["[API] 正在读".encode("utf-8")]

        def readAllStandardOutput(self):
            return self.chunks.pop(0) if self.chunks else b""

        def setProcessEnvironment(self, value):
            self.environment = value

        def setWorkingDirectory(self, value):
            self.cwd = value

        def start(self, program, arguments):
            self.command = [program, *arguments]

    process = Process()
    output = SignalRecorder()
    stages = SignalRecorder()
    finished = SignalRecorder()
    runner = SimpleNamespace(
        _process=process,
        _pending_output="",
        output=output,
        stage_changed=stages,
        finished=finished,
    )

    task_runner.QtTaskRunner._read_output(runner)
    assert output.values == []
    assert runner._pending_output == "[API] 正在读"

    task_runner.QtTaskRunner._handle_finished(runner, 0, None)
    assert output.values == ["[API] 正在读"]
    assert stages.values == ["baidu"]
    assert finished.values == [0]

    monkeypatch.setattr(task_runner, "build_process_environment", lambda: object())
    runner._pending_output = "上一任务残留"
    runner.is_running = lambda: False
    runner.failed_to_start = SignalRecorder()
    task_runner.QtTaskRunner.start(runner, ["python", "-u", "main.py"], tmp_path)
    assert runner._pending_output == ""


def test_task_runner_finished_reads_remaining_process_output_before_final_flush():
    from types import SimpleNamespace

    from gui.task_runner import QtTaskRunner

    class SignalRecorder:
        def __init__(self):
            self.values = []

        def emit(self, value):
            self.values.append(value)

    class Process:
        def readAllStandardOutput(self):
            return "[API] 完成时仍未读取的末行\n".encode("utf-8")

    output = SignalRecorder()
    stages = SignalRecorder()
    finished = SignalRecorder()
    runner = SimpleNamespace(
        _process=Process(),
        _pending_output="",
        output=output,
        stage_changed=stages,
        finished=finished,
    )

    QtTaskRunner._handle_finished(runner, 0, None)

    assert output.values == ["[API] 完成时仍未读取的末行"]
    assert stages.values == ["baidu"]
    assert finished.values == [0]


def test_task_runner_preserves_utf8_characters_split_across_process_chunks():
    from types import SimpleNamespace

    from gui.task_runner import QtTaskRunner

    class SignalRecorder:
        def __init__(self):
            self.values = []

        def emit(self, value):
            self.values.append(value)

    class Process:
        def __init__(self):
            self.chunks = [b"[API] \xe6\xad", b"\xa3\xe5\x9c\xa8\xe8\xaf\xbb\xe5\x8f\x96\n"]

        def readAllStandardOutput(self):
            return self.chunks.pop(0)

    output = SignalRecorder()
    runner = SimpleNamespace(
        _process=Process(),
        _pending_output="",
        output=output,
        stage_changed=SignalRecorder(),
    )

    QtTaskRunner._read_output(runner)
    QtTaskRunner._read_output(runner)

    assert output.values == ["[API] 正在读取"]


def test_desktop_gui_requirements_keep_updater_ui_in_runtime_and_packager_in_dev():
    root = Path(__file__).resolve().parents[1]
    requirements = (root / "requirements-dev.txt").read_text(encoding="utf-8")
    runtime = (root / "requirements-runtime.txt").read_text(encoding="utf-8")

    assert "PySide6" in requirements
    assert "pyinstaller" in requirements
    assert "PySide6" not in runtime
    assert "pyinstaller" not in runtime
    assert "playwright" in runtime


def test_desktop_gui_resolves_project_root_from_workspace():
    from gui.app import GUI_SCALE_FACTOR, resolve_app_root

    root = Path(__file__).resolve().parents[1]

    assert resolve_app_root() == root
    assert GUI_SCALE_FACTOR == "1.0"


def test_desktop_gui_rejects_standalone_online_update_folder(tmp_path):
    import pytest

    from gui.app import IncompleteInstallationError, resolve_app_root

    update_root = tmp_path / "Hourlyreport_automation_v2026.7.16.102"
    update_root.mkdir()
    (update_root / "main.py").write_text("print('update only')\n", encoding="utf-8")

    with pytest.raises(IncompleteInstallationError, match="在线更新包.*首次安装包"):
        resolve_app_root([update_root])


def test_desktop_gui_main_reports_incomplete_install_without_traceback(monkeypatch):
    import gui.app as gui_app

    messages = []
    error = gui_app.IncompleteInstallationError("程序文件不完整，请使用首次安装包。")
    monkeypatch.setattr(gui_app, "resolve_app_root", lambda: (_ for _ in ()).throw(error))
    monkeypatch.setattr(gui_app, "show_startup_error", lambda message: messages.append(message))

    assert gui_app.main() == 2
    assert messages == [str(error)]


def test_desktop_gui_progress_lives_below_project_selector(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert window.progress.objectName() == "taskProgress"
    assert window.progress_text.objectName() == "taskProgressText"
    assert window.progress.maximum() == 8
    assert window.progress.isHidden()
    assert "任务" in window.progress_text.text()
    window.close()


def test_desktop_gui_normal_window_is_fixed_with_standard_controls(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtGui import QFont
    from PySide6.QtWidgets import QApplication, QFrame, QGraphicsDropShadowEffect
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert (window.width(), window.height()) == (966, 700)
    assert (window.minimumWidth(), window.minimumHeight()) == (966, 700)
    assert (window.maximumWidth(), window.maximumHeight()) == (966, 700)
    window.resize(1200, 900)
    assert (window.width(), window.height()) == (966, 700)
    assert window.left_panel.minimumWidth() == 372
    assert window.left_panel.maximumWidth() == 372
    assert window.status_title.isHidden()
    assert window.status_detail.isHidden()
    assert bool(window.windowFlags() & Qt.WindowType.FramelessWindowHint)
    assert bool(window.windowFlags() & Qt.WindowType.WindowMinimizeButtonHint)
    assert window.title_bar.objectName() == "titleBar"
    assert window.title_bar.height() == 39
    assert window.title_layout.spacing() == 2
    assert 40 <= window.spinner.width() <= 52
    assert window.spinner.height() <= 26
    assert window.spinner.objectName() == "clawdAnimator"
    assert window.windowTitle() == "蚁之力-竞价数据自动化"
    assert window.title_label.text() == "蚁之力-竞价数据自动化"
    assert window.title_label.font().pointSize() == 10
    assert window.system_config_button.text() == "系统"
    assert window.system_config_button.height() == window.title_label.height()
    assert window.title_layout.itemAt(2).spacerItem().sizeHint().width() == 10
    assert window.hourly_title.text() == "小时报"
    assert window.daily_title.text() == "日报"
    assert [action.text() for action in window.system_config_menu.actions() if not action.isSeparator()] == [
        "项目配置检查", "导入授权配置", "导出授权配置", "恢复备份", "Excel 路径配置", "Excel 自动打开", "桌面宠物", "退出程序"
    ]
    assert window.minimize_button.toolTip() == "最小化"
    assert window.maximize_button.toolTip() == "最大化"
    assert not hasattr(window, "size_grip")
    assert all(button.width() == 34 and button.height() == 32 for button in (
        window.minimize_button, window.maximize_button, window.close_button
    ))
    assert window.close_button.toolTip() == "关闭界面"
    assert not window.close_button.icon().isNull()
    assert not window.windowIcon().isNull()
    assert window.font().family() == "Microsoft YaHei UI"
    assert window.font().weight() == QFont.Weight.Normal
    assert window.font().pointSize() == 9
    assert window.left_layout.spacing() == 14
    assert window.task_control_card.minimumHeight() == window.task_control_card.maximumHeight() == 208
    assert window.hourly_card.minimumHeight() == window.hourly_card.maximumHeight() == 202
    assert window.daily_card.minimumHeight() == window.daily_card.maximumHeight() == 204
    assert window.stage_panel.isHidden()
    assert window.current_flow_panel.height() == 224
    assert window.log_view.height() >= 300
    assert window.shell_surface.objectName() == "shellSurface"
    assert window.shell_surface.frameShape() == QFrame.Shape.NoFrame
    assert window.testAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
    assert window.window_surface.objectName() == "windowSurface"
    window_shadow = window.window_surface.graphicsEffect()
    assert isinstance(window_shadow, QGraphicsDropShadowEffect)
    assert window_shadow.blurRadius() == 8
    assert window_shadow.offset().isNull()
    assert window_shadow.color().alpha() == 38
    root_margins = window.centralWidget().layout().contentsMargins()
    assert (root_margins.left(), root_margins.top(), root_margins.right(), root_margins.bottom()) == (4, 4, 4, 4)
    shell_margins = window.shell_layout.contentsMargins()
    assert (shell_margins.left(), shell_margins.top(), shell_margins.right(), shell_margins.bottom()) == (4, 4, 4, 5)
    assert "QFrame#shellSurface" in window.styleSheet()
    assert "QFrame#windowSurface" in window.styleSheet()
    assert "border: 1px solid #e6e9ee" in window.styleSheet()
    assert "#aeb6c1" not in window.styleSheet()
    assert "border-radius: 0" in window.styleSheet()
    window.close()


def test_data_source_control_defaults_to_api_and_emits_browser(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import DataSourceModeControl

    app = QApplication.instance() or QApplication([])
    control = DataSourceModeControl()
    changes = []
    control.preference_changed.connect(changes.append)

    assert (control.width(), control.height()) == (94, 26)
    assert control.preference() == "api"
    assert control.api_button.isChecked()
    assert not control.browser_button.isChecked()
    assert control.prefix_label.text() == "模式："
    assert control.browser_button.text() == "B"
    assert control.api_button.text() == "A"
    assert control.display_order() == "A>B"
    assert control.api_button.x() < control.browser_button.x()
    assert control.segment_frame.objectName() == "dataSourceModeSegment"
    assert control.segment_frame.geometry().contains(control.api_button.geometry().center())
    assert control.segment_frame.geometry().contains(control.browser_button.geometry().center())
    assert "QFrame#dataSourceModeSegment" in control.styleSheet()
    assert control.api_button.width() == control.api_button.height() == 22
    assert control.browser_button.width() == control.browser_button.height() == 22
    assert "border-radius: 11px" in control.styleSheet()
    assert "border: 1px solid #b9d8ff" in control.styleSheet()
    assert "background: #3f83f8" in control.styleSheet()
    assert "color: #ffffff" in control.styleSheet()

    control.set_preference("browser", animate=False)

    assert control.preference() == "browser"
    assert changes == ["browser"]
    assert control.browser_button.isChecked()
    assert not control.api_button.isChecked()
    assert control.display_order() == "B>A"
    assert control.browser_button.x() < control.api_button.x()
    control.close()


def test_data_source_control_animates_selected_mode_to_the_left(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import QApplication
    from gui.main_window import DataSourceModeControl

    app = QApplication.instance() or QApplication([])
    control = DataSourceModeControl()
    control.show()
    app.processEvents()

    control.set_preference("browser", animate=True)
    QTest.qWait(220)

    assert control.display_order() == "B>A"
    assert control.browser_button.x() < control.api_button.x()
    control.close()


def test_desktop_gui_data_mode_control_lives_in_project_header(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)

    assert not hasattr(window, "data_mode_button")
    assert not hasattr(window, "data_mode_menu")
    assert not hasattr(window.inline_config_menu, "data_source_control")
    control = window.data_source_control
    assert (control.width(), control.height()) == (94, 26)
    assert control.preference() == "api"
    assert window.project_combo.popup.search.placeholderText() == "输入 B/N 快速检索对应项目"
    assert control.parentWidget() is not window.title_bar

    window._quitting = True
    window.close()


def test_desktop_gui_data_source_preference_persists_and_locks_during_tasks(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow
    from modules.project_config import get_data_source_preference

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    control = window.data_source_control

    assert control.preference() == "api"
    control.set_preference("browser", animate=False)
    assert get_data_source_preference(tmp_path) == "browser"

    window.runner.start = lambda _command, _root: None
    window.start_command("测试任务", ["python", "-V"])
    assert not control.isEnabled()
    window.show_task_error("测试启动失败")
    assert control.isEnabled()

    window.on_task_started()
    assert not control.isEnabled()
    window.on_task_finished(1)
    assert control.isEnabled()

    window.on_task_started()
    assert not control.isEnabled()
    window.show_task_error("测试启动失败")
    assert control.isEnabled()

    window._quitting = True
    window.close()


def test_excel_auto_open_defaults_true_migrates_old_preferences_and_persists(tmp_path):
    from gui.excel_open_settings import load_auto_open_excel, save_auto_open_excel

    assert load_auto_open_excel(tmp_path) is True
    config_path = tmp_path / "configs" / "app_config.json"
    config_path.parent.mkdir(parents=True)
    config_path.write_text(
        json.dumps({"open_excel_after_hourly": False, "open_excel_after_daily": True}),
        encoding="utf-8",
    )
    assert load_auto_open_excel(tmp_path) is False

    save_auto_open_excel(tmp_path, True)

    saved = json.loads(config_path.read_text(encoding="utf-8"))
    assert saved["open_excel_automatically"] is True
    assert load_auto_open_excel(tmp_path) is True


def _write_excel_path_project(path, project_id, project_name, excel_path, *, template=False):
    payload = {
        "project_id": project_id,
        "project_name": project_name,
        "is_template": template,
        "excel": {"path": str(excel_path), "hourly_sheet": "时段数据", "daily_sheet": "百度"},
        "keep_me": {"unchanged": True},
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def test_configure_excel_paths_updates_every_project_after_complete_validation(tmp_path):
    from modules.excel_path_config import configure_excel_paths

    projects_dir = tmp_path / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    selected_root = tmp_path / "new-drive" / "【竞价】"
    selected_root.mkdir(parents=True)
    suffixes = {
        "alpha": ("【❤甲】", "【2026年】【甲】竞价数据", "甲.xlsx"),
        "beta": ("【❤乙】", "【2026年】【乙】竞价数据", "乙.xlsm"),
    }
    originals = {}
    for project_id, suffix in suffixes.items():
        target = selected_root.joinpath(*suffix)
        target.parent.mkdir(parents=True)
        target.write_bytes(b"excel")
        old_path = Path("D:/Seafile/【竞价】").joinpath(*suffix)
        config_path = projects_dir / f"{project_id}.json"
        originals[project_id] = _write_excel_path_project(config_path, project_id, project_id, old_path)
    _write_excel_path_project(
        projects_dir / "project_template.json",
        "template",
        "模板",
        "D:/template.xlsx",
        template=True,
    )

    result = configure_excel_paths(tmp_path, selected_root)

    assert result.updated == 2
    assert result.errors == ()
    assert result.backup_dir is not None and result.backup_dir.is_dir()
    for project_id, suffix in suffixes.items():
        config_path = projects_dir / f"{project_id}.json"
        saved = json.loads(config_path.read_text(encoding="utf-8"))
        assert Path(saved["excel"]["path"]) == selected_root.joinpath(*suffix)
        assert saved["keep_me"] == originals[project_id]["keep_me"]
        assert (result.backup_dir / config_path.name).is_file()


def test_configure_excel_paths_is_all_or_nothing_when_any_target_is_missing(tmp_path):
    from modules.excel_path_config import configure_excel_paths

    projects_dir = tmp_path / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    selected_root = tmp_path / "【竞价】"
    selected_root.mkdir()
    existing = selected_root / "甲" / "甲.xlsx"
    existing.parent.mkdir()
    existing.write_bytes(b"excel")
    paths = {
        "alpha": "D:/Seafile/【竞价】/甲/甲.xlsx",
        "beta": "D:/Seafile/【竞价】/乙/乙.xlsx",
    }
    for project_id, excel_path in paths.items():
        _write_excel_path_project(projects_dir / f"{project_id}.json", project_id, project_id, excel_path)
    before = {path.name: path.read_bytes() for path in projects_dir.glob("*.json")}

    result = configure_excel_paths(tmp_path, selected_root)

    assert result.updated == 0
    assert result.backup_dir is None
    assert any("beta" in error and "乙.xlsx" in error for error in result.errors)
    assert {path.name: path.read_bytes() for path in projects_dir.glob("*.json")} == before


def test_configure_excel_paths_rejects_folder_not_named_competition(tmp_path):
    from modules.excel_path_config import configure_excel_paths

    selected_root = tmp_path / "Seafile"
    selected_root.mkdir()

    result = configure_excel_paths(tmp_path, selected_root)

    assert result.updated == 0
    assert result.backup_dir is None
    assert result.errors == ("请选择名称为【竞价】的文件夹。",)


def test_configure_excel_paths_restores_every_config_after_write_failure(tmp_path, monkeypatch):
    import modules.excel_path_config as excel_path_config

    projects_dir = tmp_path / "configs" / "projects"
    projects_dir.mkdir(parents=True)
    selected_root = tmp_path / "【竞价】"
    selected_root.mkdir()
    for project_id in ("alpha", "beta"):
        target = selected_root / project_id / f"{project_id}.xlsx"
        target.parent.mkdir()
        target.write_bytes(b"excel")
        _write_excel_path_project(
            projects_dir / f"{project_id}.json",
            project_id,
            project_id,
            f"D:/Seafile/【竞价】/{project_id}/{project_id}.xlsx",
        )
    before = {path.name: path.read_bytes() for path in projects_dir.glob("*.json")}
    real_write = excel_path_config._write_json_atomic
    calls = []

    def fail_second_write(path, data):
        calls.append(path)
        if len(calls) == 2:
            raise OSError("simulated write failure")
        real_write(path, data)

    monkeypatch.setattr(excel_path_config, "_write_json_atomic", fail_second_write)

    result = excel_path_config.configure_excel_paths(tmp_path, selected_root)

    assert result.updated == 0
    assert any("已恢复原配置" in error for error in result.errors)
    assert {path.name: path.read_bytes() for path in projects_dir.glob("*.json")} == before


def test_desktop_gui_excel_auto_open_menu_controls_all_task_completion(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.excel_open_settings import load_auto_open_excel
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    opened = []
    logs = []
    announcements = []
    window.open_current_project_excel = lambda: opened.append(window.current_task_type)
    window.append_log = logs.append
    window.desktop_pet.announce = lambda *args, **kwargs: announcements.append(args)
    window.desktop_pet.set_busy = lambda _busy: None

    assert not hasattr(window, "hourly_open_toggle")
    assert not hasattr(window, "daily_open_toggle")
    assert window.open_excel_automatically is True
    assert window.inline_config_menu.excel_start_choice.property("selected") is True
    assert window.inline_config_menu.excel_stop_choice.property("selected") is False
    window.set_excel_auto_open(False)
    assert load_auto_open_excel(tmp_path) is False

    window.current_task_type = "hourly"
    window.current_project_name = "演示项目"
    window.on_task_finished(0)
    assert opened == []
    assert any("跳过打开 Excel" in line for line in logs)

    window.set_excel_auto_open(True)
    window.current_task_type = "daily"
    window.current_project_name = "演示项目"
    window.on_task_finished(0)
    assert opened == ["daily"]

    window._quitting = True
    window.close()


def test_desktop_gui_help_menu_and_manual_update_check(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QDialog
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    starts = []
    window.update_manager.start = lambda: starts.append(True)
    monkeypatch.setattr(QDialog, "exec", lambda self: 0)

    assert window.help_button.text() == "帮助"
    assert [action.text() for action in window.help_menu.actions()] == ["关于小螃蟹", "检查版本"]

    window.check_updates_manually()
    assert starts == [True]
    assert window._manual_update_check_requested is True
    window.on_update_up_to_date()
    assert window.update_button.isHidden()
    assert window.update_button.property("updateState") == "hidden"

    window.show_about_clawd()
    assert window._last_about_dialog.windowTitle() == "关于小螃蟹"
    about_text = "\n".join(label.text() for label in window._last_about_dialog.findChildren(main_window.QLabel))
    assert "Clawd 小螃蟹" in about_text
    assert main_window.CURRENT_VERSION in about_text
    assert "179068898-dotcom" in about_text

    window._quitting = True
    window.close()


@pytest.mark.parametrize("failure_path", ["exit_code", "start_failure", "recheck_failure"])
def test_desktop_gui_environment_failures_restore_data_source_control(
    tmp_path,
    monkeypatch,
    failure_path,
):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    control = window.data_source_control
    control.setEnabled(False)

    if failure_path == "exit_code":
        window.on_environment_install_finished(7)
    elif failure_path == "start_failure":
        window.on_environment_install_failed("测试失败")
    else:
        monkeypatch.setattr(
            main_window,
            "run_environment_check",
            lambda _root: {"passed": False, "checks": []},
        )
        window.on_environment_install_finished(0)

    assert control.isEnabled()
    assert not window.hourly_button.isEnabled()
    assert not window.daily_button.isEnabled()
    window._quitting = True
    window.close()


def test_desktop_gui_data_source_save_failure_rolls_back_without_sensitive_details(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    import gui.main_window as main_window
    from modules.project_config import get_data_source_preference

    _write_minimal_gui_project(tmp_path)
    config_path = tmp_path / "configs" / "app_config.json"
    original_disk = config_path.read_text(encoding="utf-8")
    logs = []
    monkeypatch.setattr(
        main_window,
        "save_data_source_preference",
        lambda _root, _preference: (_ for _ in ()).throw(
            OSError(r"D:\private\secrets.json token=do-not-log")
        ),
    )

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.append_log = logs.append
    control = window.data_source_control
    control.set_preference("browser", animate=False)

    assert window.data_source_preference == "api"
    assert control.preference() == "api"
    assert control.api_button.isChecked()
    assert get_data_source_preference(tmp_path) == "api"
    assert config_path.read_text(encoding="utf-8") == original_disk
    assert logs == ["[失败] 数据源设置未保存，已继续使用原设置。"]
    assert "private" not in window.progress_text.text().lower()
    assert "token" not in window.progress_text.text().lower()

    window._quitting = True
    window.close()


def test_gui_uses_unified_hourlyreport_technical_names(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.desktop_pet import ClawdDesktopPet
    from gui.main_window import MainWindow
    from gui.update_manager import APP_EXE_NAME, GITHUB_LATEST_RELEASE_URL, UPDATE_ASSET_PATTERN
    from tools.build_desktop_exe import APP_NAME
    from tools.build_release import DESKTOP_EXE, release_name

    app = QApplication.instance() or QApplication([])
    root = Path(__file__).resolve().parents[1]
    window = MainWindow(root)
    pet = ClawdDesktopPet(root, lambda: None, lambda _x, _y: None)

    assert window.tray_icon.toolTip() == "蚁之力 · 竞价数据自动化"
    assert "蚁之力 · 竞价数据自动化" in pet.toolTip()
    assert APP_EXE_NAME == "hourlyreport_automation.exe"
    assert APP_NAME == "hourlyreport_automation"
    assert DESKTOP_EXE == APP_EXE_NAME
    assert GITHUB_LATEST_RELEASE_URL == "https://api.github.com/repos/179068898-dotcom/Hourlyreport-Automation/releases/latest"
    assert UPDATE_ASSET_PATTERN.fullmatch("Hourlyreport_automation_v2026.7.17.103.zip")
    assert release_name("2026.7.17.103", internal=True).startswith("hourly_report_bot_internal_")
    assert release_name("2026.7.17.103") == "hourly_report_bot_release_v2026.7.17.103.zip"

    pet.close_pet()
    window._quitting = True
    window.tray_icon.hide()
    window.close()


def test_desktop_gui_log_bottom_aligns_with_daily_card(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    daily_bottom = window.daily_card.mapTo(window, window.daily_card.rect().bottomLeft()).y()
    content_bottom = window.content_panel.mapTo(window, window.content_panel.rect().bottomLeft()).y()
    assert content_bottom == daily_bottom

    window._quitting = True
    window.close()


def test_desktop_gui_maximize_button_toggles_standard_and_maximized_states(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    window.toggle_maximize()
    app.processEvents()
    assert window.isMaximized()
    assert window.maximize_button.toolTip() == "还原"
    window.toggle_maximize()
    app.processEvents()
    assert not window.isMaximized()
    assert (window.width(), window.height()) == (966, 700)
    assert window.maximize_button.toolTip() == "最大化"
    window._quitting = True
    window.close()


def test_desktop_gui_matches_reference_dashboard_structure(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert window.task_control_card.objectName() == "dashboardCard"
    assert window.hourly_card.objectName() == "dashboardCard"
    assert window.daily_card.objectName() == "dashboardCard"
    assert window.current_flow_panel.objectName() == "currentFlowPanel"
    assert window.current_flow_title.text() == "当前流程"
    assert window.current_task_title.text() == "暂无运行任务"
    assert window.current_task_subtitle.text() == "请选择左侧任务开始执行"
    assert window.current_status_badge.text() == "空闲"
    assert window.flow_header_icon.objectName() == "flowHeaderIcon"
    assert window.flow_header_icon.pixmap().width() >= 22
    assert window.flow_header_icon.property("iconKind") == "flow"
    assert window.flow_crab.width() == 138
    assert not window.flow_crab.isHidden()
    assert window.flow_crab._mode == "idle"
    assert window.flow_idle_icon.isHidden()
    assert window.flow_spinner.objectName() == "clawdAnimator"
    assert window.stage_buttons == []
    assert window.stage_labels == {}
    assert window.log_ready_badge.text() == "已就绪"
    assert window.log_ready_badge.parent() is window.log_view.ready_overlay
    window.close()


def test_desktop_gui_daily_date_uses_project_card_style(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert window.date_card.objectName() == "dashboardCard"
    assert window.date_button.objectName() == "datePickerButton"
    assert window.date_button.minimumHeight() >= window.date_button.fontMetrics().height() + 2
    assert window.selected_daily_date() in window.date_button.text()
    assert "QPushButton#datePickerButton" in window.styleSheet()
    assert "text-align: center" in window.styleSheet()
    window.close()


def test_desktop_gui_calendar_uses_codex_style_popup_and_double_click_accepts(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from datetime import date
    from PySide6.QtCore import QDate, QEvent, QCoreApplication, Qt
    from PySide6.QtWidgets import QApplication, QDialog
    from gui.main_window import ModernCalendarDialog

    app = QApplication.instance() or QApplication([])
    dialog = ModernCalendarDialog(date(2026, 7, 11))

    assert dialog.objectName() == "calendarPopup"
    assert dialog.calendar.objectName() == "modernCalendar"
    assert not dialog.calendar.isNavigationBarVisible()
    assert dialog.month_label.text() == "2026年 7月"
    assert "#ffffff" in dialog.styleSheet()
    assert dialog.surface.graphicsEffect() is None
    assert bool(dialog.windowFlags() & Qt.WindowType.NoDropShadowWindowHint)
    dialog.show()
    app.processEvents()
    assert dialog.testAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
    QCoreApplication.sendEvent(dialog, QEvent(QEvent.Type.Leave))
    assert dialog.result() == QDialog.DialogCode.Rejected
    assert not dialog.isVisible()

    dialog = ModernCalendarDialog(date(2026, 7, 11))
    dialog.calendar.activated.emit(QDate(2026, 7, 11))
    assert dialog.result() == QDialog.DialogCode.Accepted
    dialog.close()


def test_desktop_gui_date_button_toggles_single_calendar_popup(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    window.pick_daily_date()
    app.processEvents()
    assert window.calendar_dialog is not None
    assert window.calendar_dialog.isVisible()

    dialog = window.calendar_dialog
    button_global = window.date_button.mapToGlobal(window.date_button.rect().topLeft())
    anchor_offset = dialog.pos() - button_global
    window.move(window.pos().x() + 40, window.pos().y() + 30)
    app.processEvents()
    moved_button_global = window.date_button.mapToGlobal(window.date_button.rect().topLeft())
    assert dialog.pos() - moved_button_global == anchor_offset

    window.pick_daily_date()
    app.processEvents()
    assert window.calendar_dialog is None

    window._quitting = True
    window.close()


def test_desktop_gui_flow_uses_clawd_idle_and_dance_modes(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    assert window.flow_crab._mode == "idle"
    window.set_current_flow("hourly", "运行小时报", "昆明牛 15点", "运行中")
    assert window.flow_crab._mode == "dance"
    window.set_current_flow_idle()
    assert window.flow_crab._mode == "idle"
    window.close()


def test_desktop_gui_uses_vista_yahei_bold_with_regular_secondary_text(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtGui import QFont
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow, MAIN_FONT_PT, SUB_FONT_PT

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert MAIN_FONT_PT == 9
    assert SUB_FONT_PT == 8
    assert window.font().pointSize() == MAIN_FONT_PT
    assert window.progress_text.font().pointSize() == SUB_FONT_PT
    assert window.font().family() == "Microsoft YaHei UI"
    assert window.font().weight() == QFont.Weight.Normal
    assert window.title_label.font().family() == "Microsoft YaHei"
    assert window.title_label.font().weight() == QFont.Weight.Bold
    assert window.system_config_button.font().family() == "Microsoft YaHei UI"
    assert window.system_config_button.font().weight() == QFont.Weight.Normal
    assert window.data_source_control.prefix_label.font().weight() == QFont.Weight.Normal
    assert window.log_view.font().family() == "Microsoft YaHei UI"
    assert window.log_view.font().weight() == QFont.Weight.Normal
    assert window.log_view.font().pointSize() == MAIN_FONT_PT
    assert 'font-family: "Microsoft YaHei UI", "Microsoft YaHei", "Segoe UI", sans-serif;' in window.styleSheet()
    assert 'font-family: "Microsoft YaHei", "Microsoft YaHei UI", "Segoe UI", sans-serif;' in window.styleSheet()
    assert "QLabel#cardTitle" in window.styleSheet()
    assert "QLabel#dailyCardTitle" in window.styleSheet()
    assert "font-weight: 600" not in window.styleSheet()
    assert "font-weight: 700" in window.styleSheet()
    assert "font-weight: 400" in window.styleSheet()
    assert "QScrollBar:vertical" in window.styleSheet()
    window.close()


def test_desktop_gui_config_actions_live_in_title_menu(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert not hasattr(window, "excel_config_button")
    assert not hasattr(window, "credentials_config_button")
    assert window.system_config_button.text() == "系统"
    assert window.system_config_button.font().pointSize() == window.title_label.font().pointSize() - 1
    assert window.system_config_button.width() <= window.system_config_button.fontMetrics().horizontalAdvance("系统") + 12
    assert window.system_config_button.height() <= window.system_config_button.fontMetrics().height() + 10
    assert [action.text() for action in window.system_config_menu.actions() if not action.isSeparator()] == [
        "项目配置检查", "导入授权配置", "导出授权配置", "恢复备份", "Excel 路径配置", "Excel 自动打开", "桌面宠物", "退出程序"
    ]
    from gui.main_window import InlineMenuRow
    inline_labels = [row.text() for row in window.inline_config_menu.findChildren(InlineMenuRow)]
    assert "更新 Excel 路径" not in inline_labels
    assert "更新账号密码" not in inline_labels
    assert "导入授权配置" in inline_labels
    assert "导出授权配置" in inline_labels
    assert "Excel 路径配置" in inline_labels
    assert "Excel 自动打开" in inline_labels
    assert [action.text() for action in window.excel_auto_open_menu.actions()] == ["启动", "停止"]
    assert [action.text() for action in window.pet_menu.actions() if not action.isSeparator()] == [
        "Clawd 小螃蟹", "隐藏宠物"
    ]
    from gui.pet_settings import PET_CLAWD, PET_HIDDEN, load_pet_mode, load_pet_scale
    configured_pet_mode = load_pet_mode(Path(__file__).resolve().parents[1])
    assert window.clawd_pet_action.isChecked() == (configured_pet_mode == PET_CLAWD)
    assert window.hidden_pet_action.isChecked() == (configured_pet_mode == PET_HIDDEN)
    assert window.inline_config_menu.size_slider.minimum() == 50
    assert window.inline_config_menu.size_slider.maximum() == 120
    assert window.inline_config_menu.size_slider.value() == round(
        load_pet_scale(Path(__file__).resolve().parents[1]) * 100
    )
    assert window.project_check_action.text() == "项目配置检查"
    assert window.system_config_button.icon().isNull()
    assert all(action.icon().isNull() for action in window.system_config_menu.actions())
    assert all(action.icon().isNull() for action in window.pet_menu.actions())
    assert not hasattr(window, "environment_check_button")
    assert not hasattr(window, "guide_button")
    assert not hasattr(window, "refresh_button")
    assert not hasattr(window, "preflight_hourly_button")
    assert not hasattr(window, "preflight_daily_button")
    assert not hasattr(window, "command_line")
    assert window.selected_project_config_path().name.endswith(".json")
    assert window.credentials_config_path() == Path(__file__).resolve().parents[1] / "secrets" / "secrets.json"
    assert window.update_button.isHidden()
    window.on_update_checking()
    assert not window.update_button.isHidden()
    assert window.update_button.property("updateState") == "checking"
    window.on_update_failed("network unavailable")
    assert window.update_button.isHidden()
    window._manual_update_check_requested = True
    window.on_update_checking()
    window.on_update_failed("network unavailable")
    assert not window.update_button.isHidden()
    assert window.update_button.property("updateState") == "failed"
    assert window.update_button.text() == "重试"

    from gui.update_manager import ReleaseUpdate
    update = ReleaseUpdate(
        version="2026.7.22.108",
        download_url="https://example/update.zip",
        asset_name="Hourlyreport_automation_v2026.7.22.108.zip",
        sha256="a" * 64,
        size=123,
    )
    window.on_update_available(update)
    assert not window.update_button.isHidden()
    assert window.update_button.property("updateState") == "available"
    assert window.update_button.text() == "更新"
    download_calls = []
    monkeypatch.setattr(window.update_manager, "start_download", lambda value: download_calls.append(value) or True)
    window.handle_update_button_clicked()
    assert download_calls == [update]
    assert window.update_button.property("updateState") == "downloading"
    assert window.update_button.text() == ""
    assert window.update_button.width() <= window.system_config_button.width()
    window.on_update_download_progress(41)
    assert not window.update_button.isHidden()
    assert window.update_button.property("updateState") == "downloading"
    assert window.update_button.text() == "41%"
    assert "41%" in window.update_button.toolTip()
    window.on_update_failed("network unavailable")
    assert not window.update_button.isHidden()
    assert window.update_button.property("updateState") == "available"
    assert window.update_button.text() == "更新"
    window.handle_update_button_clicked()
    assert download_calls == [update, update]
    window.on_update_download_progress(100)
    assert window.update_button.text() == "100%"
    window.on_update_ready("2026.7.22.108", Path("update.zip"))
    assert window.update_button.property("updateState") == "ready"
    assert window.update_button.text() == "重启"
    assert window.update_button.icon().isNull()
    assert "2026.7.22.108" in window.update_button.toolTip()
    window.on_update_up_to_date()
    assert window.update_button.isHidden()
    window.close()


def test_desktop_gui_excel_path_configuration_uses_selected_competition_root(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from types import SimpleNamespace
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    selected_root = tmp_path / "new-drive" / "【竞价】"
    selected_root.mkdir(parents=True)
    calls = []
    refreshed = []
    notices = []
    monkeypatch.setattr(QFileDialog, "getExistingDirectory", lambda *args, **kwargs: str(selected_root))
    monkeypatch.setattr(
        main_window,
        "configure_excel_paths",
        lambda root, selected: calls.append((Path(root), Path(selected)))
        or SimpleNamespace(updated=9, errors=(), backup_dir=tmp_path / "backups" / "excel-paths"),
    )
    monkeypatch.setattr(window, "refresh_projects", lambda: refreshed.append(True))
    monkeypatch.setattr(QMessageBox, "information", lambda *args: notices.append(args) or QMessageBox.StandardButton.Ok)

    window.configure_excel_paths_from_folder()

    assert calls == [(tmp_path, selected_root)]
    assert refreshed == [True]
    assert notices and "9 个项目" in notices[0][2]

    window._quitting = True
    window.close()


def test_private_ui_font_prefers_bundled_vista_face(tmp_path, monkeypatch):
    from gui import font_manager

    font_path = tmp_path / "assets" / "fonts" / "microsoft_yahei_vista_bold.ttf"
    font_path.parent.mkdir(parents=True)
    font_path.write_bytes(b"licensed-font-placeholder")
    loaded_data = []
    monkeypatch.setattr(
        font_manager.QFontDatabase,
        "addApplicationFontFromData",
        lambda data: loaded_data.append(bytes(data)) or 9,
    )
    monkeypatch.setattr(
        font_manager.QFontDatabase,
        "applicationFontFamilies",
        lambda _font_id: ["Microsoft YaHei"],
    )

    assert font_manager.load_private_ui_font(tmp_path) == font_path
    assert loaded_data == [b"licensed-font-placeholder"]


def _write_minimal_gui_project(root: Path) -> None:
    projects = root / "configs" / "projects"
    projects.mkdir(parents=True)
    (root / "configs" / "app_config.json").write_text(json.dumps({
        "default_project_id": "demo",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
    }, ensure_ascii=False), encoding="utf-8")
    (projects / "demo.json").write_text(json.dumps({
        "project_id": "demo",
        "project_name": "演示项目",
        "excel": {"path": str(root / "target.xlsx")},
    }, ensure_ascii=False), encoding="utf-8")


def test_desktop_gui_exports_authorization_config_with_standard_extension(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    secrets_path = tmp_path / "secrets" / "secrets.json"
    secrets_path.parent.mkdir(parents=True)
    secrets_path.write_text('{"baidu": {}}', encoding="utf-8")
    selected = tmp_path / "百度授权配置"
    calls = []
    messages = []
    monkeypatch.setattr(QFileDialog, "getSaveFileName", lambda *args, **kwargs: (str(selected), "授权配置"))
    monkeypatch.setattr(
        main_window,
        "export_secrets_package",
        lambda source, output: calls.append((Path(source), Path(output))) or {"package_path": str(output)},
        raising=False,
    )
    monkeypatch.setattr(QMessageBox, "information", lambda *args, **kwargs: messages.append(args[2]))

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.export_authorization_config()

    assert calls == [(secrets_path, Path(str(selected) + ".baidu-secrets"))]
    assert messages and "明文" in messages[0]
    window.close()


def test_desktop_gui_imports_authorization_config_then_runs_project_check(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    package = tmp_path / "team.baidu-secrets"
    package.write_text("{}", encoding="utf-8")
    import_calls = []
    preflight_calls = []
    information_calls = []
    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *args, **kwargs: (str(package), "授权配置"))
    monkeypatch.setattr(
        main_window,
        "import_secrets_package",
        lambda source, target, backups: import_calls.append((Path(source), Path(target), Path(backups))) or {
            "package_path": str(source),
        },
        raising=False,
    )
    monkeypatch.setattr(QMessageBox, "information", lambda *args, **kwargs: information_calls.append(args))

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.project_combo.set_multi_mode(True)
    window.run_environment_preflight = lambda allow_multi=False: preflight_calls.append(allow_multi)
    window.import_authorization_config()

    assert import_calls == [(package, tmp_path / "secrets" / "secrets.json", tmp_path / "backups")]
    assert preflight_calls == [True]
    assert information_calls == []
    window.close()


def test_desktop_gui_real_authorization_import_replaces_secrets_then_checks(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog
    import gui.main_window as main_window
    from modules.secrets_package import export_secrets_package

    _write_minimal_gui_project(tmp_path)
    source = tmp_path / "admin-secrets.json"
    package = tmp_path / "team.baidu-secrets"
    target = tmp_path / "secrets" / "secrets.json"
    source_payload = {
        "baidu": {"shared": {"username": "fake-user", "password": "fake-password"}},
        "baidu_api": {"shared": {"access_token": "fake.access.token"}},
    }
    source.write_text(json.dumps(source_payload, ensure_ascii=False), encoding="utf-8")
    target.parent.mkdir(parents=True)
    target.write_text('{"baidu":{"old":{}}}', encoding="utf-8")
    export_secrets_package(source, package)
    preflight_calls = []
    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *args, **kwargs: (str(package), "授权配置"))

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.run_environment_preflight = lambda allow_multi=False: preflight_calls.append(allow_multi)
    window.import_authorization_config()

    assert json.loads(target.read_text(encoding="utf-8")) == source_payload
    assert len(list((tmp_path / "backups").glob("secrets_before_package_import_*.json"))) == 1
    assert preflight_calls == [True]
    window.close()


def test_desktop_gui_import_failure_explains_reason_and_allows_retry(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    import gui.main_window as main_window
    from modules.secrets_package import SecretsPackageError

    _write_minimal_gui_project(tmp_path)
    bad_package = tmp_path / "bad.baidu-secrets"
    good_package = tmp_path / "good.baidu-secrets"
    selections = iter(((str(bad_package), "授权配置"), (str(good_package), "授权配置")))
    import_calls = []
    preflight_calls = []
    warnings = []
    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *args, **kwargs: next(selections))

    def fake_import(source, target, backups):
        import_calls.append(Path(source))
        if Path(source) == bad_package:
            raise SecretsPackageError("授权配置包校验失败")
        return {"package_path": str(source)}

    monkeypatch.setattr(main_window, "import_secrets_package", fake_import, raising=False)
    monkeypatch.setattr(
        QMessageBox,
        "warning",
        lambda *args, **kwargs: warnings.append(args[2]) or QMessageBox.StandardButton.Retry,
    )

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.run_environment_preflight = lambda allow_multi=False: preflight_calls.append(allow_multi)
    window.import_authorization_config()

    assert import_calls == [bad_package, good_package]
    assert warnings == ["授权配置包校验失败"]
    assert preflight_calls == [True]
    window.close()


def test_desktop_gui_blocks_authorization_import_while_task_is_running(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    import gui.main_window as main_window

    _write_minimal_gui_project(tmp_path)
    dialog_calls = []
    warnings = []
    monkeypatch.setattr(
        QFileDialog,
        "getOpenFileName",
        lambda *args, **kwargs: dialog_calls.append(True) or ("", ""),
    )
    monkeypatch.setattr(QMessageBox, "warning", lambda *args, **kwargs: warnings.append(args[2]))

    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(tmp_path)
    window.runner.is_running = lambda: True
    window.import_authorization_config()

    assert dialog_calls == []
    assert warnings and "任务" in warnings[0]
    window.close()


def test_desktop_gui_defaults_to_kunming_project(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert window.selected_project_id() == "kunming_niu"
    assert window.project_combo.currentText() == "昆明牛"
    assert "(" not in window.project_combo.currentText()
    assert len(window.projects) == 9
    assert all(project.project_id != "hefei_bai" for project in window.projects)
    window.close()


def test_desktop_gui_project_selector_supports_multi_project_ui_without_running(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QApplication
    import gui.main_window as main_window

    saved_selections = []
    monkeypatch.setattr(
        main_window,
        "save_multi_project_selection",
        lambda _root, project_ids: saved_selections.append(list(project_ids)),
    )
    monkeypatch.setattr(
        main_window,
        "load_multi_project_selection",
        lambda _root, available_ids, fallback_id="": [fallback_id or list(available_ids)[0]],
    )
    MainWindow = main_window.MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    assert window.single_project_button.isChecked()
    assert not window.project_combo.is_multi_mode()
    assert not hasattr(window.project_combo, "count_label")
    assert window.project_combo.summary_label.text() == "昆明牛"
    assert window.project_combo.border_overlay.geometry() == window.project_combo.rect()
    assert window.project_combo.border_overlay.isVisible()
    assert window.project_mode_segment.animation.duration() == 150

    window.multi_project_button.click()
    app.processEvents()
    assert window.project_combo.is_multi_mode()
    assert not window.multi_preview_hint.isVisible()
    assert window.project_mode_segment.indicator.property("preview") is False
    assert window.project_combo.selected_project_ids() == ["kunming_niu"]
    assert "已按顺序选择 1 个项目" in window.progress_text.text()
    assert window.project_combo.summary_label.isHidden()

    second_project = next(project.project_id for project in window.projects if project.project_id != "kunming_niu")
    popup = window.project_combo.popup
    popup.set_data(
        [(project.project_name, project.project_id) for project in window.projects],
        window.project_combo.selected_project_ids(),
        True,
        "kunming_niu",
    )
    assert popup.rows_layout.itemAtPosition(0, 0) is not None
    assert popup.rows_layout.itemAtPosition(0, 2) is not None
    assert popup.rows_layout.itemAtPosition(2, 2) is not None
    assert popup.rows_layout.itemAtPosition(3, 0) is None
    assert popup.scroll.verticalScrollBarPolicy() == Qt.ScrollBarPolicy.ScrollBarAlwaysOff
    assert popup.height() == 196
    assert all(not button.text()[:1].isdigit() for button in popup._project_buttons.values())

    popup.search.setText("sy")
    assert popup._project_buttons
    assert all("沈阳" in button.text() for button in popup._project_buttons.values())
    popup.search.setText("b")
    assert popup._project_buttons
    assert all(button.text().endswith("白") for button in popup._project_buttons.values())
    popup.search.setText("n")
    assert popup._project_buttons
    assert all(button.text().endswith("牛") for button in popup._project_buttons.values())
    popup.search.clear()

    popup._toggle_project(second_project, True)
    assert popup.confirm_button.isEnabled()
    popup._confirm()
    assert window.project_combo.selected_project_ids() == ["kunming_niu", second_project]
    assert saved_selections[-1] == ["kunming_niu", second_project]
    assert window.project_combo.border_overlay.geometry() == window.project_combo.rect()
    assert window.project_combo.border_overlay.isVisible()
    chip_names = [
        window.project_combo.chips_layout.itemAt(index).widget().text()
        for index in range(window.project_combo.chips_layout.count())
        if window.project_combo.chips_layout.itemAt(index).widget()
    ]
    assert chip_names == ["昆明牛", next(project.project_name for project in window.projects if project.project_id == second_project)]
    chip_widgets = [
        window.project_combo.chips_layout.itemAt(index).widget()
        for index in range(window.project_combo.chips_layout.count())
        if window.project_combo.chips_layout.itemAt(index).widget()
    ]
    assert all(chip.height() == 22 and chip.height() > chip.fontMetrics().height() for chip in chip_widgets)
    assert "已按顺序选择 2 个项目" in window.progress_text.text()

    popup.set_data(
        [(project.project_name, project.project_id) for project in window.projects],
        window.project_combo.selected_project_ids(),
        True,
        "kunming_niu",
    )
    popup._clear_selection()
    assert popup._selected_ids == ["kunming_niu"]
    assert popup.confirm_button.isEnabled()
    extra_projects = [project.project_id for project in window.projects if project.project_id != "kunming_niu"][:3]
    for project_id in extra_projects:
        popup._toggle_project(project_id, True)
    assert len(popup._selected_ids) == 3
    assert popup.summary.text() == "最多选择 3 个项目"

    starts = []
    monkeypatch.setattr(window, "start_command", lambda *args: starts.append(args))
    window.run_hourly()
    assert len(starts) == 1
    assert "run-multi" in starts[0][1]
    assert starts[0][1][starts[0][1].index("--projects") + 1] == f"kunming_niu,{second_project}"

    selector_bottom = window.project_combo.mapTo(window.task_control_card, window.project_combo.rect().bottomLeft()).y()
    assert selector_bottom < window.task_control_card.height()
    popup.hide()
    window._quitting = True
    window.close()


def test_desktop_gui_multi_project_completion_opens_successful_excels_in_order(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    first.write_bytes(b"first")
    second.write_bytes(b"second")
    reports = tmp_path / "reports"
    reports.mkdir(exist_ok=True)
    (reports / "multi_project_run_report.json").write_text(
        json.dumps(
            {
                "summary": {"success": 3, "failed": 1, "stopped": 0},
                "projects": [
                    {"project_id": "a", "status": "success", "excel_path": str(first)},
                    {"project_id": "b", "status": "failed", "excel_path": str(tmp_path / "failed.xlsx")},
                    {"project_id": "c", "status": "success", "excel_path": str(first)},
                    {"project_id": "d", "status": "success", "excel_path": str(second)},
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    opened = []
    window.open_path = lambda path: opened.append(path)
    window.desktop_pet.announce = lambda *_args, **_kwargs: None
    window.current_task_type = "hourly"
    window._multi_task_active = True

    window.finish_multi_project_task()

    assert opened == [first, second]
    assert window.status_title.text() == "多项目任务部分完成"
    assert window.current_status_badge.text() == "部分完成"
    window._quitting = True
    window.close()


def test_desktop_gui_multi_project_stop_remains_available_during_excel_stage(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    window._multi_task_active = True
    window._task_active = True
    window.current_task_type = "hourly"
    window.runner.is_running = lambda: True
    window._task_stop_locked = False
    window._task_stop_requested = False

    window.mark_stage("excel")
    window.set_stop_controls()

    assert window.hourly_stop_button.isEnabled()
    assert window._task_stop_locked is False
    window._quitting = True
    window.close()


def test_desktop_gui_multi_project_uses_queue_stop_gate_not_single_pipeline_gate(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow
    from modules.multi_project_stop import MULTI_QUEUE_STOP_GATE_ENV
    from modules.task_stop_gate import STOP_GATE_ENV, read_task_stop_decision

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    starts = []
    window.runner.start = lambda command, root, extra_env=None: starts.append((command, root, extra_env))
    window.runner.is_running = lambda: True
    window.project_mode_segment.set_multi_mode(True, animate=False)

    window.run_hourly()
    environment = starts[-1][2]
    gate = Path(environment[MULTI_QUEUE_STOP_GATE_ENV])

    assert STOP_GATE_ENV not in environment
    window.on_task_started()
    window.mark_stage("excel")
    assert window.hourly_stop_button.isEnabled()
    window.stop_current_task()
    assert read_task_stop_decision(gate) == "cancel"
    assert "当前项目完成后停止队列" in window.progress_text.text()
    window._quitting = True
    window.close()


def test_desktop_gui_restore_backup_overwrites_project_excel_after_safety_backup(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox
    from gui.main_window import MainWindow

    target_excel = tmp_path / "target.xlsx"
    selected_backup = tmp_path / "backups" / "target_backup.xlsx"
    selected_backup.parent.mkdir()
    target_excel.write_text("current excel", encoding="utf-8")
    selected_backup.write_text("restored excel", encoding="utf-8")
    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text(json.dumps({
        "default_project_id": "demo",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
    }, ensure_ascii=False), encoding="utf-8")
    (tmp_path / "configs" / "projects" / "demo.json").write_text(json.dumps({
        "project_id": "demo",
        "project_name": "演示项目",
        "excel": {"path": str(target_excel)},
    }, ensure_ascii=False), encoding="utf-8")

    monkeypatch.setattr(QFileDialog, "getOpenFileName", lambda *args, **kwargs: (str(selected_backup), "Excel"))
    monkeypatch.setattr(QMessageBox, "question", lambda *args, **kwargs: QMessageBox.StandardButton.Yes)
    monkeypatch.setattr(QMessageBox, "information", lambda *args, **kwargs: None)

    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    window.restore_backup()

    safety_backups = list((tmp_path / "backups").glob("target_before_manual_restore_*.xlsx"))
    assert target_excel.read_text(encoding="utf-8") == "restored excel"
    assert len(safety_backups) == 1
    assert safety_backups[0].read_text(encoding="utf-8") == "current excel"
    while window.has_pending_log_display():
        window.drain_log_display()
    assert "恢复备份完成" in window.log_view.toPlainText()
    window.close()


def test_desktop_gui_period_selection_marks_checked_green(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert 'QPushButton#periodButton:checked' in window.styleSheet()
    assert 'background: #ffffff' in window.styleSheet()
    assert all(not button.autoDefault() for button in window.period_buttons)
    assert window.period_buttons[0].text() == "11:00"
    assert window.period_buttons[0].icon().isNull() is False
    assert window.selected_period() == "11点"
    window.show()
    app.processEvents()
    assert len({button.width() for button in [window.hourly_action_control, *window.period_buttons]}) == 1
    image = window.period_buttons[0].grab().toImage()
    check_color = image.pixelColor(window.period_buttons[0].width() - 17, 13)
    assert check_color.green() > check_color.red()
    window.close()


def test_run_stop_split_control_uses_seventy_thirty_diagonal_geometry(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import RunStopSplitControl

    app = QApplication.instance() or QApplication([])
    control = RunStopSplitControl("运行小时报")
    control.setFixedSize(160, 48)
    control.show()
    app.processEvents()

    split_x = round(control.width() * 0.70)
    assert abs(control.run_button.width() / control.width() - 0.70) <= 0.03
    assert control.stop_button.x() == split_x - 2
    assert control.stop_button.y() == 2
    assert control.stop_button.text() == "停止"
    assert control.stop_button.icon().isNull()
    assert not control.stop_button.isEnabled()
    assert control.diagonal_gap_width == 2
    control.close()


def test_desktop_gui_safe_stop_is_available_only_before_excel(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    from modules.task_stop_gate import (
        STOP_GATE_ENV,
        TASK_CANCELLED_EXIT_CODE,
        claim_excel_write,
        read_task_stop_decision,
    )

    starts = []
    window.runner.start = lambda command, root, extra_env=None: starts.append((command, root, extra_env))
    window.runner.is_running = lambda: True

    assert not window.hourly_stop_button.isEnabled()
    assert not window.daily_stop_button.isEnabled()

    window.run_hourly()
    hourly_gate = Path(starts[-1][2][STOP_GATE_ENV])
    window.on_task_started()
    assert window.hourly_stop_button.isEnabled()
    assert not window.daily_stop_button.isEnabled()

    window.stop_current_task()
    assert read_task_stop_decision(hourly_gate) == "cancel"
    assert not window.hourly_stop_button.isEnabled()
    assert window._task_stop_requested is True

    opened = []
    window.open_current_project_excel = lambda: opened.append(True)
    window.on_task_finished(TASK_CANCELLED_EXIT_CODE)
    assert window.current_status_badge.text() == "已停止"
    assert window.progress_text.text() == "任务已停止，未继续写入 Excel。"
    assert opened == []
    assert window.hourly_button.isEnabled()
    assert not window.hourly_stop_button.isEnabled()

    window.run_daily()
    daily_gate = Path(starts[-1][2][STOP_GATE_ENV])
    window.on_task_started()
    assert window.daily_stop_button.isEnabled()
    assert claim_excel_write(daily_gate) is True
    window.mark_stage("excel")
    assert not window.daily_stop_button.isEnabled()
    window.stop_current_task()
    assert read_task_stop_decision(daily_gate) == "excel"
    window.on_task_finished(0)
    assert window.current_status_badge.text() == "已完成"
    assert opened == [True]

    window._quitting = True
    window.close()


def test_desktop_gui_non_report_task_never_enables_stop(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    window.runner.start = lambda _command, _root: None

    window.run_environment_preflight()
    window.on_task_started()

    assert not window.hourly_stop_button.isEnabled()
    assert not window.daily_stop_button.isEnabled()
    window._quitting = True
    window.close()


def test_desktop_gui_defers_application_exit_while_task_is_running(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    running = {"value": True}
    window.runner.is_running = lambda: running["value"]

    window.exit_application()

    assert window._quitting is False
    assert window._quit_after_task is True
    running["value"] = False
    window.exit_application()
    assert window._quitting is True


def test_qt_task_runner_killed_process_is_not_reported_as_start_failure():
    from types import SimpleNamespace

    from PySide6.QtCore import QProcess
    from gui.task_runner import QtTaskRunner

    messages = []
    runner = SimpleNamespace(failed_to_start=SimpleNamespace(emit=messages.append))

    QtTaskRunner._handle_error(runner, QProcess.ProcessError.Crashed)
    assert messages == []

    QtTaskRunner._handle_error(runner, QProcess.ProcessError.FailedToStart)
    assert len(messages) == 1


def test_desktop_gui_current_flow_updates_for_hourly_and_daily(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.runner.start = lambda command, root, extra_env=None: None

    window.period_buttons[0].setChecked(True)
    window.update_period_button_texts()
    window.run_hourly()
    assert window.current_task_title.text() == "运行小时报"
    assert "11点" in window.current_task_subtitle.text()
    assert window.current_status_badge.text() == "运行中"
    assert window.current_start_time_label.text().startswith("开始时间：")
    assert window.flow_idle_icon.isHidden()
    assert not window.flow_spinner.isHidden()
    assert window.flow_crab._mode == "dance"

    window.run_daily()
    assert window.current_task_title.text() == "运行日报"
    assert "月" in window.current_task_subtitle.text()
    assert "日" in window.current_task_subtitle.text()
    window.close()


def test_desktop_gui_opens_project_excel_after_success(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    opened = []
    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.open_path = lambda path: opened.append(path)
    window.runner.start = lambda command, root, extra_env=None: None

    window.run_hourly()
    window.on_task_finished(0)

    assert opened
    assert opened[-1].suffix.lower() in {".xlsx", ".xlsm", ".xls"}
    window.close()


def test_desktop_gui_creates_local_secrets_template_when_missing(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    (tmp_path / "configs" / "projects").mkdir(parents=True)
    (tmp_path / "configs" / "app_config.json").write_text(json.dumps({
        "default_project_id": "demo",
        "projects_dir": "configs/projects",
        "secrets_file": "secrets/secrets.json",
    }, ensure_ascii=False), encoding="utf-8")
    (tmp_path / "configs" / "projects" / "demo.json").write_text(json.dumps({
        "project_id": "demo",
        "project_name": "演示项目",
    }, ensure_ascii=False), encoding="utf-8")
    (tmp_path / "secrets").mkdir()
    (tmp_path / "secrets" / "secrets.example.json").write_text(json.dumps({"baidu": {}}, ensure_ascii=False), encoding="utf-8")

    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    path = window.ensure_credentials_file()

    assert path == tmp_path / "secrets" / "secrets.json"
    assert json.loads(path.read_text(encoding="utf-8")) == {"baidu": {}}
    window.close()


def test_desktop_gui_log_formatter_highlights_key_content():
    from gui.log_formatter import format_log_html

    html = format_log_html("项目 青岛白 [通知] 通过，Excel D:\\数据\\青岛.xlsx，报告 reports/final_run_report.json")

    assert "log-pass" in html
    assert "log-path" in html
    assert "log-project" in html
    assert "青岛白" in html


def test_gui_log_history_is_timestamped_appended_and_redacted(tmp_path):
    from datetime import datetime

    from gui.log_history import append_history_line

    first = append_history_line(
        tmp_path,
        "登录完成 password=plain-text accessToken=token-value",
        now=datetime(2026, 7, 19, 15, 30, 5),
    )
    second = append_history_line(
        tmp_path,
        "项目写入完成",
        now=datetime(2026, 7, 19, 15, 30, 6),
    )

    assert first == second == tmp_path / "logs" / "gui_history.log"
    content = first.read_text(encoding="utf-8")
    assert "[2026-07-19 15:30:05] 登录完成 password=*** accessToken=***" in content
    assert "[2026-07-19 15:30:06] 项目写入完成" in content
    assert "plain-text" not in content
    assert "token-value" not in content


def test_gui_log_typewriter_scales_with_backlog_and_does_not_render_all_at_once(tmp_path, monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication

    from gui.log_history import typewriter_batch_size
    from gui.main_window import MainWindow

    _write_minimal_gui_project(tmp_path)
    app = QApplication.instance() or QApplication([])
    window = MainWindow(tmp_path)
    window.reset_log_display()
    text = "正在快速读取并核对百度推广账户数据，界面将跟随任务速度逐字显示。"

    window.append_log(text)

    assert window.log_view.toPlainText() == ""
    assert text in (tmp_path / "logs" / "gui_history.log").read_text(encoding="utf-8")
    window.drain_log_display()
    partial = window.log_view.toPlainText()
    assert 0 < len(partial) < len(text)
    assert typewriter_batch_size(5000) > typewriter_batch_size(50)
    while window.has_pending_log_display():
        window.drain_log_display()
    assert window.log_view.toPlainText() == text
    window.close()


def test_desktop_gui_environment_check_reports_missing_python(tmp_path):
    from gui.environment_check import run_environment_check

    report = run_environment_check(tmp_path)

    python_check = next(item for item in report["checks"] if item["name"] == "Python environment")
    assert report["passed"] is False
    assert python_check["passed"] is False
    assert python_check["severity"] == "error"


def test_first_startup_creates_all_project_kst_directories_and_marker(tmp_path):
    from gui.environment_check import KST_PROJECT_DIR_NAMES, initialize_kst_directories_once

    data_root = tmp_path / "商务通数据"
    marker = tmp_path / "local-app-data" / "kst_directories_v1.json"

    result = initialize_kst_directories_once(data_root=data_root, marker_path=marker)

    assert result["passed"] is True
    assert result["status"] == "initialized"
    assert marker.is_file()
    assert sorted(path.name for path in data_root.iterdir() if path.is_dir()) == sorted(KST_PROJECT_DIR_NAMES)
    assert "合肥白" not in KST_PROJECT_DIR_NAMES


def test_first_startup_marker_skips_future_directory_checks(tmp_path):
    from gui.environment_check import initialize_kst_directories_once

    data_root = tmp_path / "商务通数据"
    marker = tmp_path / "local-app-data" / "kst_directories_v1.json"
    first = initialize_kst_directories_once(data_root=data_root, marker_path=marker)
    removed = data_root / "昆明牛"
    removed.rmdir()

    second = initialize_kst_directories_once(data_root=data_root, marker_path=marker)

    assert first["status"] == "initialized"
    assert second["passed"] is True
    assert second["status"] == "skipped"
    assert not removed.exists()


def test_failed_first_startup_does_not_write_completion_marker(tmp_path):
    from gui.environment_check import initialize_kst_directories_once

    data_root = tmp_path / "商务通数据"
    data_root.write_text("not a directory", encoding="utf-8")
    marker = tmp_path / "local-app-data" / "kst_directories_v1.json"

    result = initialize_kst_directories_once(data_root=data_root, marker_path=marker)

    assert result["passed"] is False
    assert result["status"] == "failed"
    assert not marker.exists()


def test_first_startup_rejects_project_name_occupied_by_file(tmp_path):
    from gui.environment_check import initialize_kst_directories_once

    data_root = tmp_path / "商务通数据"
    data_root.mkdir()
    (data_root / "南京牛").write_text("not a directory", encoding="utf-8")
    marker = tmp_path / "local-app-data" / "kst_directories_v1.json"

    result = initialize_kst_directories_once(data_root=data_root, marker_path=marker)

    assert result["passed"] is False
    assert result["status"] == "failed"
    assert "南京牛" in result["detail"]
    assert not marker.exists()


def test_desktop_window_entry_runs_first_startup_directory_initialization(tmp_path, monkeypatch):
    import gui.main_window as main_window

    calls = []
    report = {"passed": True, "status": "initialized", "detail": "ready"}

    class FakeWindow:
        def __init__(self, root):
            self.root = Path(root)
            self.startup_kst_initialization = None
            self.shown = False
            self.update_check_started = False

        def show(self):
            self.shown = True

        def start_update_check(self):
            self.update_check_started = True

    monkeypatch.setattr(main_window, "initialize_kst_directories_once", lambda: calls.append(True) or report)
    monkeypatch.setattr(main_window, "MainWindow", FakeWindow)

    window = main_window.create_window(tmp_path)

    assert calls == [True]
    assert window.startup_kst_initialization == report
    assert window.shown is True
    assert window.update_check_started is True


def test_online_update_selects_newer_github_release_asset():
    from gui.update_manager import (
        CURRENT_VERSION,
        GITHUB_LATEST_RELEASE_URL,
        parse_release_version,
        parse_version,
        select_release_update,
    )

    assert CURRENT_VERSION == "2026.7.22.109"
    assert GITHUB_LATEST_RELEASE_URL == (
        "https://api.github.com/repos/179068898-dotcom/Hourlyreport-Automation/releases/latest"
    )
    assert parse_version("v2026.7.19.104") == (2026, 7, 19, 104)
    assert parse_release_version("v2026.7.19.105") == "2026.7.19.105"
    assert parse_release_version("Hourlyreport_v2026.7.19.105") == "2026.7.19.105"
    payload = {
        "tag_name": "v2026.7.22.110",
        "draft": False,
        "prerelease": False,
        "assets": [
            {"name": "notes.txt", "browser_download_url": "https://example/notes.txt"},
            {
                "name": "Hourlyreport_automation_v2026.7.22.110.zip",
                "browser_download_url": "https://example/update.zip",
                "digest": "sha256:" + "a" * 64,
                "size": 123,
            },
        ],
    }

    update = select_release_update(payload, CURRENT_VERSION)

    assert update is not None
    assert update.version == "2026.7.22.110"
    assert update.download_url == "https://example/update.zip"
    assert update.sha256 == "a" * 64
    assert select_release_update(payload, "2026.7.22.110") is None

    for invalid in (
        {**payload, "draft": True},
        {**payload, "prerelease": True},
        {**payload, "tag_name": "release-2026.7.19.105"},
        {**payload, "assets": [{**payload["assets"][1], "digest": ""}]},
        {**payload, "assets": [{**payload["assets"][1], "digest": "sha256:not-a-hash"}]},
        {**payload, "assets": [{**payload["assets"][1], "size": 0}]},
    ):
        assert select_release_update(invalid, CURRENT_VERSION) is None


def test_online_update_104_accepts_published_105_release_shape():
    from gui.update_manager import select_release_update

    payload = {
        "tag_name": "v2026.7.19.105",
        "draft": False,
        "prerelease": False,
        "assets": [
            {
                "name": "Hourlyreport_automation_v2026.7.19.105.zip",
                "browser_download_url": (
                    "https://github.com/179068898-dotcom/Hourlyreport-Automation/releases/"
                    "download/v2026.7.19.105/Hourlyreport_automation_v2026.7.19.105.zip"
                ),
                "digest": (
                    "sha256:861dc0a764690eb7d2f11e423bc1ed3f1894e13beec94e3a813436aedfb3a5a6"
                ),
                "size": 57497183,
            }
        ],
    }

    update = select_release_update(payload, "2026.7.19.104")

    assert update is not None
    assert update.version == "2026.7.19.105"
    assert update.asset_name == "Hourlyreport_automation_v2026.7.19.105.zip"


def test_online_update_archive_rejects_path_traversal(tmp_path):
    import pytest
    import zipfile

    from gui.update_manager import validate_update_archive

    valid = tmp_path / "valid.zip"
    with zipfile.ZipFile(valid, "w") as archive:
        archive.writestr("gui/main_window.py", "ok")
        archive.writestr("gui/version.py", "CURRENT_VERSION = 'test'")
        archive.writestr("main.py", "ok")
        archive.writestr("hourlyreport_automation.exe", "exe")
    assert validate_update_archive(valid) == [
        "gui/main_window.py", "gui/version.py", "main.py", "hourlyreport_automation.exe"
    ]

    malicious = tmp_path / "malicious.zip"
    with zipfile.ZipFile(malicious, "w") as archive:
        archive.writestr("../secrets/secrets.json", "bad")
    with pytest.raises(ValueError, match="不安全"):
        validate_update_archive(malicious)


def test_online_update_archive_rejects_windows_alias_and_device_paths(tmp_path):
    import pytest
    import zipfile

    from gui.update_manager import validate_update_archive

    dangerous_names = (
        "configs./app_config.json",
        "configs /app_config.json",
        "configs:stream/app_config.json",
        "C:/configs/app_config.json",
        "CON/payload.txt",
    )
    for index, name in enumerate(dangerous_names):
        archive_path = tmp_path / f"dangerous-{index}.zip"
        with zipfile.ZipFile(archive_path, "w") as archive:
            archive.writestr(name, "bad")
        with pytest.raises(ValueError, match="不安全|受保护"):
            validate_update_archive(archive_path)


def test_online_update_check_emits_available_without_downloading(monkeypatch):
    import io
    import json

    import gui.update_manager as update_manager

    payload = {
        "tag_name": "v2026.7.22.110",
        "draft": False,
        "prerelease": False,
        "assets": [
            {
                "name": "Hourlyreport_automation_v2026.7.22.110.zip",
                "browser_download_url": "https://example/update.zip",
                "digest": "sha256:" + "a" * 64,
                "size": 123,
            }
        ],
    }

    class Response(io.BytesIO):
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            self.close()

    monkeypatch.setattr(
        update_manager.urllib.request,
        "urlopen",
        lambda *_args, **_kwargs: Response(json.dumps(payload).encode("utf-8")),
    )
    manager = update_manager.GitHubUpdateManager()
    available = []
    ready = []
    manager.available.connect(available.append)
    manager.ready.connect(lambda *args: ready.append(args))

    manager._check_for_update()

    assert [item.version for item in available] == ["2026.7.22.110"]
    assert ready == []


def test_online_update_uses_bundled_ca_context_for_github_requests(monkeypatch):
    import io
    import json
    from types import SimpleNamespace

    import gui.update_manager as update_manager

    payload = {
        "tag_name": "v2026.7.22.109",
        "draft": False,
        "prerelease": False,
        "assets": [
            {
                "name": "Hourlyreport_automation_v2026.7.22.109.zip",
                "browser_download_url": "https://example/update.zip",
                "digest": "sha256:" + "a" * 64,
                "size": 123,
            }
        ],
    }
    contexts = []
    calls = []
    ca_bundle = r"C:\app\certifi\cacert.pem"
    context = object()

    class Response(io.BytesIO):
        def __enter__(self):
            return self

        def __exit__(self, *_args):
            self.close()

    monkeypatch.setattr(update_manager, "certifi", SimpleNamespace(where=lambda: ca_bundle))
    monkeypatch.setattr(
        update_manager.ssl,
        "create_default_context",
        lambda *, cafile=None: contexts.append(cafile) or context,
    )
    monkeypatch.setattr(
        update_manager.urllib.request,
        "urlopen",
        lambda request, **kwargs: calls.append(kwargs) or Response(json.dumps(payload).encode("utf-8")),
    )

    update_manager.GitHubUpdateManager()._check_for_update()

    assert contexts == [ca_bundle]
    assert calls and calls[0]["context"] is context


def test_online_update_download_rejects_size_mismatch(tmp_path, monkeypatch):
    import io
    import pytest

    import gui.update_manager as update_manager

    class Response(io.BytesIO):
        headers = {"Content-Length": "4"}

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            self.close()

    monkeypatch.setattr(update_manager, "_update_storage_dir", lambda: tmp_path)
    monkeypatch.setattr(update_manager.urllib.request, "urlopen", lambda *_args, **_kwargs: Response(b"data"))
    update = update_manager.ReleaseUpdate(
        version="2026.7.19.105",
        download_url="https://example/update.zip",
        asset_name="Hourlyreport_automation_v2026.7.19.105.zip",
        sha256="a" * 64,
        size=5,
    )

    with pytest.raises(ValueError, match="大小"):
        update_manager.GitHubUpdateManager()._download(update)
    assert not (tmp_path / update.asset_name).exists()


def test_update_install_dialog_matches_compact_codex_style(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication

    from gui.update_dialog import UpdateInstallDialog

    app = QApplication.instance() or QApplication([])
    dialog = UpdateInstallDialog("2026.7.19.105")
    dialog.set_progress(35, "正在备份当前程序…")

    assert dialog.windowTitle() == "正在安装更新"
    assert dialog.title_label.text() == "正在安装更新"
    assert "安装完成后" in dialog.detail_label.text()
    assert dialog.progress_bar.value() == 35
    assert dialog.stage_label.text() == "正在备份当前程序…"
    assert dialog.width() <= 460
    assert dialog.height() <= 150
    assert "border-radius" in dialog.styleSheet()
    dialog.close()


def test_update_helper_uses_progress_dialog_and_canonical_exe():
    from gui.update_manager import UPDATE_HELPER_SOURCE

    assert "run_update_install_dialog" in UPDATE_HELPER_SOURCE
    assert "hourlyreport_automation.exe" in UPDATE_HELPER_SOURCE
    assert "rollback" in UPDATE_HELPER_SOURCE


def test_frozen_update_launcher_copies_itself_and_waits_for_ready_marker(tmp_path, monkeypatch):
    import sys
    import zipfile

    from gui import update_manager

    root = tmp_path / "app"
    storage = tmp_path / "updates"
    root.mkdir()
    storage.mkdir()
    launcher = root / update_manager.APP_EXE_NAME
    launcher.write_bytes(b"frozen-exe")
    archive = tmp_path / "update.zip"
    with zipfile.ZipFile(archive, "w") as package:
        package.writestr(update_manager.APP_EXE_NAME, b"new-exe")
        package.writestr("main.py", "print('new')")
        package.writestr("gui/version.py", "CURRENT_VERSION='2026.7.23.109'")

    commands = []

    class Process:
        def poll(self):
            return None

    def fake_popen(command, **_kwargs):
        commands.append(command)
        Path(command[-1]).write_text("ready\n", encoding="utf-8")
        return Process()

    monkeypatch.setattr(update_manager, "_update_storage_dir", lambda: storage)
    monkeypatch.setattr(update_manager.subprocess, "Popen", fake_popen)
    monkeypatch.setattr(sys, "frozen", True, raising=False)
    monkeypatch.setattr(sys, "executable", str(launcher))

    update_manager.launch_update_helper(root, archive, "2026.7.23.109", launcher)

    helper = storage / "hourlyreport_update_helper.exe"
    assert helper.read_bytes() == b"frozen-exe"
    assert commands[0][0] == str(helper)
    assert commands[0][1] == "--apply-update-helper"
    assert not (root / ".venv" / "Scripts" / "pythonw.exe").exists()


def _load_update_helper_namespace(root: Path, monkeypatch):
    import sys

    from gui.update_manager import UPDATE_HELPER_SOURCE

    monkeypatch.setattr(sys, "argv", ["apply_update.py", str(root)])
    namespace = {"__name__": "update_helper_test", "__file__": "apply_update.py"}
    exec(compile(UPDATE_HELPER_SOURCE, "apply_update.py", "exec"), namespace)
    namespace["wait_for_exit"] = lambda _pid: None
    return namespace


def test_update_helper_retries_short_lived_restart_and_logs_success(tmp_path, monkeypatch):
    root = tmp_path / "app"
    storage = tmp_path / "storage"
    root.mkdir()
    storage.mkdir()
    launcher = root / "hourlyreport_automation.exe"
    launcher.write_bytes(b"exe")
    namespace = _load_update_helper_namespace(root, monkeypatch)
    processes = [
        type("Process", (), {"pid": 101, "poll": lambda self: 1})(),
        type("Process", (), {"pid": 202, "poll": lambda self: None})(),
    ]
    calls = []

    def fake_popen(command, **kwargs):
        calls.append((command, kwargs))
        return processes.pop(0)

    monkeypatch.setattr(namespace["subprocess"], "Popen", fake_popen)
    monkeypatch.setattr(namespace["time"], "sleep", lambda _seconds: None)

    pid = namespace["restart_updated_app"](launcher, root, storage, attempts=2, settle_seconds=0)

    assert pid == 202
    assert len(calls) == 2
    log = (storage / "update_apply.log").read_text(encoding="utf-8")
    assert "restart=ok" in log
    assert "pid=202" in log


def test_gui_app_dispatches_frozen_update_helper_before_normal_startup(monkeypatch, tmp_path):
    import gui.app as desktop_app

    calls = []
    monkeypatch.setattr(
        desktop_app,
        "run_update_helper_from_argv",
        lambda argv: calls.append(list(argv)) or 7,
        raising=False,
    )
    monkeypatch.setattr(
        desktop_app,
        "resolve_app_root",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("normal startup must not run")),
    )
    argv = [
        "hourlyreport_update_helper.exe",
        "--apply-update-helper",
        str(tmp_path),
    ]
    monkeypatch.setattr(desktop_app.sys, "argv", argv)

    assert desktop_app.main() == 7
    assert calls == [argv]


def test_update_helper_applies_program_files_without_touching_config(tmp_path, monkeypatch):
    import zipfile

    root = tmp_path / "app"
    storage = tmp_path / "storage"
    root.mkdir()
    storage.mkdir()
    (root / "hourlyreport_automation.exe").write_bytes(b"old-exe")
    (root / "main.py").write_text("old-main", encoding="utf-8")
    (root / "configs").mkdir()
    (root / "configs" / "app_config.json").write_text("private-config", encoding="utf-8")
    package = tmp_path / "update.zip"
    with zipfile.ZipFile(package, "w") as archive:
        archive.writestr("hourlyreport_automation.exe", b"new-exe")
        archive.writestr("main.py", "new-main")
        archive.writestr("gui/version.py", "CURRENT_VERSION='next'")

    namespace = _load_update_helper_namespace(root, monkeypatch)

    class Emitter:
        def emit(self, *_args):
            pass

    signals = type("Signals", (), {"progress": Emitter()})()
    ok, message, launcher = namespace["apply_update"](
        root, package, "2026.7.19.105", 123, storage, signals
    )

    assert ok is True
    assert message == ""
    assert launcher == root / "hourlyreport_automation.exe"
    assert launcher.read_bytes() == b"new-exe"
    assert (root / "main.py").read_text(encoding="utf-8") == "new-main"
    assert (root / "configs" / "app_config.json").read_text(encoding="utf-8") == "private-config"
    backups = list((storage / "backups").glob("2026.7.19.105_*"))
    assert len(backups) == 1
    assert (backups[0] / "hourlyreport_automation.exe").read_bytes() == b"old-exe"


def test_update_helper_rolls_back_partial_install_failure(tmp_path, monkeypatch):
    import zipfile

    root = tmp_path / "app"
    storage = tmp_path / "storage"
    root.mkdir()
    storage.mkdir()
    (root / "hourlyreport_automation.exe").write_bytes(b"old-exe")
    (root / "main.py").write_text("old-main", encoding="utf-8")
    package = tmp_path / "update.zip"
    with zipfile.ZipFile(package, "w") as archive:
        archive.writestr("hourlyreport_automation.exe", b"new-exe")
        archive.writestr("main.py", "new-main")
        archive.writestr("gui/version.py", "CURRENT_VERSION='next'")

    namespace = _load_update_helper_namespace(root, monkeypatch)
    original_replace = namespace["replace_with_retry"]
    failed = {"value": False}

    def fail_once(source, target):
        if Path(target).name == "main.py" and not failed["value"]:
            failed["value"] = True
            raise OSError("simulated apply failure")
        return original_replace(source, target)

    namespace["replace_with_retry"] = fail_once

    class Emitter:
        def emit(self, *_args):
            pass

    signals = type("Signals", (), {"progress": Emitter()})()
    ok, message, _launcher = namespace["apply_update"](
        root, package, "2026.7.19.105", 123, storage, signals
    )

    assert ok is False
    assert "simulated apply failure" in message
    assert (root / "hourlyreport_automation.exe").read_bytes() == b"old-exe"
    assert (root / "main.py").read_text(encoding="utf-8") == "old-main"
    assert not (root / "gui" / "version.py").exists()
    assert "rollback=ok" in (storage / "update_apply.log").read_text(encoding="utf-8")


def test_update_helper_rejects_windows_alias_paths(tmp_path, monkeypatch):
    namespace = _load_update_helper_namespace(tmp_path, monkeypatch)
    for name in ("configs./app.json", "configs /app.json", "file:stream", "C:/file", "NUL.txt"):
        with pytest.raises(ValueError, match="不安全|受保护"):
            namespace["safe_member_path"](name)


def test_update_helper_rollback_continues_after_one_restore_error(tmp_path, monkeypatch):
    root = tmp_path / "app"
    backup = tmp_path / "backup"
    root.mkdir()
    backup.mkdir()
    for name in ("a.txt", "b.txt"):
        (root / name).write_text(f"new-{name}", encoding="utf-8")
        (backup / name).write_text(f"old-{name}", encoding="utf-8")

    namespace = _load_update_helper_namespace(root, monkeypatch)
    original_replace = namespace["replace_with_retry"]

    def fail_a_only(source, target):
        if Path(target).name == "a.txt":
            raise OSError("a is locked")
        return original_replace(source, target)

    namespace["replace_with_retry"] = fail_a_only

    with pytest.raises(RuntimeError, match="a is locked"):
        namespace["rollback"](root, backup, set(), {Path("a.txt"), Path("b.txt")})

    assert (root / "a.txt").read_text(encoding="utf-8") == "new-a.txt"
    assert (root / "b.txt").read_text(encoding="utf-8") == "old-b.txt"


def test_runtime_dependencies_exclude_gui_toolkit_bundled_in_exe():
    root = Path(__file__).resolve().parents[1]
    requirements = (root / "requirements-runtime.txt").read_text(encoding="utf-8").lower()
    assert "pyside6" not in requirements


def test_online_release_refuses_source_or_exe_version_mismatch(tmp_path):
    import json
    import pytest

    from tools.build_desktop_exe import write_build_manifest
    from tools.build_release import build_release

    (tmp_path / "dist").mkdir()
    executable = tmp_path / "dist" / "hourlyreport_automation.exe"
    executable.write_bytes(b"exe-104")
    (tmp_path / "gui").mkdir()
    (tmp_path / "gui" / "version.py").write_text(
        'CURRENT_VERSION = "2026.7.19.104"\n', encoding="utf-8"
    )
    (tmp_path / "main.py").write_text("print('ok')\n", encoding="utf-8")
    write_build_manifest(tmp_path, executable, "2026.7.19.104")

    with pytest.raises(ValueError, match="源码版本"):
        build_release(tmp_path, version="2026.7.19.105", online_update=True)

    manifest_path = tmp_path / "dist" / "hourlyreport_automation.build.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["sha256"] = "0" * 64
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(ValueError, match="构建清单|SHA-256"):
        build_release(tmp_path, version="2026.7.19.104", online_update=True)


def test_desktop_gui_startup_check_attempts_hidden_install_when_environment_missing(tmp_path, monkeypatch):
    import gui.environment_check as environment_check

    install = tmp_path / "install_env.bat"
    install.write_text("@echo off\n", encoding="utf-8")
    calls = []

    monkeypatch.setattr(
        environment_check.subprocess,
        "run",
        lambda *args, **kwargs: calls.append((args, kwargs)) or type("R", (), {"returncode": 0, "stdout": "ok", "stderr": ""})(),
    )

    result = environment_check.repair_environment_if_needed(tmp_path, {"passed": False, "checks": []})

    assert result["attempted"] is True
    assert result["returncode"] == 0
    assert calls
    assert calls[0][1]["cwd"] == tmp_path
    assert calls[0][0][0] == ["cmd.exe", "/d", "/c", str(install)]


def test_desktop_gui_python_bootstrap_uses_isolated_nuget_runtime():
    root = Path(__file__).resolve().parents[1]
    bootstrap = (root / "tools" / "bootstrap_python.ps1").read_text(encoding="utf-8")
    installer = (root / "install_env.bat").read_text(encoding="utf-8")

    assert 'Version = "3.14.5"' in bootstrap
    assert "www.nuget.org/api/v2/package/python/$Version" in bootstrap
    assert "03ad5810986afd8273a34a28c15cb594300ba7f4749f24362d69206fa1b6ac15" in bootstrap
    assert "Get-FileHash" in bootstrap
    assert "Expand-Archive" in bootstrap
    assert 'Join-Path $StagingDir "tools"' in bootstrap
    assert "runtime\\python" in installer
    assert "bootstrap_python.ps1" in installer
    assert "requirements-runtime.txt" in installer


def test_install_env_uses_verified_private_python_instead_of_system_python():
    root = Path(__file__).resolve().parents[1]
    installer = (root / "install_env.bat").read_text(encoding="ascii")

    assert "%LocalAppData%\\Programs\\Python" not in installer
    assert "%ProgramFiles%\\Python" not in installer
    assert "py -3 --version" not in installer
    assert "python --version" not in installer
    assert "python3 --version" not in installer
    assert installer.index("bootstrap_python.ps1") < installer.index("-m venv .venv")


def test_windows_user_entry_bats_are_ascii_and_use_crlf():
    root = Path(__file__).resolve().parents[1]
    entry_names = [
        "install_env.bat",
        "run_menu.bat",
        "run_desktop_gui.bat",
        "run_hermes_hourly.bat",
        "run_hermes_daily.bat",
    ]

    for name in entry_names:
        raw = (root / name).read_bytes()
        raw.decode("ascii")
        assert b"\n" in raw, name
        assert raw.count(b"\n") == raw.count(b"\r\n"), name


def test_install_env_check_mode_is_parseable_and_has_no_side_effects():
    if os.name != "nt":
        return

    import subprocess

    root = Path(__file__).resolve().parents[1]
    result = subprocess.run(
        ["cmd.exe", "/d", "/c", "install_env.bat", "--check"],
        cwd=root,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=30,
    )

    assert result.returncode == 0, result.stdout + result.stderr
    assert "[ENV][CHECK] PYTHON_READY=" in result.stdout
    assert "Environment setup completed" not in result.stdout


def test_desktop_gui_task_runner_infers_progress_stages():
    from gui.task_runner import infer_pet_event, infer_stage

    assert infer_stage("[HERMES] Running hourly quick preflight...") == "preflight"
    assert infer_stage("[通知] 百度账号登录完成") == "login"
    assert infer_stage("fetch-baidu-auto started") == "baidu"
    assert infer_stage("parse-kst-export completed") == "kst"
    assert infer_stage("Excel 写入完成") == "excel"
    assert infer_stage("[ERROR] Preflight failed") == "error"
    assert infer_stage("[API] 正在读取百度数据") == "baidu"
    assert infer_stage("[降级] API 读取仍未完成，准备切换浏览器") == "baidu"
    assert infer_stage("[浏览器] 正在启动浏览器降级流程") == "login"
    assert infer_stage("[实际来源] API") == "baidu"
    assert infer_pet_event("已填写百度登录字段：username") == "login"
    assert infer_pet_event("顶部用户名已匹配项目账号") == "login_ready"
    assert infer_pet_event("[1/4] 读取百度搜索推广数据") == "baidu"
    assert infer_pet_event("百度日报表格数据已稳定") == "baidu_ready"
    assert infer_pet_event("[API] 正在读取项目账户数据") == "baidu"
    assert infer_pet_event("[API] 百度数据读取完成") == "baidu_ready"
    assert infer_pet_event("[浏览器] 正在启动浏览器降级流程") == "login"
    assert infer_pet_event("[实际来源] API") == "baidu_ready"
    assert infer_pet_event("[2/4] 解析快商通导出文件") == "kst"
    assert infer_pet_event("[3/4] 合并百度与快商通数据") == "merge"
    assert infer_pet_event("[4/4] 写入 Excel 并复核") == "excel"
    assert infer_pet_event("[失败] 百度数据读取异常") == "failed"
    assert infer_pet_event("百度数据读取未通过") == "failed"


def test_api_routing_log_safe_api_fallback_transitions(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    class Logger:
        def __init__(self):
            self.messages = []

        def _record(self, message, *args):
            self.messages.append(message % args if args else message)

        info = _record
        warning = _record

    logger = Logger()
    result = fetch_baidu_resilient_hourly(
        {
            "project_id": "demo",
            "project_name": "示例项目",
            "baidu": {"data_source_mode": "api_preferred"},
        },
        tmp_path,
        logger,
        "15点",
        api_fetcher=lambda **_kwargs: (_ for _ in ()).throw(
            BaiduReportApiError(
                "https://api.example.invalid/?accessToken=do-not-log",
                category="network_error",
            )
        ),
        browser_fetcher=lambda **_kwargs: _route_report("baidu_auto_overview"),
        sleep=lambda _seconds: None,
    )

    text = "\n".join(logger.messages)
    assert result["data_source"] == "browser_fallback"
    assert "[数据源] 当前模式：API 优先" in text
    assert "[API] 正在读取示例项目百度数据" in text
    assert "[降级] API 读取仍未完成，准备切换浏览器：network_error" in text
    assert "[浏览器] 正在启动浏览器降级流程" in text
    assert "[实际来源] 浏览器降级" in text
    assert "accessToken" not in text
    assert "https://" not in text


def test_api_routing_log_sanitizes_unrecognized_api_failure_category(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly
    from modules.baidu_report_api import BaiduReportApiError

    class Logger:
        def __init__(self):
            self.messages = []

        def warning(self, message, *args):
            self.messages.append(message % args if args else message)

        def info(self, message, *args):
            self.messages.append(message % args if args else message)

    logger = Logger()
    fetch_baidu_resilient_hourly(
        {"baidu": {"data_source_mode": "api_preferred"}},
        tmp_path,
        logger,
        "15点",
        api_fetcher=lambda **_kwargs: (_ for _ in ()).throw(
            BaiduReportApiError("请求失败", category="https://api.example.invalid/?token=do-not-log")
        ),
        browser_fetcher=lambda **_kwargs: _route_report("baidu_auto_overview"),
    )

    text = "\n".join(logger.messages)
    assert "[降级] API 读取仍未完成，准备切换浏览器：api_error" in text
    assert "https://" not in text
    assert "token=" not in text


def test_task4_attempt_report_redacts_sensitive_exception_message(tmp_path):
    from modules.baidu_report_api import _write_attempt_report

    _write_attempt_report(
        tmp_path,
        config={"project_id": "demo", "project_name": "示例项目"},
        selected_date="2026-07-17",
        period="15点",
        category="network_error",
        message=(
            "https://api.example.invalid/?accessToken=secret-token&refreshToken=refresh-token"
            "&token=token-value&secret=secret-value&password=password-value&header=X-Api-Key"
        ),
    )

    attempt = json.loads((tmp_path / "reports" / "baidu_api_attempt_report.json").read_text("utf-8"))
    serialized = json.dumps(attempt, ensure_ascii=False)
    assert attempt["error_category"] == "network_error"
    assert attempt["errors"] == ["百度 API 网络请求失败，请稍后重试"]
    for sensitive in ("https://", "accessToken", "refreshToken", "token=", "secret", "password", "header", "X-Api-Key"):
        assert sensitive not in serialized


def test_api_routing_log_multi_source_api_success_logs_each_source_and_actual_source(tmp_path):
    from modules.baidu_data_source import fetch_baidu_resilient_hourly

    class Logger:
        def __init__(self):
            self.messages = []

        def info(self, message, *args):
            self.messages.append(message % args if args else message)

    logger = Logger()

    result = fetch_baidu_resilient_hourly(
        _two_source_route_config(),
        tmp_path,
        logger,
        "15点",
        api_fetcher=lambda *, config, **_kwargs: {
            "accounts": {
                next(iter(config["accounts"])): {"展现": 1, "点击": 1, "消费": 1.0}
            },
            "errors": [],
        },
        browser_fetcher=lambda **_kwargs: (_ for _ in ()).throw(AssertionError("不应降级浏览器")),
        sleep=lambda _seconds: None,
    )

    text = "\n".join(logger.messages)
    assert result["data_source"] == "api"
    assert "[API 1/2]" in text
    assert "[API 2/2]" in text
    assert "[实际来源] API" in text


def test_desktop_pet_settings_default_to_clawd_and_persist_hidden(tmp_path):
    from gui.pet_settings import (
        PET_CLAWD,
        PET_HIDDEN,
        load_pet_mode,
        load_pet_position,
        load_pet_scale,
        save_pet_mode,
        save_pet_position,
        save_pet_scale,
    )

    config = tmp_path / "configs" / "app_config.json"
    config.parent.mkdir(parents=True)
    config.write_text(json.dumps({"default_project_id": "demo"}, ensure_ascii=False), encoding="utf-8")

    assert load_pet_mode(tmp_path) == PET_CLAWD
    save_pet_mode(tmp_path, PET_HIDDEN)

    saved = json.loads(config.read_text(encoding="utf-8"))
    assert saved["default_project_id"] == "demo"
    assert saved["desktop_pet"] == PET_HIDDEN
    assert load_pet_mode(tmp_path) == PET_HIDDEN
    assert load_pet_scale(tmp_path) == 1.0
    assert load_pet_position(tmp_path) is None

    save_pet_scale(tmp_path, 0.73)
    save_pet_position(tmp_path, 321, 654)
    assert load_pet_scale(tmp_path) == 0.73
    assert load_pet_position(tmp_path) == (321, 654)
    save_pet_scale(tmp_path, 1.4)
    assert load_pet_scale(tmp_path) == 1.2
    save_pet_scale(tmp_path, 0.2)
    assert load_pet_scale(tmp_path) == 0.5


def test_desktop_pet_uses_the_approved_clawd_v2_asset(monkeypatch):
    import hashlib

    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.desktop_pet import (
        ANIMATIONS,
        CLASSIC_PASSAGES,
        CLASSIC_RECITATION_MS,
        IDLE_ROUTINES,
        ClawdDesktopPet,
    )

    root = Path(__file__).resolve().parents[1]
    manifest = json.loads((root / "assets" / "clawd" / "pet.json").read_text(encoding="utf-8"))
    sprite = root / "assets" / "clawd" / "spritesheet.webp"
    assert manifest["id"] == "clawd"
    assert manifest["spriteVersionNumber"] == 2
    assert hashlib.sha256(sprite.read_bytes()).hexdigest() == "08f1c320c535dfb63667af7b844b6af75af2a4a71f768d93c7be86d9679534a1"

    app = QApplication.instance() or QApplication([])
    pet = ClawdDesktopPet(root, lambda: None)
    assert pet.available
    assert set(ANIMATIONS) == {
        "idle", "walk_right", "walk_left", "waving", "jumping", "failed", "waiting", "running", "review",
        "look_a", "look_b",
    }
    assert len(IDLE_ROUTINES) == 8
    assert len({routine.name for routine in IDLE_ROUTINES}) == 8
    expected_titles = {
        "洛神赋", "滕王阁序", "前出师表", "陈情表", "岳阳楼记",
        "醉翁亭记", "兰亭集序", "桃花源记", "前赤壁赋", "逍遥游",
    }
    assert len(CLASSIC_PASSAGES) >= 30
    assert {passage.title for passage in CLASSIC_PASSAGES} == expected_titles
    assert all(len(passage.lines) == 2 for passage in CLASSIC_PASSAGES)
    assert all(all(line.strip() for line in passage.lines) for passage in CLASSIC_PASSAGES)
    recitation = next(routine for routine in IDLE_ROUTINES if routine.recites_classic)
    assert sum(step.duration_ms for step in recitation.steps) == CLASSIC_RECITATION_MS == 10_000
    pet.set_enabled(True)
    pet.announce("正在读取百度数据", "running")
    assert pet._state == "running"
    assert pet._bubble.label.text() == "正在读取百度数据"
    base_size = pet.size()
    pet.move(240, 180)
    original_position = pet.pos()
    pet.set_pet_scale(1.4)
    assert pet.pet_scale() == 1.2
    assert pet.width() > base_size.width()
    assert pet.height() > base_size.height()
    assert pet.pos() == original_position
    pet.close_pet()


def test_desktop_pet_idle_routines_are_randomized_and_preempted_by_tasks(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.desktop_pet import CLASSIC_PASSAGES, ClawdDesktopPet

    app = QApplication.instance() or QApplication([])
    pet = ClawdDesktopPet(Path(__file__).resolve().parents[1], lambda: None)
    pet.set_enabled(True)

    assert pet.trigger_idle_routine("诵读名篇") is True
    assert pet._routine is not None
    assert pet._routine.name == "诵读名篇"
    bubble_text = pet._bubble.label.text()
    assert "\n" not in bubble_text
    assert not bubble_text.startswith("《")
    assert bubble_text in {"".join(passage.lines) for passage in CLASSIC_PASSAGES}
    assert all(passage.title not in bubble_text for passage in CLASSIC_PASSAGES)
    assert pet._bubble.label.wordWrap() is False
    assert pet._bubble.height() == 86

    pet._passage_queue.clear()
    pet._rng.seed(20260714)
    shuffled_cycle = [pet._next_classic_passage() for _ in range(len(CLASSIC_PASSAGES))]
    assert len({(passage.title, passage.lines) for passage in shuffled_cycle}) == len(CLASSIC_PASSAGES)
    assert shuffled_cycle != list(CLASSIC_PASSAGES)
    assert pet._bubble.isVisible()

    pet.set_busy(True)
    assert pet._routine is None
    assert not pet._idle_timer.isActive()
    assert not pet._bubble.isVisible()
    assert pet.trigger_idle_routine("桌面巡逻") is False

    pet.set_busy(False)
    assert pet._idle_timer.isActive()
    pet.close_pet()


def test_desktop_pet_left_click_toggles_right_click_does_nothing_and_drag_saves(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import QPoint, Qt
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import QApplication
    from gui.desktop_pet import ClawdDesktopPet

    app = QApplication.instance() or QApplication([])
    toggles = []
    positions = []
    pet = ClawdDesktopPet(
        Path(__file__).resolve().parents[1],
        lambda: toggles.append(True),
        lambda x, y: positions.append((x, y)),
    )
    pet.set_enabled(True)
    app.processEvents()

    QTest.mouseClick(pet, Qt.MouseButton.RightButton, pos=QPoint(30, 30))
    assert toggles == []
    QTest.mouseClick(pet, Qt.MouseButton.LeftButton, pos=QPoint(30, 30))
    assert toggles == [True]

    start = pet.pos()
    QTest.mousePress(pet, Qt.MouseButton.LeftButton, pos=QPoint(30, 30))
    QTest.mouseMove(pet, QPoint(75, 70), delay=30)
    QTest.mouseRelease(pet, Qt.MouseButton.LeftButton, pos=QPoint(75, 70))
    app.processEvents()
    assert pet.pos() != start
    assert positions and positions[-1] == (pet.x(), pet.y())
    assert toggles == [True]
    pet.close_pet()


def test_system_menu_only_opens_after_click_not_hover(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()
    QTest.mouseMove(window.system_config_button)
    QTest.qWait(100)
    assert not window.inline_config_menu.isVisible()
    QTest.mouseClick(window.system_config_button, Qt.MouseButton.LeftButton)
    QTest.qWait(100)
    assert window.inline_config_menu.isVisible()
    assert not window.system_config_menu.isVisible()
    window.inline_config_menu.hide()
    window._quitting = True
    window.close()


def test_system_submenus_expand_inline_without_flyout_or_shadow(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtTest import QTest
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    menu = window.inline_config_menu
    collapsed_height = menu.height()

    assert menu.width() == 224
    assert bool(menu.windowFlags() & Qt.WindowType.NoDropShadowWindowHint)
    assert menu.pet_section.isHidden()
    assert menu.size_section.isHidden()
    assert menu.pet_toggle.property("expandable") is True
    assert menu.size_toggle.property("expandable") is True
    assert menu.pet_layout.spacing() == 3
    assert window.task_title.text() == "项目控制台"
    assert not menu.pet_toggle.is_expanded()
    assert not menu.size_toggle.is_expanded()

    menu.pet_toggle.click()
    assert not menu.pet_section.isHidden()
    assert menu.pet_toggle.is_expanded()
    pet_height = menu.height()
    assert pet_height > collapsed_height
    menu.size_toggle.click()
    assert not menu.size_section.isHidden()
    assert menu.size_toggle.is_expanded()
    assert menu.height() > pet_height
    assert not window.pet_menu.isVisible()
    assert menu.size_slider.minimum() == 50
    assert menu.size_slider.maximum() == 120
    menu.size_slider.setValue(73)
    assert menu.size_value_label.text() == "73%"
    assert window.desktop_pet.pet_scale() == 0.73
    window._pet_scale_save_timer.stop()

    menu.hide()
    window._quitting = True
    window.close()


def test_desktop_gui_single_instance_guard_rejects_second_owner(tmp_path):
    import uuid

    from PySide6.QtCore import QCoreApplication
    from PySide6.QtTest import QTest
    from gui.single_instance import SingleInstanceGuard

    app = QCoreApplication.instance() or QCoreApplication([])
    instance_id = "hourly_report_test_" + uuid.uuid4().hex
    first = SingleInstanceGuard(instance_id, tmp_path)
    second = SingleInstanceGuard(instance_id, tmp_path)
    activations = []
    first.activate_requested.connect(lambda: activations.append(True))
    try:
        assert first.acquire() is True
        assert second.acquire() is False
        assert second.notify_existing() is True
        QTest.qWait(100)
        assert activations == [True]
    finally:
        second.close()
        first.close()


def test_desktop_pet_keeps_running_when_console_is_hidden(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    import gui.main_window as main_window

    monkeypatch.setattr(main_window, "load_pet_mode", lambda root: main_window.PET_CLAWD)
    MainWindow = main_window.MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    window.request_console_close()
    app.processEvents()
    assert not window.isVisible()
    assert window.desktop_pet.isVisible()

    window.toggle_console_visibility()
    app.processEvents()
    assert window.isVisible()

    window.toggle_console_visibility()
    app.processEvents()
    assert window.isVisible()

    window._quitting = True
    window.tray_icon.hide()
    window.desktop_pet.close_pet()
    window.close()


def test_desktop_gui_tray_icon_opens_console_and_exposes_exit(monkeypatch):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QApplication, QSystemTrayIcon
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])

    assert not window.tray_icon.icon().isNull()
    assert [action.text() for action in window.tray_menu.actions()] == ["打开控制台", "退出程序"]
    assert window.tray_icon.contextMenu() is window.tray_menu
    assert window.tray_menu.testAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
    assert bool(window.tray_menu.windowFlags() & Qt.WindowType.NoDropShadowWindowHint)
    assert "border-radius: 10px" in window.tray_menu.styleSheet()

    window.hide()
    window.tray_icon.activated.emit(QSystemTrayIcon.ActivationReason.Trigger)
    app.processEvents()
    assert window.isVisible()

    window._quitting = True
    window.tray_icon.hide()
    window.desktop_pet.close_pet()
    window.close()


@pytest.mark.parametrize("pet_mode", ["clawd", "hidden"])
def test_desktop_gui_close_only_hides_for_every_pet_mode(monkeypatch, pet_mode):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    import gui.main_window as main_window

    monkeypatch.setattr(main_window, "load_pet_mode", lambda root: pet_mode)
    app = QApplication.instance() or QApplication([])
    window = main_window.MainWindow(Path(__file__).resolve().parents[1])
    window.show()
    app.processEvents()

    window.close_button.click()
    app.processEvents()

    assert not window.isVisible()
    assert window._quitting is False
    assert window.tray_icon.isVisible()

    window.show()
    window.close()
    app.processEvents()
    assert not window.isVisible()
    assert window._quitting is False
    assert window.tray_icon.isVisible()

    window._quitting = True
    window.tray_icon.hide()
    window.desktop_pet.close_pet()
    window.close()


@pytest.mark.parametrize("exit_source", ["tray", "system_menu"])
def test_desktop_gui_exit_actions_clean_tray_and_pet(monkeypatch, exit_source):
    monkeypatch.setenv("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication
    from gui.main_window import MainWindow

    app = QApplication.instance() or QApplication([])
    window = MainWindow(Path(__file__).resolve().parents[1])
    quit_calls = []
    pet_close_calls = []
    original_close_pet = window.desktop_pet.close_pet
    monkeypatch.setattr(app, "quit", lambda: quit_calls.append(True))
    monkeypatch.setattr(
        window.desktop_pet,
        "close_pet",
        lambda: (pet_close_calls.append(True), original_close_pet()),
    )

    if exit_source == "tray":
        window.tray_exit_action.trigger()
    else:
        window.inline_config_menu.exit_requested.emit()

    assert window._quitting is True
    assert not window.tray_icon.isVisible()
    assert pet_close_calls == [True]
    assert quit_calls == [True]
    window.close()


def test_desktop_gui_app_icon_assets_and_build_icon_are_configured():
    root = Path(__file__).resolve().parents[1]
    icon = root / "assets" / "app_icon.ico"
    runtime_source = root / "assets" / "app_icon.png"
    exe_source = root / "assets" / "app_icon_exe.png"
    build_script = (root / "tools" / "build_desktop_exe.py").read_text(encoding="utf-8")
    spec_source = (root / "tools" / "hourlyreport_automation.spec").read_text(encoding="utf-8")

    assert icon.exists()
    assert runtime_source.exists()
    assert exe_source.exists()
    assert "hourlyreport_automation.spec" in build_script
    assert "app_icon.ico" in spec_source
    assert "exe = EXE(" in spec_source
    assert "console=False" in spec_source

    from PySide6.QtGui import QImage

    runtime_image = QImage(str(runtime_source))
    assert not runtime_image.isNull()
    runtime_corner = runtime_image.pixelColor(
        max(1, runtime_image.width() // 32), max(1, runtime_image.height() // 32)
    )
    assert runtime_corner.alpha() == 0
    runtime_center = runtime_image.pixelColor(runtime_image.width() // 2, runtime_image.height() // 2)
    assert runtime_center.alpha() == 255

    exe_image = QImage(str(exe_source))
    assert not exe_image.isNull()
    assert exe_image.pixelColor(1, 1).alpha() == 0
    dark_background = exe_image.pixelColor(exe_image.width() // 2, exe_image.height() // 16)
    assert dark_background.alpha() == 255
    assert max(dark_background.red(), dark_background.green(), dark_background.blue()) < 40


def test_desktop_build_spec_filters_unused_qt_components():
    root = Path(__file__).resolve().parents[1]
    spec = root / "tools" / "hourlyreport_automation.spec"
    build_script = (root / "tools" / "build_desktop_exe.py").read_text(encoding="utf-8")

    assert spec.exists()
    source = spec.read_text(encoding="utf-8")
    for marker in (
        "qt6quick",
        "qt6qml",
        "qt6pdf",
        "qt6virtualkeyboard",
        "qt6opengl",
        "opengl32sw.dll",
    ):
        assert marker in source.lower()
    assert "qwindows.dll" in source.lower()
    assert "hourlyreport_automation.spec" in build_script


def test_desktop_gui_windows_app_id_changes_when_icon_changes(tmp_path):
    from gui.app import windows_app_user_model_id

    assets = tmp_path / "assets"
    assets.mkdir()
    icon = assets / "app_icon.png"
    icon.write_bytes(b"first-icon")
    first_id = windows_app_user_model_id(tmp_path)

    icon.write_bytes(b"second-icon")
    second_id = windows_app_user_model_id(tmp_path)

    assert first_id.startswith("HourlyreportAutomation.Console.")
    assert first_id != second_id
    source = (Path(__file__).resolve().parents[1] / "gui" / "app.py").read_text(encoding="utf-8")
    assert source.index("configure_windows_app_identity(root)") < source.index("QApplication(sys.argv)")


def test_clawd_animator_uses_reference_geometry_and_six_phase_loop():
    from gui.clawd import CLAWD_BODY, CLAWD_EYE, PHASE_DURATIONS, _dance_pose, _idle_phase_at, _phase_at, _pose_for

    assert CLAWD_BODY.name().upper() == "#D4634A"
    assert CLAWD_EYE.name().upper() == "#2A1810"
    assert PHASE_DURATIONS == (5.0, 3.0, 3.0, 3.0, 6.0, 3.0)
    assert _phase_at(0.0)[0] == 0
    assert _phase_at(5.2)[0] == 1
    assert _phase_at(14.2)[0] == 4
    assert _pose_for(4, 0.25).x > 0
    assert _pose_for(5, 0.5).sparkle is True
    assert _idle_phase_at(0.0)[0] == 1
    assert _dance_pose(0.15).x < 0
    assert _dance_pose(0.50).x > 0


# ── 百度登录状态守卫测试 (CAS 兜底版) ─────────────────────


def test_load_login_state_returns_empty_when_file_missing(tmp_path):
    from modules.baidu_session import load_browser_login_state
    state = load_browser_login_state(tmp_path)
    assert state["last_profile"] is None


def test_mark_login_success_and_read_profile(tmp_path):
    from modules.baidu_session import (
        get_browser_login_profile, load_browser_login_state, mark_browser_login_success,
    )
    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_browser_login_success(
        tmp_path, credential_profile="kunming_niu_baidu",
        project_id="kunming_niu", project_name="昆明牛",
        task="run-daily", url="https://cc.baidu.com/report",
    )
    state = load_browser_login_state(tmp_path)
    assert state["last_profile"] == "kunming_niu_baidu"
    assert state["last_project_id"] == "kunming_niu"
    assert state["last_login_at"] is not None
    assert "username" not in state
    assert "password" not in state


def test_get_browser_login_profile(tmp_path):
    from modules.baidu_session import get_browser_login_profile, mark_browser_login_success
    (tmp_path / "reports").mkdir(exist_ok=True)
    assert get_browser_login_profile(tmp_path) is None
    mark_browser_login_success(tmp_path, "nanjing_niu_baidu", project_id="nanjing_niu")
    assert get_browser_login_profile(tmp_path) == "nanjing_niu_baidu"


def test_clear_browser_login_state(tmp_path):
    from modules.baidu_session import clear_browser_login_state, get_browser_login_profile, mark_browser_login_success
    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_browser_login_success(tmp_path, "test_profile", project_id="test")
    assert get_browser_login_profile(tmp_path) == "test_profile"
    clear_browser_login_state(tmp_path)
    assert get_browser_login_profile(tmp_path) is None


def test_get_current_project_credential_profile():
    from modules.baidu_session import get_current_project_credential_profile
    assert get_current_project_credential_profile({"baidu": {"credential_profile": "kunming_niu_baidu"}}) == "kunming_niu_baidu"
    assert get_current_project_credential_profile({"baidu": {"credential_project": "nanjing_niu_baidu"}}) == "nanjing_niu_baidu"
    assert get_current_project_credential_profile({}) == ""


def test_browser_login_state_no_passwords(tmp_path):
    from modules.baidu_session import load_browser_login_state, save_browser_login_state
    (tmp_path / "reports").mkdir(exist_ok=True)
    save_browser_login_state(tmp_path, {"last_profile": "test", "username": "strip", "password": "strip"})
    state = load_browser_login_state(tmp_path)
    assert "username" not in state
    assert "password" not in state


def test_menu_no_longer_pre_saves_login_state():
    root = Path(__file__).resolve().parents[1]
    content = (root / "menu.py").read_text(encoding="utf-8")
    assert "save_login_state(" not in content
    assert "check_profile_match(" not in content
    assert "baidu_session" in content


# ── CAS 登录页测试 ────────────────────────────────────────


def test_goto_baidu_login_page_uses_cas_url(monkeypatch):
    from unittest.mock import MagicMock
    from modules.baidu_session import goto_baidu_login_page, BAIDU_CAS_LOGIN_URL
    fake_page = MagicMock()
    fake_page.url = "about:blank"
    result = goto_baidu_login_page(fake_page)
    assert result["success"] is True
    fake_page.goto.assert_called_once()
    assert "cas.baidu.com" in fake_page.goto.call_args[0][0]


def test_goto_baidu_login_page_returns_error_on_exception():
    from unittest.mock import MagicMock
    from modules.baidu_session import goto_baidu_login_page
    fake_page = MagicMock()
    fake_page.goto.side_effect = Exception("timeout")
    result = goto_baidu_login_page(fake_page)
    assert result["success"] is False


def test_session_usable_page_skips_cas(tmp_path, monkeypatch):
    """已在可用搜索推广数据页 -> 不跳 CAS。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")
    monkeypatch.setattr("modules.baidu_session._page_is_usable_search_promotion", lambda p, r, c: True)
    cas_calls = []
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    result = ensure_baidu_profile_session(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result.get("passed") is True
    assert len(cas_calls) == 0, "可用页面不应跳 CAS"


def test_session_profile_mismatch_triggers_cas_login(tmp_path, monkeypatch):
    """last_profile 不一致且用户名无法识别时触发 CAS 登录。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session, mark_browser_login_success
    (tmp_path / "reports").mkdir(exist_ok=True)
    mark_browser_login_success(tmp_path, "old_profile", project_id="old")
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")
    monkeypatch.setattr("modules.baidu_session._page_is_usable_search_promotion", lambda p, r, c: False)
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda p: False)
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: None)
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: None)
    cas_calls = []
    login_calls = []
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: login_calls.append(1) or True)
    result = ensure_baidu_profile_session(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result.get("passed") is True
    assert len(cas_calls) >= 1
    assert len(login_calls) >= 1


def test_session_last_profile_none_triggers_cas(tmp_path, monkeypatch):
    """last_profile=None 且无法识别用户名时触发 CAS 登录。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")
    monkeypatch.setattr("modules.baidu_session._page_is_usable_search_promotion", lambda p, r, c: False)
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda p: False)
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: None)
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: None)
    cas_calls = []
    login_calls = []
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: login_calls.append(1) or True)
    result = ensure_baidu_profile_session(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result.get("passed") is True
    assert len(cas_calls) >= 1
    assert len(login_calls) >= 1


def test_username_match_skips_cas_login(tmp_path, monkeypatch):
    """当前账号匹配时直接通过，不调用 CAS 登录。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")
    monkeypatch.setattr("modules.baidu_session._page_is_usable_search_promotion", lambda p, r, c: False)
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "test_user")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "test_user")
    cas_calls = []
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    result = ensure_baidu_profile_session(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result.get("passed") is True
    assert len(cas_calls) == 0, "账号匹配时不应调用 CAS 登录"


def test_no_profile_skips_check(tmp_path):
    """无 credential_profile 时不做检查。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session
    fake_page = MagicMock()
    logger = logging.getLogger("test")
    result = ensure_baidu_profile_session(tmp_path, {}, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result.get("passed") is True


def test_force_relogin_uses_cas_page(tmp_path, monkeypatch):
    """force_relogin_current_project 退出成功 → 进入 CAS 登录页。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "about:blank"
    logger = logging.getLogger("test")
    cas_calls = []
    login_calls = []
    monkeypatch.setattr("modules.baidu_session.logout_baidu_account", lambda page: {"success": True, "message": "ok"})
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: login_calls.append(1) or True)
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "test_user")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "test_user")
    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is True
    assert len(cas_calls) >= 1
    assert len(login_calls) >= 1


def test_force_relogin_does_not_show_chrome_during_automatic_switch(tmp_path, monkeypatch):
    """自动切换百度账号时不应把 Chrome 拉到前台。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"
    logger = logging.getLogger("test")

    monkeypatch.setattr("modules.baidu_session.logout_baidu_account", lambda page: {"success": True, "message": "ok"})
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)
    monkeypatch.setattr(
        "modules.browser_manager.show_browser_page_for_manual_intervention",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("automatic relogin must stay silent")),
    )

    assert force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None) is True


def test_force_relogin_login_failure_returns_false(tmp_path, monkeypatch):
    """CAS 登录失败时 force_relogin 返回 False。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "test"}, "project_id": "p"}
    fake_page = MagicMock()
    logger = logging.getLogger("test")
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: False)
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "test_user")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "test_user")
    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is False


def test_login_flow_username_mismatch_allows_proceeding(tmp_path, monkeypatch):
    """用户名不匹配时：不写 browser_login_state，返回 needs_project_account_check。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import _do_login_flow, load_browser_login_state

    (tmp_path / "reports").mkdir(exist_ok=True)
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")

    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "user_a")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "user_b")
    monkeypatch.setattr("modules.baidu_session.get_current_project_credential_profile", lambda c: "test_prof")
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = _do_login_flow(tmp_path, {"baidu": {"credential_profile": "test_prof"}},
                            fake_page, logger, "p", "n", None, lambda _: None)
    assert result["success"] is True
    assert result["account_verified"] is False
    assert result["needs_project_account_check"] is True
    state = load_browser_login_state(tmp_path)
    assert state.get("last_profile") is None, "用户名不匹配时不应写 login_state"


def test_login_flow_undetectable_user_allows_proceeding(tmp_path, monkeypatch):
    """无法识别用户名时：不写 browser_login_state，返回 needs_project_account_check。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import _do_login_flow, load_browser_login_state

    (tmp_path / "reports").mkdir(exist_ok=True)
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")

    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "user_a")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: None)
    monkeypatch.setattr("modules.baidu_session.get_current_project_credential_profile", lambda c: "test_prof")
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = _do_login_flow(tmp_path, {"baidu": {"credential_profile": "test_prof"}},
                            fake_page, logger, "p", "n", None, lambda _: None)
    assert result["success"] is True
    assert result["account_verified"] is False
    assert result["needs_project_account_check"] is True
    state = load_browser_login_state(tmp_path)
    assert state.get("last_profile") is None, "无法识别时不应写 login_state"


def test_login_flow_username_match_returns_verified(tmp_path, monkeypatch):
    """用户名匹配时 account_verified=True，但 _do_login_flow 本身不写状态。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import _do_login_flow, load_browser_login_state

    (tmp_path / "reports").mkdir(exist_ok=True)
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    logger = logging.getLogger("test")

    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "test_user")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "test_user")
    monkeypatch.setattr("modules.baidu_session.get_current_project_credential_profile", lambda c: "kunming_niu_baidu")
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = _do_login_flow(tmp_path, {"baidu": {"credential_profile": "kunming_niu_baidu"}},
                            fake_page, logger, "kunming_niu", "昆明牛", None, lambda _: None)
    assert result["success"] is True
    assert result["account_verified"] is True
    assert result["needs_project_account_check"] is False
    # _do_login_flow 本身不写状态
    state = load_browser_login_state(tmp_path)
    assert state.get("last_profile") is None, "_do_login_flow 不写状态，由项目账户复核后写"


def test_logout_baidu_account_prioritizes_click(monkeypatch):
    """logout_baidu_account 优先通过页面点击退出。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import logout_baidu_account

    fake_page = MagicMock()
    fake_page.url = "about:blank"
    fake_el = MagicMock()
    fake_el.count.return_value = 1
    fake_el.is_visible.return_value = True
    fake_page.locator.return_value.first = fake_el

    # wait_until_logged_out 返回 True
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: True)

    result = logout_baidu_account(fake_page, root=".")
    assert result["success"] is True, "应退出成功"


def test_logout_baidu_account_accepts_cas_after_account_click(monkeypatch):
    """点击账号入口后若已进入 CAS，直接视为退出成功，不继续反复找退出按钮。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import logout_baidu_account

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    fake_page.viewport_size = {"width": 1920, "height": 1080}
    fake_el = MagicMock()
    fake_el.count.return_value = 1
    fake_el.is_visible.return_value = True
    fake_el.bounding_box.return_value = {"x": 1700, "y": 0, "width": 120, "height": 48}
    fake_page.locator.return_value.first = fake_el

    def fake_click(page, box):
        page.url = "https://cas.baidu.com/?tpl=www2"
        return True

    dump_calls = []
    monkeypatch.setattr("modules.baidu_session._click_element_center", fake_click)
    monkeypatch.setattr("modules.baidu_session._dump_candidates_to", lambda page, root, filename: dump_calls.append(filename) or [])

    result = logout_baidu_account(fake_page, root=".")

    assert result["success"] is True
    assert result["message"] == "点击账号入口后已进入百度登录页"
    assert dump_calls == ["baidu_logout_candidates_before.json", "baidu_logout_candidates_after_account_click.json"]


def test_logout_candidate_filter_ignores_large_container_and_accepts_menu_item():
    """退出候选只接受明确菜单项，避免点到包含“退出”的整页大容器。"""
    from modules.baidu_session import _is_logout_candidate

    assert _is_logout_candidate({
        "text": "首页 数据报告 退出登录 账户管理",
        "tag": "DIV",
        "class": "root",
        "box": {"x": 0, "y": 0, "width": 1920, "height": 900},
    }) is False
    assert _is_logout_candidate({
        "text": "退出登录",
        "tag": "LI",
        "class": "one-menu-item",
        "box": {"x": 1700, "y": 160, "width": 120, "height": 36},
    }) is True


def test_click_element_center_uses_candidate_width_and_height():
    """候选矩形带宽高时点击真实中心，而不是默认猜 40x24。"""
    from modules.baidu_session import _click_element_center

    class FakeMouse:
        def __init__(self):
            self.moves = []
            self.clicks = []

        def move(self, x, y):
            self.moves.append((x, y))

        def click(self, x, y):
            self.clicks.append((x, y))

    class FakePage:
        def __init__(self):
            self.mouse = FakeMouse()

        def wait_for_timeout(self, timeout):
            pass

    page = FakePage()

    assert _click_element_center(page, {"x": 100, "y": 50, "width": 160, "height": 40}) is True
    assert page.mouse.clicks == [(180, 70)]


def test_logout_baidu_account_no_entry_returns_false():
    """找不到退出入口时不 traceback，返回 success=False。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import logout_baidu_account

    fake_page = MagicMock()
    fake_el = MagicMock()
    fake_el.count.return_value = 0
    fake_page.locator.return_value.first = fake_el

    result = logout_baidu_account(fake_page, root=".")
    assert isinstance(result, dict)
    assert result["success"] is False


def test_force_relogin_tries_logout_then_cas(tmp_path, monkeypatch):
    """force_relogin 先调 logout，再进 CAS。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "about:blank"
    logger = logging.getLogger("test")

    logout_calls = []
    cas_calls = []
    login_calls = []
    monkeypatch.setattr("modules.baidu_session.logout_baidu_account",
                        lambda page: logout_calls.append(1) or {"success": True, "message": "ok"})
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page",
                        lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "test_user")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: "test_user")
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed",
                        lambda p, r, c, l: login_calls.append(1) or True)

    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is True
    assert len(logout_calls) >= 1, "应先调用 logout"
    assert len(cas_calls) >= 1, "应进入 CAS"
    assert len(login_calls) >= 1, "应登录"


def test_logout_failure_resets_cookies_then_uses_cas(tmp_path, monkeypatch):
    """logout 失败且页面已登录时，清 cookie 后继续进入 CAS。"""
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}
    fake_page = MagicMock()
    fake_page.url = "about:blank"
    logger = logging.getLogger("test")

    cas_calls = []
    monkeypatch.setattr("modules.baidu_session.logout_baidu_account",
                        lambda page: {"success": False, "message": "未找到"})
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda page: True)
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page",
                        lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is True
    assert len(cas_calls) >= 1
    fake_page.context.clear_cookies.assert_called()


def test_wait_until_cas_only_accepts_cas_url():
    """wait_until_cas_login_page 只在 URL 含 cas.baidu.com 时返回 True。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import wait_until_cas_login_page

    fake_page = MagicMock()
    fake_page.url = "https://cas.baidu.com/?tpl=www2"
    assert wait_until_cas_login_page(fake_page, timeout_ms=500) is True

    fake_page.url = "https://www2.baidu.com/"
    assert wait_until_cas_login_page(fake_page, timeout_ms=500) is False


def test_logout_success_requires_cas_verification(monkeypatch):
    """退出成功必须验证 CAS 页面。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import logout_baidu_account

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    fake_el = MagicMock()
    fake_el.count.return_value = 0
    fake_page.locator.return_value.first = fake_el

    # CAS 验证失败
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)

    result = logout_baidu_account(fake_page, root=".")
    assert result["success"] is False, "未到 CAS 不应返回成功"


def test_baidu_session_has_no_mark_login_calls():
    """baidu_session.py 中不再有 mark_browser_login_success 调用。"""
    root = Path(__file__).resolve().parents[1]
    source = (root / "modules" / "baidu_session.py").read_text(encoding="utf-8")
    lines = source.split("\n")
    # 找到 def mark_browser_login_success 之后的所有调用
    in_def = False
    calls = []
    for line in lines:
        if "def mark_browser_login_success" in line:
            in_def = True
            continue
        if in_def and "mark_browser_login_success(" in line and "def " not in line:
            calls.append(line.strip())
    assert len(calls) == 0, f"baidu_session.py 不应调用 mark_browser_login_success：{calls}"


# ── 未知百度账户报告测试 ──────────────────────────────────


def test_build_unknown_report_structure():
    """build_unknown_baidu_accounts_report 生成完整结构。"""
    from modules.baidu_unknown_accounts import build_unknown_baidu_accounts_report

    config = {"project_id": "kunming_niu", "project_name": "昆明牛"}
    parsed = {
        "unknown_accounts": [
            {"account_name": "未知A", "展现": 100, "点击": 10, "消费": 50.5,
             "reason": "百度后台抓到该账户，但当前项目配置 accounts 中未配置"},
        ]
    }

    report = build_unknown_baidu_accounts_report(
        config, parsed, task="hourly", date="2026-05-14", period="15点",
    )

    assert report["date"] == "2026-05-14"
    assert report["task"] == "hourly"
    assert report["period"] == "15点"
    assert report["project_id"] == "kunming_niu"
    assert report["project_name"] == "昆明牛"
    assert len(report["unknown_accounts"]) == 1
    assert report["unknown_accounts"][0]["account_name"] == "未知A"
    assert "suggestion" in report["unknown_accounts"][0]


def test_write_unknown_report_when_empty(tmp_path):
    """unknown_accounts 为空时不写文件。"""
    from modules.baidu_unknown_accounts import write_unknown_baidu_accounts_report

    report = {"unknown_accounts": []}
    result = write_unknown_baidu_accounts_report(tmp_path, report)
    assert result is None


def test_write_unknown_report_when_non_empty(tmp_path):
    """unknown_accounts 非空时写入文件。"""
    from modules.baidu_unknown_accounts import write_unknown_baidu_accounts_report

    (tmp_path / "reports").mkdir(exist_ok=True)
    report = {
        "unknown_accounts": [
            {"account_name": "未知A", "展现": 100, "点击": 10, "消费": 50},
        ]
    }
    result = write_unknown_baidu_accounts_report(tmp_path, report)
    assert result is not None
    assert (tmp_path / "reports" / "unknown_baidu_accounts.json").exists()


def test_build_unknown_report_includes_suggestion():
    """unknown report 每项包含 suggestion 字段。"""
    from modules.baidu_unknown_accounts import build_unknown_baidu_accounts_report

    config = {"project_id": "test_proj"}
    parsed = {"unknown_accounts": [{"account_name": "X", "展现": 1, "点击": 0, "消费": 0}]}

    report = build_unknown_baidu_accounts_report(config, parsed, task="daily")
    assert "suggestion" in report["unknown_accounts"][0]
    assert "configs/projects" in report["unknown_accounts"][0]["suggestion"]


def test_auto_report_includes_unknown_accounts():
    """baidu_auto.py 报告结构包含 unknown_accounts 和 ignored_unknown_accounts。"""
    root = Path(__file__).resolve().parents[1]
    source = (root / "modules" / "baidu_auto.py").read_text(encoding="utf-8")
    assert '"unknown_accounts"' in source
    assert '"ignored_unknown_accounts"' in source
    assert "unknown_baidu_accounts_report" in source


def test_daily_report_includes_unknown_accounts():
    """baidu_daily.py 报告结构包含 unknown_accounts 和 ignored_unknown_accounts。"""
    root = Path(__file__).resolve().parents[1]
    source = (root / "modules" / "baidu_daily.py").read_text(encoding="utf-8")
    assert '"unknown_accounts"' in source
    assert '"ignored_unknown_accounts"' in source
    assert "unknown_baidu_accounts_report" in source


def test_build_release_excludes_unknown_accounts_json():
    """build_release 不包含 reports/unknown_baidu_accounts.json。"""
    root = Path(__file__).resolve().parents[1]
    release = build_release(root, version="0.4.17")

    import zipfile
    with zipfile.ZipFile(release) as archive:
        names = set(archive.namelist())
    assert "reports/unknown_baidu_accounts.json" not in names


# ── 未知账户终端提醒测试 ──────────────────────────────────


def test_has_unknown_accounts_detects_presence():
    """has_unknown_baidu_accounts 正确判断未知账户有无。"""
    from modules.baidu_unknown_accounts import has_unknown_baidu_accounts

    assert has_unknown_baidu_accounts({"unknown_accounts": [{"account_name": "X"}]}) is True
    assert has_unknown_baidu_accounts({"unknown_accounts": []}) is False
    assert has_unknown_baidu_accounts({}) is False


def test_format_notice_includes_metrics():
    """提醒行包含账户名、展现、点击、消费。"""
    from modules.baidu_unknown_accounts import format_unknown_baidu_accounts_notice

    report = {
        "unknown_accounts": [
            {"account_name": "未知A", "展现": 100, "点击": 10, "消费": 50.5},
        ]
    }
    lines = format_unknown_baidu_accounts_notice(report)
    text = " ".join(lines)
    assert "未知A" in text
    assert "100" in text
    assert "10" in text
    assert "50.5" in text
    assert "已单独隔离" in text
    assert "不影响本次写入" in text


def test_format_notice_no_report_path():
    """默认提醒不包含 unknown_baidu_accounts.json 路径。"""
    from modules.baidu_unknown_accounts import format_unknown_baidu_accounts_notice

    report = {
        "unknown_accounts": [{"account_name": "X", "展现": 1, "点击": 0, "消费": 0}],
        "unknown_accounts_report": "reports/unknown_baidu_accounts.json",
    }
    lines = format_unknown_baidu_accounts_notice(report)
    text = " ".join(lines)
    assert "unknown_baidu_accounts.json" not in text


def test_format_notice_empty_returns_empty():
    """unknown_accounts 为空时 format_notice 返回空列表。"""
    from modules.baidu_unknown_accounts import format_unknown_baidu_accounts_notice

    assert format_unknown_baidu_accounts_notice({"unknown_accounts": []}) == []


def test_ignored_unknown_does_not_trigger_notice():
    """ignored_unknown_accounts 非空但 unknown_accounts 为空时不触发提醒。"""
    from modules.baidu_unknown_accounts import has_unknown_baidu_accounts, format_unknown_baidu_accounts_notice

    report = {
        "unknown_accounts": [],
        "ignored_unknown_accounts": [{"account_name": "零账户", "展现": 0, "点击": 0, "消费": 0}],
    }
    assert has_unknown_baidu_accounts(report) is False
    assert format_unknown_baidu_accounts_notice(report) == []


def test_run_pipeline_continues_after_unknown_notice(tmp_path):
    """未知账户提醒不阻断 pipeline 后续步骤。"""
    import logging

    kst_file = tmp_path / "kst.xlsx"
    kst_file.write_text("placeholder", encoding="utf-8")
    excel_file = tmp_path / "target.xlsx"
    excel_file.write_text("placeholder", encoding="utf-8")

    def ok_baidu(**kwargs):
        return {
            "date": "2026-05-07", "period": "15",
            "accounts": {
                "银康01": {"展现": 1, "点击": 1, "消费": 1.0},
                "银康银屑02": {"展现": 2, "点击": 2, "消费": 2.0},
                "银康03": {"展现": 3, "点击": 3, "消费": 3.0},
            },
            "unknown_accounts": [{"account_name": "未知A", "展现": 100, "点击": 10, "消费": 50}],
            "errors": [],
        }

    def ok_kst(export_file, config, root, period):
        return {"parse_report": {"passed": True, "errors": []}, "outputs": {}}

    def ok_merge(**kwargs):
        return {"merged": {"date": "2026-05-07", "period": "15点"}, "validate_report": {"passed": True, "errors": []}, "outputs": {}}

    def ok_write(**kwargs):
        return {
            "date": "2026-05-07", "period": "15点", "excel_path": str(excel_file),
            "writes": [{"cell": "P5"}], "self_check": {"verification_passed": True}, "errors": [],
        }

    report = run_half_auto_pipeline(
        config={"excel_path": str(excel_file)},
        root=tmp_path, logger=logging.getLogger("test"),
        period="15点", kst_file=kst_file, assume_yes=True,
        fetch_baidu_func=ok_baidu, parse_kst_func=ok_kst,
        merge_func=ok_merge, write_func=ok_write,
    )
    # pipeline 应成功完成，不因未知账户中断
    assert report["passed"] is True
    assert report["failed_step"] is None


def test_verbose_shows_report_path(capsys):
    """verbose 模式显示 unknown_baidu_accounts.json 路径。"""
    from modules.baidu_unknown_accounts import print_unknown_baidu_accounts_notice
    from modules.console_ui import set_verbose

    set_verbose(True)
    report = {
        "unknown_accounts": [{"account_name": "X", "展现": 1, "点击": 0, "消费": 0}],
        "unknown_accounts_report": "reports/unknown_baidu_accounts.json",
    }
    print_unknown_baidu_accounts_notice(report)

    captured = capsys.readouterr()
    assert "unknown_baidu_accounts.json" in captured.out
    set_verbose(False)


def test_goto_report_non_noauth_passes(tmp_path):
    """report URL 且非 noauth → _goto_report_page 直接返回 True。"""
    from unittest.mock import MagicMock
    from modules.baidu_overview import _goto_report_page
    import logging
    logger = logging.getLogger("test")
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    assert _goto_report_page(fake_page, logger) is True


def test_goto_report_noauth_relogin_menu_success(monkeypatch):
    """noauth → 重登 → 菜单三步成功 → 最终验证搜索推广通过 → True。"""
    from unittest.mock import MagicMock
    from modules.baidu_overview import _goto_report_page
    import logging
    logger = logging.getLogger("test")

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"
    fake_page.locator("body").inner_text.return_value = "搜索推广"

    # 动态 mock：先 noauth=True 进菜单路径，之后 noauth=False
    results = [True, False]
    def dynamic_noauth(page):
        return results.pop(0) if results else False

    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", dynamic_noauth)
    monkeypatch.setattr("modules.baidu_overview._ensure_baidu_home_rendered",
                        lambda page, config, logger: "")
    monkeypatch.setattr("modules.baidu_overview._click_by_text_or_role",
                        lambda page, labels, logger: "clicked")
    monkeypatch.setattr("modules.baidu_detector.classify_baidu_page",
                        lambda url, text: {"login_status": "logged_in", "signals": {"has_search_promotion": True}})

    result = _goto_report_page(fake_page, logger,
                               root=".", config={"baidu": {"credential_profile": "test"}, "project_id": "p"})
    assert result is True


def test_goto_report_menu_click_fails_returns_false(monkeypatch):
    """菜单路径点击失败 → _goto_report_page 返回 False。"""
    from unittest.mock import MagicMock
    from modules.baidu_overview import _goto_report_page
    import logging
    logger = logging.getLogger("test")

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"

    # 动态 mock：前2次 True 进入 noauth 分支
    results = [True, False]
    def dynamic_noauth(page):
        return results.pop(0) if results else False

    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", dynamic_noauth)
    monkeypatch.setattr("modules.baidu_session.force_relogin_current_project",
                        lambda root, config, page, logger, **kw: True)
    monkeypatch.setattr("modules.baidu_overview._ensure_baidu_home_rendered",
                        lambda page, config, logger: "")
    # 第一个菜单点击返回 None（失败）
    monkeypatch.setattr("modules.baidu_overview._click_by_text_or_role",
                        lambda page, labels, logger: None)

    result = _goto_report_page(fake_page, logger,
                               root=".", config={"baidu": {"credential_profile": "test"}, "project_id": "p"})
    assert result is False, "菜单点击失败应返回 False"


def test_page_usable_false_when_username_mismatch(tmp_path, monkeypatch):
    """_page_is_usable 在 detected_user != expected_user 时返回 False。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import _page_is_usable_search_promotion

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"

    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", lambda page: False)
    monkeypatch.setattr("modules.baidu_detector.classify_baidu_page",
                        lambda url, text: {"login_status": "logged_in", "signals": {"has_search_promotion": True}})
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda root, config: "user_a")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda page: "user_b")

    result = _page_is_usable_search_promotion(fake_page, tmp_path, {})
    assert result is False, "用户名不匹配应返回 False"


def test_page_usable_true_when_username_match(tmp_path, monkeypatch):
    """_page_is_usable 在 detected_user == expected_user 时返回 True。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import _page_is_usable_search_promotion

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"

    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", lambda page: False)
    monkeypatch.setattr("modules.baidu_detector.classify_baidu_page",
                        lambda url, text: {"login_status": "logged_in", "signals": {"has_search_promotion": True}})
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda root, config: "user_a")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda page: "user_a")

    result = _page_is_usable_search_promotion(fake_page, tmp_path, {})
    assert result is True


def test_page_usable_false_when_last_profile_empty_no_detect(tmp_path, monkeypatch):
    """_page_is_usable 在无法识别且 last_profile 为空时返回 False。"""
    from unittest.mock import MagicMock
    from modules.baidu_session import _page_is_usable_search_promotion

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/report"

    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", lambda page: False)
    monkeypatch.setattr("modules.baidu_detector.classify_baidu_page",
                        lambda url, text: {"login_status": "logged_in", "signals": {"has_search_promotion": True}})
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda root, config: None)
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda page: None)
    monkeypatch.setattr("modules.baidu_session.get_browser_login_profile", lambda root: None)
    monkeypatch.setattr("modules.baidu_session.get_current_project_credential_profile", lambda config: "proj_a")

    result = _page_is_usable_search_promotion(fake_page, tmp_path, {})
    assert result is False


def test_logger_error_no_playwright_call_log(capsys):
    """默认终端不输出 Playwright call log。"""
    import logging
    logger = logging.getLogger("test_plog")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    import sys
    handler = logging.StreamHandler(sys.stdout)
    handler.setLevel(logging.WARNING)
    logger.addHandler(handler)

    logger.error("进入百度报告页异常")
    captured = capsys.readouterr()
    assert "Call log:" not in captured.err, "默认不应有 Playwright call log"
    assert "navigating to" not in captured.err

    logger.handlers.clear()


def test_goto_report_homepage_calls_menu_path(monkeypatch):
    """非 report URL（如 homepage）→ 走菜单路径。"""
    from unittest.mock import MagicMock
    from modules.baidu_overview import _goto_report_page
    import logging
    logger = logging.getLogger("test")

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"
    fake_page.locator("body").inner_text.return_value = ""

    click_calls = []
    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", lambda page: False)
    monkeypatch.setattr("modules.baidu_overview._ensure_baidu_home_rendered",
                        lambda page, config, logger: "")
    monkeypatch.setattr("modules.baidu_overview._click_by_text_or_role",
                        lambda page, labels, logger: click_calls.append(labels) or "clicked")
    monkeypatch.setattr("modules.baidu_detector.classify_baidu_page",
                        lambda url, text: {"login_status": "logged_in", "signals": {"has_search_promotion": True}})

    result = _goto_report_page(fake_page, logger, root=".", config={"project_id": "p"})
    assert result is True
    assert len(click_calls) == 3


def test_baidu_prepare_overview_returns_early_on_open_errors():
    """open_report 有 errors 时立即返回，不继续 validate_overview_ready。"""
    root = Path(__file__).resolve().parents[1]
    source = (root / "modules" / "baidu_overview.py").read_text(encoding="utf-8")
    assert "baidu-prepare-overview 中断" in source
    assert "搜索推广页打开失败" in source

def test_tentative_bypass_missing_accounts_triggers_relogin(tmp_path, monkeypatch):
    import logging
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}

    def fake_open(config, root, logger):
        return {"final_url": "https://cc.baidu.com/report", "errors": [], "session_check": {"decision": "tentative_bypass", "passed": True}}

    validate_calls = []
    def fake_validate(visible_text, target_date, config):
        validate_calls.append(1)
        if len(validate_calls) == 1:
            return {"passed": False, "errors": ["页面未看到目标账户：银康01"]}
        return {"passed": True, "errors": []}

    clear_calls = []
    mark_calls = []
    monkeypatch.setattr("modules.baidu_overview.baidu_open_overview", fake_open)
    monkeypatch.setattr("modules.baidu_overview.validate_overview_ready", fake_validate)
    monkeypatch.setattr("modules.baidu_session.clear_browser_login_state", lambda root: clear_calls.append(1))
    monkeypatch.setattr("modules.baidu_session.mark_browser_login_success", lambda root, credential_profile, **kw: mark_calls.append(1))

    from modules.baidu_overview import baidu_prepare_overview
    report = baidu_prepare_overview(config, tmp_path, logger=logging.getLogger("test"))
    assert report.get("validation_retry_triggered") is True
    assert report.get("validation_retry_passed") is True
    assert len(clear_calls) >= 1
    assert len(mark_calls) >= 1
    assert len(validate_calls) == 2
    assert "银康01" not in str(report.get("errors", []))


def test_tentative_bypass_retry_still_fails_error_converges(tmp_path, monkeypatch):
    import logging
    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu", "project_name": "昆明牛"}

    def fake_open(config, root, logger):
        return {"final_url": "https://cc.baidu.com/report", "errors": [], "session_check": {"decision": "tentative_bypass", "passed": True}}

    def fake_validate(visible_text, target_date, config):
        return {"passed": False, "errors": ["页面未看到目标账户：银康01"]}

    mark_calls = []
    monkeypatch.setattr("modules.baidu_overview.baidu_open_overview", fake_open)
    monkeypatch.setattr("modules.baidu_overview.validate_overview_ready", fake_validate)
    monkeypatch.setattr("modules.baidu_session.clear_browser_login_state", lambda root: None)
    monkeypatch.setattr("modules.baidu_session.mark_browser_login_success", lambda root, credential_profile, **kw: mark_calls.append(1))

    from modules.baidu_overview import baidu_prepare_overview
    report = baidu_prepare_overview(config, tmp_path, logger=logging.getLogger("test"))
    assert report.get("validation_retry_triggered") is True
    assert len(report["errors"]) > 0
    assert "未看到当前项目账户" in " ".join(report["errors"])
    assert "银康01" not in " ".join(report["errors"])
    assert len(mark_calls) == 0
def test_force_relogin_retries_logout_once_before_cas(tmp_path, monkeypatch):
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"
    logger = logging.getLogger("test")

    logout_results = [
        {"success": False, "message": "first"},
        {"success": True, "message": "second"},
    ]
    cas_calls = []
    monkeypatch.setattr("modules.baidu_session.logout_baidu_account", lambda page: logout_results.pop(0))
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda page: True)
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is True
    assert len(cas_calls) == 1


def test_force_relogin_double_logout_failure_falls_back_to_cookie_reset(tmp_path, monkeypatch):
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import force_relogin_current_project

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"
    logger = logging.getLogger("test")

    cas_calls = []
    monkeypatch.setattr("modules.baidu_session.logout_baidu_account", lambda page: {"success": False, "message": "nope"})
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda page: True)
    monkeypatch.setattr("modules.baidu_session.wait_until_cas_login_page", lambda page, timeout_ms=5000: False)
    monkeypatch.setattr("modules.baidu_session.goto_baidu_login_page", lambda page: cas_calls.append(1) or {"success": True})
    monkeypatch.setattr("modules.baidu_overview._auto_login_if_needed", lambda p, r, c, l: True)

    result = force_relogin_current_project(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result is True
    assert len(cas_calls) >= 1
    fake_page.context.clear_cookies.assert_called()


def test_session_detected_none_logged_in_uses_tentative_bypass(tmp_path, monkeypatch):
    import logging
    from unittest.mock import MagicMock
    from modules.baidu_session import ensure_baidu_profile_session

    (tmp_path / "reports").mkdir(exist_ok=True)
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"
    logger = logging.getLogger("test")

    relogin_calls = []
    monkeypatch.setattr("modules.baidu_session.get_expected_baidu_username", lambda r, c: "target")
    monkeypatch.setattr("modules.baidu_session.detect_current_baidu_username", lambda p: None)
    monkeypatch.setattr("modules.baidu_session.is_baidu_logged_in", lambda p: True)
    monkeypatch.setattr("modules.baidu_session._page_is_usable_search_promotion", lambda p, r, c: False)
    monkeypatch.setattr("modules.baidu_session.force_relogin_current_project", lambda *args, **kwargs: relogin_calls.append(1) or True)

    result = ensure_baidu_profile_session(tmp_path, config, fake_page, logger, input_func=lambda _: "", output_func=lambda _: None)
    assert result["passed"] is True
    assert result["decision"] == "tentative_bypass"
    assert relogin_calls == []


def test_goto_report_homepage_failure_returns_false_without_parse(monkeypatch):
    from unittest.mock import MagicMock
    from modules.baidu_overview import _goto_report_page
    import logging
    logger = logging.getLogger("test")

    fake_page = MagicMock()
    fake_page.url = "https://cc.baidu.com/homepage"

    monkeypatch.setattr("modules.baidu_overview._ensure_baidu_home_rendered", lambda page, config, logger: "")
    monkeypatch.setattr("modules.baidu_overview._click_by_text_or_role", lambda page, labels, logger: None)
    monkeypatch.setattr("modules.baidu_session.is_baidu_noauth_page", lambda page: False)

    result = _goto_report_page(fake_page, logger, root=".", config={"project_id": "p"})
    assert result is False


def test_baidu_prepare_overview_does_not_write_state_when_validate_fails(tmp_path, monkeypatch):
    import logging

    (tmp_path / "reports").mkdir(exist_ok=True)
    (tmp_path / "reports" / "baidu_visible_text.txt").write_text("homepage", encoding="utf-8")
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}

    monkeypatch.setattr(
        "modules.baidu_overview.baidu_open_overview",
        lambda config, root, logger: {
            "final_url": "https://cc.baidu.com/homepage",
            "errors": [],
            "session_check": {"decision": "bypass", "passed": True},
        },
    )
    monkeypatch.setattr(
        "modules.baidu_overview.validate_overview_ready",
        lambda visible_text, target_date, config: {
            "passed": False,
            "errors": ["百度搜索推广数据页打开失败，请检查百度后台页面状态"],
        },
    )
    mark_calls = []
    monkeypatch.setattr("modules.baidu_session.mark_browser_login_success", lambda root, credential_profile, **kw: mark_calls.append(1))

    from modules.baidu_overview import baidu_prepare_overview
    report = baidu_prepare_overview(config, tmp_path, logger=logging.getLogger("test"))
    assert report["errors"]
    assert mark_calls == []


def test_prepare_baidu_daily_report_page_checks_session_before_single_goto(tmp_path, monkeypatch):
    import logging
    from unittest.mock import MagicMock

    from modules.baidu_daily import _prepare_baidu_daily_report_page

    (tmp_path / "reports").mkdir(exist_ok=True)
    page = MagicMock()
    page.url = "https://cc.baidu.com/homepage"
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    report = {"errors": []}
    calls = []

    def fake_session(root, config, page, logger, task=None):
        calls.append("session")
        return {"passed": True, "decision": "relogin", "reason": "state_unknown"}

    def fake_goto(page, logger, root=None, config=None):
        calls.append("goto")
        page.url = "https://cc.baidu.com/report"
        return True

    monkeypatch.setattr("modules.baidu_session.ensure_baidu_profile_session", fake_session)
    monkeypatch.setattr("modules.baidu_daily._goto_report_page", fake_goto)

    ok = _prepare_baidu_daily_report_page(page, config, tmp_path, logging.getLogger("test"), report)

    assert ok is True
    assert calls == ["session", "goto"]
    assert report["session_check"]["decision"] == "relogin"
    assert report["errors"] == []


def test_prepare_baidu_daily_report_page_retries_once_after_login_redirect(tmp_path, monkeypatch):
    import logging
    from unittest.mock import MagicMock

    from modules.baidu_daily import _prepare_baidu_daily_report_page

    (tmp_path / "reports").mkdir(exist_ok=True)
    page = MagicMock()
    page.url = "https://cc.baidu.com/homepage"
    config = {"baidu": {"credential_profile": "kunming_niu_baidu"}, "project_id": "kunming_niu"}
    report = {"errors": []}
    calls = []

    monkeypatch.setattr(
        "modules.baidu_session.ensure_baidu_profile_session",
        lambda root, config, page, logger, task=None: calls.append("session") or {"passed": True, "decision": "tentative_bypass"},
    )

    def fake_goto(page, logger, root=None, config=None):
        calls.append("goto")
        if calls.count("goto") == 1:
            page.url = "https://cas.baidu.com/?tpl=www2&fromu=https%3A%2F%2Fcc.baidu.com%2Freport"
            return False
        page.url = "https://cc.baidu.com/report"
        return True

    def fake_login(page, root, config, logger):
        calls.append("login")
        page.url = "https://cc.baidu.com/report"
        return True

    monkeypatch.setattr("modules.baidu_daily._goto_report_page", fake_goto)
    monkeypatch.setattr("modules.baidu_daily._auto_login_if_needed", fake_login)

    ok = _prepare_baidu_daily_report_page(page, config, tmp_path, logging.getLogger("test"), report)

    assert ok is True
    assert calls == ["session", "goto", "login", "goto"]
    assert report["errors"] == []


def test_daily_search_promotion_check_does_not_navigate_report_again(monkeypatch):
    import logging
    from unittest.mock import MagicMock

    from modules.baidu_daily import _ensure_search_promotion_before_daily_date

    page = MagicMock()
    page.url = "https://cc.baidu.com/report"
    monkeypatch.setattr("modules.baidu_daily._read_page_text", lambda page: "数据报告 搜索推广 账户 展现 点击 消费")
    monkeypatch.setattr("modules.baidu_daily.classify_baidu_page", lambda url, text: {"page_type": "搜索推广数据页"})
    monkeypatch.setattr("modules.baidu_daily.is_search_promotion_overview", lambda classification: True)
    monkeypatch.setattr(
        "modules.baidu_daily._goto_report_page",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("不应重复打开 report")),
    )

    ok, visible_text = _ensure_search_promotion_before_daily_date(page, {}, logging.getLogger("test"), root=".")

    assert ok is True
    assert "搜索推广" in visible_text


def test_refresh_report_skips_networkidle_when_table_is_already_parseable(monkeypatch):
    import logging

    from modules.baidu_overview import _refresh_report_and_wait_for_data

    class FakePage:
        def __init__(self):
            self.load_state_calls = []
            self.reload_calls = []
            self.wait_calls = []

        def reload(self, **kwargs):
            self.reload_calls.append(kwargs)

        def wait_for_load_state(self, state, **kwargs):
            self.load_state_calls.append((state, kwargs))

        def wait_for_timeout(self, timeout):
            self.wait_calls.append(timeout)

    page = FakePage()
    monkeypatch.setattr("modules.baidu_overview._read_page_text", lambda _page: "账户 展现 点击 消费")
    monkeypatch.setattr(
        "modules.baidu_overview.extract_baidu_rows_from_page",
        lambda _page, _config: {"debug": {"parse_ready": True}},
    )

    text = _refresh_report_and_wait_for_data(
        page,
        {"baidu": {"report_table_wait_seconds": 30}},
        logging.getLogger("test"),
    )

    assert text == "账户 展现 点击 消费"
    assert len(page.reload_calls) == 1
    assert page.load_state_calls == []
    assert page.wait_calls == [300, 300]


def test_extract_baidu_rows_from_page_prefers_dom_table_and_parses_currency_formats():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return """
数据报告
搜索推广
账户
展现
点击
消费
"""

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return [
                {"账户": "竞网CS博润241209", "展现": "1,234", "点击": "56", "消费": "￥1,234.56"},
                {"账户": "竞网CS博润240304", "展现": "2,345", "点击": "--", "消费": "123.45元"},
                {"账户": "竞网CS博润251218", "展现": "-", "点击": "78", "消费": "0"},
                {"账户": "总计-3", "展现": "3,579", "点击": "134", "消费": "1358.01"},
            ]

    config = {
        "project_id": "changsha_niu",
        "accounts": {
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209", "aliases": ["竞网CS博润241209"]},
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304", "aliases": ["竞网CS博润240304"]},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218", "aliases": ["竞网CS博润251218"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)
    parsed = parse_baidu_table(result["rows"], config)

    assert result["extraction_method"] == "dom"
    assert result["detected_headers"] == ["账户", "展现", "点击", "消费"]
    assert result["debug"]["parsed_account_count"] == 3
    assert parsed["accounts"]["竞网CS博润241209"]["点击"] == 56
    assert parsed["accounts"]["竞网CS博润241209"]["消费"] == 1234.56
    assert parsed["accounts"]["竞网CS博润240304"]["点击"] == 0
    assert parsed["accounts"]["竞网CS博润251218"]["展现"] == 0


def test_extract_baidu_rows_from_page_falls_back_to_visible_text_when_dom_unusable():
    from modules.baidu_parser import extract_baidu_rows_from_page

    text = """
数据报告
搜索推广
账户
账户ID
展现
点击
消费
总计-3
-
-
-
-
宁波博润1
1001
1,001
11
￥101.50
宁波博润2
1002
2,002
22
202.25元
宁波博润12
1003
3,003
33
303.75
20条/页
"""

    class FakeLocator:
        def __init__(self, value):
            self._value = value

        def inner_text(self, timeout=None):
            return self._value

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator(text)

        def evaluate(self, script):
            return [{"列1": "未知", "列2": "N/A"}]

    config = {
        "project_id": "ningbo_niu",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)
    parsed = parse_baidu_table(result["rows"], config)

    assert result["extraction_method"] == "visible_text"
    assert set(parsed["accounts"].keys()) == {"宁波博润1", "宁波博润2", "宁波博润12"}
    assert parsed["accounts"]["宁波博润12"]["消费"] == 303.75


def test_parse_baidu_table_reports_raw_value_and_extraction_method_for_non_numeric_fields():
    config = {
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
        }
    }
    rows = [
        {
            "账户": "宁波博润1",
            "展现": "1,234",
            "点击": "N/A",
            "消费": "abc",
            "__source__": "dom",
            "__row_sample_id__": "row-2",
        }
    ]

    parsed = parse_baidu_table(rows, config)

    assert any("raw_value=N/A" in error and "extraction_method=dom" in error and "row_sample_id=row-2" in error for error in parsed["errors"])
    assert any("raw_value=abc" in error and "account_name=宁波博润1" in error for error in parsed["errors"])


def test_baidu_prepare_overview_fails_when_report_table_not_parseable(tmp_path, monkeypatch):
    import logging

    (tmp_path / "reports").mkdir(exist_ok=True)
    (tmp_path / "reports" / "baidu_visible_text.txt").write_text(
        "数据报告 搜索推广 2026/05/07 账户 展现 点击 消费",
        encoding="utf-8",
    )
    (tmp_path / "reports" / "baidu_table_parse_debug.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "required_accounts": ["A", "B", "C"],
                "extraction_method": "dom",
                "detected_headers": ["账户", "展现", "点击", "消费"],
                "parsed_account_count": 0,
                "parsed_accounts": [],
                "missing_accounts": ["A", "B", "C"],
                "non_numeric_fields": [],
                "sample_rows": [],
                "row_cell_count": [4],
                "parse_ready": False,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    config = {
        "project_id": "changsha_niu",
        "project_name": "长沙牛",
        "accounts": {
            "A": {"baidu_name": "A", "aliases": ["A"]},
            "B": {"baidu_name": "B", "aliases": ["B"]},
            "C": {"baidu_name": "C", "aliases": ["C"]},
        },
    }

    monkeypatch.setattr(
        "modules.baidu_overview.baidu_open_overview",
        lambda config, root, logger: {
            "final_url": "https://cc.baidu.com/report",
            "final_page_type": "搜索推广",
            "errors": [],
            "session_check": {"decision": "bypass", "passed": True},
        },
    )

    from modules.baidu_overview import baidu_prepare_overview
    report = baidu_prepare_overview(config, tmp_path, logger=logging.getLogger("test"))

    assert report["errors"]
    assert any("百度搜索推广账户表格未完整加载" in error for error in report["errors"])


def test_fetch_baidu_auto_uses_dom_candidates_from_prepare_artifacts(tmp_path, monkeypatch):
    import logging
    from datetime import date

    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    (reports_dir / "baidu_visible_text.txt").write_text(date.today().strftime("%Y/%m/%d"), encoding="utf-8")
    dom_rows = [
        {"账户": "竞网CS博润241209", "展现": "1,001", "点击": "11", "消费": "101.5", "__source__": "dom", "__row_sample_id__": "dom-row-1"},
        {"账户": "竞网CS博润240304", "展现": "2,002", "点击": "22", "消费": "202.5", "__source__": "dom", "__row_sample_id__": "dom-row-2"},
        {"账户": "竞网CS博润251218", "展现": "3,003", "点击": "33", "消费": "303.5", "__source__": "dom", "__row_sample_id__": "dom-row-3"},
    ]
    (reports_dir / "baidu_table_candidates.json").write_text(
        json.dumps({"source": "dom", "rows": dom_rows, "row_count": 3, "detected_headers": ["账户", "展现", "点击", "消费"]}, ensure_ascii=False),
        encoding="utf-8",
    )
    (reports_dir / "baidu_table_parse_debug.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "required_accounts": ["竞网CS博润241209", "竞网CS博润240304", "竞网CS博润251218"],
                "extraction_method": "dom",
                "detected_headers": ["账户", "展现", "点击", "消费"],
                "parsed_account_count": 3,
                "parsed_accounts": ["竞网CS博润241209", "竞网CS博润240304", "竞网CS博润251218"],
                "missing_accounts": [],
                "non_numeric_fields": [],
                "sample_rows": [{"row_sample_id": "dom-row-1", "source": "dom", "cells": {"账户": "竞网CS博润241209"}}],
                "parse_ready": True,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "modules.baidu_auto.baidu_prepare_overview",
        lambda config, root, logger: {"errors": [], "open_report": {"final_url": "https://cc.baidu.com/report", "final_page_type": "搜索推广"}},
    )
    monkeypatch.setattr(
        "modules.baidu_auto.extract_baidu_rows_from_visible_text",
        lambda text: (_ for _ in ()).throw(AssertionError("不应回退到 visible_text")),
    )

    config = {
        "project_id": "changsha_niu",
        "project_name": "长沙牛",
        "accounts": {
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209", "aliases": ["竞网CS博润241209"]},
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304", "aliases": ["竞网CS博润240304"]},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218", "aliases": ["竞网CS博润251218"]},
        },
    }

    report = fetch_baidu_auto(config, tmp_path, logging.getLogger("test"), "15点")

    assert report["errors"] == []
    assert report["parse_source"] == "dom"
    assert set(report["accounts"].keys()) == {"竞网CS博润241209", "竞网CS博润240304", "竞网CS博润251218"}


def test_baidu_prepare_overview_fails_early_on_visible_text_percent_misalignment(tmp_path, monkeypatch):
    import logging
    from modules.baidu_overview import baidu_prepare_overview

    (tmp_path / "reports").mkdir(exist_ok=True)
    (tmp_path / "reports" / "baidu_visible_text.txt").write_text(
        "数据报告 搜索推广 2026/05/07 账户 展现 点击 消费",
        encoding="utf-8",
    )
    (tmp_path / "reports" / "baidu_table_parse_debug.json").write_text(
        json.dumps(
            {
                "project_id": "ningbo_niu",
                "required_accounts": ["宁波博润1", "宁波博润2", "宁波博润12"],
                "extraction_method": "visible_text",
                "detected_headers": ["账户", "展现", "点击", "消费"],
                "parsed_account_count": 1,
                "parsed_accounts": ["宁波博润1"],
                "missing_accounts": ["宁波博润2", "宁波博润12"],
                "non_numeric_fields": [
                    {
                        "account_name": "宁波博润1",
                        "field": "消费",
                        "raw_value": "8.77%",
                        "extraction_method": "visible_text",
                        "row_sample_id": "visible_text-row-2",
                    }
                ],
                "sample_rows": [
                    {"row_sample_id": "visible_text-row-2", "source": "visible_text", "cells": {"账户": "宁波博润1", "消费": "8.77%"}}
                ],
                "parse_ready": True,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        "modules.baidu_overview.baidu_open_overview",
        lambda config, root, logger: {
            "final_url": "https://cc.baidu.com/report",
            "final_page_type": "搜索推广",
            "errors": [],
            "session_check": {"decision": "bypass", "passed": True},
        },
    )

    config = {
        "project_id": "ningbo_niu",
        "project_name": "宁波牛",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    report = baidu_prepare_overview(config, tmp_path, logger=logging.getLogger("test"))

    assert report["errors"]
    assert "百度搜索推广账户表格未完整加载或列解析异常，请刷新后重试" in report["errors"]
    assert report["parse_check"]["debug"]["missing_accounts"] == ["宁波博润2", "宁波博润12"]


def test_extract_baidu_rows_from_page_supports_grid_rows_for_changsha_accounts():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return "数据报告 搜索推广"

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return [
                {"账户": "竞网CS博润241209", "展现": "1,234", "点击": "56", "消费": "123.45", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-1"},
                {"账户": "竞网CS博润240304", "展现": "2,345", "点击": "67", "消费": "234.56", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-2"},
                {"账户": "竞网CS博润251218", "展现": "3,456", "点击": "78", "消费": "345.67", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-3"},
            ]

    config = {
        "project_id": "changsha_niu",
        "accounts": {
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209", "aliases": ["竞网CS博润241209"]},
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304", "aliases": ["竞网CS博润240304"]},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218", "aliases": ["竞网CS博润251218"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)
    parsed = parse_baidu_table(result["rows"], config)

    assert result["extraction_method"] == "dom"
    assert result["debug"]["missing_accounts"] == []
    assert parsed["errors"] == []


def test_extract_baidu_rows_from_page_supports_grid_rows_for_ningbo_accounts():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return "数据报告 搜索推广"

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return [
                {"账户": "宁波博润1", "展现": "1,111", "点击": "11", "消费": "111.11", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-1"},
                {"账户": "宁波博润2", "展现": "2,222", "点击": "22", "消费": "222.22", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-2"},
                {"账户": "宁波博润12", "展现": "3,333", "点击": "33", "消费": "333.33", "__source__": "dom_grid", "__row_sample_id__": "dom_grid-row-3"},
            ]

    config = {
        "project_id": "ningbo_niu",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)
    parsed = parse_baidu_table(result["rows"], config)

    assert result["extraction_method"] == "dom"
    assert result["debug"]["missing_accounts"] == []
    assert parsed["errors"] == []


def test_extract_baidu_rows_from_page_ignores_total_row_like_header_candidate_and_uses_split_grid_candidate():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return "数据报告 搜索推广"

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return {
                "table_like_found": True,
                "candidates": [
                    {
                        "table_root_selector": "div.summary-grid",
                        "header_source": "row",
                        "header_cells": ["总计-5", "7,908", "3,379.4", "9.57%"],
                        "body_rows": [["总计-5", "7,908", "3,379.4", "9.57%"]],
                        "body_row_count": 1,
                    },
                    {
                        "table_root_selector": "div.report-grid",
                        "header_source": "split_grid",
                        "header_cells": ["账户", "展现", "点击", "消费"],
                        "body_rows": [
                            ["竞网CS博润241209", "1,234", "56", "123.45"],
                            ["竞网CS博润240304", "2,345", "67", "234.56"],
                            ["竞网CS博润251218", "3,456", "78", "345.67"],
                        ],
                        "body_row_count": 3,
                        "scroll_attempts": 1,
                        "accounts_seen_each_scroll": [["竞网CS博润241209", "竞网CS博润240304", "竞网CS博润251218"]],
                        "reached_bottom": False,
                    },
                ],
            }

    config = {
        "project_id": "changsha_niu",
        "accounts": {
            "竞网CS博润241209": {"baidu_name": "竞网CS博润241209", "aliases": ["竞网CS博润241209"]},
            "竞网CS博润240304": {"baidu_name": "竞网CS博润240304", "aliases": ["竞网CS博润240304"]},
            "竞网CS博润251218": {"baidu_name": "竞网CS博润251218", "aliases": ["竞网CS博润251218"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)

    assert result["extraction_method"] == "dom"
    assert result["debug"]["table_root_selector"] == "div.report-grid"
    assert result["debug"]["header_source"] == "split_grid"
    assert result["debug"]["header_cells"] == ["账户", "展现", "点击", "消费"]
    assert result["debug"]["missing_accounts"] == []
    assert any(
        attempt.get("header_cells") == ["总计-5", "7,908", "3,379.4", "9.57%"] and attempt.get("header_valid") is False
        for attempt in result["debug"]["dom_attempts"]
    )


def test_extract_baidu_rows_from_page_supports_scroll_aggregated_accounts_for_ningbo():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return "数据报告 搜索推广"

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return {
                "table_like_found": True,
                "candidates": [
                    {
                        "table_root_selector": "div.virtual-grid",
                        "header_source": "columnheader",
                        "header_cells": ["账户", "展现", "点击", "消费"],
                        "body_rows": [
                            ["宁波博润1", "1,111", "11", "111.11"],
                            ["宁波博润2", "2,222", "22", "222.22"],
                            ["宁波博润12", "3,333", "33", "333.33"],
                        ],
                        "body_row_count": 3,
                        "scroll_attempts": 3,
                        "accounts_seen_each_scroll": [
                            ["宁波博润1"],
                            ["宁波博润1", "宁波博润2"],
                            ["宁波博润1", "宁波博润2", "宁波博润12"],
                        ],
                        "reached_bottom": True,
                    }
                ],
            }

    config = {
        "project_id": "ningbo_niu",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)

    assert result["extraction_method"] == "dom"
    assert result["debug"]["scroll_attempts"] == 3
    assert result["debug"]["accounts_seen_each_scroll"][-1] == ["宁波博润1", "宁波博润2", "宁波博润12"]
    assert result["debug"]["required_accounts_found"] == ["宁波博润1", "宁波博润2", "宁波博润12"]
    assert result["debug"]["missing_accounts"] == []


def test_extract_baidu_rows_from_page_does_not_treat_rate_or_ratio_columns_as_click_or_cost():
    from modules.baidu_parser import extract_baidu_rows_from_page

    class FakeLocator:
        def inner_text(self, timeout=None):
            return "数据报告 搜索推广"

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return {
                "table_like_found": True,
                "candidates": [
                    {
                        "table_root_selector": "div.bad-grid",
                        "header_source": "columnheader",
                        "header_cells": ["账户", "展现", "点击率", "消费占比"],
                        "body_rows": [["宁波博润1", "1,111", "8.86%", "11.08%"]],
                        "body_row_count": 1,
                    }
                ],
            }

    config = {
        "project_id": "ningbo_niu",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)

    assert result["extraction_method"] == "fallback_failed"
    assert result["debug"]["header_valid"] is False
    assert result["debug"]["invalid_header_reason"] in {"missing_required_headers", "invalid_metric_columns"}


def test_extract_baidu_rows_from_page_rejects_visible_text_fallback_when_accounts_incomplete_or_percent_misaligned():
    from modules.baidu_parser import extract_baidu_rows_from_page

    text = """
数据报告
搜索推广
账户
账户ID
展现
点击
消费
总计-3
-
-
-
-
宁波博润1
1001
1,001
8.86%
101.50
宁波博润2
1002
2,002
22
11.08%
20条/页
"""

    class FakeLocator:
        def inner_text(self, timeout=None):
            return text

    class FakePage:
        def locator(self, selector):
            assert selector == "body"
            return FakeLocator()

        def evaluate(self, script):
            return []

    config = {
        "project_id": "ningbo_niu",
        "accounts": {
            "宁波博润1": {"baidu_name": "宁波博润1", "aliases": ["宁波博润1"]},
            "宁波博润2": {"baidu_name": "宁波博润2", "aliases": ["宁波博润2"]},
            "宁波博润12": {"baidu_name": "宁波博润12", "aliases": ["宁波博润12"]},
        },
    }

    result = extract_baidu_rows_from_page(FakePage(), config)

    assert result["extraction_method"] == "fallback_failed"
    assert result["debug"]["extraction_method"] == "fallback_failed"
    assert result["debug"]["percent_misalignment"] is True
    assert "宁波博润12" in result["debug"]["missing_accounts"]


def test_write_table_parse_artifacts_writes_latest_and_timestamped_files(tmp_path):
    from modules.baidu_overview import _write_table_parse_artifacts

    extraction = {
        "extraction_method": "dom",
        "rows": [{"账户": "宁波博润1", "展现": "1,111", "点击": "11", "消费": "111.11"}],
        "detected_headers": ["账户", "展现", "点击", "消费"],
        "debug": {
            "project_id": "ningbo_niu",
            "required_accounts": ["宁波博润1", "宁波博润2", "宁波博润12"],
            "extraction_method": "dom",
            "detected_headers": ["账户", "展现", "点击", "消费"],
            "parsed_account_count": 1,
            "parsed_accounts": ["宁波博润1"],
            "missing_accounts": ["宁波博润2", "宁波博润12"],
            "non_numeric_fields": [],
            "parse_ready": False,
        },
    }

    _write_table_parse_artifacts(tmp_path, extraction)

    reports_dir = tmp_path / "reports"
    assert (reports_dir / "baidu_table_parse_debug.json").exists()
    assert (reports_dir / "baidu_table_parse_debug_latest.json").exists()
    assert len(list(reports_dir.glob("baidu_table_parse_debug_ningbo_niu_*.json"))) == 1
    assert (reports_dir / "baidu_table_candidates.json").exists()
    assert (reports_dir / "baidu_table_candidates_latest.json").exists()
    assert len(list(reports_dir.glob("baidu_table_candidates_ningbo_niu_*.json"))) == 1


def test_doctor_uses_runtime_excel_path_when_project_file_is_stale(tmp_path):
    from openpyxl import Workbook

    configs_dir = tmp_path / "configs"
    projects_dir = configs_dir / "projects"
    projects_dir.mkdir(parents=True)
    (configs_dir / "app_config.json").write_text(
        json.dumps(
            {
                "default_project_id": "changsha_niu",
                "projects_dir": "configs/projects",
                "secrets_file": "secrets/secrets.json",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    stale_path = tmp_path / "missing.xlsx"
    actual_path = tmp_path / "actual.xlsx"
    wb = Workbook()
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.save(actual_path)
    (projects_dir / "changsha_niu.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "project_name": "长沙牛",
                "excel": {"path": str(stale_path), "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
                "kst": {"export_dir": "kst_exports", "auto_pick_latest": True, "max_file_age_hours": 2},
                "baidu": {"credential_profile": "changsha_niu_baidu", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
                "accounts": [
                    {"standard_name": "A", "baidu_names": ["A"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
                    {"standard_name": "B", "baidu_names": ["B"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
                    {"standard_name": "C", "baidu_names": ["C"], "excel_name": "C", "kst_ids": ["3"], "kst_names": ["C"]},
                ],
                "hourly": {"periods": ["11点", "15点", "18点"]},
                "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    config = {
        "project_id": "changsha_niu",
        "project_name": "长沙牛",
        "excel_path": str(actual_path),
        "sheet_name": "时段数据",
        "daily_sheet_name": "百度",
        "excel_engine": "openpyxl",
        "kst": {"export_dir": "kst_exports"},
        "browser": {"cdp_endpoint": "http://127.0.0.1:9", "auto_start_debug_chrome": False},
        "accounts": {"A": {}, "B": {}, "C": {}},
    }

    report = run_doctor(tmp_path, config)

    assert report["checks"]["target_excel"]["passed"] is True
    assert "actual.xlsx" in report["checks"]["target_excel"]["message"]


def test_build_runtime_config_and_doctor_resolve_same_changsha_excel_path(tmp_path):
    from openpyxl import Workbook
    from modules.project_config import load_project_config, build_runtime_config_from_project
    from modules.doctor import _runtime_excel_path

    configs_dir = tmp_path / "configs"
    projects_dir = configs_dir / "projects"
    projects_dir.mkdir(parents=True)
    (configs_dir / "app_config.json").write_text(
        json.dumps(
            {"default_project_id": "changsha_niu", "projects_dir": "configs/projects", "secrets_file": "secrets/secrets.json"},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    excel_path = tmp_path / "长沙" / "【长沙npx】2026竞价数据.xlsx"
    excel_path.parent.mkdir(parents=True)
    wb = Workbook()
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.save(excel_path)
    (projects_dir / "changsha_niu.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "project_name": "长沙牛",
                "excel": {"path": str(excel_path), "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
                "kst": {"export_dir": "kst_exports", "auto_pick_latest": True, "max_file_age_hours": 2},
                "baidu": {"credential_profile": "changsha_niu_baidu", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
                "accounts": [
                    {"standard_name": "A", "baidu_names": ["A"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
                    {"standard_name": "B", "baidu_names": ["B"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
                    {"standard_name": "C", "baidu_names": ["C"], "excel_name": "C", "kst_ids": ["3"], "kst_names": ["C"]},
                ],
                "hourly": {"periods": ["11点", "15点", "18点"]},
                "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    project = load_project_config(tmp_path, "changsha_niu")
    runtime = build_runtime_config_from_project(project, {})
    doctor_path = _runtime_excel_path(tmp_path, runtime, project)

    assert runtime["excel_path"] == str(excel_path)
    assert str(doctor_path) == str(excel_path)


def test_doctor_reports_full_missing_excel_path(tmp_path):
    configs_dir = tmp_path / "configs"
    projects_dir = configs_dir / "projects"
    projects_dir.mkdir(parents=True)
    (configs_dir / "app_config.json").write_text(
        json.dumps(
            {"default_project_id": "changsha_niu", "projects_dir": "configs/projects", "secrets_file": "secrets/secrets.json"},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    missing_path = tmp_path / "missing-dir" / "【长沙npx】2026竞价数据.xlsx"
    (projects_dir / "changsha_niu.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "project_name": "长沙牛",
                "excel": {"path": str(missing_path), "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
                "kst": {"export_dir": "kst_exports", "auto_pick_latest": True, "max_file_age_hours": 2},
                "baidu": {"credential_profile": "changsha_niu_baidu", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
                "accounts": [
                    {"standard_name": "A", "baidu_names": ["A"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
                    {"standard_name": "B", "baidu_names": ["B"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
                    {"standard_name": "C", "baidu_names": ["C"], "excel_name": "C", "kst_ids": ["3"], "kst_names": ["C"]},
                ],
                "hourly": {"periods": ["11点", "15点", "18点"]},
                "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    report = run_doctor(tmp_path, {"browser": {"cdp_endpoint": "http://127.0.0.1:9", "auto_start_debug_chrome": False}, "kst": {}})

    assert report["checks"]["target_excel"]["passed"] is False
    assert str(missing_path) in report["checks"]["target_excel"]["message"]


def test_doctor_suggests_similar_excel_filenames_when_parent_exists(tmp_path):
    from openpyxl import Workbook

    configs_dir = tmp_path / "configs"
    projects_dir = configs_dir / "projects"
    projects_dir.mkdir(parents=True)
    (configs_dir / "app_config.json").write_text(
        json.dumps(
            {"default_project_id": "changsha_niu", "projects_dir": "configs/projects", "secrets_file": "secrets/secrets.json"},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    target_dir = tmp_path / "长沙"
    target_dir.mkdir(parents=True)
    similar_file = target_dir / "【长沙npx】2026竞价数据.xlsx"
    wb = Workbook()
    wb.active.title = "时段数据"
    wb.create_sheet("百度")
    wb.save(similar_file)
    missing_path = target_dir / "【长沙N】2026竞价数据.xlsx"
    (projects_dir / "changsha_niu.json").write_text(
        json.dumps(
            {
                "project_id": "changsha_niu",
                "project_name": "长沙牛",
                "excel": {"path": str(missing_path), "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
                "kst": {"export_dir": "kst_exports", "auto_pick_latest": True, "max_file_age_hours": 2},
                "baidu": {"credential_profile": "changsha_niu_baidu", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
                "accounts": [
                    {"standard_name": "A", "baidu_names": ["A"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
                    {"standard_name": "B", "baidu_names": ["B"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
                    {"standard_name": "C", "baidu_names": ["C"], "excel_name": "C", "kst_ids": ["3"], "kst_names": ["C"]},
                ],
                "hourly": {"periods": ["11点", "15点", "18点"]},
                "daily": {"write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"], "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"]},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    report = run_doctor(tmp_path, {"browser": {"cdp_endpoint": "http://127.0.0.1:9", "auto_start_debug_chrome": False}, "kst": {}})

    detail = report["checks"]["target_excel"]["detail"]
    assert report["checks"]["target_excel"]["passed"] is False
    assert detail["parent_exists"] is True
    assert "【长沙npx】2026竞价数据.xlsx" in detail["similar_files"]


def test_resolve_baidu_sources_wraps_legacy_project_as_single_source():
    from modules.baidu_multi_source import resolve_baidu_sources

    project = {
        "project_id": "legacy",
        "project_name": "旧项目",
        "baidu": {"credential_profile": "legacy_baidu"},
        "accounts": [
            {"standard_name": "A", "baidu_names": ["BA"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
        ],
    }

    sources = resolve_baidu_sources(project)

    assert sources == [
        {
            "source_id": "default",
            "source_name": "旧项目",
            "credential_profile": "legacy_baidu",
            "accounts": project["accounts"],
            "required": True,
        }
    ]


def test_nanjing_niu_config_maps_new_npx6_account_across_sources():
    from modules.project_config import build_runtime_config_from_project, load_project_config, validate_project_config

    project = load_project_config(Path.cwd(), "nanjing_niu")
    runtime = build_runtime_config_from_project(project, {})

    assert validate_project_config(project) == []
    assert list(runtime["accounts"])[-1] == "华厦npx6"
    assert runtime["accounts"]["华厦npx6"]["baidu_names"] == ["baidu-华厦npx6", "华厦npx6"]
    assert runtime["accounts"]["华厦npx6"]["excel_name"] == "baidu-华厦npx6"
    assert runtime["kst"]["promotion_id_accounts"]["85492975"] == "华厦npx6"

    daily_titles = _find_account_titles(
        [
            {
                "row": 1,
                "col": 1,
                "address": "A1",
                "raw_text": "baidu-华厦npx6（85492975）",
                "normalized_text": normalize_text("baidu-华厦npx6（85492975）"),
            }
        ],
        runtime,
    )
    assert daily_titles["华厦npx6"]["found"] is True

    baidu_rows = [
        {"账户": account["baidu_name"], "展现": "1", "点击": "1", "消费": "1"}
        for account in runtime["accounts"].values()
    ]
    baidu_report = parse_baidu_table(baidu_rows, runtime)
    assert baidu_report["errors"] == []
    assert baidu_report["accounts"]["华厦npx6"]["消费"] == 1

    kst_report = aggregate_kst_export_rows(
        [{"备注说明": "推广ID：85492975", "名片标签": "转潜-有效", "访客消息数": "1"}],
        runtime,
    )
    assert kst_report["errors"] == []
    assert kst_report["accounts"]["华厦npx6"]["总对话"] == 1


def test_shenyang_niu_config_resolves_two_baidu_sources():
    from modules.project_config import build_runtime_config_from_project, load_project_config
    from modules.baidu_multi_source import resolve_baidu_sources

    project = load_project_config(Path.cwd(), "shenyang_niu")
    runtime = build_runtime_config_from_project(project, {})
    sources = resolve_baidu_sources(runtime)

    assert project["project_id"] == "shenyang_niu"
    assert [source["source_id"] for source in sources] == ["shenyang_niu_zhongya", "shenyang_niu_yinkang"]
    assert [source["credential_profile"] for source in sources] == [
        "shenyang_niu_zhongya_baidu",
        "shenyang_niu_yinkang_baidu",
    ]
    assert list(runtime["accounts"]) == ["沈阳中亚02", "沈阳银康01", "沈阳中亚01"]
    assert [account["standard_name"] for account in project["excel_accounts"]] == ["沈阳中亚02", "沈阳银康01", "沈阳中亚01"]
    assert [account["standard_name"] for source in sources for account in source["accounts"]] == [
        "沈阳中亚01",
        "沈阳中亚02",
        "沈阳中亚03",
        "沈阳银康01",
        "沈阳银康02",
        "沈阳银康03",
    ]
    assert runtime["kst"]["promotion_id_accounts"]["37084684"] == "沈阳中亚01"
    assert runtime["kst"]["promotion_id_accounts"] == {
        "37084945": "沈阳中亚02",
        "47190166": "沈阳银康01",
        "37084684": "沈阳中亚01",
    }
    assert sources[1]["accounts"][2]["kst_ids"] == ["47190275"]
    assert sources[1]["accounts"][2]["baidu_names"] == ["沈阳银康银屑病3"]


def test_shenyang_bai_config_has_verified_dual_source_mappings_and_example_profiles():
    from modules.project_config import load_project_config

    root = Path(__file__).resolve().parents[1]
    project = load_project_config(root, "shenyang_bai")
    sources = project["baidu_sources"]

    assert [source["source_id"] for source in sources] == [
        "shenyang_bai_source_a",
        "shenyang_bai_source_b",
    ]
    assert [account["standard_name"] for account in project["accounts"]] == [
        "沈阳中亚白癜风1",
        "中亚白癜风3",
        "中亚白癜风5",
        "沈阳银康银屑病6",
    ]
    assert [account["kst_ids"] for account in sources[0]["accounts"]] == [
        ["36607542"],
        ["51639657"],
        ["54566968"],
    ]
    assert [account["kst_ids"] for account in sources[1]["accounts"]] == [["69976114"]]

    example = json.loads((root / "secrets" / "secrets.example.json").read_text(encoding="utf-8"))
    for profile in ["shenyang_bai_source_a_baidu", "shenyang_bai_source_b_baidu"]:
        assert example["baidu"][profile] == {"username": "", "password": ""}


def test_retired_hefei_bai_project_and_example_profiles_are_absent():
    root = Path(__file__).resolve().parents[1]
    data = json.loads((root / "secrets" / "secrets.example.json").read_text(encoding="utf-8"))

    assert not (root / "configs" / "projects" / "hefei_bai.json").exists()
    for profile in ["hefei_bai_huaxia_baidu", "hefei_bai_xinhuaxia_baidu"]:
        assert profile not in data["baidu"]


def test_validate_project_config_rejects_duplicate_baidu_name_inside_same_source():
    from modules.project_config import validate_project_config

    project = {
        "project_id": "bad_multi",
        "project_name": "坏多来源",
        "excel": {"path": "target.xlsx", "hourly_sheet": "时段数据", "daily_sheet": "百度", "engine": "openpyxl"},
        "kst": {"export_dir": "exports", "auto_pick_latest": True, "max_file_age_hours": 2},
        "baidu": {"credential_profile": "unused", "data_path": ["首页", "数据报告", "数据概览", "搜索推广"]},
        "baidu_sources": [
            {
                "source_id": "source_a",
                "source_name": "来源A",
                "credential_profile": "profile_a",
                "accounts": [
                    {"standard_name": "A", "baidu_names": ["重复账户"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
                    {"standard_name": "B", "baidu_names": ["重复账户"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
                ],
            }
        ],
        "accounts": [
            {"standard_name": "A", "baidu_names": ["重复账户"], "excel_name": "A", "kst_ids": ["1"], "kst_names": ["A"]},
            {"standard_name": "B", "baidu_names": ["重复账户"], "excel_name": "B", "kst_ids": ["2"], "kst_names": ["B"]},
        ],
        "hourly": {"periods": ["11点", "15点", "18点"]},
        "daily": {
            "write_fields": ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"],
            "do_not_write_fields": ["总对话", "预约", "到诊", "就诊"],
        },
    }

    errors = validate_project_config(project)

    assert any("source_a 百度账户名重复" in error for error in errors)


def test_aggregate_baidu_source_reports_sums_metrics_and_keeps_source_details():
    from modules.baidu_multi_source import aggregate_baidu_source_reports

    config = {"project_id": "demo", "project_name": "演示", "accounts": {"A": {}, "B": {}}}
    reports = [
        {
            "source_id": "source_a",
            "source_name": "来源A",
            "report": {
                "date": "2026-05-22",
                "period": "15点",
                "accounts": {
                    "A": {"展现": 10, "点击": 1, "消费": 2.5},
                    "B": {"展现": 20, "点击": 2, "消费": 3.5},
                },
                "unknown_accounts": [{"account_name": "未知", "展现": 99, "点击": 9, "消费": 9}],
                "ignored_unknown_accounts": [{"account_name": "空账户", "展现": 0, "点击": 0, "消费": 0}],
                "errors": [],
            },
        },
        {
            "source_id": "source_b",
            "source_name": "来源B",
            "report": {
                "date": "2026-05-22",
                "period": "15点",
                "accounts": {
                    "A": {"展现": 3, "点击": 4, "消费": 5.25},
                },
                "unknown_accounts": [],
                "ignored_unknown_accounts": [],
                "errors": [],
            },
        },
    ]

    report = aggregate_baidu_source_reports(config, reports, period="15点")

    assert report["accounts"]["A"]["展现"] == 13
    assert report["accounts"]["A"]["点击"] == 5
    assert report["accounts"]["A"]["消费"] == 7.75
    assert report["accounts"]["B"]["展现"] == 20
    assert report["unknown_accounts"] == [{"account_name": "未知", "展现": 99, "点击": 9, "消费": 9, "source_id": "source_a", "source_name": "来源A"}]
    assert "未知" not in report["accounts"]
    assert len(report["source_reports"]) == 2


def test_aggregate_baidu_source_reports_outputs_only_excel_accounts_and_classifies_candidate_only_rows():
    from modules.baidu_multi_source import aggregate_baidu_source_reports

    config = {
        "project_id": "shenyang_niu",
        "accounts": {
            "沈阳中亚02": {},
            "沈阳银康01": {},
            "沈阳中亚01": {},
        },
    }
    reports = [
        {
            "source_id": "shenyang_niu_zhongya",
            "source_name": "沈阳中亚",
            "report": {
                "date": "2026-05-22",
                "period": "15点",
                "accounts": {
                    "沈阳中亚01": {"展现": 1, "点击": 2, "消费": 3},
                    "沈阳中亚02": {"展现": 4, "点击": 5, "消费": 6},
                    "沈阳中亚03": {"展现": 0, "点击": 0, "消费": 0},
                },
                "errors": [],
            },
        },
        {
            "source_id": "shenyang_niu_yinkang",
            "source_name": "沈阳银康",
            "report": {
                "date": "2026-05-22",
                "period": "15点",
                "accounts": {
                    "沈阳银康01": {"展现": 7, "点击": 8, "消费": 9},
                    "沈阳银康02": {"展现": 10, "点击": 0, "消费": 0},
                    "沈阳银康03": {"展现": 0, "点击": 0, "消费": 0},
                },
                "errors": [],
            },
        },
    ]

    report = aggregate_baidu_source_reports(config, reports, period="15点")

    assert list(report["accounts"]) == ["沈阳中亚01", "沈阳中亚02", "沈阳银康01"]
    assert set(report["accounts"]) == {"沈阳中亚02", "沈阳银康01", "沈阳中亚01"}
    assert [item["account_name"] for item in report["ignored_inactive_accounts"]] == ["沈阳中亚03", "沈阳银康03"]
    assert [item["account_name"] for item in report["skipped_unmapped_accounts"]] == ["沈阳银康02"]
    assert report["errors"] == []


def test_aggregate_baidu_source_reports_fails_when_any_source_failed():
    from modules.baidu_multi_source import aggregate_baidu_source_reports

    config = {"project_id": "demo", "accounts": {"A": {}}}
    reports = [
        {"source_id": "source_a", "source_name": "来源A", "report": {"accounts": {"A": {"展现": 1, "点击": 1, "消费": 1}}, "errors": []}},
        {"source_id": "source_b", "source_name": "来源B", "report": {"accounts": {}, "errors": ["登录失败"]}},
    ]

    report = aggregate_baidu_source_reports(config, reports, period="15点")

    assert report["accounts"] == {}
    assert any("source_b" in error and "登录失败" in error for error in report["errors"])


def test_doctor_reports_multi_baidu_sources_and_secret_profiles(tmp_path, monkeypatch):
    import modules.doctor as doctor
    from modules.project_config import load_project_config, build_runtime_config_from_project

    configs_dir = tmp_path / "configs"
    projects_dir = configs_dir / "projects"
    secrets_dir = tmp_path / "secrets"
    projects_dir.mkdir(parents=True)
    secrets_dir.mkdir()
    (configs_dir / "app_config.json").write_text(
        json.dumps({"default_project_id": "shenyang_niu", "projects_dir": "configs/projects", "secrets_file": "secrets/secrets.json"}, ensure_ascii=False),
        encoding="utf-8",
    )
    source_project = Path.cwd() / "configs" / "projects" / "shenyang_niu.json"
    (projects_dir / "shenyang_niu.json").write_text(source_project.read_text(encoding="utf-8"), encoding="utf-8")
    (secrets_dir / "secrets.json").write_text(
        json.dumps(
            {
                "baidu": {
                    "shenyang_niu_zhongya_baidu": {"username": "user-a", "password": "pass-a"},
                    "shenyang_niu_yinkang_baidu": {"username": "user-b", "password": "pass-b"},
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(doctor, "_check_cdp", lambda root, config: {"passed": True, "level": "ok", "message": "skip"})
    monkeypatch.setattr(doctor, "_check_chrome", lambda config: {"passed": True, "level": "ok", "message": "skip"})
    monkeypatch.setattr(doctor, "_check_hourly_structure", lambda root, config, excel_path: {"passed": True, "level": "ok", "message": "skip"})

    project = load_project_config(tmp_path, "shenyang_niu")
    config = build_runtime_config_from_project(project, {"browser": {"auto_start_debug_chrome": False}})
    report = run_doctor(tmp_path, config)

    baidu_sources = report["checks"]["baidu_sources"]
    secrets = report["checks"]["secrets_json"]
    assert baidu_sources["passed"] is True
    assert baidu_sources["detail"]["source_count"] == 2
    assert secrets["passed"] is True
    assert "user-a" not in json.dumps(report, ensure_ascii=False)
    assert "pass-a" not in json.dumps(report, ensure_ascii=False)


def test_doctor_baidu_source_detail_marks_candidate_only_accounts_for_shenyang():
    from modules.project_config import load_project_config
    from modules.doctor import _check_baidu_sources

    project = load_project_config(Path.cwd(), "shenyang_niu")
    report = _check_baidu_sources(project)

    assert report["passed"] is True
    assert report["detail"]["excel_accounts"] == ["沈阳中亚02", "沈阳银康01", "沈阳中亚01"]
    assert report["detail"]["candidate_only_accounts"] == ["沈阳中亚03", "沈阳银康02", "沈阳银康03"]


@pytest.mark.parametrize(
    ("task", "config_key", "filename", "output_key"),
    [
        ("hourly", "output_path", "staged_hourly.json", "account_data"),
        ("daily", "daily_output_path", "staged_daily.json", "daily_data"),
    ],
)
def test_baidu_multi_source_respects_configured_unified_output_path(
    tmp_path, task, config_key, filename, output_key
):
    import logging
    from copy import deepcopy

    from modules.baidu_multi_source import fetch_baidu_multi_source

    config = deepcopy(_two_source_route_config())
    config["baidu"][config_key] = f"reports/{filename}"

    def fake_fetch(*, config, **_kwargs):
        account = next(iter(config["accounts"]))
        return {
            "date": "2026-07-16",
            "accounts": {account: {"展现": 1, "点击": 1, "消费": 1.0}},
            "errors": [],
        }

    report = fetch_baidu_multi_source(
        config,
        tmp_path,
        logging.getLogger("configured-multi-output"),
        period="18点",
        fetch_source_func=fake_fetch,
        task=task,
        target_date="2026-07-16" if task == "daily" else None,
    )

    expected = tmp_path / "reports" / filename
    canonical_name = "baidu_daily_data.json" if task == "daily" else "baidu_account_data.json"
    assert report["errors"] == []
    assert expected.exists()
    assert report["outputs"][output_key] == str(expected)
    assert not (tmp_path / "reports" / canonical_name).exists()


def test_baidu_multi_source_writes_human_readable_markdown_report(tmp_path):
    import logging

    from modules.baidu_multi_source import fetch_baidu_multi_source

    config = {
        "project_id": "shenyang_niu",
        "project_name": "沈阳牛",
        "accounts": {"沈阳中亚02": {}, "沈阳银康01": {}},
        "baidu_sources": [
            {
                "source_id": "zhongya",
                "source_name": "沈阳中亚",
                "credential_profile": "secret_profile_a",
                "accounts": [{"standard_name": "沈阳中亚02", "baidu_names": ["百度中亚02"]}, {"standard_name": "沈阳中亚03", "baidu_names": ["百度中亚03"]}],
            },
            {
                "source_id": "yinkang",
                "source_name": "沈阳银康",
                "credential_profile": "secret_profile_b",
                "accounts": [{"standard_name": "沈阳银康01", "baidu_names": ["百度银康01"]}, {"standard_name": "沈阳银康02", "baidu_names": ["百度银康02"]}],
            },
        ],
    }

    def fake_fetch(*, config, root, logger, period):
        if config["baidu_source"]["source_id"] == "zhongya":
            return {
                "date": "2026-05-24",
                "accounts": {"沈阳中亚02": {"展现": 10, "点击": 2, "消费": 5.5}, "沈阳中亚03": {"展现": 0, "点击": 0, "消费": 0}},
                "unknown_accounts": [{"account_name": "未知A", "展现": 1, "点击": 0, "消费": 0, "reason": "未配置"}],
                "ignored_unknown_accounts": [],
                "errors": [],
            }
        return {
            "date": "2026-05-24",
            "accounts": {"沈阳银康01": {"展现": 20, "点击": 3, "消费": 6.5}, "沈阳银康02": {"展现": 1, "点击": 1, "消费": 1}},
            "unknown_accounts": [],
            "ignored_unknown_accounts": [{"account_name": "空账户", "展现": 0, "点击": 0, "消费": 0, "reason": "已忽略"}],
            "errors": [],
        }

    report = fetch_baidu_multi_source(config, tmp_path, logging.getLogger("test"), "15点", fake_fetch)
    md_path = tmp_path / "reports" / "baidu_multi_source_report.md"
    markdown = md_path.read_text(encoding="utf-8")

    assert report["errors"] == []
    assert md_path.exists()
    assert "沈阳牛" in markdown
    assert "日期：2026-05-24" in markdown
    assert "时段：15点" in markdown
    assert "## 来源摘要" in markdown
    assert "## 最终写入账户" in markdown
    assert "## 被忽略的未启用账户" in markdown
    assert "## 被跳过的未映射账户" in markdown
    assert "## unknown_accounts" in markdown
    assert "username" not in markdown.lower()
    assert "password" not in markdown.lower()
    assert "secret_profile" not in markdown


def test_baidu_multi_source_cost_validation_reports_matching_totals_and_source_unknown_details():
    from modules.baidu_multi_source import aggregate_baidu_source_reports

    config = {"accounts": {"A": {}, "B": {}}}
    source_reports = [
        {
            "source_id": "s1",
            "source_name": "来源1",
            "report": {
                "accounts": {"A": {"展现": 1, "点击": 1, "消费": 1.1}},
                "unknown_accounts": [{"account_name": "X", "展现": 1, "点击": 0, "消费": 2}],
                "ignored_unknown_accounts": [{"account_name": "Z", "展现": 0, "点击": 0, "消费": 0}],
                "errors": [],
            },
        },
        {"source_id": "s2", "source_name": "来源2", "report": {"accounts": {"B": {"展现": 2, "点击": 1, "消费": 2.2}}, "errors": []}},
    ]

    report = aggregate_baidu_source_reports(config, source_reports)

    assert report["source_total_cost_sum"] == 3.3
    assert report["final_total_cost"] == 3.3
    assert report["diff"] == 0
    assert report["errors"] == []
    assert report["source_reports"][0]["report"]["unknown_accounts"][0]["source_id"] == "s1"
    assert report["source_reports"][0]["report"]["ignored_unknown_accounts"][0]["source_name"] == "来源1"


def test_baidu_multi_source_cost_validation_fails_when_totals_differ():
    from modules.baidu_multi_source import build_cost_validation

    validation = build_cost_validation(10.2, 10.0)

    assert validation["passed"] is False
    assert validation["diff"] == 0.2
    assert validation["errors"]


def test_multi_source_candidate_accounts_may_be_absent_from_source_page():
    from modules.baidu_multi_source import build_source_runtime_config
    from modules.baidu_parser import parse_baidu_table

    config = {"accounts": {"写入A": {}}}
    source = {
        "source_id": "source_a",
        "source_name": "来源A",
        "credential_profile": "profile_a",
        "accounts": [
            {"standard_name": "写入A", "baidu_names": ["百度A"]},
            {"standard_name": "候选B", "baidu_names": ["百度B"]},
        ],
    }
    source_config = build_source_runtime_config(config, source)
    parsed = parse_baidu_table([{"账户": "百度A", "展现": "1", "点击": "1", "消费": "1"}], source_config)

    assert source_config["baidu"]["allow_missing_candidate_accounts"] is True
    assert parsed["errors"] == []
    assert list(parsed["accounts"]) == ["写入A"]


def test_multi_source_overview_allows_absent_candidate_accounts():
    from modules.baidu_overview import validate_overview_ready

    report = validate_overview_ready(
        "数据报告 搜索推广 2026/05/24 账户 展现 点击 消费 百度A",
        "2026-05-24",
        {
            "baidu": {"allow_missing_candidate_accounts": True},
            "accounts": {
                "写入A": {"baidu_name": "百度A"},
                "候选B": {"baidu_name": "百度B"},
            },
        },
    )

    assert report["passed"] is True
    assert report["accounts"]["候选B"] is False
    assert report["allow_missing_candidate_accounts"] is True


def test_doctor_multi_source_detail_includes_summary_counts_and_candidate_note():
    from modules.project_config import load_project_config
    from modules.doctor import _check_baidu_sources

    project = load_project_config(Path.cwd(), "shenyang_niu")
    report = _check_baidu_sources(project)
    detail = report["detail"]

    assert detail["baidu_sources_count"] == 2
    assert detail["excel_accounts_count"] == 3
    assert detail["baidu_candidate_accounts_count"] == 6
    assert "不算错误" in detail["candidate_only_note"]
    assert all("candidate_accounts_count" in source for source in detail["sources"])
    assert all("has_duplicate_baidu_name" in source for source in detail["sources"])


def test_inspect_excel_structure_writes_excel_account_regions_report(tmp_path):
    import logging

    from openpyxl import Workbook
    from modules.excel_inspector import inspect_excel_structure

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A1"] = "日期"
    ws["B1"] = "时段"
    ws["D2"] = "项目甲"
    headers = ["展现", "点击", "消费"]
    for index, header in enumerate(headers, start=4):
        ws.cell(row=3, column=index).value = header
    excel_path = tmp_path / "demo.xlsx"
    wb.save(excel_path)
    config = {
        "project_id": "demo",
        "project_name": "演示",
        "excel_path": str(excel_path),
        "sheet_name": "时段数据",
        "accounts": {"项目甲": {"aliases": ["项目甲"], "excel_name": "项目甲", "baidu_name": "项目甲"}},
        "field_aliases": {"日期": ["日期"], "时段": ["时段"], "展现": ["展现"], "点击": ["点击"], "消费": ["消费"]},
    }

    inspect_excel_structure(config, tmp_path, logging.getLogger("test"))
    report = json.loads((tmp_path / "reports" / "excel_account_regions.json").read_text(encoding="utf-8"))

    assert report["configured_excel_accounts"] == ["项目甲"]
    assert report["missing_configured_accounts"] == []
    assert report["detected_account_regions"] == [{"account_name": "项目甲", "row": 2, "column": 4, "title_cell": "D2"}]


def test_fetch_baidu_daily_single_source_keeps_existing_single_fetch_path(tmp_path, monkeypatch):
    import logging
    import modules.baidu_daily as daily

    seen = []

    def fake_single(**kwargs):
        seen.append(kwargs["target_date"])
        return {"date": kwargs["target_date"], "accounts": {"A": {}}, "errors": []}

    monkeypatch.setattr(daily, "_fetch_baidu_daily_single", fake_single)
    report = daily.fetch_baidu_daily(
        config={"project_id": "legacy", "accounts": {"A": {}}},
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date="2026-05-23",
    )

    assert seen == ["2026-05-23"]
    assert report["date"] == "2026-05-23"


def test_build_baidu_daily_report_allows_missing_multi_source_candidate_account(monkeypatch):
    import modules.baidu_daily as daily

    monkeypatch.setattr(
        daily,
        "extract_baidu_rows_from_visible_text",
        lambda text: [{"账户": "百度A", "展现": "1", "点击": "1", "消费": "1"}],
    )
    monkeypatch.setattr(daily, "_extract_selected_date_from_text", lambda text: "2026-05-23")

    report = daily.build_baidu_daily_report_from_visible_text(
        "数据报告 搜索推广",
        {
            "baidu": {"allow_missing_candidate_accounts": True},
            "accounts": {"写入A": {"baidu_name": "百度A"}, "候选B": {"baidu_name": "百度B"}},
        },
        "2026-05-23",
    )

    assert report["accounts"]["写入A"]["消费"] == 1
    assert report["errors"] == []


def test_daily_baidu_snapshot_rejects_total_row_mismatch():
    import modules.baidu_daily as daily

    report = {
        "accounts": {
            "A": {"展现": 10, "点击": 1, "消费": 1.0},
            "B": {"展现": 20, "点击": 2, "消费": 2.0},
            "C": {"展现": 30, "点击": 3, "消费": 3.0},
        }
    }
    rows = [
        {"账户": "总计-3", "展现": "60", "点击": "6", "消费": "100"},
        {"账户": "A", "展现": "10", "点击": "1", "消费": "1"},
        {"账户": "B", "展现": "20", "点击": "2", "消费": "2"},
        {"账户": "C", "展现": "30", "点击": "3", "消费": "3"},
    ]

    result = daily.validate_daily_baidu_snapshot(report, rows, {"accounts": {"A": {}, "B": {}, "C": {}}})

    assert result["passed"] is False
    assert any("总计校验失败" in error for error in result["errors"])


def test_wait_stable_daily_report_ignores_first_unstable_snapshot(monkeypatch):
    import logging
    import modules.baidu_daily as daily

    class FakePage:
        def __init__(self):
            self.read_index = 0
            self.waits = []

        def wait_for_timeout(self, timeout):
            self.waits.append(timeout)

    page = FakePage()
    texts = ["low", "correct", "correct"]

    def fake_read(_page):
        value = texts[min(_page.read_index, len(texts) - 1)]
        _page.read_index += 1
        return value

    def fake_report(text, config, target_date, visible_text_path=None):
        cost = 1.0 if text == "low" else 10.0
        return {
            "date": target_date,
            "target_date": target_date,
            "accounts": {"A": {"展现": 1, "点击": 1, "消费": cost}},
            "errors": [],
        }

    def fake_validate(report, rows, config):
        return {
            "passed": True,
            "errors": [],
            "signature": daily._daily_report_signature(report, ["A"]),
            "total_diff": {},
        }

    monkeypatch.setattr(daily, "_read_page_text", fake_read)
    monkeypatch.setattr(daily, "extract_baidu_rows_from_visible_text", lambda text: [])
    monkeypatch.setattr(daily, "build_baidu_daily_report_from_visible_text", fake_report)
    monkeypatch.setattr(daily, "validate_daily_baidu_snapshot", fake_validate)

    snapshot = daily._wait_stable_daily_report_snapshot(
        page,
        {"accounts": {"A": {}}, "baidu": {"report_table_wait_seconds": 10, "daily_stability_interval_ms": 1}},
        "2026-05-23",
        logging.getLogger("test"),
    )

    assert snapshot["stable"] is True
    assert snapshot["report"]["accounts"]["A"]["消费"] == 10.0
    assert len(snapshot["attempts"]) == 3


def test_wait_stable_daily_report_accepts_repeated_valid_snapshot_even_when_not_consecutive(monkeypatch):
    import logging
    import modules.baidu_daily as daily

    class FakePage:
        def __init__(self):
            self.read_index = 0
            self.waits = []

        def wait_for_timeout(self, timeout):
            self.waits.append(timeout)

    page = FakePage()
    texts = ["correct-a", "correct-b", "correct-a"]

    def fake_read(_page):
        value = texts[min(_page.read_index, len(texts) - 1)]
        _page.read_index += 1
        return value

    def fake_report(text, config, target_date, visible_text_path=None):
        cost = 10.0 if text == "correct-a" else 10.01
        return {
            "date": target_date,
            "target_date": target_date,
            "accounts": {"A": {"展现": 1, "点击": 1, "消费": cost}},
            "errors": [],
        }

    def fake_validate(report, rows, config):
        return {
            "passed": True,
            "errors": [],
            "signature": daily._daily_report_signature(report, ["A"]),
            "total_diff": {},
        }

    monkeypatch.setattr(daily, "_read_page_text", fake_read)
    monkeypatch.setattr(daily, "extract_baidu_rows_from_visible_text", lambda text: [])
    monkeypatch.setattr(daily, "build_baidu_daily_report_from_visible_text", fake_report)
    monkeypatch.setattr(daily, "validate_daily_baidu_snapshot", fake_validate)

    snapshot = daily._wait_stable_daily_report_snapshot(
        page,
        {"accounts": {"A": {}}, "baidu": {"report_table_wait_seconds": 10, "daily_stability_interval_ms": 1}},
        "2026-05-23",
        logging.getLogger("test"),
    )

    assert snapshot["stable"] is True
    assert snapshot["report"]["accounts"]["A"]["消费"] == 10.0
    assert len(snapshot["attempts"]) == 3


def test_fetch_baidu_daily_multi_source_uses_shared_aggregation_and_is_merge_compatible(tmp_path, monkeypatch):
    import logging

    import modules.baidu_daily as daily
    from modules.data_merger import merge_daily_files

    config = {
        "project_id": "shenyang_niu",
        "project_name": "沈阳牛",
        "accounts": {"写入A": {}, "写入B": {}},
        "baidu_sources": [
            {
                "source_id": "source_a",
                "source_name": "来源A",
                "credential_profile": "secret_a",
                "accounts": [{"standard_name": "写入A"}, {"standard_name": "候选零"}],
            },
            {
                "source_id": "source_b",
                "source_name": "来源B",
                "credential_profile": "secret_b",
                "accounts": [{"standard_name": "写入B"}, {"standard_name": "候选有量"}],
            },
        ],
    }

    def fake_daily_single(*, config, root, logger, target_date):
        assert target_date == "2026-05-23"
        assert config["baidu"]["daily_output_path"].endswith(f"{config['baidu_source']['source_id']}.json")
        if config["baidu_source"]["source_id"] == "source_a":
            return {
                "date": target_date,
                "accounts": {
                    "写入A": {"展现": 10, "点击": 2, "消费": 3.5},
                    "候选零": {"展现": 0, "点击": 0, "消费": 0},
                },
                "unknown_accounts": [{"account_name": "未知账户", "展现": 1, "点击": 0, "消费": 0}],
                "ignored_unknown_accounts": [],
                "errors": [],
            }
        return {
            "date": target_date,
            "accounts": {
                "写入B": {"展现": 20, "点击": 4, "消费": 6.5},
                "候选有量": {"展现": 1, "点击": 1, "消费": 1},
            },
            "unknown_accounts": [],
            "ignored_unknown_accounts": [{"account_name": "空账户", "展现": 0, "点击": 0, "消费": 0}],
            "errors": [],
        }

    monkeypatch.setattr(daily, "_fetch_baidu_daily_single", fake_daily_single)
    report = daily.fetch_baidu_daily(config, tmp_path, logging.getLogger("test"), "2026-05-23")

    assert report["task"] == "daily"
    assert report["target_date"] == "2026-05-23"
    assert report["period"] is None
    assert report["multi_source"] is True
    assert report["accounts"] == {
        "写入A": {"展现": 10, "点击": 2, "消费": 3.5},
        "写入B": {"展现": 20, "点击": 4, "消费": 6.5},
    }
    assert [row["account_name"] for row in report["ignored_inactive_accounts"]] == ["候选零"]
    assert [row["account_name"] for row in report["skipped_unmapped_accounts"]] == ["候选有量"]
    assert report["unknown_accounts"][0]["source_id"] == "source_a"
    assert report["ignored_unknown_accounts"][0]["source_name"] == "来源B"
    assert report["source_total_cost_sum"] == 10
    assert report["final_total_cost"] == 10
    assert report["errors"] == []
    assert (tmp_path / "reports" / "baidu_daily_data.json").exists()
    assert (tmp_path / "reports" / "baidu_daily_validate_report.json").exists()
    markdown = (tmp_path / "reports" / "baidu_multi_source_report.md").read_text(encoding="utf-8")
    assert "日报" in markdown
    assert "日期：2026-05-23" in markdown
    assert "时段：无" in markdown
    assert "secret_a" not in markdown

    reports = tmp_path / "reports"
    (reports / "kst_daily_data.json").write_text(
        json.dumps({
            "date": "2026-05-23",
            "accounts": {
                "写入A": {"总对话": 1, "有效对话": 1, "无效对话": 0, "一般有效对话": 0, "有效转潜": 1, "总转潜": 1},
                "写入B": {"总对话": 0, "有效对话": 0, "无效对话": 0, "一般有效对话": 0, "有效转潜": 0, "总转潜": 0},
            },
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    merged = merge_daily_files(config, tmp_path, logging.getLogger("test"), "2026-05-23")
    assert merged["validate_report"]["passed"] is True
    assert merged["merged"]["accounts"]["写入A"]["展现"] == 10


def test_fetch_baidu_daily_multi_source_fails_when_any_source_fails(tmp_path, monkeypatch):
    import logging

    import modules.baidu_daily as daily

    config = {
        "accounts": {"A": {}},
        "baidu_sources": [
            {"source_id": "source_a", "source_name": "来源A", "accounts": [{"standard_name": "A"}]},
            {"source_id": "source_b", "source_name": "来源B", "accounts": [{"standard_name": "B"}]},
        ],
    }

    def fake_daily_single(*, config, root, logger, target_date):
        if config["baidu_source"]["source_id"] == "source_b":
            return {"date": target_date, "accounts": {}, "errors": ["登录失败"]}
        return {"date": target_date, "accounts": {"A": {"展现": 1, "点击": 1, "消费": 1}}, "errors": []}

    monkeypatch.setattr(daily, "_fetch_baidu_daily_single", fake_daily_single)
    report = daily.fetch_baidu_daily(config, tmp_path, logging.getLogger("test"), "2026-05-23")

    assert report["accounts"] == {}
    assert any("source_b" in error and "登录失败" in error for error in report["errors"])
    assert report["self_check"]["all_sources_passed"] is False


def test_run_daily_pipeline_stops_when_multi_source_baidu_step_fails(tmp_path):
    import logging

    export = tmp_path / "daily.xlsx"
    export.write_text("placeholder", encoding="utf-8")

    def bad_baidu(**kwargs):
        return {"date": "2026-05-23", "multi_source": True, "errors": ["source_b: 登录失败"]}

    def should_not_parse(*args, **kwargs):
        raise AssertionError("百度多来源失败后不应解析商务通日报")

    report = run_daily_pipeline(
        config={"excel_path": "target.xlsx"},
        root=tmp_path,
        logger=logging.getLogger("test"),
        target_date="2026-05-23",
        kst_file=export,
        fetch_baidu_func=bad_baidu,
        parse_kst_func=should_not_parse,
    )

    assert report["passed"] is False
    assert report["failed_step"] == "fetch-baidu-daily"
    assert report["errors"] == ["source_b: 登录失败"]


def test_write_merged_hourly_data_finds_date_and_period_above_account_titles(tmp_path):
    import logging
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A1"] = "日期"
    ws["B1"] = "时段"
    ws["F2"] = "华厦npx1"
    ws.merge_cells("F2:M2")
    ws["N2"] = "华厦npx3"
    ws.merge_cells("N2:U2")
    ws["V2"] = "华厦npx5"
    ws.merge_cells("V2:AC2")
    headers = ["展现", "点击", "消费", "总对话", "有效对话", "一般有效", "有效转潜", "总转潜"]
    for offset, header in enumerate(headers):
        ws.cell(row=3, column=6 + offset).value = header
        ws.cell(row=3, column=14 + offset).value = header
        ws.cell(row=3, column=22 + offset).value = header
    ws["A4"] = "2026-05-07"
    ws.merge_cells("A4:A6")
    ws["B4"] = "11点"
    ws["B5"] = "3点"
    ws["B6"] = "6点"
    excel_path = tmp_path / "nanjing.xlsx"
    wb.save(excel_path)

    reports = tmp_path / "reports"
    reports.mkdir()
    (reports / "merged_hourly_data.json").write_text(
        """
{
  "date": "2026-05-07",
  "period": "15点",
  "accounts": {
    "华厦npx1": {"展现": 101, "点击": 11, "消费": 12.5, "总对话": 3, "有效对话": 2, "一般有效": 0, "有效转潜": 1, "总转潜": 1},
    "华厦npx3": {"展现": 202, "点击": 22, "消费": 23.5, "总对话": 4, "有效对话": 2, "一般有效": 1, "有效转潜": 1, "总转潜": 1},
    "华厦npx5": {"展现": 303, "点击": 33, "消费": 34.5, "总对话": 5, "有效对话": 3, "一般有效": 1, "有效转潜": 1, "总转潜": 2}
  }
}
""",
        encoding="utf-8",
    )

    config = {
        "excel_path": str(excel_path),
        "sheet_name": "时段数据",
        "accounts": {
            "华厦npx1": {"aliases": ["华厦npx1"], "excel_name": "华厦npx1", "baidu_name": "华厦npx1"},
            "华厦npx3": {"aliases": ["华厦npx3"], "excel_name": "华厦npx3", "baidu_name": "华厦npx3"},
            "华厦npx5": {"aliases": ["华厦npx5"], "excel_name": "华厦npx5", "baidu_name": "华厦npx5"},
        },
        "field_aliases": {
            "日期": ["日期"],
            "时段": ["时段"],
            "展现": ["展现"],
            "点击": ["点击"],
            "消费": ["消费"],
            "总对话": ["总对话"],
            "有效对话": ["有效对话"],
            "一般有效": ["一般有效"],
            "有效转潜": ["有效转潜"],
            "总转潜": ["总转潜"]
        },
    }

    report = write_merged_hourly_data(config, tmp_path, logging.getLogger("test"), "15点")

    assert report["errors"] == []
    assert report["self_check"]["verification_passed"] is True


def test_inspector_and_writer_share_global_date_period_field_locations(tmp_path):
    import logging
    from openpyxl import Workbook, load_workbook
    from modules.excel_inspector import inspect_excel_structure

    wb = Workbook()
    ws = wb.active
    ws.title = "时段数据"
    ws["A1"] = "每日时段统计数据"
    ws["B3"] = "日期"
    ws["C3"] = "时段"
    ws["F5"] = "南京账户1"
    ws.merge_cells("F5:M5")
    ws["N5"] = "南京账户2"
    ws.merge_cells("N5:U5")
    ws["V5"] = "南京账户3"
    ws.merge_cells("V5:AC5")
    headers = ["展现", "点击", "消费", "总对话", "有效对话", "一般有效", "有效转潜", "总转潜"]
    for offset, header in enumerate(headers):
        ws.cell(row=6, column=6 + offset).value = header
        ws.cell(row=6, column=14 + offset).value = header
        ws.cell(row=6, column=22 + offset).value = header
    ws["B7"] = "2026-05-07"
    ws.merge_cells("B7:B9")
    ws["C7"] = "11点"
    ws["C8"] = "3点"
    ws["C9"] = "6点"
    excel_path = tmp_path / "nanjing_global.xlsx"
    wb.save(excel_path)

    reports = tmp_path / "reports"
    reports.mkdir()
    (reports / "merged_hourly_data.json").write_text(
        json.dumps(
            {
                "date": "2026-05-07",
                "period": "15点",
                "accounts": {
                    "南京账户1": {"展现": 101, "点击": 11, "消费": 12.5, "总对话": 3, "有效对话": 2, "一般有效": 0, "有效转潜": 1, "总转潜": 1},
                    "南京账户2": {"展现": 202, "点击": 22, "消费": 23.5, "总对话": 4, "有效对话": 3, "一般有效": 1, "有效转潜": 1, "总转潜": 1},
                    "南京账户3": {"展现": 303, "点击": 33, "消费": 34.5, "总对话": 5, "有效对话": 4, "一般有效": 0, "有效转潜": 2, "总转潜": 2},
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    config = {
        "excel_path": str(excel_path),
        "sheet_name": "时段数据",
        "accounts": {
            "南京账户1": {"aliases": ["南京账户1"], "excel_name": "南京账户1", "baidu_name": "南京账户1"},
            "南京账户2": {"aliases": ["南京账户2"], "excel_name": "南京账户2", "baidu_name": "南京账户2"},
            "南京账户3": {"aliases": ["南京账户3"], "excel_name": "南京账户3", "baidu_name": "南京账户3"},
        },
        "field_aliases": {
            "日期": ["日期"],
            "时段": ["时段"],
            "展现": ["展现"],
            "点击": ["点击"],
            "消费": ["消费"],
            "总对话": ["总对话"],
            "有效对话": ["有效对话"],
            "一般有效": ["一般有效"],
            "有效转潜": ["有效转潜"],
            "总转潜": ["总转潜"],
        },
    }

    structure = inspect_excel_structure(config=config, root=tmp_path, logger=logging.getLogger("test"))
    assert structure["errors"] == []
    assert structure["global_fields"]["日期"]["header_cell"] == "B3"
    assert structure["global_fields"]["时段"]["header_cell"] == "C3"

    report = write_merged_hourly_data(config, tmp_path, logging.getLogger("test"), "15点")

    assert report["errors"] == []
    assert report["self_check"]["verification_passed"] is True

    verify_wb = load_workbook(excel_path, data_only=False, read_only=False)
    verify_ws = verify_wb["时段数据"]
    assert verify_ws["F8"].value == 101
    assert verify_ws["N8"].value == 202
    assert verify_ws["V8"].value == 303


_API_READINESS_SINGLE_PROFILES = {
    "changsha_niu": "changsha_niu_baidu",
    "kunming_niu": "kunming_niu_baidu",
    "nanjing_bai": "nanjing_bai_baidu",
    "nanjing_niu": "nanjing_niu_baidu",
    "ningbo_niu": "ningbo_niu_baidu",
    "qingdao_bai": "qingdao_bai_baidu",
    "shenzhen_bai": "shenzhen_bai_baidu",
}
_API_READINESS_MULTI_PROFILES = {
    "shenyang_bai": [
        ("source_a", "沈阳白来源A", "shenyang_bai_source_a_baidu"),
        ("source_b", "沈阳白来源B", "shenyang_bai_source_b_baidu"),
    ],
    "shenyang_niu": [
        ("zhongya", "沈阳牛中亚来源", "shenyang_niu_zhongya_baidu"),
        ("yinkang", "沈阳牛银康来源", "shenyang_niu_yinkang_baidu"),
    ],
}


def _api_readiness_inventory():
    from copy import deepcopy

    projects = {}
    for project_id, profile_id in _API_READINESS_SINGLE_PROFILES.items():
        account_name = f"{project_id}_account"
        projects[project_id] = {
            "project_id": project_id,
            "project_name": project_id,
            "baidu": {
                "api_profile": profile_id,
                "credential_profile": f"{project_id}_credential",
            },
            "accounts": {
                account_name: {
                    "baidu_name": account_name,
                    "baidu_names": [account_name],
                    "excel_name": account_name,
                }
            },
        }
    for project_id, sources in _API_READINESS_MULTI_PROFILES.items():
        source_configs = []
        for index, (source_id, source_name, profile_id) in enumerate(sources, start=1):
            account_name = f"{project_id}_account_{index}"
            source_configs.append(
                {
                    "source_id": source_id,
                    "source_name": source_name,
                    "credential_profile": f"{project_id}_credential_{index}",
                    "api_profile": profile_id,
                    "accounts": [
                        {
                            "standard_name": account_name,
                            "baidu_names": [account_name],
                            "excel_name": account_name,
                        }
                    ],
                }
            )
        projects[project_id] = {
            "project_id": project_id,
            "project_name": project_id,
            "baidu": {},
            "accounts": {},
            "baidu_sources": source_configs,
        }
    return deepcopy(projects)


def _install_api_readiness_inventory(monkeypatch, readiness, projects):
    from copy import deepcopy

    monkeypatch.setattr(
        readiness,
        "list_projects",
        lambda _root: [
            {
                "project_id": project_id,
                "project_name": str(project.get("project_name") or project_id),
                "path": f"configs/projects/{project_id}.json",
            }
            for project_id, project in sorted(projects.items())
        ],
    )
    monkeypatch.setattr(
        readiness,
        "load_project_config",
        lambda _root, project_id: deepcopy(projects[project_id]),
    )
    monkeypatch.setattr(
        readiness,
        "build_runtime_config_from_project",
        lambda project, _base: deepcopy(project),
    )
    monkeypatch.setattr(readiness, "load_config", lambda *_args, **_kwargs: {})


def test_api_readiness_checks_nine_projects_and_eleven_independent_profiles(tmp_path, monkeypatch):
    import logging
    from modules import baidu_api_readiness as readiness

    projects = _api_readiness_inventory()
    _install_api_readiness_inventory(monkeypatch, readiness, projects)
    calls = []

    def fetch_func(**kwargs):
        config = kwargs["config"]
        calls.append(kwargs)
        assert kwargs["root"] == tmp_path
        assert kwargs["commit_standard_report"] is False
        assert kwargs["commit_attempt_report"] is False
        assert kwargs["target_date"] == "2026-07-16"
        assert kwargs["period"] == "15点"
        return {"accounts": dict(config["accounts"]), "errors": []}

    report = readiness.run_baidu_api_readiness(
        tmp_path,
        logging.getLogger("api-readiness"),
        fetch_func=fetch_func,
        target_date="2026-07-16",
        period="15点",
    )

    expected_profiles = set(_API_READINESS_SINGLE_PROFILES.values()) | {
        profile_id
        for sources in _API_READINESS_MULTI_PROFILES.values()
        for _source_id, _source_name, profile_id in sources
    }
    assert report["passed"] is True
    assert report["project_count"] == 9
    assert report["profile_count"] == 11
    assert report["unique_profile_count"] == 11
    assert {item["api_profile"] for item in report["results"]} == expected_profiles
    assert len(calls) == 11
    assert len({id(call["task_context"]) for call in calls}) == 11
    assert all(item["passed"] for item in report["results"])
    assert all(item["account_count"] == 1 for item in report["results"])
    output = tmp_path / "reports" / "baidu_api_readiness_report.json"
    assert json.loads(output.read_text(encoding="utf-8")) == report
    assert not output.with_suffix(output.suffix + ".tmp").exists()


def test_api_readiness_continues_after_failures_and_never_serializes_raw_errors(tmp_path, monkeypatch):
    from modules import baidu_api_readiness as readiness
    from modules.baidu_report_api import BaiduReportApiError

    class CapturingLogger:
        def __init__(self):
            self.messages = []

        def info(self, message, *args):
            self.messages.append(message % args if args else message)

        def warning(self, message, *args):
            self.messages.append(message % args if args else message)

    projects = _api_readiness_inventory()
    _install_api_readiness_inventory(monkeypatch, readiness, projects)
    seen = []

    def fetch_func(**kwargs):
        profile_id = kwargs["config"]["baidu"]["api_profile"]
        seen.append(profile_id)
        if profile_id == "changsha_niu_baidu":
            raise BaiduReportApiError(
                "https://private.example/?token=header.payload.signature&secretKey=do-not-write"
                "&password=do-not-write&header=X-Api-Key",
                category="authorization_error",
            )
        if profile_id == "kunming_niu_baidu":
            raise RuntimeError("https://private.example/callback?access_token=do-not-write")
        return {"accounts": dict(kwargs["config"]["accounts"]), "errors": []}

    logger = CapturingLogger()
    report = readiness.run_baidu_api_readiness(
        tmp_path,
        logger,
        fetch_func=fetch_func,
        target_date="2026-07-16",
        period="18点",
    )

    assert len(seen) == 11
    assert report["passed"] is False
    failed = {item["api_profile"]: item for item in report["results"] if not item["passed"]}
    assert failed["changsha_niu_baidu"]["error_category"] == "authorization_error"
    assert failed["kunming_niu_baidu"]["error_category"] == "api_error"
    serialized = json.dumps(report, ensure_ascii=False)
    logged = "\n".join(logger.messages)
    for forbidden in [
        "https://",
        "header.payload.signature",
        "do-not-write",
        "private.example",
        "access_token",
        "secretKey",
        "password=",
        "X-Api-Key",
    ]:
        assert forbidden not in serialized
        assert forbidden not in logged


@pytest.mark.parametrize(
    "invalid_shape",
    ["not_dict", "empty_accounts", "missing_errors", "errors_not_list", "missing_account", "extra_account", "wrong_account_name"],
)
def test_api_readiness_rejects_incomplete_or_mismatched_fetch_reports(tmp_path, monkeypatch, invalid_shape):
    import logging
    from modules import baidu_api_readiness as readiness

    projects = _api_readiness_inventory()
    projects["changsha_niu"]["accounts"]["changsha_niu_account_2"] = {
        "baidu_name": "changsha_niu_account_2",
        "baidu_names": ["changsha_niu_account_2"],
        "excel_name": "changsha_niu_account_2",
    }
    _install_api_readiness_inventory(monkeypatch, readiness, projects)

    def fetch_func(**kwargs):
        expected = dict(kwargs["config"]["accounts"])
        if kwargs["config"]["baidu"]["api_profile"] != "changsha_niu_baidu":
            return {"accounts": expected, "errors": []}
        if invalid_shape == "not_dict":
            return []
        if invalid_shape == "empty_accounts":
            return {"accounts": {}, "errors": []}
        if invalid_shape == "missing_errors":
            return {"accounts": expected}
        if invalid_shape == "errors_not_list":
            return {"accounts": expected, "errors": None}
        if invalid_shape == "missing_account":
            expected.pop(next(iter(expected)))
        elif invalid_shape == "extra_account":
            expected["unexpected_account"] = {}
        else:
            expected.pop(next(iter(expected)))
            expected["wrong_account_name"] = {}
        return {"accounts": expected, "errors": []}

    report = readiness.run_baidu_api_readiness(
        tmp_path,
        logging.getLogger("api-readiness-shape"),
        fetch_func=fetch_func,
        target_date="2026-07-16",
    )

    target = next(
        item for item in report["results"] if item["api_profile"] == "changsha_niu_baidu"
    )
    assert report["passed"] is False
    assert target["passed"] is False
    assert target["error_category"] == "integrity_error"
    assert target["account_count"] == 0


@pytest.mark.parametrize("fetcher_name", ["hourly", "daily"])
def test_api_readiness_production_fetcher_can_suppress_attempt_report(tmp_path, fetcher_name):
    import logging
    from modules.baidu_report_api import (
        BaiduReportApiError,
        fetch_baidu_api_daily,
        fetch_baidu_api_hourly,
    )

    config = _api_production_config(tmp_path)
    fetcher = fetch_baidu_api_hourly if fetcher_name == "hourly" else fetch_baidu_api_daily
    kwargs = {
        "config": config,
        "root": tmp_path,
        "logger": logging.getLogger(f"api-readiness-no-attempt-{fetcher_name}"),
        "target_date": "2026-07-16",
        "token_provider": lambda *_args, **_kwargs: ("header.payload.signature", {}),
        "transport": lambda *_args: (_ for _ in ()).throw(
            BaiduReportApiError("token=do-not-write", category="network_error")
        ),
        "commit_standard_report": False,
        "commit_attempt_report": False,
    }
    if fetcher_name == "hourly":
        kwargs["period"] = "15点"

    with pytest.raises(BaiduReportApiError):
        fetcher(**kwargs)

    assert not (tmp_path / "reports" / "baidu_api_attempt_report.json").exists()


def test_api_readiness_real_production_failure_does_not_create_attempt_report(tmp_path, monkeypatch):
    import logging
    import modules.baidu_report_api as report_api
    from modules import baidu_api_readiness as readiness

    projects = _api_readiness_inventory()
    _install_api_readiness_inventory(monkeypatch, readiness, projects)
    monkeypatch.setattr(
        report_api,
        "_load_api_identity",
        lambda _root, config: ("manager", config["baidu"]["api_profile"], {}),
    )
    monkeypatch.setattr(
        report_api,
        "_account_user_ids",
        lambda config: ([1], {1: next(iter(config["accounts"]))}),
    )
    attempt_flags = []

    def production_fetcher(**kwargs):
        attempt_flags.append(kwargs.get("commit_attempt_report", "missing"))
        return report_api.fetch_baidu_api_hourly(
            **kwargs,
            token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
            transport=lambda *_args: (_ for _ in ()).throw(
                report_api.BaiduReportApiError(
                    "https://private.example/?token=do-not-write",
                    category="network_error",
                )
            ),
        )

    report = readiness.run_baidu_api_readiness(
        tmp_path,
        logging.getLogger("api-readiness-real-production"),
        fetch_func=production_fetcher,
        target_date="2026-07-16",
    )

    assert report["passed"] is False
    assert attempt_flags == [False] * 11
    assert not (tmp_path / "reports" / "baidu_api_attempt_report.json").exists()


def test_api_readiness_production_fetcher_keeps_attempt_report_enabled_by_default(tmp_path):
    import logging
    from modules.baidu_report_api import BaiduReportApiError, fetch_baidu_api_hourly

    config = _api_production_config(tmp_path)
    with pytest.raises(BaiduReportApiError):
        fetch_baidu_api_hourly(
            config,
            tmp_path,
            logging.getLogger("api-readiness-default-attempt"),
            "15点",
            target_date="2026-07-16",
            token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
            transport=lambda *_args: (_ for _ in ()).throw(
                BaiduReportApiError("token=do-not-write", category="network_error")
            ),
            commit_standard_report=False,
        )

    attempt_path = tmp_path / "reports" / "baidu_api_attempt_report.json"
    assert attempt_path.exists()
    attempt = json.loads(attempt_path.read_text(encoding="utf-8"))
    assert attempt["error_category"] == "network_error"
    assert "do-not-write" not in json.dumps(attempt, ensure_ascii=False)


@pytest.mark.parametrize("inventory_case", ["missing_project", "extra_project", "missing_profile", "duplicate_profile", "extra_profile"])
def test_api_readiness_rejects_invalid_production_inventory_before_network_calls(tmp_path, monkeypatch, inventory_case):
    import logging
    from copy import deepcopy
    from modules import baidu_api_readiness as readiness

    projects = _api_readiness_inventory()
    if inventory_case == "missing_project":
        projects.pop("nanjing_bai")
    elif inventory_case == "extra_project":
        extra = deepcopy(projects["nanjing_bai"])
        extra["project_id"] = "unexpected_project"
        extra["project_name"] = "unexpected_project"
        extra["baidu"]["api_profile"] = "unexpected_project_baidu"
        projects["unexpected_project"] = extra
    elif inventory_case == "missing_profile":
        projects["nanjing_bai"]["baidu"]["api_profile"] = ""
    elif inventory_case == "duplicate_profile":
        projects["nanjing_bai"]["baidu"]["api_profile"] = projects["nanjing_niu"]["baidu"]["api_profile"]
    else:
        projects["shenyang_bai"]["baidu_sources"].append(
            {
                "source_id": "unexpected",
                "source_name": "unexpected",
                "credential_profile": "unexpected",
                "api_profile": "unexpected_profile",
                "accounts": [{"standard_name": "unexpected", "baidu_names": ["unexpected"], "excel_name": "unexpected"}],
            }
        )
    _install_api_readiness_inventory(monkeypatch, readiness, projects)
    calls = []

    report = readiness.run_baidu_api_readiness(
        tmp_path,
        logging.getLogger("api-readiness-inventory"),
        fetch_func=lambda **kwargs: calls.append(kwargs) or {"accounts": {}, "errors": []},
        target_date="2026-07-16",
    )

    assert report["passed"] is False
    assert calls == []
    assert report["inventory_errors"]
    assert all(item["error_category"] == "configuration_error" for item in report["inventory_errors"])


@pytest.mark.parametrize("passed, expected_code", [(True, 0), (False, 1)])
def test_api_readiness_cli_bypasses_current_project_and_returns_report_status(tmp_path, monkeypatch, passed, expected_code):
    import main as cli_main

    calls = []
    monkeypatch.setattr(cli_main, "ROOT", tmp_path)
    monkeypatch.setattr(cli_main, "load_config", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("不应加载默认运行配置")))
    monkeypatch.setattr(cli_main, "get_current_project", lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("不应加载当前项目")))
    monkeypatch.setattr(
        cli_main,
        "run_baidu_api_readiness",
        lambda root, logger, **kwargs: calls.append((root, kwargs)) or {"passed": passed},
    )
    monkeypatch.setattr(
        "sys.argv",
        ["main.py", "--mode", "test-baidu-api-readiness", "--date", "2026-07-16", "--period", "18点"],
    )

    result = cli_main.main()

    assert result == expected_code
    assert calls == [(tmp_path, {"target_date": "2026-07-16", "period": "18点"})]


def test_api_readiness_hourly_reader_accepts_explicit_safe_target_date(tmp_path):
    import logging
    from modules.baidu_report_api import fetch_baidu_api_hourly

    config = _api_production_config(tmp_path)
    seen_dates = []

    def transport(_url, payload, _timeout):
        selected_date = payload["body"]["startDate"]
        seen_dates.append(selected_date)
        return _api_success_response(selected_date)

    report = fetch_baidu_api_hourly(
        config,
        tmp_path,
        logging.getLogger("api-readiness-explicit-date"),
        "15点",
        token_provider=lambda *_args, **_kwargs: ("header.payload.signature", {}),
        transport=transport,
        commit_standard_report=False,
        target_date="2026-07-16",
    )

    assert seen_dates == ["2026-07-16"]
    assert report["date"] == "2026-07-16"
    assert report["self_check"]["production_output_replaced"] is False
