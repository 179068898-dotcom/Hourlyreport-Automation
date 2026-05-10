from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from modules.excel_inspector import (
    _build_account_ranges,
    _build_merged_bounds_map,
    _build_merged_value_map,
    _find_account_titles,
    _get_merged_value,
    _scan_non_empty_cells,
    _write_sheet_dump,
)
from modules.text_normalizer import normalize_for_display, normalize_text


DAILY_SHEET_NAME = "百度"
ALLOWED_FIELDS = ["展现", "点击", "消费", "有效对话", "无效对话", "一般有效对话", "有效转潜", "总转潜"]
FORBIDDEN_FIELDS = ["总对话", "预约", "到诊", "就诊"]
FIELD_ALIASES = {
    "展现": ["展现", "展现量"],
    "点击": ["点击", "点击量"],
    "消费": ["消费", "花费"],
    "有效对话": ["有效对话"],
    "无效对话": ["无效对话"],
    "一般有效对话": ["一般有效对话"],
    "有效转潜": ["有效转潜"],
    "总转潜": ["总转潜"],
    "总对话": ["总对话"],
    "预约": ["预约"],
    "到诊": ["到诊"],
    "就诊": ["就诊"],
}


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _coerce_excel_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalize_for_display(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _find_date_column(ws: Worksheet, merged_values: dict[tuple[int, int], Any]) -> dict[str, Any]:
    for row in range(1, min(ws.max_row, 10) + 1):
        for col in range(1, min(ws.max_column, 10) + 1):
            value = _get_merged_value(ws, row, col, merged_values)
            if normalize_text(value) == "日期":
                return {
                    "found": True,
                    "header_cell": ws.cell(row=row, column=col).coordinate,
                    "header_row": row,
                    "col": col,
                    "column_letter": get_column_letter(col),
                }
    return {"found": False}


def _find_date_rows(ws: Worksheet, date_col: int | None) -> dict[str, Any]:
    if not date_col:
        return {"count": 0, "rows": []}
    rows = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=date_col).value
        date_text = _coerce_excel_date(value)
        if not date_text:
            continue
        rows.append({
            "row": row,
            "cell": ws.cell(row=row, column=date_col).coordinate,
            "date": date_text,
            "raw_value": normalize_for_display(value),
        })
    return {
        "count": len(rows),
        "first_date": rows[0]["date"] if rows else None,
        "last_date": rows[-1]["date"] if rows else None,
        "rows": rows[:400],
    }


def _field_match(norm: str, aliases: list[str]) -> bool:
    return any(norm == normalize_text(alias) for alias in aliases)


def _find_daily_fields_for_account(
    ws: Worksheet,
    meta: dict[str, Any],
    merged_values: dict[tuple[int, int], Any],
) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    if not meta.get("found") or "range" not in meta:
        return fields
    r = meta["range"]
    header_min_row = int(r["min_row"]) + 1
    header_max_row = min(ws.max_row, int(r["min_row"]) + 5)
    for field_name, aliases in FIELD_ALIASES.items():
        matches = []
        for row in range(header_min_row, header_max_row + 1):
            for col in range(int(r["min_col"]), int(r["max_col"]) + 1):
                value = _get_merged_value(ws, row, col, merged_values)
                norm = normalize_text(value)
                if norm and _field_match(norm, aliases):
                    matches.append({
                        "row": row,
                        "col": col,
                        "address": ws.cell(row=row, column=col).coordinate,
                        "raw_text": normalize_for_display(value),
                    })
        if matches:
            matches.sort(key=lambda item: (item["row"], item["col"]))
            cell = matches[0]
            fields[field_name] = {
                "found": True,
                "header_cell": cell["address"],
                "header_row": cell["row"],
                "header_col": cell["col"],
                "write_allowed": field_name in ALLOWED_FIELDS,
                "raw_text": cell["raw_text"],
                "all_matches": matches[:20],
            }
        else:
            fields[field_name] = {
                "found": False,
                "write_allowed": field_name in ALLOWED_FIELDS,
                "all_matches": [],
            }
    return fields


def _validate_daily_report(report: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not report.get("sheet_found"):
        errors.append(f"找不到日报目标 sheet：{report.get('sheet_name')}")
        return errors
    if not report.get("date_column", {}).get("found"):
        errors.append("百度 sheet 未识别到日期列")
    if report.get("date_rows", {}).get("count", 0) <= 0:
        errors.append("百度 sheet 未识别到任何日期行")
    for account, meta in report.get("accounts", {}).items():
        if not meta.get("found"):
            errors.append(f"百度 sheet 未识别到账户区域：{account}")
            continue
        fields = meta.get("fields", {})
        for field in ALLOWED_FIELDS:
            if not fields.get(field, {}).get("found"):
                errors.append(f"账户 {account} 未识别到允许写入字段：{field}")
        for field in ["预约", "到诊", "就诊"]:
            if fields.get(field, {}).get("found") and fields[field].get("write_allowed"):
                errors.append(f"人工字段被误标记为可写：{account} {field}")
    return errors


def inspect_daily_worksheet(ws: Worksheet, config: dict[str, Any], excel_path: str) -> dict[str, Any]:
    merged_values = _build_merged_value_map(ws)
    merged_bounds = _build_merged_bounds_map(ws)
    rows = _scan_non_empty_cells(ws, merged_values)
    date_column = _find_date_column(ws, merged_values)
    accounts = _find_account_titles(rows, config)
    accounts = _build_account_ranges(accounts, ws, merged_bounds)
    for account, meta in accounts.items():
        meta["fields"] = _find_daily_fields_for_account(ws, meta, merged_values)
    report: dict[str, Any] = {
        "mode": "inspect-daily-excel",
        "excel_path": excel_path,
        "sheet_name": ws.title,
        "sheet_found": True,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_range_count": len(ws.merged_cells.ranges),
        "allowed_fields": ALLOWED_FIELDS,
        "forbidden_fields": FORBIDDEN_FIELDS,
        "date_column": date_column,
        "date_rows": _find_date_rows(ws, date_column.get("col") if date_column.get("found") else None),
        "accounts": accounts,
        "errors": [],
    }
    report["errors"] = _validate_daily_report(report)
    return report


def inspect_daily_excel_structure(config: dict[str, Any], root: Path, logger) -> dict[str, Any]:
    excel_path = Path(config["excel_path"])
    if not excel_path.is_absolute():
        excel_path = root / excel_path
    report_path = root / "reports" / "daily_excel_structure_report.json"
    dump_path = root / "reports" / "daily_sheet_text_dump.csv"
    sheet_name = config.get("daily_sheet_name", DAILY_SHEET_NAME)
    report: dict[str, Any] = {
        "mode": "inspect-daily-excel",
        "excel_path": str(excel_path),
        "sheet_name": sheet_name,
        "sheet_found": False,
        "allowed_fields": ALLOWED_FIELDS,
        "forbidden_fields": FORBIDDEN_FIELDS,
        "accounts": {},
        "outputs": {
            "report": str(report_path),
            "sheet_text_dump": str(dump_path),
        },
        "errors": [],
    }
    if not excel_path.exists():
        report["errors"].append(f"找不到 Excel 文件：{excel_path}")
        _write_json(report_path, report)
        return report
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    report["available_sheets"] = wb.sheetnames
    if sheet_name not in wb.sheetnames:
        report["errors"].append(f"找不到 sheet：{sheet_name}")
        _write_json(report_path, report)
        return report
    ws = wb[sheet_name]
    report = inspect_daily_worksheet(ws, config, str(excel_path))
    report["outputs"] = {
        "report": str(report_path),
        "sheet_text_dump": str(dump_path),
    }
    _write_sheet_dump(_scan_non_empty_cells(ws, _build_merged_value_map(ws)), dump_path)
    _write_json(report_path, report)
    if report["errors"]:
        logger.warning("日报 Excel 结构识别存在问题：%s", report["errors"])
    else:
        logger.info("日报 Excel 结构识别通过。")
    return report
