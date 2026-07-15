from __future__ import annotations

import json
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile
import re
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from modules.console_ui import print_check_result, verbose_print
from modules.excel_engine import format_openpyxl_save_error
from modules.excel_inspector import (
    _build_merged_value_map,
    _get_merged_value,
    _write_excel_account_regions_report,
    inspect_excel_structure,
    inspect_excel_worksheet,
)
from modules.daily_excel_inspector import ALLOWED_FIELDS as DAILY_WRITE_FIELDS
from modules.daily_excel_inspector import DAILY_SHEET_NAME, _write_daily_structure_outputs, inspect_daily_worksheet
from modules.text_normalizer import normalize_text
from modules.validators import get_required_accounts, validate_merged_daily_data, validate_merged_hourly_data


WRITE_FIELDS = ["展现", "点击", "消费", "总对话", "有效", "有效转潜", "总转潜"]
PERIOD_TO_EXCEL = {
    "11点": "11点",
    "15点": "3点",
    "18点": "6点",
    "3点": "3点",
    "6点": "6点",
    "11": "11点",
    "15": "3点",
    "18": "6点",
    "3": "3点",
    "6": "6点",
    "11dian": "11点",
    "15dian": "3点",
    "18dian": "6点",
}


def _resolve_excel_path(config: dict[str, Any], root: Path) -> Path:
    if not config.get("excel_path"):
        raise ValueError("config.json 中缺少 excel_path，请先配置目标 Excel 路径")
    excel_path = Path(config["excel_path"])
    if not excel_path.is_absolute():
        excel_path = root / excel_path
    return excel_path


def _normalize_period_for_excel(period: str | None) -> str:
    raw = normalize_text(period or "15点")
    period_map = {normalize_text(k): v for k, v in PERIOD_TO_EXCEL.items()}
    if raw not in period_map:
        raise ValueError(f"不支持的时段：{period}，只支持 11点 / 15点 / 18点")
    return period_map[raw]


def _cell_date_matches(value: Any, target_date: date) -> bool:
    if isinstance(value, datetime):
        return value.date() == target_date
    if isinstance(value, date):
        return value == target_date
    text = normalize_text(value)
    return text in {
        target_date.isoformat(),
        target_date.strftime("%Y/%m/%d"),
        target_date.strftime("%Y年%m月%d日"),
        target_date.strftime("%m月%d日"),
    }


def _find_target_row(
    ws: Worksheet,
    date_col: int,
    period_col: int,
    target_date: date,
    period: str,
    merged_values: dict[tuple[int, int], Any],
) -> int | None:
    excel_period = normalize_text(_normalize_period_for_excel(period))
    for row in range(1, (ws.max_row or 1) + 1):
        date_value = _get_merged_value(ws, row, date_col, merged_values)
        period_value = _get_merged_value(ws, row, period_col, merged_values)
        if _cell_date_matches(date_value, target_date) and normalize_text(period_value) == excel_period:
            return row
    return None


def _resolve_global_row_fields(structure: dict[str, Any], account_meta: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    global_fields = structure.get("global_fields", {})
    account_fields = account_meta.get("fields", {})
    date_field = global_fields.get("日期") or account_fields.get("日期", {})
    period_field = global_fields.get("时段") or account_fields.get("时段", {})
    return date_field or {}, period_field or {}


def _range_contains(cell_range: dict[str, int], row: int, col: int) -> bool:
    return (
        int(cell_range["min_row"]) <= row <= int(cell_range["max_row"])
        and int(cell_range["min_col"]) <= col <= int(cell_range["max_col"])
    )


def _validate_write_target(
    account: str,
    field: str,
    row: int,
    col: int,
    account_meta: dict[str, Any],
    protected_regions: list[dict[str, Any]],
) -> list[str]:
    errors: list[str] = []
    account_range = account_meta.get("range")
    if not account_range or not _range_contains(account_range, row, col):
        errors.append(f"账户 {account} 字段 {field} 目标单元格不在账户分户区域内")
    for region in protected_regions:
        region_range = region.get("range", {})
        if region_range and _range_contains(region_range, row, col):
            errors.append(f"账户 {account} 字段 {field} 目标单元格落入受保护区域")
    return errors


def _make_backup(excel_path: Path, root: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = root / "backups" / f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}"
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(excel_path, backup_path)
    return backup_path


def _sheet_xml_path(zf: ZipFile, sheet_name: str) -> str | None:
    workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
    for sheet in workbook_root.findall(".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        if sheet.attrib.get("name") != sheet_name:
            continue
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        if not rel_id or rel_id not in rel_targets:
            return None
        target = rel_targets[rel_id].lstrip("/")
        if target.startswith("xl/"):
            return target
        return f"xl/{target}"
    return None


def _replace_xml_node(xml: str, tag: str, replacement: str | None) -> tuple[str, bool]:
    pattern = re.compile(rf"<{tag}\b[^>]*(?:/>|>.*?</{tag}>)", re.S)
    if replacement:
        if pattern.search(xml):
            return pattern.sub(replacement, xml, count=1), True
        if tag == "sheetProtection":
            insert_after = re.search(r"</sheetViews>", xml)
            if insert_after:
                pos = insert_after.end()
                return xml[:pos] + replacement + xml[pos:], True
        if tag == "autoFilter":
            insert_before = re.search(
                r"<(?:sortState|dataConsolidate|customSheetViews|mergeCells|phoneticPr|conditionalFormatting|"
                r"dataValidations|hyperlinks|printOptions|pageMargins|pageSetup|headerFooter|rowBreaks|"
                r"colBreaks|customProperties|cellWatches|ignoredErrors|smartTags|drawing|legacyDrawing|"
                r"legacyDrawingHF|picture|oleObjects|controls|webPublishItems|tableParts|extLst)\b",
                xml,
            )
            if insert_before:
                pos = insert_before.start()
                return xml[:pos] + replacement + xml[pos:], True
            insert_before_close = re.search(r"</worksheet>", xml)
            if insert_before_close:
                pos = insert_before_close.start()
                return xml[:pos] + replacement + xml[pos:], True
        insert_before = re.search(r"<sheetViews\b", xml)
        if insert_before:
            pos = insert_before.start()
            return xml[:pos] + replacement + xml[pos:], True
        return xml, False
    return pattern.sub("", xml, count=1), bool(pattern.search(xml))


def _restore_sheet_filter_protection_metadata(
    excel_path: Path,
    backup_path: Path,
    sheet_names: list[str],
    logger,
) -> bool:
    """恢复 openpyxl 保存时容易丢失的保护/筛选 UI 元数据。

    只替换目标 sheet 的 sheetProtection 和 autoFilter 节点，不改单元格值、公式、样式或其他 sheet。
    """
    if not backup_path.exists() or not excel_path.exists():
        return False

    changed_files: dict[str, bytes] = {}
    processed_sheets: list[str] = []
    try:
        with ZipFile(backup_path, "r") as backup_zip, ZipFile(excel_path, "r") as current_zip:
            for sheet_name in sheet_names:
                backup_sheet_path = _sheet_xml_path(backup_zip, sheet_name)
                current_sheet_path = _sheet_xml_path(current_zip, sheet_name)
                if not backup_sheet_path or not current_sheet_path:
                    continue

                processed_sheets.append(sheet_name)
                backup_xml = backup_zip.read(backup_sheet_path).decode("utf-8")
                current_xml = current_zip.read(current_sheet_path).decode("utf-8")
                restored_xml = current_xml
                for tag in ("sheetProtection", "autoFilter"):
                    backup_node_match = re.search(
                        rf"<{tag}\b[^>]*(?:/>|>.*?</{tag}>)",
                        backup_xml,
                        re.S,
                    )
                    replacement = backup_node_match.group(0) if backup_node_match else None
                    restored_xml, _ = _replace_xml_node(restored_xml, tag, replacement)
                if restored_xml != current_xml:
                    changed_files[current_sheet_path] = restored_xml.encode("utf-8")

            if changed_files:
                with tempfile.NamedTemporaryFile(delete=False, suffix=excel_path.suffix) as tmp:
                    tmp_path = Path(tmp.name)
                with ZipFile(tmp_path, "w", ZIP_DEFLATED) as out_zip:
                    for item in current_zip.infolist():
                        out_zip.writestr(item, changed_files.get(item.filename, current_zip.read(item.filename)))
        if changed_files:
            shutil.move(str(tmp_path), excel_path)
            logger.info("已恢复 Excel 筛选/保护 UI 元数据：%s", ", ".join(processed_sheets))
        elif processed_sheets:
            logger.info("Excel 筛选/保护 UI 元数据保持不变：%s", ", ".join(processed_sheets))
        return bool(processed_sheets)
    except Exception as exc:
        logger.warning("恢复 Excel 筛选/保护 UI 元数据失败：%s", exc)
        return False


def _parse_report_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        raise ValueError("merged_hourly_data.json 中缺少 date")
    return datetime.strptime(text, "%Y-%m-%d").date()


def _find_daily_target_row(ws: Worksheet, date_col: int, target_date: date, merged_values: dict[tuple[int, int], Any]) -> int | None:
    for row in range(1, (ws.max_row or 1) + 1):
        date_value = _get_merged_value(ws, row, date_col, merged_values)
        if _cell_date_matches(date_value, target_date):
            return row
    return None


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _values_match(actual: Any, expected: Any) -> bool:
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return abs(float(actual) - float(expected)) < 0.000001
    return actual == expected


def _read_back_values(excel_path: Path, sheet_name: str, coordinates: list[str]) -> dict[str, Any]:
    requested = list(dict.fromkeys(coordinates))
    if not requested:
        return {}
    positions = {coordinate_to_tuple(coordinate): coordinate for coordinate in requested}
    row_numbers = [row for row, _col in positions]
    column_numbers = [col for _row, col in positions]
    workbook = load_workbook(excel_path, data_only=False, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"找不到 sheet：{sheet_name}")
        worksheet = workbook[sheet_name]
        values: dict[str, Any] = {coordinate: None for coordinate in requested}
        min_row = min(row_numbers)
        min_col = min(column_numbers)
        rows = worksheet.iter_rows(
            min_row=min_row,
            max_row=max(row_numbers),
            min_col=min_col,
            max_col=max(column_numbers),
        )
        for row_index, row_cells in enumerate(rows, start=min_row):
            for column_index, cell in enumerate(row_cells, start=min_col):
                coordinate = positions.get((row_index, column_index))
                if coordinate:
                    values[coordinate] = cell.value
        return values
    finally:
        workbook.close()


def build_mock_account_data(config: dict[str, Any] | None = None) -> dict[str, dict[str, Any]]:
    accounts = get_required_accounts(config)
    return {
        account: {
            "展现": 1000 + index,
            "点击": 50 + index,
            "消费": 300 + index + 0.17,
            "总对话": 10 + index,
            "有效": 4 + index,
            "有效转潜": 2 + index,
            "总转潜": 3 + index,
        }
        for index, account in enumerate(accounts, start=1)
    }


def mock_write_excel(
    config: dict[str, Any],
    root: Path,
    logger,
    period: str | None = None,
    target_date: date | None = None,
) -> dict[str, Any]:
    target_date = target_date or date.today()
    period = period or "15点"
    excel_path = _resolve_excel_path(config, root)
    sheet_name = config.get("sheet_name", "时段数据")
    out_path = root / "reports" / "mock_write_report.json"
    report: dict[str, Any] = {
        "mode": "mock-write",
        "project_id": config.get("project_id"),
        "project_name": config.get("project_name"),
        "excel_path": str(excel_path),
        "sheet_name": sheet_name,
        "date": target_date.isoformat(),
        "period": period,
        "excel_period": _normalize_period_for_excel(period),
        "backup_path": None,
        "accounts": {},
        "errors": [],
        "self_check": {
            "backup_created": False,
            "structure_passed": False,
            "wrote_protected_region": False,
            "filter_ui_metadata_restored": False,
            "verification_passed": False,
        },
    }

    if not excel_path.exists():
        report["errors"].append(f"找不到 Excel 文件：{excel_path}")
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report

    structure = inspect_excel_structure(config=config, root=root, logger=logger)
    structure_errors = structure.get("errors", [])
    if structure_errors:
        report["errors"].extend(structure_errors)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report
    report["self_check"]["structure_passed"] = True

    backup_path = _make_backup(excel_path, root)
    report["backup_path"] = str(backup_path)
    report["self_check"]["backup_created"] = backup_path.exists()
    logger.info("写入前备份已创建：%s", backup_path)

    wb = load_workbook(excel_path, data_only=False, read_only=False)
    if sheet_name not in wb.sheetnames:
        report["errors"].append(f"找不到 sheet：{sheet_name}")
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        wb.close()
        return report
    ws = wb[sheet_name]
    merged_values = _build_merged_value_map(ws)
    mock_data = build_mock_account_data(config)
    protected_regions = structure.get("summary_regions", [])
    write_ops: list[dict[str, Any]] = []

    for account, values in mock_data.items():
        account_meta = structure["accounts"].get(account, {})
        fields = account_meta.get("fields", {})
        date_field, period_field = _resolve_global_row_fields(structure, account_meta)
        row = _find_target_row(
            ws,
            int(date_field.get("header_col", 0)),
            int(period_field.get("header_col", 0)),
            target_date,
            period,
            merged_values,
        )
        account_report = {"target_row": row, "writes": [], "errors": []}
        if row is None:
            account_report["errors"].append(f"找不到日期 {target_date.isoformat()} 时段 {_normalize_period_for_excel(period)}")
            report["accounts"][account] = account_report
            continue

        for field in WRITE_FIELDS:
            field_meta = fields.get(field, {})
            if not field_meta.get("found"):
                account_report["errors"].append(f"找不到字段：{field}")
                continue
            col = int(field_meta["header_col"])
            target_errors = _validate_write_target(account, field, int(row), col, account_meta, protected_regions)
            if target_errors:
                account_report["errors"].extend(target_errors)
                continue
            cell = ws.cell(row=int(row), column=col)
            old_value = cell.value
            new_value = values[field]
            cell.value = new_value
            op = {
                "field": field,
                "cell": cell.coordinate,
                "old_value": old_value,
                "new_value": new_value,
                "verified": False,
            }
            account_report["writes"].append(op)
            write_ops.append({"account": account, **op})
        report["accounts"][account] = account_report

    for account, account_report in report["accounts"].items():
        report["errors"].extend(account_report["errors"])

    if report["errors"]:
        wb.close()
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.error("模拟写入中断：%s", report["errors"])
        return report

    try:
        wb.save(excel_path)
    except OSError as exc:
        wb.close()
        report["errors"].append(format_openpyxl_save_error(excel_path, exc))
        logger.error("模拟保存 Excel 失败：%s", exc)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report
    wb.close()
    report["self_check"]["filter_ui_metadata_restored"] = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        [sheet_name],
        logger,
    )
    logger.info("模拟数据已写入并保存：%s", excel_path)

    read_back_values = _read_back_values(excel_path, sheet_name, [op["cell"] for op in write_ops])
    for op in write_ops:
        value = read_back_values.get(op["cell"])
        op["verified"] = value == op["new_value"]
        account_writes = report["accounts"][op["account"]]["writes"]
        for item in account_writes:
            if item["cell"] == op["cell"] and item["field"] == op["field"]:
                item["verified"] = op["verified"]
                item["read_back_value"] = value
        if not op["verified"]:
            report["errors"].append(f"写入后复核不一致：{op['account']} {op['field']} {op['cell']}")

    report["self_check"]["verification_passed"] = not report["errors"] and all(op["verified"] for op in write_ops)
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("模拟写入报告已输出：%s", out_path)
    return report


def write_merged_hourly_data(
    config: dict[str, Any],
    root: Path,
    logger,
    period: str | None = None,
) -> dict[str, Any]:
    source_path = root / "reports" / "merged_hourly_data.json"
    out_path = root / "reports" / "write_report.json"
    final_path = root / "reports" / "final_run_report.json"
    report: dict[str, Any] = {
        "mode": "write-excel",
        "project_id": config.get("project_id"),
        "project_name": config.get("project_name"),
        "source": str(source_path),
        "excel_path": None,
        "sheet_name": config.get("sheet_name", "时段数据"),
        "date": None,
        "period": period,
        "excel_period": None,
        "backup_path": None,
        "writes": [],
        "accounts": {},
        "overwrite_summary": {
            "overwrite_count": 0,
            "items": [],
        },
        "errors": [],
        "self_check": {
            "merged_data_exists": source_path.exists(),
            "merged_validation_passed": False,
            "structure_passed": False,
            "backup_created": False,
            "wrote_protected_region": False,
            "filter_ui_metadata_restored": False,
            "verification_passed": False,
        },
    }

    def finish() -> dict[str, Any]:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        final_report = {
            "passed": not report["errors"] and report["self_check"]["verification_passed"],
            "date": report.get("date"),
            "period": report.get("period"),
            "outputs": {
                "write_report": str(out_path),
                "final_run_report": str(final_path),
            },
            "self_check": report["self_check"],
            "errors": report["errors"],
        }
        final_path.write_text(json.dumps(final_report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report

    if not source_path.exists():
        report["errors"].append(f"找不到合并数据文件：{source_path}")
        return finish()

    try:
        merged_data = json.loads(source_path.read_text(encoding="utf-8"))
        target_date = _parse_report_date(merged_data.get("date"))
        write_period = period or merged_data.get("period") or "15点"
        excel_period = _normalize_period_for_excel(write_period)
        excel_path = _resolve_excel_path(config, root)
    except Exception as exc:
        report["errors"].append(str(exc))
        return finish()

    report["excel_path"] = str(excel_path)
    report["date"] = target_date.isoformat()
    report["period"] = write_period
    report["excel_period"] = excel_period

    required_accounts = get_required_accounts(config)
    merge_errors = validate_merged_hourly_data(merged_data, required_accounts=required_accounts)
    if merge_errors:
        report["errors"].extend(merge_errors)
        return finish()
    report["self_check"]["merged_validation_passed"] = True

    if not excel_path.exists():
        report["errors"].append(f"找不到 Excel 文件：{excel_path}")
        return finish()

    sheet_name = config.get("sheet_name", "时段数据")
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    if sheet_name not in wb.sheetnames:
        report["errors"].append(f"找不到 sheet：{sheet_name}")
        wb.close()
        return finish()
    ws = wb[sheet_name]
    structure = inspect_excel_worksheet(ws, config, excel_path)
    _write_excel_account_regions_report(config, root, structure)
    structure_errors = structure.get("errors", [])
    if structure_errors:
        report["errors"].extend(structure_errors)
        wb.close()
        return finish()
    report["self_check"]["structure_passed"] = True
    logger.info("Excel 结构识别通过")

    backup_path = _make_backup(excel_path, root)
    report["backup_path"] = str(backup_path)
    report["self_check"]["backup_created"] = backup_path.exists()
    logger.info("正式写入前备份已创建：%s", backup_path)

    merged_values = _build_merged_value_map(ws)
    protected_regions = structure.get("summary_regions", [])
    account_values = merged_data.get("accounts", {})
    planned_writes: list[dict[str, Any]] = []

    for account in required_accounts:
        values = account_values.get(account, {})
        account_meta = structure["accounts"].get(account, {})
        fields = account_meta.get("fields", {})
        account_report = {"target_row": None, "writes": [], "errors": []}
        date_field, period_field = _resolve_global_row_fields(structure, account_meta)
        if not date_field.get("found"):
            account_report["errors"].append("找不到字段：日期")
        if not period_field.get("found"):
            account_report["errors"].append("找不到字段：时段")
        if account_report["errors"]:
            report["accounts"][account] = account_report
            continue

        row = _find_target_row(
            ws,
            int(date_field.get("header_col", 0)),
            int(period_field.get("header_col", 0)),
            target_date,
            write_period,
            merged_values,
        )
        account_report["target_row"] = row
        if row is None:
            account_report["errors"].append(f"找不到日期 {target_date.isoformat()} 时段 {excel_period}")
            report["accounts"][account] = account_report
            continue

        for field in WRITE_FIELDS:
            field_meta = fields.get(field, {})
            if not field_meta.get("found"):
                account_report["errors"].append(f"找不到字段：{field}")
                continue
            if field not in values:
                account_report["errors"].append(f"合并数据缺少字段：{field}")
                continue
            col = int(field_meta["header_col"])
            target_errors = _validate_write_target(account, field, int(row), col, account_meta, protected_regions)
            if target_errors:
                account_report["errors"].extend(target_errors)
                continue

            cell = ws.cell(row=int(row), column=col)
            old_value = _json_safe(cell.value)
            op = {
                "date": target_date.isoformat(),
                "period": write_period,
                "excel_period": excel_period,
                "account": account,
                "field": field,
                "cell": cell.coordinate,
                "old_value": old_value,
                "new_value": values[field],
                "verified": False,
            }
            if old_value not in (None, ""):
                report["overwrite_summary"]["items"].append({
                    "account": account,
                    "field": field,
                    "cell": cell.coordinate,
                    "old_value": old_value,
                    "new_value": values[field],
                })
            account_report["writes"].append(op)
            planned_writes.append(op)
        report["accounts"][account] = account_report

    for account_report in report["accounts"].values():
        report["errors"].extend(account_report["errors"])

    if report["errors"]:
        wb.close()
        logger.error("正式写入中断：%s", report["errors"])
        return finish()

    report["overwrite_summary"]["overwrite_count"] = len(report["overwrite_summary"]["items"])
    if report["overwrite_summary"]["overwrite_count"]:
        print_check_result("写入覆盖", "warn", f"本次会覆盖 {report['overwrite_summary']['overwrite_count']} 个已有值（旧值已记录在 write_report.json）")
        logger.warning("本次写入会覆盖 %s 个已有值。", report["overwrite_summary"]["overwrite_count"])

    for op in planned_writes:
        ws[op["cell"]].value = op["new_value"]
    try:
        wb.save(excel_path)
    except OSError as exc:
        wb.close()
        report["errors"].append(format_openpyxl_save_error(excel_path, exc))
        logger.error("保存小时报 Excel 失败：%s", exc)
        return finish()
    wb.close()
    report["self_check"]["filter_ui_metadata_restored"] = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        [sheet_name],
        logger,
    )
    logger.info("合并数据已写入并保存：%s", excel_path)

    read_back_values = _read_back_values(excel_path, sheet_name, [op["cell"] for op in planned_writes])
    for op in planned_writes:
        read_back = read_back_values.get(op["cell"])
        op["read_back_value"] = _json_safe(read_back)
        op["verified"] = _values_match(read_back, op["new_value"])
        if not op["verified"]:
            report["errors"].append(f"写入后复核不一致：{op['account']} {op['field']} {op['cell']}")

    report["writes"] = planned_writes
    for account in required_accounts:
        account_writes = [op for op in planned_writes if op["account"] == account]
        if account in report["accounts"]:
            report["accounts"][account]["writes"] = account_writes
    report["self_check"]["verification_passed"] = not report["errors"] and all(op["verified"] for op in planned_writes)
    logger.info("正式写入报告已输出：%s", out_path)
    return finish()


def write_merged_daily_data(
    config: dict[str, Any],
    root: Path,
    logger,
    target_date: str | None = None,
) -> dict[str, Any]:
    source_path = root / "reports" / "merged_daily_data.json"
    out_path = root / "reports" / "daily_write_report.json"
    report: dict[str, Any] = {
        "mode": "write-daily",
        "project_id": config.get("project_id"),
        "project_name": config.get("project_name"),
        "source": str(source_path),
        "excel_path": None,
        "sheet_name": DAILY_SHEET_NAME,
        "date": target_date,
        "backup_path": None,
        "writes": [],
        "accounts": {},
        "overwrite_summary": {"overwrite_count": 0, "items": []},
        "skipped_fields": ["总对话", "预约", "到诊", "就诊"],
        "errors": [],
        "self_check": {
            "merged_data_exists": source_path.exists(),
            "merged_validation_passed": False,
            "structure_passed": False,
            "backup_created": False,
            "wrote_forbidden_field": False,
            "filter_ui_metadata_restored": False,
            "verification_passed": False,
        },
    }

    def finish() -> dict[str, Any]:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return report

    if not source_path.exists():
        report["errors"].append(f"找不到日报合并数据文件：{source_path}")
        return finish()

    try:
        merged_data = json.loads(source_path.read_text(encoding="utf-8"))
        merged_date = _parse_report_date(merged_data.get("date"))
        if target_date and merged_date.isoformat() != target_date:
            report["errors"].append(f"日报合并数据日期不匹配：目标 {target_date}，文件 {merged_date.isoformat()}")
            return finish()
        excel_path = _resolve_excel_path(config, root)
    except Exception as exc:
        report["errors"].append(str(exc))
        return finish()

    report["excel_path"] = str(excel_path)
    report["date"] = merged_date.isoformat()

    required_accounts = get_required_accounts(config)
    merge_errors = validate_merged_daily_data(merged_data, required_accounts=required_accounts)
    if merge_errors:
        report["errors"].extend(merge_errors)
        return finish()
    report["self_check"]["merged_validation_passed"] = True

    if not excel_path.exists():
        report["errors"].append(f"找不到 Excel 文件：{excel_path}")
        return finish()

    daily_sheet_name = config.get("daily_sheet_name", DAILY_SHEET_NAME)
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    if daily_sheet_name not in wb.sheetnames:
        report["errors"].append(f"找不到 sheet：{daily_sheet_name}")
        wb.close()
        return finish()
    ws = wb[daily_sheet_name]
    structure = inspect_daily_worksheet(ws, config, str(excel_path))
    _write_daily_structure_outputs(structure, root, ws)
    structure_errors = structure.get("errors", [])
    if structure_errors:
        report["errors"].extend(structure_errors)
        wb.close()
        return finish()
    report["self_check"]["structure_passed"] = True
    logger.info("日报 Excel 结构识别通过。")

    backup_path = _make_backup(excel_path, root)
    report["backup_path"] = str(backup_path)
    report["self_check"]["backup_created"] = backup_path.exists()
    logger.info("日报正式写入前备份已创建：%s", backup_path)

    merged_values = _build_merged_value_map(ws)
    date_col = int(structure.get("date_column", {}).get("col", 0))
    row = _find_daily_target_row(ws, date_col, merged_date, merged_values)
    if row is None:
        report["errors"].append(f"百度 sheet 找不到日期行：{merged_date.isoformat()}")
        wb.close()
        return finish()

    account_values = merged_data.get("accounts", {})
    planned_writes: list[dict[str, Any]] = []
    for account in required_accounts:
        account_meta = structure.get("accounts", {}).get(account, {})
        fields = account_meta.get("fields", {})
        values = account_values.get(account, {})
        account_report = {"target_row": row, "writes": [], "errors": []}
        for field in DAILY_WRITE_FIELDS:
            field_meta = fields.get(field, {})
            if not field_meta.get("found"):
                account_report["errors"].append(f"找不到字段：{field}")
                continue
            if not field_meta.get("write_allowed"):
                account_report["errors"].append(f"字段不允许写入：{field}")
                continue
            if field not in values:
                account_report["errors"].append(f"日报合并数据缺少字段：{field}")
                continue
            col = int(field_meta["header_col"])
            target_errors = _validate_write_target(account, field, row, col, account_meta, [])
            if target_errors:
                account_report["errors"].extend(target_errors)
                continue
            cell = ws.cell(row=row, column=col)
            old_value = _json_safe(cell.value)
            op = {
                "date": merged_date.isoformat(),
                "account": account,
                "field": field,
                "cell": cell.coordinate,
                "old_value": old_value,
                "new_value": values[field],
                "verified": False,
            }
            if old_value not in (None, ""):
                report["overwrite_summary"]["items"].append({
                    "account": account,
                    "field": field,
                    "cell": cell.coordinate,
                    "old_value": old_value,
                    "new_value": values[field],
                })
            account_report["writes"].append(op)
            planned_writes.append(op)
        report["accounts"][account] = account_report

    for account_report in report["accounts"].values():
        report["errors"].extend(account_report["errors"])
    forbidden_written = {op["field"] for op in planned_writes} & set(report["skipped_fields"])
    if forbidden_written:
        report["self_check"]["wrote_forbidden_field"] = True
        report["errors"].append(f"检测到禁止写入字段：{', '.join(sorted(forbidden_written))}")
    if report["errors"]:
        wb.close()
        logger.error("日报正式写入中断：%s", report["errors"])
        return finish()

    report["overwrite_summary"]["overwrite_count"] = len(report["overwrite_summary"]["items"])
    if report["overwrite_summary"]["overwrite_count"]:
        print_check_result("日报写入覆盖", "warn", f"本次会覆盖 {report['overwrite_summary']['overwrite_count']} 个已有值（旧值已记录在 daily_write_report.json）")
        logger.warning("日报写入会覆盖 %s 个已有值。", report["overwrite_summary"]["overwrite_count"])

    for op in planned_writes:
        ws[op["cell"]].value = op["new_value"]
    try:
        wb.save(excel_path)
    except OSError as exc:
        wb.close()
        report["errors"].append(format_openpyxl_save_error(excel_path, exc))
        logger.error("保存日报 Excel 失败：%s", exc)
        return finish()
    wb.close()
    report["self_check"]["filter_ui_metadata_restored"] = _restore_sheet_filter_protection_metadata(
        excel_path,
        backup_path,
        [daily_sheet_name],
        logger,
    )
    logger.info("日报合并数据已写入并保存：%s", excel_path)

    read_back_values = _read_back_values(excel_path, daily_sheet_name, [op["cell"] for op in planned_writes])
    for op in planned_writes:
        read_back = read_back_values.get(op["cell"])
        op["read_back_value"] = _json_safe(read_back)
        op["verified"] = _values_match(read_back, op["new_value"])
        if not op["verified"]:
            report["errors"].append(f"日报写入后复核不一致：{op['account']} {op['field']} {op['cell']}")

    report["writes"] = planned_writes
    for account in required_accounts:
        if account in report["accounts"]:
            report["accounts"][account]["writes"] = [op for op in planned_writes if op["account"] == account]
    report["self_check"]["verification_passed"] = not report["errors"] and all(op["verified"] for op in planned_writes)
    logger.info("日报正式写入报告已输出：%s", out_path)
    return finish()


def write_account_data(*args, **kwargs):
    return write_merged_hourly_data(*args, **kwargs)
