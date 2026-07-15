from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from modules.baidu_report_api import fetch_baidu_api_probe
from modules.data_merger import build_merged_hourly_data, normalize_period_for_report
from modules.excel_inspector import inspect_excel_structure
from modules.excel_writer import (
    WRITE_FIELDS,
    _build_merged_value_map,
    _find_target_row,
    _normalize_period_for_excel,
    _resolve_excel_path,
    _resolve_global_row_fields,
    _validate_write_target,
)
from modules.kst_export_parser import (
    find_latest_kst_export,
    parse_kst_export_file,
    write_empty_kst_export_result,
)
from modules.validators import get_required_accounts, validate_merged_hourly_data


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _preview_excel_targets(
    config: dict[str, Any],
    root: Path,
    logger,
    merged: dict[str, Any],
    period: str,
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    excel_path = _resolve_excel_path(config, root)
    structure = inspect_excel_structure(config=config, root=root, logger=logger)
    errors = list(structure.get("errors") or [])
    preview: list[dict[str, Any]] = []
    if errors:
        return preview, errors, structure

    sheet_name = config.get("sheet_name", "时段数据")
    workbook = load_workbook(excel_path, data_only=False, read_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            return preview, [f"找不到 sheet：{sheet_name}"], structure
        worksheet = workbook[sheet_name]
        merged_values = _build_merged_value_map(worksheet)
        protected_regions = structure.get("summary_regions", [])
        target_date = date.fromisoformat(str(merged.get("date")))

        for account in get_required_accounts(config):
            account_meta = (structure.get("accounts") or {}).get(account, {})
            fields = account_meta.get("fields", {})
            date_field, period_field = _resolve_global_row_fields(structure, account_meta)
            row = _find_target_row(
                worksheet,
                int(date_field.get("header_col", 0)),
                int(period_field.get("header_col", 0)),
                target_date,
                period,
                merged_values,
            )
            if row is None:
                errors.append(f"账户 {account} 找不到日期 {target_date.isoformat()} 时段 {_normalize_period_for_excel(period)}")
                continue

            values = (merged.get("accounts") or {}).get(account, {})
            for field in WRITE_FIELDS:
                field_meta = fields.get(field, {})
                if not field_meta.get("found"):
                    errors.append(f"账户 {account} 找不到字段：{field}")
                    continue
                column = int(field_meta["header_col"])
                target_errors = _validate_write_target(
                    account,
                    field,
                    int(row),
                    column,
                    account_meta,
                    protected_regions,
                )
                if target_errors:
                    errors.extend(target_errors)
                    continue
                cell = worksheet.cell(row=int(row), column=column)
                preview.append({
                    "account": account,
                    "field": field,
                    "cell": cell.coordinate,
                    "old_value": cell.value,
                    "new_value": values.get(field),
                })
    finally:
        workbook.close()
    return preview, errors, structure


def simulate_baidu_api_hourly(
    config: dict[str, Any],
    root: Path,
    logger,
    period: str | None,
    target_date: str | None = None,
) -> dict[str, Any]:
    normalized_period = normalize_period_for_report(period)
    selected_date = target_date or date.today().isoformat()
    output_path = root / "reports" / "baidu_api_hourly_simulation_report.json"
    merged_path = root / "reports" / "baidu_api_hourly_simulated_merged.json"
    report: dict[str, Any] = {
        "mode": "simulate-baidu-api-hourly",
        "project_id": config.get("project_id"),
        "project_name": config.get("project_name"),
        "date": selected_date,
        "period": normalized_period,
        "started_at": datetime.now().isoformat(timespec="seconds"),
        "finished_at": None,
        "api": None,
        "kst": None,
        "merged": None,
        "planned_writes": [],
        "errors": [],
        "self_check": {
            "api_passed": False,
            "kst_passed": False,
            "merge_passed": False,
            "excel_structure_passed": False,
            "wrote_excel": False,
            "created_backup": False,
            "production_output_replaced": False,
        },
        "outputs": {
            "simulation_report": str(output_path),
            "simulated_merged": str(merged_path),
        },
    }

    api_report = fetch_baidu_api_probe(
        config=config,
        root=root,
        logger=logger,
        target_date=selected_date,
        period=normalized_period,
    )
    report["api"] = {
        "source": api_report.get("source"),
        "account_count": len(api_report.get("accounts") or {}),
        "errors": api_report.get("errors") or [],
        "diagnostics": api_report.get("diagnostics") or {},
    }
    if api_report.get("errors"):
        report["errors"].extend(api_report["errors"])
    else:
        report["self_check"]["api_passed"] = True

    export_file = find_latest_kst_export(root, config)
    if export_file is None:
        kst_result = write_empty_kst_export_result(
            config,
            root,
            normalized_period,
            "未找到 30 分钟内的快商通导出文件，按 0 对话处理",
        )
    else:
        kst_result = parse_kst_export_file(export_file, config, root, normalized_period)
    kst_report = _read_json(root / "reports" / "kst_dialog_data.json")
    parse_report = kst_result.get("parse_report") or {}
    report["kst"] = {
        "export_file": str(export_file or ""),
        "no_export_file": bool((kst_report.get("summary") or {}).get("no_export_file")),
        "account_count": len(kst_report.get("accounts") or {}),
        "parse_passed": bool(parse_report.get("passed")),
    }
    if not parse_report.get("passed"):
        report["errors"].extend(parse_report.get("errors") or ["快商通数据解析失败"])
    else:
        report["self_check"]["kst_passed"] = True

    if not report["errors"]:
        merged = build_merged_hourly_data(
            api_report,
            kst_report,
            period=normalized_period,
            baidu_source="reports/baidu_api_probe_report.json",
            kst_source="reports/kst_dialog_data.json",
            required_accounts=get_required_accounts(config),
        )
        merged["project_id"] = config.get("project_id")
        merged["project_name"] = config.get("project_name")
        merge_errors = validate_merged_hourly_data(
            merged,
            api_report,
            kst_report,
            get_required_accounts(config),
        )
        report["merged"] = merged
        report["errors"].extend(merge_errors)
        if not merge_errors:
            report["self_check"]["merge_passed"] = True
            _write_json(merged_path, merged)

    if report["self_check"]["merge_passed"]:
        preview, preview_errors, structure = _preview_excel_targets(
            config,
            root,
            logger,
            report["merged"],
            normalized_period,
        )
        report["planned_writes"] = preview
        report["errors"].extend(preview_errors)
        report["excel_preview"] = {
            "excel_path": structure.get("excel_path"),
            "sheet_name": structure.get("sheet_name"),
            "planned_write_count": len(preview),
        }
        report["self_check"]["excel_structure_passed"] = not preview_errors

    report["finished_at"] = datetime.now().isoformat(timespec="seconds")
    report["self_check"]["passed"] = not report["errors"]
    _write_json(output_path, report)
    logger.info("百度 API 小时报模拟完成：%s；结果：%s", output_path, "通过" if not report["errors"] else "失败")
    return report
