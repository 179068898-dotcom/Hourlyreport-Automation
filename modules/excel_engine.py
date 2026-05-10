from __future__ import annotations

import importlib.util
import shutil
from pathlib import Path
from typing import Any


SUPPORTED_ENGINES = {"openpyxl", "excel_com"}


def get_excel_engine(project: dict[str, Any] | None = None, config: dict[str, Any] | None = None) -> str:
    project = project or {}
    config = config or {}
    engine = (
        project.get("excel", {}).get("engine")
        if isinstance(project.get("excel"), dict)
        else None
    )
    engine = engine or project.get("excel_engine") or config.get("excel_engine") or "openpyxl"
    return str(engine).strip().lower() or "openpyxl"


def get_project_excel_path(project: dict[str, Any], root: Path) -> Path | None:
    excel = project.get("excel") if isinstance(project.get("excel"), dict) else {}
    value = excel.get("path") or project.get("excel_path")
    if not value:
        return None
    path = Path(value)
    return path if path.is_absolute() else root / path


def get_project_sheet_name(project: dict[str, Any], report_type: str) -> str:
    excel = project.get("excel") if isinstance(project.get("excel"), dict) else {}
    sheets = project.get("sheets") if isinstance(project.get("sheets"), dict) else {}
    if report_type == "daily":
        return str(excel.get("daily_sheet") or sheets.get("daily") or "百度")
    return str(excel.get("hourly_sheet") or sheets.get("hourly") or "时段数据")


def is_openpyxl_installed() -> bool:
    return importlib.util.find_spec("openpyxl") is not None


def test_openpyxl_save_copy(excel_path: Path, root: Path) -> dict[str, Any]:
    if not excel_path.exists():
        return {"passed": False, "message": f"目标 xlsx 不存在：{excel_path}"}
    if excel_path.suffix.lower() != ".xlsx":
        return {"passed": False, "message": f"openpyxl 模式只支持 .xlsx 文件：{excel_path.name}"}

    tmp_dir = root / "reports" / "_doctor_tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    tmp_path = tmp_dir / f"{excel_path.stem}.openpyxl_test.xlsx"
    try:
        from openpyxl import load_workbook

        shutil.copy2(excel_path, tmp_path)
        wb = load_workbook(tmp_path, data_only=False, read_only=False)
        wb.save(tmp_path)
        verify_wb = load_workbook(tmp_path, data_only=False, read_only=True)
        verify_wb.close()
        return {"passed": True, "message": f"openpyxl 临时副本写入保存测试通过：{tmp_path}", "path": str(tmp_path)}
    except Exception as exc:
        return {"passed": False, "message": format_openpyxl_save_error(excel_path, exc), "error": str(exc)}
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
            if tmp_dir.exists() and not any(tmp_dir.iterdir()):
                tmp_dir.rmdir()
        except OSError:
            pass


def format_openpyxl_save_error(excel_path: Path, exc: BaseException) -> str:
    return (
        f"保存 xlsx 失败：{excel_path}。"
        f"请关闭 WPS 中的目标文件后重试；如果你用 Microsoft Excel 打开了它，也请先关闭。"
        f"原始错误：{exc}"
    )
