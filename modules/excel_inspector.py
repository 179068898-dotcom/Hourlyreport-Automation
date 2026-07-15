from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from modules.text_normalizer import normalize_for_display, normalize_text
from modules.validators import validate_excel_report, validate_excel_report_v2


def _write_excel_account_regions_report(config: dict[str, Any], root: Path, structure_report: dict[str, Any]) -> dict[str, Any]:
    configured_accounts = list((config.get("accounts") or {}).keys())
    detected_regions = []
    for account_name, meta in (structure_report.get("accounts") or {}).items():
        if not meta.get("found"):
            continue
        detected_regions.append({
            "account_name": account_name,
            "row": meta.get("title_row"),
            "column": meta.get("title_col"),
            "title_cell": meta.get("title_cell"),
        })
    detected_names = [item["account_name"] for item in detected_regions]
    report = {
        "project_id": config.get("project_id"),
        "project_name": config.get("project_name"),
        "excel_path": structure_report.get("excel_path") or config.get("excel_path"),
        "sheet_name": structure_report.get("sheet_name") or config.get("sheet_name", "时段数据"),
        "detected_account_regions": detected_regions,
        "configured_excel_accounts": configured_accounts,
        "missing_configured_accounts": [name for name in configured_accounts if name not in detected_names],
        "extra_detected_accounts": [name for name in detected_names if name not in configured_accounts],
    }
    path = root / "reports" / "excel_account_regions.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def _build_merged_value_map(ws: Worksheet) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        if value is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value
    return merged_values


def _build_merged_bounds_map(ws: Worksheet) -> dict[tuple[int, int], dict[str, int]]:
    merged_bounds: dict[tuple[int, int], dict[str, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        bounds = {
            "min_row": merged_range.min_row,
            "max_row": merged_range.max_row,
            "min_col": merged_range.min_col,
            "max_col": merged_range.max_col,
        }
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_bounds[(row, col)] = bounds
    return merged_bounds


def _get_merged_value(ws: Worksheet, row: int, col: int, merged_values: dict[tuple[int, int], Any] | None = None) -> Any:
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    if merged_values is not None:
        return merged_values.get((row, col))
    coord = cell.coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None


def _scan_non_empty_cells(ws: Worksheet, merged_values: dict[tuple[int, int], Any] | None = None) -> list[dict[str, Any]]:
    if merged_values is None:
        merged_values = _build_merged_value_map(ws)
    rows: list[dict[str, Any]] = []
    materialized_cells = getattr(ws, "_cells", None)
    if not isinstance(materialized_cells, dict):
        coordinates = (
            (row, col)
            for row in range(1, (ws.max_row or 1) + 1)
            for col in range(1, (ws.max_column or 1) + 1)
        )
    else:
        coordinates = iter(sorted(set(materialized_cells) | set(merged_values)))

    for row, col in coordinates:
        if isinstance(materialized_cells, dict):
            cell = materialized_cells.get((row, col))
            value = cell.value if cell is not None else None
            if value is None:
                value = merged_values.get((row, col))
        else:
            value = _get_merged_value(ws, row, col, merged_values)
        if value is None:
            continue
        raw_text = normalize_for_display(value)
        normalized_text = normalize_text(value)
        if not normalized_text:
            continue
        rows.append(
            {
                "row": row,
                "col": col,
                "address": f"{get_column_letter(col)}{row}",
                "raw_text": raw_text,
                "normalized_text": normalized_text,
            }
        )
    return rows


def _write_sheet_dump(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["row", "col", "address", "raw_text", "normalized_text"])
        writer.writeheader()
        writer.writerows(rows)


def _find_account_titles(rows: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for account, info in config["accounts"].items():
        aliases = info.get("aliases", []) + [account, info.get("excel_name", ""), info.get("baidu_name", "")]
        normalized_aliases = [normalize_text(alias) for alias in aliases if normalize_text(alias)]
        matches = [
            row
            for row in rows
            if any(alias in row["normalized_text"] for alias in normalized_aliases)
        ]
        if matches:
            matches.sort(key=lambda item: (item["row"], item["col"]))
            first = matches[0]
            result[account] = {
                "found": True,
                "title_cell": first["address"],
                "title_row": first["row"],
                "title_col": first["col"],
                "raw_text": first["raw_text"],
                "all_matches": matches[:20],
            }
        else:
            result[account] = {"found": False, "all_matches": []}
    return result


def _build_account_ranges(account_titles: dict[str, dict[str, Any]], ws: Worksheet, merged_bounds: dict[tuple[int, int], dict[str, int]] | None = None) -> dict[str, dict[str, Any]]:
    found = [(account, meta) for account, meta in account_titles.items() if meta.get("found")]
    if not found:
        return account_titles
    if merged_bounds is None:
        merged_bounds = _build_merged_bounds_map(ws)

    title_rows = {int(meta["title_row"]) for _account, meta in found}
    if len(title_rows) == 1:
        found.sort(key=lambda item: item[1]["title_col"])
        for index, (_account, meta) in enumerate(found):
            title_row = int(meta["title_row"])
            title_col = int(meta["title_col"])
            bounds = merged_bounds.get((title_row, title_col), {})
            min_col = int(bounds.get("min_col", title_col))
            if bounds.get("max_col"):
                max_col = int(bounds["max_col"])
            elif index + 1 < len(found):
                max_col = int(found[index + 1][1]["title_col"]) - 1
            else:
                max_col = ws.max_column
            meta["layout"] = "column_block"
            meta["range"] = {
                "min_row": title_row,
                "max_row": ws.max_row,
                "min_col": min_col,
                "max_col": max_col,
            }
        return account_titles

    found.sort(key=lambda item: item[1]["title_row"])
    for index, (_account, meta) in enumerate(found):
        start = int(meta["title_row"])
        end = int(found[index + 1][1]["title_row"]) - 1 if index + 1 < len(found) else ws.max_row
        meta["layout"] = "row_block"
        meta["range"] = {"min_row": start, "max_row": end, "min_col": 1, "max_col": ws.max_column}
    return account_titles


def _match_field_header(text: str, aliases: list[str]) -> tuple[bool, bool]:
    normalized_aliases = [normalize_text(alias) for alias in aliases]
    return _match_normalized_field_header(text, normalized_aliases)


def _match_normalized_field_header(text: str, normalized_aliases: list[str]) -> tuple[bool, bool]:
    if any(alias and text == alias for alias in normalized_aliases):
        return True, True
    if any(alias and len(alias) >= 2 and alias in text for alias in normalized_aliases):
        return True, False
    return False, False


def _find_fields_in_range(
    ws: Worksheet,
    aliases_config: dict[str, list[str]],
    merged_values: dict[tuple[int, int], Any],
    *,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    for field_name, aliases in aliases_config.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases]
        exact_cells = []
        partial_cells = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                value = _get_merged_value(ws, row, col, merged_values)
                normalized = normalize_text(value)
                if not normalized:
                    continue
                matched, exact = _match_normalized_field_header(normalized, normalized_aliases)
                if matched:
                    target = exact_cells if exact else partial_cells
                    target.append(
                        {
                            "row": row,
                            "col": col,
                            "address": ws.cell(row=row, column=col).coordinate,
                            "raw_text": normalize_for_display(value),
                        }
                    )
        matches = exact_cells or partial_cells
        if matches:
            matches.sort(key=lambda item: (item["row"], item["col"]))
            first = matches[0]
            fields[field_name] = {
                "found": True,
                "header_cell": first["address"],
                "header_row": first["row"],
                "header_col": first["col"],
                "raw_text": first["raw_text"],
                "all_matches": matches[:20],
            }
        else:
            fields[field_name] = {"found": False, "all_matches": []}
    return fields


def _find_fields_in_rows(
    rows: list[dict[str, Any]],
    aliases_config: dict[str, list[str]],
) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    for field_name, aliases in aliases_config.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases]
        exact_cells = []
        partial_cells = []
        for row in rows:
            matched, exact = _match_normalized_field_header(row["normalized_text"], normalized_aliases)
            if matched:
                (exact_cells if exact else partial_cells).append(
                    {
                        "row": row["row"],
                        "col": row["col"],
                        "address": row["address"],
                        "raw_text": row["raw_text"],
                    }
                )
        matches = exact_cells or partial_cells
        if matches:
            matches.sort(key=lambda item: (item["row"], item["col"]))
            first = matches[0]
            fields[field_name] = {
                "found": True,
                "header_cell": first["address"],
                "header_row": first["row"],
                "header_col": first["col"],
                "raw_text": first["raw_text"],
                "all_matches": matches[:20],
            }
        else:
            fields[field_name] = {"found": False, "all_matches": []}
    return fields


def _find_global_fields(
    ws: Worksheet,
    config: dict[str, Any],
    merged_values: dict[tuple[int, int], Any],
    rows: list[dict[str, Any]] | None = None,
) -> dict[str, dict[str, Any]]:
    aliases_config = config.get("field_aliases", {})
    global_aliases = {name: aliases for name, aliases in aliases_config.items() if name in {"日期", "时段"}}
    if not global_aliases:
        return {}
    if rows is not None:
        return _find_fields_in_rows(rows, global_aliases)
    return _find_fields_in_range(
        ws,
        global_aliases,
        merged_values,
        min_row=1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
    )


def _find_fields_for_account(ws: Worksheet, meta: dict[str, Any], config: dict[str, Any], merged_values: dict[tuple[int, int], Any]) -> dict[str, dict[str, Any]]:
    fields: dict[str, dict[str, Any]] = {}
    if not meta.get("found") or "range" not in meta:
        return fields

    block = meta["range"]
    aliases_config = config.get("field_aliases", {})
    for field_name, aliases in aliases_config.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases]
        exact_cells = []
        partial_cells = []
        if field_name in {"日期", "时段"}:
            header_min_row = 1
            header_max_row = min(ws.max_row, max(int(block["max_row"]), int(block["min_row"]) + 10))
            min_col, max_col = 1, ws.max_column
        else:
            header_min_row = int(block["min_row"]) + 1
            header_max_row = min(ws.max_row, int(block["min_row"]) + 10)
            min_col, max_col = int(block["min_col"]), int(block["max_col"])

        for row in range(header_min_row, header_max_row + 1):
            for col in range(min_col, max_col + 1):
                value = _get_merged_value(ws, row, col, merged_values)
                normalized = normalize_text(value)
                if not normalized:
                    continue
                matched, exact = _match_normalized_field_header(normalized, normalized_aliases)
                if matched:
                    target = exact_cells if exact else partial_cells
                    target.append(
                        {
                            "row": row,
                            "col": col,
                            "address": ws.cell(row=row, column=col).coordinate,
                            "raw_text": normalize_for_display(value),
                        }
                    )
        matches = exact_cells or partial_cells
        if matches:
            matches.sort(key=lambda item: (item["row"], item["col"]))
            first = matches[0]
            fields[field_name] = {
                "found": True,
                "header_cell": first["address"],
                "header_row": first["row"],
                "header_col": first["col"],
                "raw_text": first["raw_text"],
                "all_matches": matches[:20],
            }
        else:
            fields[field_name] = {"found": False, "all_matches": []}
    return fields


def _find_fields_for_account_v2(ws: Worksheet, meta: dict[str, Any], config: dict[str, Any], merged_values: dict[tuple[int, int], Any]) -> dict[str, dict[str, Any]]:
    if not meta.get("found") or "range" not in meta:
        return {}
    block = meta["range"]
    aliases_config = config.get("field_aliases", {})
    account_aliases = {name: aliases for name, aliases in aliases_config.items() if name not in {"日期", "时段"}}
    return _find_fields_in_range(
        ws,
        account_aliases,
        merged_values,
        min_row=int(block["min_row"]) + 1,
        max_row=min(ws.max_row, int(block["min_row"]) + 10),
        min_col=int(block["min_col"]),
        max_col=int(block["max_col"]),
    )


def _find_summary_regions(rows: list[dict[str, Any]], merged_bounds: dict[tuple[int, int], dict[str, int]]) -> list[dict[str, Any]]:
    summary_keywords = [normalize_text(text) for text in ["每日时段统计数据", "汇总"]]
    regions = []
    seen: set[tuple[int, int, int, int]] = set()
    for row in rows:
        text = row["normalized_text"]
        if not any(keyword and keyword in text for keyword in summary_keywords):
            continue
        bounds = merged_bounds.get((row["row"], row["col"]), {"min_row": row["row"], "max_row": row["row"], "min_col": row["col"], "max_col": row["col"]})
        key = (bounds["min_row"], bounds["max_row"], bounds["min_col"], bounds["max_col"])
        if key in seen:
            continue
        seen.add(key)
        regions.append({"title_cell": row["address"], "raw_text": row["raw_text"], "range": bounds, "write_allowed": False})
    return regions


def _find_suspicious_titles(rows: list[dict[str, Any]], config: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    config = config or {}
    account_keywords = [
        str(item.get("excel_name") or item.get("standard_name") or "")
        for item in config.get("accounts", {}).values()
        if isinstance(item, dict)
    ]
    keywords = [*account_keywords, "账户", "数据", "时段", "展现", "点击", "消费", "对话", "转潜"]
    normalized_keywords = [normalize_text(keyword) for keyword in keywords]
    suspicious = []
    for row in rows:
        text = row["normalized_text"]
        if any(keyword and keyword in text for keyword in normalized_keywords):
            suspicious.append(row)
    return suspicious[:300]


def dump_sheet_text(config: dict[str, Any], root: Path, logger) -> Path:
    excel_path = Path(config["excel_path"])
    sheet_name = config.get("sheet_name", "时段数据")
    if not excel_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{excel_path}")
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    if sheet_name not in wb.sheetnames:
        available_sheets = wb.sheetnames
        wb.close()
        raise ValueError(f"找不到 sheet：{sheet_name}，当前 sheet：{available_sheets}")
    ws = wb[sheet_name]
    merged_values = _build_merged_value_map(ws)
    rows = _scan_non_empty_cells(ws, merged_values)
    out_path = root / "reports" / "sheet_text_dump.csv"
    _write_sheet_dump(rows, out_path)
    wb.close()
    logger.info("sheet 文本扫描完成，共 %s 个非空文本单元格：%s", len(rows), out_path)
    return out_path


def inspect_excel_worksheet(ws: Worksheet, config: dict[str, Any], excel_path: str | Path) -> dict[str, Any]:
    report: dict[str, Any] = {
        "excel_path": str(excel_path),
        "sheet_name": ws.title,
        "sheet_found": True,
        "accounts": {},
        "suspicious_titles": [],
        "errors": [],
        "max_row": ws.max_row,
        "max_col": ws.max_column,
    }
    merged_values = _build_merged_value_map(ws)
    merged_bounds = _build_merged_bounds_map(ws)
    report["merged_range_count"] = len(ws.merged_cells.ranges)
    rows = _scan_non_empty_cells(ws, merged_values)

    report["summary_regions"] = _find_summary_regions(rows, merged_bounds)
    accounts = _find_account_titles(rows, config)
    accounts = _build_account_ranges(accounts, ws, merged_bounds)
    for account, meta in accounts.items():
        meta["fields"] = _find_fields_for_account_v2(ws, meta, config, merged_values)
    report["accounts"] = accounts
    report["global_fields"] = _find_global_fields(ws, config, merged_values, rows=rows)
    report["suspicious_titles"] = _find_suspicious_titles(rows, config)
    report["errors"].extend(validate_excel_report_v2(report, required_accounts=list(accounts.keys())))
    return report


def inspect_excel_structure(config: dict[str, Any], root: Path, logger) -> dict[str, Any]:
    excel_path = Path(config["excel_path"])
    sheet_name = config.get("sheet_name", "时段数据")
    report: dict[str, Any] = {
        "excel_path": str(excel_path),
        "sheet_name": sheet_name,
        "sheet_found": False,
        "accounts": {},
        "suspicious_titles": [],
        "errors": [],
    }
    if not excel_path.exists():
        report["errors"].append(f"找不到 Excel 文件：{excel_path}")
        _write_excel_account_regions_report(config, root, report)
        return report

    wb = load_workbook(excel_path, data_only=False, read_only=False)
    available_sheets = wb.sheetnames
    if sheet_name not in available_sheets:
        report["available_sheets"] = available_sheets
        report["errors"].append(f"找不到 sheet：{sheet_name}")
        _write_excel_account_regions_report(config, root, report)
        wb.close()
        return report

    report = inspect_excel_worksheet(wb[sheet_name], config, excel_path)
    report["available_sheets"] = available_sheets
    _write_excel_account_regions_report(config, root, report)

    if report["errors"]:
        logger.warning("Excel 结构识别存在问题：%s", report["errors"])
    else:
        logger.info("Excel 结构识别通过")
    wb.close()
    return report
