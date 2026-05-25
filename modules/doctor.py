from __future__ import annotations

import importlib.metadata
import json
import platform
import sys
from pathlib import Path
from typing import Any

from modules.browser_manager import get_browser_settings
from modules.chrome_debug import ensure_chrome_debug_ready, find_chrome_executable
from modules.excel_engine import (
    get_excel_engine,
    get_project_excel_path,
    get_project_sheet_name,
    is_openpyxl_installed,
    test_openpyxl_save_copy,
)
from modules.excel_inspector import inspect_excel_structure
from modules.kst_export_parser import find_latest_kst_export
from modules.project_config import get_current_project, load_app_config, validate_project_config
from modules.baidu_multi_source import resolve_baidu_sources


REQUIREMENT_NAME_MAP = {
    "python-dateutil": "python-dateutil",
    "pywin32": "pywin32",
}
EXCEL_COM_ONLY_REQUIREMENTS = {"xlwings", "pywin32"}


def _ok(message: str, detail: Any = None) -> dict[str, Any]:
    return {"passed": True, "level": "ok", "message": message, "detail": detail}


def _warn(message: str, detail: Any = None) -> dict[str, Any]:
    return {"passed": False, "level": "warning", "message": message, "detail": detail}


def _resolve(root: Path, value: str | Path | None) -> Path | None:
    if value in (None, ""):
        return None
    path = Path(value)
    return path if path.is_absolute() else root / path


def _similar_excel_names(excel_path: Path | None) -> list[str]:
    if not excel_path:
        return []
    parent = excel_path.parent
    if not parent.exists():
        return []
    keywords = [item for item in ["竞价数据", "长沙", "npx", excel_path.suffix.lower()] if item]
    matches: list[str] = []
    for candidate in parent.iterdir():
        if not candidate.is_file():
            continue
        name = candidate.name
        lowered = name.lower()
        if candidate == excel_path:
            continue
        if any(keyword == excel_path.suffix.lower() and lowered.endswith(keyword) or keyword in name for keyword in keywords):
            matches.append(name)
    return sorted(set(matches))[:20]


def _build_missing_excel_message(excel_path: Path | None, project_id: str) -> tuple[str, dict[str, Any]]:
    if not excel_path:
        hint = f"请修改 configs/projects/{project_id}.json 的 excel.path" if project_id else "请在项目配置中设置 excel.path"
        return hint, {"excel_path": None, "parent_exists": False, "similar_files": []}
    similar_files = _similar_excel_names(excel_path)
    detail = {
        "excel_path": str(excel_path),
        "parent": str(excel_path.parent),
        "parent_exists": excel_path.parent.exists(),
        "similar_files": similar_files,
    }
    message = f"找不到目标 Excel：{excel_path}"
    if excel_path.parent.exists():
        message += f"；父目录存在：{excel_path.parent}"
    if similar_files:
        message += f"；相似文件：{', '.join(similar_files)}"
    return message, detail


def _check_python() -> dict[str, Any]:
    version = platform.python_version()
    if sys.version_info >= (3, 10):
        return _ok(f"Python 版本正常：{version}")
    return _warn(f"Python 版本偏低：{version}，建议使用 3.10 或以上")


def _parse_requirement_name(line: str) -> str | None:
    line = line.strip()
    if not line or line.startswith("#"):
        return None
    line = line.split(";", 1)[0].strip()
    for sep in [">=", "==", "<=", "~=", ">", "<"]:
        if sep in line:
            line = line.split(sep, 1)[0].strip()
    return line or None


def _check_requirements(root: Path, excel_engine: str = "openpyxl") -> dict[str, Any]:
    req = root / "requirements.txt"
    if not req.exists():
        return _warn("未找到 requirements.txt")
    requirement_files = [req]
    optional_req = root / "requirements-excel-com.txt"
    if excel_engine == "excel_com" and optional_req.exists():
        requirement_files.append(optional_req)
    missing = []
    checked = []
    skipped_optional = []
    for requirement_file in requirement_files:
        for line in requirement_file.read_text(encoding="utf-8").splitlines():
            name = _parse_requirement_name(line)
            if not name:
                continue
            if excel_engine == "openpyxl" and name in EXCEL_COM_ONLY_REQUIREMENTS:
                skipped_optional.append(name)
                continue
            package_name = REQUIREMENT_NAME_MAP.get(name, name)
            try:
                importlib.metadata.version(package_name)
                checked.append(name)
            except importlib.metadata.PackageNotFoundError:
                missing.append(name)
    if missing:
        return _warn(
            f"依赖未安装完整，缺少：{', '.join(missing)}",
            {"checked": checked, "missing": missing, "skipped_optional": skipped_optional},
        )
    message = f"requirements 依赖检查通过，共 {len(checked)} 项"
    if skipped_optional:
        message += f"（openpyxl 模式已跳过 Excel COM 备用依赖：{', '.join(skipped_optional)}）"
    return _ok(message, {"checked": checked, "skipped_optional": skipped_optional})


def _check_excel_available() -> dict[str, Any]:
    if platform.system() != "Windows":
        return _warn("当前不是 Windows，无法检查 Microsoft Excel COM")
    candidates = [
        Path("C:/Program Files/Microsoft Office/root/Office16/EXCEL.EXE"),
        Path("C:/Program Files (x86)/Microsoft Office/root/Office16/EXCEL.EXE"),
        Path("C:/Program Files/Microsoft Office/Office16/EXCEL.EXE"),
        Path("C:/Program Files (x86)/Microsoft Office/Office16/EXCEL.EXE"),
    ]
    for path in candidates:
        if path.exists():
            return _ok(f"Microsoft Excel 已找到：{path}", {"path": str(path)})
    return _warn("未找到 Microsoft Excel。如使用 openpyxl 则可忽略此项。")


def _check_excel_engine(engine: str) -> dict[str, Any]:
    if engine == "openpyxl":
        return _ok("当前 Excel 写入引擎：openpyxl，适合 WPS 环境，不要求安装 Microsoft Excel", {"engine": engine})
    if engine == "excel_com":
        return _ok("当前 Excel 写入引擎：excel_com", {"engine": engine})
    return _warn(f"不支持的 Excel 写入引擎：{engine}", {"engine": engine})


def _check_openpyxl_installed() -> dict[str, Any]:
    if is_openpyxl_installed():
        return _ok("openpyxl 已安装，可直接读写 xlsx 文件")
    return _warn("openpyxl 未安装，请先运行 install_env.bat 安装依赖")


def _check_secret_profile(secrets_file: Path | None, project: dict[str, Any]) -> dict[str, Any]:
    sources = resolve_baidu_sources(project) if project else []
    profiles = [str(source.get("credential_profile") or "") for source in sources if source.get("credential_profile")]
    if not secrets_file or not secrets_file.exists():
        return _warn("百度账号未配置，如需自动登录请联系管理员")
    try:
        data = json.loads(secrets_file.read_text(encoding="utf-8-sig"))
    except Exception:
        return _warn("百度凭据文件无法读取")
    missing = []
    configured = []
    for profile in profiles:
        item = data.get("baidu", {}).get(profile)
        if isinstance(item, dict) and item.get("username") and item.get("password"):
            configured.append(profile)
        else:
            missing.append(profile)
    if missing:
        return _warn(f"百度账号未配置：{', '.join(missing)}", {"configured_profiles": configured, "missing_profiles": missing})
    return _ok(f"百度账号已配置：{', '.join(configured)}", {"configured_profiles": configured})


def _check_baidu_sources(project: dict[str, Any]) -> dict[str, Any]:
    sources = resolve_baidu_sources(project) if project else []
    details = []
    errors: list[str] = []
    excel_accounts = [
        str(account.get("standard_name") or "")
        for account in project.get("excel_accounts", []) or project.get("accounts", [])
        if account.get("standard_name")
    ]
    candidate_names: list[str] = []
    for source in sources:
        source_id = str(source.get("source_id") or "")
        accounts = source.get("accounts") or []
        baidu_name_owner: dict[str, str] = {}
        duplicate_names = []
        missing_kst_ids = []
        for account in accounts:
            standard = str(account.get("standard_name") or "")
            if standard:
                candidate_names.append(standard)
            if not account.get("kst_ids"):
                missing_kst_ids.append(standard)
            for baidu_name in account.get("baidu_names") or []:
                name = str(baidu_name)
                if name in baidu_name_owner:
                    duplicate_names.append({"baidu_name": name, "first": baidu_name_owner[name], "second": standard})
                else:
                    baidu_name_owner[name] = standard
        if not source.get("credential_profile"):
            errors.append(f"{source_id} 缺少 credential_profile")
        if not accounts:
            errors.append(f"{source_id} accounts 未配置")
        if duplicate_names:
            errors.append(f"{source_id} 存在重复 baidu_name")
        if missing_kst_ids:
            errors.append(f"{source_id} 缺少商务通推广备注 ID：{', '.join(missing_kst_ids)}")
        details.append({
            "source_id": source_id,
            "source_name": source.get("source_name"),
            "credential_profile": source.get("credential_profile"),
            "credential_profile_exists": bool(source.get("credential_profile")),
            "account_count": len(accounts),
            "candidate_accounts_count": len(accounts),
            "has_duplicate_baidu_name": bool(duplicate_names),
            "duplicate_baidu_names": duplicate_names,
            "missing_kst_ids": missing_kst_ids,
        })
    candidate_only = [name for name in dict.fromkeys(candidate_names) if name not in set(excel_accounts)]
    detail = {
        "source_count": len(sources),
        "baidu_sources_count": len(sources),
        "excel_accounts_count": len(excel_accounts),
        "baidu_candidate_accounts_count": len(dict.fromkeys(candidate_names)),
        "sources": details,
        "excel_accounts": excel_accounts,
        "candidate_only_accounts": candidate_only,
        "candidate_only_note": "候选账户不在 Excel 中，不算错误。",
    }
    if errors:
        return _warn("百度来源配置不完整：" + "；".join(errors), detail)
    return _ok(f"百度来源配置通过：{len(sources)} 个", detail)


def _check_chrome(config: dict[str, Any]) -> dict[str, Any]:
    chrome_exe = find_chrome_executable(config)
    if chrome_exe:
        settings = get_browser_settings(config)
        profile_dir = settings.get("browser_profile_dir", "browser_profile/chrome")
        return _ok(f"Google Chrome 已找到：{chrome_exe}；项目专用用户目录：{profile_dir}", {"path": str(chrome_exe)})
    return _warn("未找到 Google Chrome，请确认已安装或在 browser.managed.executable_path 指定路径")


def _check_cdp(root: Path, config: dict[str, Any]) -> dict[str, Any]:
    settings = get_browser_settings(config)
    host = settings.get("remote_debugging_host", "127.0.0.1")
    port = settings.get("remote_debugging_port", 9222)
    endpoint = settings["cdp_endpoint"]
    auto_start = settings.get("auto_start_debug_chrome", True)

    result = ensure_chrome_debug_ready(root, config, host=host, port=port, auto_start=auto_start)
    if result["ready"]:
        if result.get("port_already_open"):
            return _ok(f"Chrome 调试端口已就绪：{endpoint}（已存在 Chrome 实例）")
        return _ok(f"Chrome 调试端口已就绪：{endpoint}（已启动项目专用 Chrome）", result)
    message = f"Chrome 调试端口未就绪：{endpoint}"
    if result.get("error"):
        message += f"。{result['error']}"
    return _warn(message)


def _runtime_excel_path(root: Path, config: dict[str, Any], project: dict[str, Any]) -> Path | None:
    if config.get("excel_path"):
        return _resolve(root, config.get("excel_path"))
    return get_project_excel_path(project, root)


def _check_sheet(excel_path: Path | None, sheet_name: str, label: str) -> dict[str, Any]:
    if not excel_path or not excel_path.exists():
        return _warn(f"无法检查 {label}，因为目标 Excel 不存在")
    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, read_only=True, data_only=False)
        if sheet_name in wb.sheetnames:
            return _ok(f"{label} 存在：{sheet_name}")
        return _warn(f"{label} 不存在：{sheet_name}；当前 sheet：{', '.join(wb.sheetnames)}")
    except Exception as exc:
        return _warn(f"检查 {label} 失败：{exc}")


class _SilentLogger:
    def info(self, *args, **kwargs):
        return None

    def warning(self, *args, **kwargs):
        return None


def _check_hourly_structure(root: Path, config: dict[str, Any], excel_path: Path | None) -> dict[str, Any]:
    if not excel_path or not excel_path.exists():
        return _warn("无法检查小时报结构，因为目标 Excel 不存在")
    inspect_config = dict(config)
    inspect_config["excel_path"] = str(excel_path)
    try:
        report = inspect_excel_structure(inspect_config, root, logger=_SilentLogger())
    except Exception as exc:
        return _warn(f"小时报结构检查失败：{exc}")
    if report.get("errors"):
        return _warn("小时报结构缺少必要字段：" + "；".join(report["errors"]), report)
    return _ok("小时报结构识别通过", report)


def run_doctor(root: str | Path, config: dict[str, Any]) -> dict[str, Any]:
    root_path = Path(root)
    checks: dict[str, dict[str, Any]] = {}
    project: dict[str, Any] = {}
    app_config: dict[str, Any] = {}

    checks["python"] = _check_python()
    checks["chrome"] = _check_chrome(config)
    checks["chrome_debug_port"] = _check_cdp(root_path, config)

    try:
        app_config = load_app_config(root_path)
        checks["app_config"] = _ok("app_config.json 存在并可读取", app_config)
    except Exception as exc:
        checks["app_config"] = _warn(f"app_config 检查失败：{exc}")

    try:
        project = get_current_project(root_path)
        errors = validate_project_config(project)
        if errors:
            checks["project_config"] = _warn("当前项目配置不完整", errors)
        else:
            checks["project_config"] = _ok(f"当前项目配置完整：{project.get('project_id')} - {project.get('project_name')}")
        checks["baidu_sources"] = _check_baidu_sources(project)
    except Exception as exc:
        checks["project_config"] = _warn(f"项目配置读取失败：{exc}")

    excel_engine = get_excel_engine(project, config)
    checks["requirements"] = _check_requirements(root_path, excel_engine=excel_engine)
    checks["excel_engine"] = _check_excel_engine(excel_engine)
    if excel_engine == "openpyxl":
        checks["openpyxl"] = _check_openpyxl_installed()
    elif excel_engine == "excel_com":
        checks["excel_app"] = _check_excel_available()

    pid = str(config.get("project_id") or project.get("project_id", ""))
    excel_path = _runtime_excel_path(root_path, config, project)
    if excel_path and excel_path.exists():
        checks["target_excel"] = _ok(
            f"已找到：{excel_path.name}",
            {
                "project_id": pid,
                "excel_path": str(excel_path),
                "hourly_sheet": str(config.get("sheet_name") or get_project_sheet_name(project, "hourly") if project else "时段数据"),
                "daily_sheet": str(config.get("daily_sheet_name") or get_project_sheet_name(project, "daily") if project else "百度"),
            },
        )
    else:
        message, detail = _build_missing_excel_message(excel_path, pid)
        checks["target_excel"] = _warn(message, detail)

    if excel_engine == "openpyxl" and checks.get("target_excel", {}).get("passed"):
        save_test = test_openpyxl_save_copy(excel_path, root_path) if excel_path else {"passed": False, "message": "目标 Excel 未配置"}
        checks["openpyxl_save_test"] = _ok(save_test["message"], save_test) if save_test.get("passed") else _warn(save_test["message"], save_test)

    daily_sheet_name = str(config.get("daily_sheet_name") or get_project_sheet_name(project, "daily") if project else "百度")
    hourly_sheet_name = str(config.get("sheet_name") or get_project_sheet_name(project, "hourly") if project else "时段数据")
    checks["daily_sheet"] = _check_sheet(excel_path, daily_sheet_name, "日报 sheet")
    checks["hourly_sheet"] = _check_sheet(excel_path, hourly_sheet_name, "小时报 sheet")
    checks["hourly_structure"] = _check_hourly_structure(root_path, dict(config, sheet_name=hourly_sheet_name), excel_path)

    export_dir = _resolve(root_path, config.get("kst", {}).get("export_dir") or project.get("kst", {}).get("export_dir") if project else config.get("kst", {}).get("export_dir", "kst_exports"))
    if export_dir and export_dir.exists():
        checks["kst_export_dir"] = _ok(f"已找到：{export_dir}")
    else:
        hint = f"请修改 configs/projects/{pid}.json 的 kst.export_dir" if pid else "请在项目配置中设置 kst.export_dir"
        checks["kst_export_dir"] = _warn(hint)

    secrets_file = _resolve(root_path, app_config.get("secrets_file")) if app_config else root_path / "secrets" / "secrets.json"
    checks["secrets_json"] = _check_secret_profile(secrets_file, project)

    latest = find_latest_kst_export(root_path, {"kst": {"export_dir": str(export_dir)}}) if export_dir else None
    if latest:
        checks["latest_kst_export"] = _ok(f"已找到：{latest.name}")
    else:
        checks["latest_kst_export"] = _warn("目录中未找到 Excel/CSV 导出表，请把商务通导出表放到该目录")

    passed_count = sum(1 for item in checks.values() if item.get("passed"))
    report = {
        "mode": "doctor",
        "project_id": config.get("project_id") or project.get("project_id") if project else config.get("project_id"),
        "project_name": config.get("project_name") or project.get("project_name") if project else config.get("project_name"),
        "checks": checks,
        "summary": {
            "total": len(checks),
            "passed": passed_count,
            "failed": len(checks) - passed_count,
            "all_passed": passed_count == len(checks),
        },
    }
    out = root_path / "reports" / "doctor_report.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def print_doctor_report(report: dict[str, Any]) -> None:
    from modules.console_ui import print_check_table

    checks: list[dict[str, Any]] = []
    for key, item in report.get("checks", {}).items():
        name = _DOCTOR_CHECK_LABELS.get(key, key)
        status = "pass" if item.get("passed") else "warn" if item.get("level") == "warning" else "fail"
        checks.append({"name": name, "status": status, "message": item.get("message", "")})

    title = f"文件合格校验  -  {report.get('project_name') or report.get('project_id') or '未知'}"
    print_check_table(title, checks)


_DOCTOR_CHECK_LABELS = {
    "python": "Python 版本",
    "chrome": "Google Chrome",
    "chrome_debug_port": "Chrome 调试端口",
    "app_config": "应用配置",
    "project_config": "项目配置",
    "baidu_sources": "百度来源配置",
    "requirements": "依赖包",
    "excel_engine": "Excel 写入引擎",
    "openpyxl": "openpyxl 安装",
    "excel_app": "Microsoft Excel",
    "target_excel": "目标 Excel",
    "openpyxl_save_test": "openpyxl 保存测试",
    "daily_sheet": "日报 sheet",
    "hourly_sheet": "小时报 sheet",
    "hourly_structure": "小时报结构",
    "kst_export_dir": "商务通导出目录",
    "secrets_json": "百度凭据文件",
    "latest_kst_export": "最新商务通导出文件",
}
